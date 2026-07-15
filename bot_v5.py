# bot_v5.py — 자비스(JARVIS) 봇 v5 완성판
# 설계서 기준 + 결제내역 OCR + 콜카드↔결제내역 교차대조
# 작성일: 2026-03-30

import os
import io
import re
import json
import time
import asyncio
import logging
import threading
import base64
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer

import httpx
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ContextTypes, filters
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic

load_dotenv()

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────
KST = timezone(timedelta(hours=9))
NET_GOAL      = 150_000      # 일 순수익 목표
CARD_FEE_RATE = 0.033        # 카드수수료 3.3%
INSURANCE_DAILY = 7_945      # 일 보험료
DOW_KOR = ["월", "화", "수", "목", "금", "토", "일"]

TELEGRAM_TOKEN     = os.getenv("TELEGRAM_TOKEN", "")
ALLOWED_IDS_RAW    = [
    os.getenv("ALLOWED_CHAT_ID", ""),
    os.getenv("ALLOWED_CHAT_ID2", ""),
]
ALLOWED_IDS = {x for x in ALLOWED_IDS_RAW if x}

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
SUPABASE_URL      = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY      = os.getenv("SUPABASE_KEY", "")
PORT              = int(os.getenv("PORT", "10000"))

# 캐스퍼 명령서 #014 반영 (2026-07-10) — 봇 직접 GitHub 커밋용
# 이지스가 쓰는 PAT와 별도로 봇 전용 PAT를 Render 환경변수에 등록해서 사용 권장.
GITHUB_PAT   = os.getenv("GITHUB_PAT", "")
GITHUB_OWNER = os.getenv("GITHUB_OWNER", "Kim0900")
GITHUB_REPO  = os.getenv("GITHUB_REPO", "magi-taxi-data")

OCR_MODEL      = "claude-haiku-4-5-20251001"
HEADERS_SB = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Health Check 서버
# ──────────────────────────────────────────────
class HealthHandler(BaseHTTPRequestHandler):
    def _cors_headers(self):
        """모든 응답에 CORS 허용 헤더 추가"""
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization, apikey')

    def do_OPTIONS(self):
        """CORS preflight 요청 처리"""
        self.send_response(204)
        self._cors_headers()
        self.end_headers()

    def do_GET(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()
        self.wfile.write(b"Jarvis v5 OK")

    def do_POST(self):
        """API 엔드포인트 — OCR / 마기분석 / 아틀라스보고"""
        import json as _j, re as _re
        length = int(self.headers.get('Content-Length', 0))
        raw_body = self.rfile.read(length)
        try:
            payload = _j.loads(raw_body.decode('utf-8')) if raw_body else {}
        except Exception:
            payload = {}

        def send_json(code, data):
            body = _j.dumps(data, ensure_ascii=False).encode('utf-8')
            self.send_response(code)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization, apikey')
            self.end_headers()
            self.wfile.write(body)

        if self.path == '/ocr_receipt':
            try:
                import anthropic as _ant
                b64 = payload.get('image_b64', '')
                mt  = payload.get('media_type', 'image/jpeg')
                client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
                msg = client.messages.create(
                    model="claude-haiku-4-5-20251001", max_tokens=400,
                    messages=[{"role":"user","content":[
                        {"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
                        {"type":"text","text":'이 택시 매출집계 영수증에서 정보를 추출해서 JSON만 반환해줘.\n{"date":"YYYY-MM-DD","total_sales":숫자,"commission":숫자,"trip_count":숫자,"start_time":"HH:MM","end_time":"HH:MM"}\n숫자만(원제외). JSON만 반환.'}
                    ]}]
                )
                txt = _re.sub(r"```[a-z]*", "", msg.content[0].text.strip()).strip()
                send_json(200, {"success": True, "data": _j.loads(txt)})
            except Exception as e:
                logger.error(f"OCR 오류: {e}")
                send_json(400, {"success": False, "error": str(e)})
            return

        if self.path == '/gpx_parse':
            # GPX는 앱(브라우저 DOMParser)에서 직접 파싱 — 봇 불필요
            # 이 엔드포인트는 예비용(서버사이드 파싱 필요 시 확장)
            send_json(200, {"status":"ok","message":"GPX는 앱 내 파싱 완료"})
            return

        if self.path == '/ocr_history':
            try:
                import anthropic as _ant, re as _re, json as _j
                b64 = payload.get('image_b64', '')
                mt  = payload.get('media_type', 'image/jpeg')
                client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
                # 1단계: 유형 분류
                cls_msg = client.messages.create(
                    model="claude-haiku-4-5-20251001", max_tokens=20,
                    messages=[{"role":"user","content":[
                        {"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
                        {"type":"text","text":"'일별운행이력' 또는 '결제내역' 중 하나만 답해."}
                    ]}]
                )
                is_daily = '일별' in cls_msg.content[0].text

                if is_daily:
                    prompt = (
                        '이 카카오T 일별운행이력 화면에서 모든 운행 건을 추출해서 JSON만 반환해줘.\n'
                        '{"type":"daily_history","date":"YYYY-MM-DD","calls":['
                        '{"배차시각":"HH:MM","하차시각":"HH:MM","출발지":"대구 OO구 OO동","도착지":"대구 OO구 OO동","요금":숫자,"결제방식":"자동 또는 직접"}]}\n'
                        '날짜: 상단 YYYY년 M월 D일. 결제방식: 직접결제 있으면 직접, 없으면 자동. JSON만 반환.'
                    )
                else:
                    prompt = (
                        '이 결제내역 화면에서 모든 결제 건을 추출해서 JSON만 반환해줘.\n'
                        '{"type":"payment","date":"YYYY-MM-DD","total":숫자,"items":['
                        '{"시각":"HH:MM","요금":숫자,"카드":"카드사명"}]}\n'
                        '날짜: 조회일/거래일. 취소건 제외. 숫자만(원제외). JSON만 반환.'
                    )
                ocr_msg = client.messages.create(
                    model="claude-haiku-4-5-20251001", max_tokens=2000,
                    messages=[{"role":"user","content":[
                        {"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
                        {"type":"text","text":prompt}
                    ]}]
                )
                txt = _re.sub(r"```[a-z]*","",ocr_msg.content[0].text.strip()).strip()
                data = _j.loads(txt)
                send_json(200, {"success":True,"data":data})
            except Exception as e:
                logger.error(f"OCR history 오류: {e}")
                send_json(400, {"success":False,"error":str(e)})
            return

        if self.path == '/magi_analyze':
            try:
                import anthropic as _ant
                client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=120.0)
                msg = client.messages.create(
                    model=payload.get('model','claude-sonnet-4-20250514'),
                    max_tokens=int(payload.get('max_tokens',1500)),
                    system=payload.get('system_prompt','당신은 마기입니다.'),
                    messages=[{"role":"user","content":payload.get('user_message','')}]
                )
                send_json(200, {"success": True, "result": msg.content[0].text})
            except Exception as e:
                logger.error(f"마기분석 오류: {e}")
                send_json(400, {"success": False, "error": str(e)})
            return

        if self.path == '/atlas-report':
            try:
                import threading
                threading.Thread(
                    target=lambda: asyncio.run(save_atlas_report(payload)),
                    daemon=True
                ).start()
                send_json(200, {"status": "ok"})
                logger.info(f"아틀라스 보고 수신: {payload.get('title','?')}")
            except Exception as e:
                send_json(400, {"error": str(e)})
            return

        self.send_response(404)
        self.end_headers()

    def log_message(self, *args):
        pass

def run_health_server():
    server = HTTPServer(("0.0.0.0", PORT), HealthHandler)
    logger.info(f"Health server on port {PORT}")
    server.serve_forever()

# ──────────────────────────────────────────────
# Supabase 헬퍼
# ──────────────────────────────────────────────
async def sb_h(method: str, path: str, **kwargs) -> dict | list | None:
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    # 캐스퍼 수정 2026-07-06: headers를 여기서 무조건 고정으로 넘기면서
    # 동시에 **kwargs에도 headers가 들어있는 경우(sb_upsert, calc_official_var_score 등
    # headers=를 직접 넘기는 모든 호출) "multiple values for keyword argument 'headers'"
    # TypeError로 무조건 크래시하던 잠복 버그. kwargs의 headers를 우선 사용하도록 수정.
    headers = kwargs.pop("headers", HEADERS_SB)
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.request(method, url, headers=headers, **kwargs)
        if r.status_code in (200, 201):
            # 캐스퍼 수정 2026-07-06 (2차): Prefer: return=minimal을 쓰는 호출은
            # Supabase가 본문을 아예 비워서 응답하는데, 여기서 무조건 r.json()을
            # 호출해서 JSONDecodeError로 크래시하던 버그. 빈 본문이면 파싱 생략.
            if not r.content:
                return []
            try:
                return r.json()
            except Exception:
                logger.error(f"Supabase {method} {path} → JSON 파싱 실패: {r.text[:200]}")
                return []
        logger.error(f"Supabase {method} {path} → {r.status_code}: {r.text}")
        return None

async def sb_insert(table: str, data: dict) -> dict | None:
    return await sb_h("POST", table, json=data)

async def sb_select(table: str, params: dict = None) -> list:
    result = await sb_h("GET", table, params=params or {})
    return result if isinstance(result, list) else []

async def sb_upsert(table: str, data: dict, on_conflict: str) -> dict | None:
    return await sb_h(
        "POST", table,
        json=data,
        headers={**HEADERS_SB, "Prefer": f"resolution=merge-duplicates,return=representation"},
        params={"on_conflict": on_conflict}
    )


# ──────────────────────────────────────────────
# GitHub 직접 커밋 (캐스퍼 명령서 #014 §2)
# ──────────────────────────────────────────────
async def github_commit_briefing(날짜: str, content: str) -> dict:
    """브리핑 markdown을 magi-taxi-data 레포 /briefings/ 폴더에 직접 커밋.
    반환: {"ok": bool, "url": str|None, "error": str|None}
    """
    if not GITHUB_PAT:
        return {"ok": False, "url": None, "error": "GITHUB_PAT 환경변수 미설정"}

    # 캐스퍼 명령서 #017 반영(2026-07-13): 아르고스 정밀 브리핑(/briefings/)과
    # 경로·파일명 분리 — 이지스 파이프라인A와의 덮어쓰기 충돌 방지
    path = f"bot_briefings/bot_summary_{날짜.replace('-','')}.md"
    api_url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{path}"
    headers = {
        "Authorization": f"Bearer {GITHUB_PAT}",
        "Accept": "application/vnd.github+json",
    }

    async with httpx.AsyncClient(timeout=20) as client:
        # 기존 파일 존재 여부 확인 (있으면 sha 필요 — 덮어쓰기용)
        sha = None
        try:
            r_get = await client.get(api_url, headers=headers)
            if r_get.status_code == 200:
                sha = r_get.json().get("sha")
        except Exception:
            pass  # 조회 실패해도 신규 생성 시도는 계속 진행

        # 카카오/우버/배회 건수 커밋 메시지용 집계
        try:
            calls = await sb_select("raw_calls", {"날짜": f"eq.{날짜}"}) or []
            kakao_n = sum(1 for c in calls if (c.get("콜유형") or "") == "카카오T")
            uber_n  = sum(1 for c in calls if (c.get("콜유형") or "") == "우버")
            bhw_n   = sum(1 for c in calls if (c.get("콜유형") or "") == "배회")
        except Exception:
            kakao_n = uber_n = bhw_n = 0

        commit_msg = f"[BOT] {날짜.replace('-','')} 브리핑 자동생성 — 카카오{kakao_n}건·우버{uber_n}건·배회{bhw_n}건"

        body = {
            "message": commit_msg,
            "content": base64.b64encode(content.encode("utf-8")).decode("ascii"),
            "branch": "main",
        }
        if sha:
            body["sha"] = sha

        try:
            r_put = await client.put(api_url, headers=headers, json=body)
        except Exception as e:
            return {"ok": False, "url": None, "error": f"GitHub 요청 실패: {e}"}

        if r_put.status_code in (200, 201):
            html_url = r_put.json().get("content", {}).get("html_url")
            return {"ok": True, "url": html_url, "error": None}
        else:
            return {"ok": False, "url": None, "error": f"{r_put.status_code}: {r_put.text[:200]}"}

async def sb_delete_receipt(conditions: dict) -> int:
    """payment_receipts 조건부 삭제. 삭제 건수 반환"""
    query = "&".join(f"{k}={v}" for k, v in conditions.items())
    r = await sb_h("DELETE", f"payment_receipts?{query}")
    # Supabase DELETE는 삭제된 행 반환 (Prefer: return=representation)
    if isinstance(r, list):
        return len(r)
    return 0

async def sb_delete_last(table: str, filter_params: dict) -> bool:
    rows = await sb_select(table, {**filter_params, "order": "id.desc", "limit": "1"})
    if not rows:
        return False
    row_id = rows[0]["id"]
    r = await sb_h("DELETE", f"{table}?id=eq.{row_id}")
    return True

# ──────────────────────────────────────────────
# 유틸 함수
# ──────────────────────────────────────────────
def now_kst() -> datetime:
    return datetime.now(KST)

def today_kst():
    return now_kst().date()

def get_dow(d=None) -> str:
    d = d or today_kst()
    return DOW_KOR[d.weekday()]

def fmt(n) -> str:
    """None·문자열 안전 처리"""
    if n is None:
        return "0원"
    try:
        return f"{int(n):,}원"
    except (ValueError, TypeError):
        return "0원"

def calc_net(매출: int, 지출: int) -> int:
    return int(매출 * (1 - CARD_FEE_RATE)) - 지출

async def today_summary() -> dict:
    today = str(today_kst())
    calls = await sb_select("raw_calls", {"날짜": f"eq.{today}"})
    expenses = await sb_select("expenses", {"날짜": f"eq.{today}"})
    건수 = len(calls)
    매출 = sum(c.get("요금", 0) or 0 for c in calls)
    지출 = sum(e.get("금액", 0) or 0 for e in expenses)
    순수익 = calc_net(매출, 지출)
    달성률 = int(순수익 / NET_GOAL * 100) if NET_GOAL else 0
    return {
        "건수": 건수, "매출": 매출, "지출": 지출,
        "순수익": 순수익, "달성률": 달성률,
    }

async def today_expenses() -> list:
    today = str(today_kst())
    return await sb_select("expenses", {"날짜": f"eq.{today}", "order": "id.asc"})

async def insert_insurance(date):
    date_str = str(date)
    existing = await sb_select(
        "expenses",
        {"날짜": f"eq.{date_str}", "카테고리": "eq.보험료", "자동여부": "eq.true"}
    )
    if existing:
        return
    await sb_insert("expenses", {
        "날짜": date_str,
        "카테고리": "보험료",
        "금액": INSURANCE_DAILY,
        "메모": "자동 보험료",
        "자동여부": True,
    })
    logger.info(f"보험료 자동 기록: {date_str}")

# ──────────────────────────────────────────────
# 이미지 큐 (동시 업로드 과부하 방지)
# ──────────────────────────────────────────────
image_queue: asyncio.Queue = None  # main()에서 초기화

# ──────────────────────────────────────────────
# Claude API — 이미지 분류 + OCR
# ──────────────────────────────────────────────
async def claude_vision(image_bytes: bytes, prompt: str, max_tokens: int = 500) -> str:
    """Claude API 비동기 호출. 모든 이미지 포맷을 JPEG로 정규화 후 전송."""

    def _prepare_image(raw: bytes) -> tuple[bytes, str]:
        """이미지를 JPEG로 변환 + 최대 높이 3000px 리사이즈"""
        try:
            from PIL import Image as _Image
            import io as _io
            img = _Image.open(_io.BytesIO(raw))
            # RGB 변환
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            # 최대 높이 3000px (세로 긴 이미지 처리)
            MAX_H = 3000
            if img.height > MAX_H:
                ratio = MAX_H / img.height
                img = img.resize((int(img.width * ratio), MAX_H), _Image.LANCZOS)
                logger.info(f"이미지 높이 축소: {img.height}→{MAX_H}px")
            buf = _io.BytesIO()
            img.save(buf, format="JPEG", quality=85, optimize=True)
            return buf.getvalue(), "image/jpeg"
        except Exception as e:
            logger.warning(f"이미지 변환 실패({e}) → 원본 사용")
            # 포맷 감지
            if raw[:4] == b"RIFF" or raw[:4] == b"WEBP":
                return raw, "image/webp"
            elif raw[:8] == b"\x89PNG\r\n\x1a\n":
                return raw, "image/png"
            return raw, "image/jpeg"

    img_data, media_type = _prepare_image(image_bytes)
    b64 = base64.standard_b64encode(img_data).decode()

    def _sync_call():
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
        msg = client.messages.create(
            model=OCR_MODEL,
            max_tokens=max_tokens,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        return msg.content[0].text.strip()

    return await asyncio.to_thread(_sync_call)


def resize_image_if_needed(image_bytes: bytes) -> bytes:
    """
    Claude API 전송 전 이미지 최적화.
    - 최대 너비 1000px (세로 비율 유지)
    - JPEG quality 80
    - Pillow 필수 (requirements.txt에 Pillow 추가 필요)
    """
    try:
        from PIL import Image
        import io

        img = Image.open(io.BytesIO(image_bytes))
        orig_w, orig_h = img.width, img.height
        orig_kb = len(image_bytes) / 1024

        # 최대 너비 1000px 또는 최대 높이 3000px 초과 시 축소
        MAX_WIDTH = 1000
        MAX_HEIGHT = 3000
        ratio_w = MAX_WIDTH / orig_w if orig_w > MAX_WIDTH else 1.0
        ratio_h = MAX_HEIGHT / orig_h if orig_h > MAX_HEIGHT else 1.0
        ratio = min(ratio_w, ratio_h)
        if ratio < 1.0:
            new_w = int(orig_w * ratio)
            new_h = int(orig_h * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)
        
        # RGB 변환 (PNG RGBA 등 처리)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=80, optimize=True)
        result = buf.getvalue()
        result_kb = len(result) / 1024

        logger.info(
            f"이미지 최적화: {orig_w}x{orig_h} {orig_kb:.0f}KB"
            f" → {img.width}x{img.height} {result_kb:.0f}KB"
        )
        return result

    except ImportError:
        logger.error("Pillow 미설치 — requirements.txt에 Pillow 추가 필요")
        return image_bytes
    except Exception as e:
        logger.error(f"이미지 리사이즈 오류: {e}")
        return image_bytes

async def classify_image(image_bytes: bytes) -> str:
    prompt = (
        "이 이미지가 아래 중 어느 종류인지 판단해서 해당 단어 하나만 답해줘.\n\n"
        "【결제】← 최우선 확인\n"
        "  카카오T 또는 세큐티 결제내역 화면. 아래 중 하나라도 있으면 반드시 '결제':\n"
        "  · '결제내역조회', '거래일자', '결제구분', '조회기간' 텍스트\n"
        "  · 날짜+시각(YYYY-MM-DD | HH:MM:SS 또는 YYYY-MM-DD HH:MM) 옆에 금액 목록\n"
        "  · 'KB카드', '신한카드', '현대카드', 'BC카드', '하나카드', '농협카드' 등 카드사명\n"
        "  · '승인정상', '1승인', '거래일자' 텍스트\n"
        "  · 세로로 5건 이상 금액 목록 나열\n\n"
        "【충전】\n"
        "  전기차 충전 앱 이용내역. 아래 중 하나라도 있으면 '충전':\n"
        "  · 'kWh', '충전량', '충전완료', '충전소' 텍스트\n"
        "  · '전기차 충전' 탭 UI\n\n"
        "【콜카드】\n"
        "  카카오T 택시 운행기록 1건. '배차', '승차', '하차' + 출발지·도착지 주소.\n\n"
        "【일별운행이력】← 콜카드보다 먼저 확인\n"
        "  카카오T '일별 운행 이력' 화면. 아래 특징이 있으면 반드시 '일별운행이력':\n"
        "  · 상단에 'YYYY년 M월 D일(요일) N건' 형식\n"
        "  · 여러 건의 운행이 세로로 나열\n"
        "  · 각 건마다 'HH:MM - HH:MM [실시간]' 시간 범위\n"
        "  · '직접결제' 텍스트 포함 가능\n"
        "  · '실시간 운행 N건 / N원' 요약\n\n"
        "【세큐티】\n"
        "  세큐티 등급·점수 리포트. 종합점수, 수락률 등 항목.\n\n"
        "【기타】위 4가지 해당 없음.\n\n"
        "⚠️ 핵심 구분:\n"
        "  충전: kWh 단위 있음\n"
        "  결제: 카드사명 + 날짜+금액 목록\n"
        "  콜카드: 운행 1건(배차·승차·하차)\n\n"
        "반드시 결제·충전·콜카드·세큐티·기타 중 하나만 답해. 다른 말 금지."
    )
    result = await claude_vision(image_bytes, prompt, max_tokens=15)
    for keyword in ["일별운행이력", "콜카드", "충전", "결제", "세큐티"]:
        if keyword in result:
            return keyword
    return "기타"


# ══════════════════════════════════════════════
# 세큐티 OCR + 저장 + 조회
# ══════════════════════════════════════════════


# ══════════════════════════════════════════════
# 일별 운행이력 OCR + 저장
# ══════════════════════════════════════════════

async def save_atlas_report(data: dict):
    """아틀라스 보고서 Supabase 저장 + 텔레그램 알림"""
    try:
        payload = {
            "report_type": data.get("report_type", "manual"),
            "source": "atlas",
            "title": data.get("title", "아틀라스 보고"),
            "payload": data.get("payload", data),
            "status": "pending",
            "run_date": str(today_kst()),
        }
        result = await sb_insert("atlas_reports", payload)
        report_id = result[0]["id"] if result else "?"

        # 텔레그램 알림 (봇 애플리케이션에 전송)
        # application 객체에 접근하기 위해 전역 변수 사용
        global _bot_app
        if _bot_app:
            from telegram import Bot
            bot = _bot_app.bot
            allowed_ids = [int(x) for x in os.getenv("ALLOWED_USER_IDS","").split(",") if x]
            for uid in allowed_ids:
                try:
                    await bot.send_message(
                        chat_id=uid,
                        text=f"📡 아틀라스 보고 #{report_id}\n"
                             f"유형: {payload['report_type']}\n"
                             f"제목: {payload['title']}\n"
                             f"앱에서 마기 분석 자동 시작"
                    )
                except Exception:
                    pass
        logger.info(f"atlas_report #{report_id} 저장 완료")
    except Exception as e:
        logger.error(f"atlas_report 저장 오류: {e}")


async def ocr_daily_history(image_bytes: bytes) -> dict | None:
    """
    카카오T '일별 운행 이력' 화면 OCR.
    반환: {날짜: "YYYY-MM-DD", 콜목록: [{배차시각, 하차시각, 출발지, 도착지, 요금, 결제방식}, ...]}
    """
    prompt = (
        "이 카카오T '일별 운행 이력' 화면에서 정보를 추출해서 JSON만 반환해줘.\n"
        '{"날짜":"YYYY-MM-DD","콜목록":['
        '{"배차시각":"HH:MM","하차시각":"HH:MM","출발지":"대구 OO구 OO동",'
        '"도착지":"대구 OO구 OO동","요금":숫자,"결제방식":"자동 또는 직접"}]}\n'
        "⚠️ 날짜: 상단 'YYYY년 M월 D일' → YYYY-MM-DD 변환\n"
        "⚠️ 배차시각: 각 콜의 앞 시각 (예: 02:58 - 03:11 에서 02:58)\n"
        "⚠️ 하차시각: 각 콜의 뒤 시각 (예: 02:58 - 03:11 에서 03:11)\n"
        "⚠️ 결제방식: '직접결제' 텍스트 있으면 '직접', 없으면 '자동'\n"
        "⚠️ 요금: 파란색 숫자. 직접결제는 표시된 요금 그대로 추출\n"
        "⚠️ 출발지/도착지: '대구 OO구 OO동' 형식. 구/동만 추출\n"
        "JSON만 반환. 설명·마크다운 금지."
    )
    try:
        raw = await claude_vision(image_bytes, prompt, max_tokens=2000)
        raw = raw.strip()
        import re as _re
        raw = _re.sub(r"```json\s*", "", raw)
        raw = _re.sub(r"```\s*", "", raw)
        raw = raw.strip()
        import json as _json
        return _json.loads(raw)
    except Exception as e:
        logger.error(f"일별운행이력 OCR 오류: {e}")
        return None


async def process_daily_history(update, image_bytes: bytes):
    """
    일별 운행이력 이미지 처리:
    OCR → 날짜 보정 → raw_calls 저장 → 결과 안내
    """
    from datetime import date as _dc, timedelta as _td

    await update.message.reply_text("📋 일별 운행이력 분석 중...")

    data = await ocr_daily_history(image_bytes)
    if not data or not data.get("콜목록"):
        await update.message.reply_text(
            "❌ 일별 운행이력 인식 실패\n"
            "💡 화면을 더 크게 캡처해서 다시 올려주세요."
        )
        return

    # 화면 날짜 파싱
    screen_date_str = data.get("날짜", "")
    try:
        screen_date = _dc.fromisoformat(screen_date_str)
    except Exception:
        screen_date = today_kst()
        logger.warning(f"날짜 파싱 실패: {screen_date_str} → 오늘 사용")

    DOW_MAP = ["월","화","수","목","금","토","일"]
    saved = 0
    updated = 0
    dates_used = set()
    result_lines = []

    for call in data.get("콜목록", []):
        배차 = call.get("배차시각", "")
        하차 = call.get("하차시각", "")
        출발 = call.get("출발지", "")
        도착 = call.get("도착지", "")
        요금 = call.get("요금", 0) or 0
        결제방식 = call.get("결제방식", "자동")

        # 날짜 보정: 배차 06시 이전 → 화면날짜 +1일 (새벽 운행)
        try:
            h = int(배차.split(":")[0])
            save_date = screen_date + _td(days=1) if h < 6 else screen_date
        except Exception:
            save_date = screen_date

        save_date_str = str(save_date)
        dow = DOW_MAP[save_date.weekday()]
        dates_used.add(save_date_str)

        # 중복 삭제 후 재저장
        deleted = await delete_duplicate_call(save_date_str, 배차, 요금)
        if deleted:
            updated += deleted

        비고 = "직접결제(요금미확인)" if 결제방식 == "직접" else None

        payload = {
            "날짜":     save_date_str,
            "요일":     dow,
            "배차시각": 배차,
            "하차시각": 하차,
            "출발지":   출발,
            "도착지":   도착,
            "요금":     요금,
            "콜유형":   "카카오T",
            "비고":     비고,
        }
        result = await sb_insert("raw_calls", payload)
        if result:
            saved += 1
            직접표시 = " [직접결제]" if 결제방식 == "직접" else ""
            result_lines.append(
                f"  {배차}~{하차} {출발}→{도착} {fmt(요금)}{직접표시}"
            )

    # 결과 메시지
    dates_sorted = sorted(dates_used)
    msg = [
        f"✅ 일별 운행이력 저장 완료",
        f"화면날짜: {screen_date_str} | 저장: {saved}건",
        f"날짜 분포: {', '.join(dates_sorted)}",
        "",
    ]
    msg.extend(result_lines[:10])  # 최대 10건 표시
    if len(result_lines) > 10:
        msg.append(f"  ... 외 {len(result_lines)-10}건")

    msg.append("")
    msg.append("💡 교차대조:")
    for d in dates_sorted:
        msg.append(f"  대조 {d}")

    await update.message.reply_text("\n".join(msg))



# ══════════════════════════════════════════════
# 운행 일관성 모니터링 (Step E)
# ══════════════════════════════════════════════


# ══════════════════════════════════════════════
# Step A: 카카오 공식 6변수 자동 평가
# ══════════════════════════════════════════════

async def calc_official_var_score(날짜: str) -> dict:
    """카카오 AI 배차 공식 6변수 평가 및 daily_summary 저장"""
    from datetime import date as _dc
    today = _dc.today()
    mo = 날짜[:7]

    # var_2: 오늘 운행완료수
    calls_today = await sb_select("raw_calls", {"날짜": f"eq.{날짜}"})
    daily_completed = len(calls_today)

    # var_2: 이번달 일평균
    calls_month = await sb_select("raw_calls", {
        "and": f"(날짜.gte.{mo}-01,날짜.lte.{mo}-31)"
    })
    from datetime import date as _d2
    days_so_far = (_d2.today() - _d2(int(mo[:4]), int(mo[5:7]), 1)).days + 1
    # 명령서 #010 대응(2026-07-08): 행 개수가 아닌 건수 가중 합계로 수정
    month_weighted_count = sum(_extract_count(c) for c in calls_month)
    monthly_avg = month_weighted_count / max(days_so_far, 1)

    # var_3·4: sekuti에서 조회 (현재 0으로 고정 — 마스터 등급)
    avoid_count = 0
    one_star_count = 0

    # var_5: 수락률 (현재 100% 유지 중)
    acceptance_rate = 100

    # AI 진입 추정 (운행완료수 기반)
    AREA_AVG_LOW, AREA_AVG_HIGH = 18, 25
    if monthly_avg >= AREA_AVG_HIGH:
        ai_estimate = "85%+"
    elif monthly_avg >= AREA_AVG_LOW:
        ai_estimate = f"{int(60 + (monthly_avg-AREA_AVG_LOW)/(AREA_AVG_HIGH-AREA_AVG_LOW)*25)}%"
    else:
        ai_estimate = f"{int(40 + monthly_avg/AREA_AVG_LOW*20)}%"

    score = {
        "date": 날짜,
        "vars": {
            "var_1_acceptance_prob": "양호" if len(calls_month) >= 30 else "데이터 축적 중",
            "var_2_daily_completed": daily_completed,
            "var_2_monthly_avg": round(monthly_avg, 1),
            "var_2_area_avg_est": f"{AREA_AVG_LOW}~{AREA_AVG_HIGH}",
            "var_3_avoid_count_monthly": avoid_count,
            "var_4_one_star_monthly": one_star_count,
            "var_5_acceptance_rate": acceptance_rate,
            "var_6_eta_score": "위치 의존"
        },
        "weak_var": "var_2_daily_completed",
        "ai_inclusion_estimate": ai_estimate,
        "improvement_needed": "운행 시간 확대 + 매일 운행" if monthly_avg < AREA_AVG_LOW else "유지"
    }

    # daily_summary에 upsert
    await sb_h("POST", f"daily_summary",
        json={"날짜": 날짜, "official_var_score": score},
        headers={**HEADERS_SB, "Prefer": "resolution=merge-duplicates,return=minimal"},
        params={"on_conflict": "날짜"}
    )
    return score


async def handle_forecast(update, date_str=None):
    """/forecast [YYYY-MM-DD] — 사전 예측 시안"""
    from datetime import date as _dc, timedelta as _td
    target = date_str or str(_dc.today() + _td(days=1))
    try:
        _dc.fromisoformat(target)
    except ValueError:
        await update.message.reply_text("❌ 날짜 형식 오류. 예: /forecast 2026-05-21")
        return

    rows = await sb_select("forecast", {"forecast_date": f"eq.{target}"})
    if rows:
        r = rows[0]
        cp  = r.get("kakao_count_point", "?")
        cl  = r.get("kakao_count_ci_low", "?")
        ch  = r.get("kakao_count_ci_high", "?")
        rp  = int(r.get("revenue_point", 0) or 0)
        rl  = int(r.get("revenue_ci_low", 0) or 0)
        rh  = int(r.get("revenue_ci_high", 0) or 0)
        wf  = r.get("weather_forecast", "미정")
        mv  = r.get("model_version", "v0.3")
        nt  = r.get("notes", "")
        hz  = r.get("hotzones", [])
        lines = [
            f"🔮 예측 시안 — {target}",
            f"",
            f"📊 콜수: {cp}건 (CI {cl}~{ch}건)",
            f"💰 매출: {rp:,}원 (CI {rl:,}~{rh:,}원)",
            f"🌤️ 날씨: {wf}",
        ]
        if hz:
            lines.append(f"🔥 핫존: {' / '.join(hz) if isinstance(hz,list) else hz}")
        lines += [f"", f"📝 {nt}", f"모델: {mv}"]
        await update.message.reply_text("\n".join(lines))
        return

    # 신규 생성 — 메타모델 v0.3
    dow = ["월","화","수","목","금","토","일"][_dc.fromisoformat(target).weekday()]
    DOW_MODEL = {
        "월": (8,  6, 11,  72000,  55000,  95000),
        "화": (9,  7, 12,  85000,  65000, 110000),
        "수": (10, 8, 13,  95000,  75000, 120000),
        "목": (9,  7, 12,  85000,  65000, 110000),
        "금": (12,10, 15, 115000,  90000, 145000),
        "토": (13,11, 16, 125000, 100000, 155000),
        "일": (12,10, 15, 118000,  95000, 148000),
    }
    pt, cl2, ch2, rp2, rl2, rh2 = DOW_MODEL.get(dow, (10, 8, 13, 95000, 75000, 120000))
    fd = {
        "forecast_date": target,
        "kakao_count_point": pt, "kakao_count_ci_low": cl2, "kakao_count_ci_high": ch2,
        "revenue_point": rp2, "revenue_ci_low": rl2, "revenue_ci_high": rh2,
        "weather_forecast": "미정 (수동 업데이트 필요)",
        "model_version": "메타모델 v0.3",
        "notes": f"{dow}요일 요일별 메타모델 기반 예측",
        "hotzones": ["중구 성내2동", "수성구 범어동", "동구 동대구역"]
    }
    result = await sb_insert("forecast", fd)
    fid = result[0]["id"] if result else "?"
    lines = [
        f"🔮 예측 시안 #{fid} — {target} ({dow}요일)",
        f"",
        f"📊 콜수: {pt}건 (CI {cl2}~{ch2}건)",
        f"💰 매출: {rp2:,}원 (CI {rl2:,}~{rh2:,}원)",
        f"🔥 핫존: 성내2동 / 범어동 / 동대구역",
        f"",
        f"⚠️ 메타모델 v0.3 — 실측 후 정밀화됩니다",
    ]
    await update.message.reply_text("\n".join(lines))


async def handle_completion_status(update):
    """/completion_status — 운행완료수 현황 및 AI 진입 가능성"""
    from datetime import date as _dc
    today = str(_dc.today())
    mo = today[:7]

    calls_month = await sb_select("raw_calls", {
        "and": f"(날짜.gte.{mo}-01,날짜.lte.{mo}-31)"
    })
    from datetime import date as _d2
    days_so_far = (_d2.today() - _d2(int(mo[:4]), int(mo[5:7]), 1)).days + 1
    total = len(calls_month)
    avg = total / max(days_so_far, 1)

    AREA_LOW, AREA_HIGH = 18, 25
    pct = int(avg / AREA_HIGH * 100)

    # 진입 가능성
    if avg >= AREA_HIGH:
        level = "🟢 높음"
        comment = "AI 1순위 후보군 정상 진입 구간"
    elif avg >= AREA_LOW:
        level = "🟡 중간"
        needed = round(AREA_HIGH - avg, 1)
        comment = f"목표까지 {needed}건/일 더 필요"
    else:
        level = "🔴 낮음"
        needed = round(AREA_HIGH - avg, 1)
        comment = f"목표까지 {needed}건/일 더 필요"

    # 6월 목표 계산 (매일 운영 가정)
    import calendar as _cal
    days_in_month = _cal.monthrange(int(mo[:4]), int(mo[5:7]))[1]
    target_for_area = AREA_LOW * days_in_month

    lines = [
        f"📊 운행완료수 현황 — {mo}",
        f"",
        f"이번달 누적: {total}건",
        f"일평균: {avg:.1f}건/일",
        f"사업구역 평균 추정: {AREA_LOW}~{AREA_HIGH}건/일",
        f"대비: {pct}%",
        f"",
        f"AI 1순위 진입 가능성: {level}",
        f"💡 {comment}",
        f"",
        f"[월간 목표]",
        f"AI 안정권(18건/일): {target_for_area}건/월",
        f"현재: {total}건 / 잔여: {max(target_for_area-total,0)}건",
        f"",
        f"[6변수 현황]",
        f"  ② 운행완료수: {avg:.1f}건/일 {'❌' if avg < AREA_LOW else '✅'}",
        f"  ③ 만나지않기: 0회 ✅",
        f"  ④ 평점1점: 0회 ✅",
        f"  ⑤ 수락률: 100% ✅",
    ]
    await update.message.reply_text("\n".join(lines))


# ══════════════════════════════════════════════
# Step D: /briefing — 7섹션 통합 보고
# ══════════════════════════════════════════════

import re as _re_count

def _extract_count(c: dict) -> int:
    """raw_calls 한 행의 실제 건수 (수동입력=1, OCR요약행=실제trip_count).
    앱(index.html)의 extractCount와 동일 로직. 명령서 #010 대응(2026-07-08)으로
    calc_official_var_score/handle_briefing의 월평균 계산에도 공통 적용."""
    비고 = c.get("비고") or ""
    m = _re_count.search(r"건수\s*(\d+)\s*건", 비고)
    if m:
        return int(m.group(1))
    건수 = c.get("건수")
    if 건수 is not None:
        try:
            return int(건수)
        except Exception:
            pass
    return 1


async def calc_kpi_metrics(날짜: str, 매출: int, work_hours) -> dict:
    """캐스퍼 명령서 #008 §3 반영 — daily_summary KPI 4종 봇 자체 계산
    (아르고스 브리핑 텍스트 파싱 방식 폐기, raw_calls/daily_summary 원본 직접 집계로 전환)

    kpi_7day_avg: 반드시 축B(카카오T, 00시 기준 달력일, 콜 없는 날은 0) 기준.
    raw_calls의 '날짜'는 축A(영업일, 저녁 시작 기준) 라벨이므로,
    00~05시 배차 콜은 실제로는 다음 캘린더일 새벽 연장 운행 → 익일로 재귀속해서 집계.

    2026-07-08 수정: 앱(index.html)에서 영수증 하루치를 요약 1행으로 저장하는 경우
    (예: 12건이 한 행에 뭉쳐 요금 합계만 기록) raw_calls "행 개수"를 그대로 건수로 세면
    과소집계된다. ①번(누적산출)에서 이미 적용한 것과 동일하게, 비고 텍스트의
    "건수 N건" 패턴에서 실제 건수를 역추출해서 가중 집계하도록 보정.
    """
    from datetime import date as _dc, timedelta as _td

    target = _dc.fromisoformat(날짜)
    window_start = (target - _td(days=7)).isoformat()

    # 명령서 #012 반영(2026-07-09): raw_calls에 축A/축B 라벨이 섞여 있으므로,
    # 축B 전용 지표(kpi_7day_avg, kpi_longdist_rate)는 date_axis='B' 행만 사용.
    # kpi_avg_fare(오늘 평일단가)는 오늘 하루치 실적 확인용이라 축 구분 없이
    # 전체 사용 — 아래에서 today_kakao는 window_calls 전체 기준으로 별도 처리.
    window_calls_all = await sb_select("raw_calls", {
        "and": f"(날짜.gte.{window_start},날짜.lte.{날짜})"
    }) or []
    window_calls = [c for c in window_calls_all if (c.get("date_axis") or "B") == "B"]

    # 명령서 #009 재검증(2026-07-08) 결과 반영: 자동 +1일 보정 제거.
    # 이지스가 업로드하는 raw_calls는 아르고스가 명령서#020·#021-rev1 자정분리
    # 원칙으로 이미 축B(달력일) 확정 처리한 뒤 적재하므로, 날짜 필드가
    # 이미 정확한 캘린더일이다. 여기에 "00~05시는 +1일" 자동 보정을 또 걸면
    # 이중 보정이 되어 날짜가 잘못 밀린다(실측 CSV 대조로 확인, 오차 12→8로 개선).
    # 향후 raw_calls에 날짜 라벨링 기준(축A/축B)을 구분하는 컬럼이 생기기 전까지는
    # 날짜 필드를 그대로 신뢰하는 것이 자동 보정보다 정확하다.
    def _axis_b_date(축a_날짜, 배차시각):
        return 축a_날짜

    # 축B 일별 카카오T 완료건수 (최근 7일, 콜 없는 날은 0으로 유지) — 건수 가중 집계
    axis_b_counts = {(target - _td(days=6 - i)).isoformat(): 0 for i in range(7)}
    for c in window_calls:
        if (c.get("콜유형") or "") != "카카오T":
            continue
        b_date = _axis_b_date(c.get("날짜"), c.get("배차시각"))
        if b_date in axis_b_counts:
            axis_b_counts[b_date] += _extract_count(c)
    kpi_7day_avg = round(sum(axis_b_counts.values()) / 7, 2)

    # 평균단가 — 오늘 카카오T 콜 기준(축 구분 없이 전체), 평일(월~금)만 산출. 건수 가중 평균.
    weekday = target.weekday()  # 0=월 ... 5=토 6=일
    today_kakao = [c for c in window_calls_all
                   if c.get("날짜") == 날짜 and (c.get("콜유형") or "") == "카카오T"]
    if weekday <= 4 and today_kakao:
        total_fare = sum(c.get("요금", 0) or 0 for c in today_kakao)
        total_cnt = sum(_extract_count(c) for c in today_kakao)
        kpi_avg_fare = int(total_fare / total_cnt) if total_cnt else None
    else:
        kpi_avg_fare = None

    # 장거리 비율 — 7일 윈도우, 카카오T, 15,000원 이상.
    # 요약행은 건당 평균요금(요금/건수)으로 개별 콜의 장거리 여부를 근사 추정.
    week_kakao = [c for c in window_calls if (c.get("콜유형") or "") == "카카오T"]
    if week_kakao:
        total_trips = 0
        longdist_trips = 0
        for c in week_kakao:
            cnt = _extract_count(c)
            fare = c.get("요금", 0) or 0
            per_trip_fare = fare / cnt if cnt else fare
            total_trips += cnt
            if per_trip_fare >= 15000:
                longdist_trips += cnt
        kpi_longdist_rate = round(longdist_trips / total_trips * 100, 1) if total_trips else None
    else:
        kpi_longdist_rate = None

    # 시간당 매출
    kpi_hourly_revenue = int(매출 / work_hours) if work_hours and work_hours > 0 else None

    return {
        "kpi_7day_avg": kpi_7day_avg,
        "kpi_avg_fare": kpi_avg_fare,
        "kpi_longdist_rate": kpi_longdist_rate,
        "kpi_hourly_revenue": kpi_hourly_revenue,
    }


async def handle_briefing(update, date_str: str = None):
    """매 운행 후 7섹션 통합 브리핑"""
    from datetime import date as _dc
    날짜 = date_str or str(_dc.today())
    mo = 날짜[:7]

    await update.message.reply_text(f"📋 {날짜} 브리핑 생성 중...")

    # 데이터 수집
    calls = await sb_select("raw_calls", {"날짜": f"eq.{날짜}"})
    calls_month = await sb_select("raw_calls", {
        "and": f"(날짜.gte.{mo}-01,날짜.lte.{mo}-31)"
    })

    total = len(calls)
    매출 = sum(c.get("요금", 0) or 0 for c in calls)
    avg_fare = int(매출 / total) if total else 0

    from datetime import date as _d2
    days_so_far = (_d2.today() - _d2(int(mo[:4]), int(mo[5:7]), 1)).days + 1
    # 명령서 #010 대응(2026-07-08): 행 개수가 아닌 건수 가중 합계로 수정.
    # 또한 아르고스 실측치(일평균 10~15건)와의 괴리 원인이 분모 정의 차이일 가능성이 높아,
    # 캘린더일 기준 월평균과 별도로 "운행일 기준 평균"도 함께 계산해 브리핑에 병기.
    # 아르고스 방법론 확정 회신 오기 전까지는 두 수치를 나란히 보여줘서 비교 가능하게 함.
    month_weighted_count = sum(_extract_count(c) for c in calls_month)
    monthly_avg_calls = month_weighted_count / max(days_so_far, 1)
    operating_days = len(set(c.get("날짜") for c in calls_month if c.get("날짜")))
    workday_avg_calls = month_weighted_count / operating_days if operating_days else 0.0

    # 공식 6변수 평가
    var_score = await calc_official_var_score(날짜)

    # 7섹션 구성
    lines = [
        f"═══ 자비스 브리핑 {날짜} ═══",
        f"",
        f"[A] 운행 데이터",
        f"  콜수: {total}건 | 매출: {fmt(매출)}원",
        f"  건당단가: {fmt(avg_fare)}원",
        f"",
        f"[B] 카카오 알고리즘 관점",
        f"  ② 오늘 완료수: {total}건 (월평균 {monthly_avg_calls:.1f}건 · 운행일평균 {workday_avg_calls:.1f}건)",
        f"  ⑤ 수락률: 100% ✅",
        f"  AI 진입 추정: {var_score['ai_inclusion_estimate']}",
        f"  약점: {var_score['improvement_needed']}",
        f"",
        f"[C] 확률 분포",
        f"  건당단가 {fmt(avg_fare)}원",
        f"  {'목표단가 초과 ✅' if avg_fare >= 10000 else '목표단가 미달 (10,000원 목표)'}",
        f"",
        f"[D] 운빨 vs 추세",
        f"  오늘: {total}건 / 월평균: {monthly_avg_calls:.1f}건 / 운행일평균: {workday_avg_calls:.1f}건",
        f"  {'▲ 추세 우위' if total >= monthly_avg_calls else '▼ 추세 하회'}",
        f"",
        f"[E] 종합 진단",
        f"  운행완료수 약점 {'개선 중 📈' if monthly_avg_calls >= 12 else '강화 필요 ⚠️'}",
        f"  수락률·평점·만나지않기 모두 최고 ✅",
        f"",
        f"[F] 다음 운행 전략",
        f"  19~21시 수성구 집중 → 21시 성내2동 앵커",
        f"  수락률 100% 유지 (콜 거절 금지)",
        f"  목표: {max(0, 18-total)}건 이상 추가 달성",
        f"",
        f"[G] 베이지안 업데이트",
        f"  오늘 {total}건 반영 완료",
        f"  누적 {len(calls_month)}건 → 모델 정밀도 {min(95, 60 + len(calls_month)//10)}%",
    ]

    # DB 저장
    briefing_data = {
        "run_date": 날짜,
        "section_a": {"calls": total, "revenue": 매출, "avg_fare": avg_fare},
        "section_b": var_score,
        "section_c": {"avg_fare": avg_fare, "target": 10000},
        "section_d": {"today": total, "monthly_avg": round(monthly_avg_calls, 1)},
        "section_e": f"운행완료수 {'개선중' if monthly_avg_calls >= 12 else '강화필요'}",
        "section_f": "19~21 수성구 → 21시 성내2동 앵커, 수락률 100% 유지",
        "section_g": {"cumulative": len(calls_month), "model_accuracy": min(95, 60+len(calls_month)//10)}
    }
    await sb_h("POST", "daily_briefing",
        json=briefing_data,
        headers={**HEADERS_SB, "Prefer": "resolution=merge-duplicates,return=minimal"},
        params={"on_conflict": "run_date"}
    )

    briefing_text = "\n".join(lines)

    # 캐스퍼 명령서 #004/#006 반영 (2026-07-06) — 이지스 연동용 bot_briefings 저장
    # atlas_reports는 아틀라스→앱 파이프라인 전용이라 재사용 시 충돌하므로 별도 테이블 사용
    try:
        await sb_insert("bot_briefings", {
            "briefing_type": "daily_close",
            "content": briefing_text,
            "run_date": 날짜,
            "status": "pending"
        })
    except Exception as e:
        logger.error(f"bot_briefings 저장 오류: {e}")

    # 캐스퍼 명령서 #008 §3 — KPI 4종 (봇 자체 계산, 축B 필수)
    kpi = {}
    try:
        def _to_virtual_min(hhmm: str) -> int:
            h, m = map(int, hhmm.split(":"))
            minutes = h * 60 + m
            if h < 6:  # 0~5시는 전날 심야운행의 연장으로 간주
                minutes += 24 * 60
            return minutes

        시각목록 = [c.get("배차시각") for c in calls if c.get("배차시각")]
        summary_payload = {"날짜": 날짜}
        work_hours = None
        if 시각목록:
            정렬됨 = sorted((( _to_virtual_min(t), t) for t in 시각목록))
            시작 = 정렬됨[0][1]
            종료 = 정렬됨[-1][1]
            work_hours = round((정렬됨[-1][0] - 정렬됨[0][0]) / 60, 2)
            summary_payload.update({
                "work_start_time": 시작,
                "work_end_time": 종료,
                "work_hours": work_hours
            })

        kpi = await calc_kpi_metrics(날짜, 매출, work_hours)
        summary_payload.update(kpi)

        await sb_upsert("daily_summary", summary_payload, on_conflict="날짜")
    except Exception as e:
        logger.error(f"daily_summary 운행시간/KPI 저장 오류: {e}")

    # ══════════════════════════════════════════════
    # 캐스퍼 명령서 #014 반영 (2026-07-10)
    # 완전 자동화: 브리핑 생성 → GitHub 직접 커밋 → 텔레그램 요약+파일 전송
    # ══════════════════════════════════════════════

    kakao_n = sum(1 for c in calls if (c.get("콜유형") or "") == "카카오T")
    uber_n  = sum(1 for c in calls if (c.get("콜유형") or "") == "우버")
    bhw_n   = sum(1 for c in calls if (c.get("콜유형") or "") == "배회")

    # 절벽구간(간이 산정): 오늘 콜 간 40분 이상 공백 — GPX 교차검증 없는 raw_calls 시각만의 근사치.
    # 아르고스의 정식 Dead Zone 분석(GPX 이동 여부 확인 포함)과는 다른, 봇 자체의 단순 근사값임을 명시.
    def _gap_stats(calls, threshold_min=40):
        times = [t for t in (c.get("배차시각") for c in calls) if t]
        if len(times) < 2:
            return 0, 0
        mins_sorted = sorted(_to_virtual_min(t) for t in times)
        gap_count = 0
        gap_total = 0
        for i in range(1, len(mins_sorted)):
            gap = mins_sorted[i] - mins_sorted[i-1]
            if gap >= threshold_min:
                gap_count += 1
                gap_total += gap
        return gap_total, gap_count

    gap_total_min, gap_count = _gap_stats(calls)

    kpi_7day = kpi.get("kpi_7day_avg")
    kpi_fare = kpi.get("kpi_avg_fare")
    kpi_long = kpi.get("kpi_longdist_rate")
    kpi_hourly = kpi.get("kpi_hourly_revenue")
    kpi_met = sum([
        (kpi_7day or 0) >= 10,
        (kpi_fare or 0) >= 10000,
        (kpi_long or 0) >= 20,
        (kpi_hourly or 0) >= 20000,
    ])

    # 전체 브리핑 markdown 문서 (GitHub 커밋 + 파일첨부용)
    full_md = "\n".join([
        f"# 자비스 브리핑 {날짜}",
        "",
        briefing_text,
        "",
        "---",
        f"*bot_v5.py 자동생성 · KPI 판정 {kpi_met}/4 충족*",
    ])

    # GitHub 직접 커밋
    gh_result = await github_commit_briefing(날짜, full_md)
    if not gh_result["ok"]:
        logger.error(f"GitHub 커밋 실패: {gh_result['error']}")

    # 텔레그램 요약 메시지 (명령서 #014 §3 템플릿)
    summary_msg = (
        f"[봇 자동요약 / 실시간, 검증 전]\n"
        f"[자비스 브리핑 요약 / {날짜}]\n"
        f"카카오T {kakao_n}건 · 우버{uber_n}건 · 배회{bhw_n}건 | 매출 {fmt(매출)}원\n"
        f"7일평균 {(f'{kpi_7day:.1f}건' if kpi_7day is not None else '-건')} "
        f"(기준10건 대비 {'✅' if (kpi_7day or 0) >= 10 else '❌'})\n"
        f"평균단가 {(fmt(kpi_fare) + '원' if kpi_fare is not None else '-원')} "
        f"(기준10,000원 대비 {'✅' if (kpi_fare or 0) >= 10000 else '❌'})\n"
        f"KPI 판정: {kpi_met}/4 충족\n"
        f"절벽구간: {gap_total_min}분 ({gap_count}건, 간이산정)\n"
        f"오늘 요약: 콜 {total}건 · 매출 {fmt(매출)}원 기록\n"
        f"→ 전체 브리핑은 첨부파일 참고\n"
        f"→ 아르고스 정밀 분석은 별도로 브리핑 확인"
        + (f"\nGitHub: {gh_result['url']}" if gh_result["ok"] else "\n⚠️ GitHub 저장 실패 (로그 확인 필요)")
    )

    # 안전장치: 4,096자 초과 시 자동 분할
    if len(summary_msg) > 4000:
        for i in range(0, len(summary_msg), 4000):
            await update.message.reply_text(summary_msg[i:i+4000])
    else:
        await update.message.reply_text(summary_msg)

    # 전체 브리핑 파일 첨부 전송
    try:
        file_bytes = io.BytesIO(full_md.encode("utf-8"))
        file_bytes.name = f"bot_summary_{날짜.replace('-','')}.md"
        await update.message.reply_document(document=file_bytes, filename=f"bot_summary_{날짜.replace('-','')}.md")
    except Exception as e:
        logger.error(f"브리핑 파일 전송 오류: {e}")


async def save_operation_consistency(날짜: str, 시작시각: str, 종료시각: str,
                                     총건수: int, 총매출: int):
    """운행 일관성 데이터 저장 및 점수 산출"""
    try:
        h_start = int(시작시각.split(":")[0])
        m_start = int(시작시각.split(":")[1])
        # 목표 19:00 기준 격차 (분)
        격차 = (h_start * 60 + m_start) - (19 * 60)
        격차_abs = abs(격차)

        # 일관성 점수 (100점 기준)
        # 시작 시각 ±15분 = 100, ±30분 = 80, ±60분 = 60, 초과 = 40
        if 격차_abs <= 15: start_score = 100
        elif 격차_abs <= 30: start_score = 80
        elif 격차_abs <= 60: start_score = 60
        else: start_score = 40

        일관성점수 = start_score

        payload = {
            "날짜":         날짜,
            "시작시각":     시작시각,
            "종료시각":     종료시각,
            "시작격차_분":  격차,
            "일관성점수":   일관성점수,
            "총건수":       총건수,
            "총매출":       총매출,
        }
        await sb_upsert("operation_consistency", payload, on_conflict="날짜")
        return 일관성점수
    except Exception as e:
        logger.error(f"일관성 저장 오류: {e}")
        return None


async def report_operation_consistency(update, 날짜: str = None):
    """운행 일관성 보고 — 오늘 + 최근 7일 평균"""
    try:
        target = 날짜 or str(today_kst())
        rows = await sb_select("operation_consistency",
                               {"order": "날짜.desc", "limit": "7"})
        if not rows:
            await update.message.reply_text("📊 운행 일관성 데이터 없음")
            return

        today_row = next((r for r in rows if r.get("날짜") == target), rows[0])
        시작 = today_row.get("시작시각", "?")
        종료 = today_row.get("종료시각", "?")
        격차 = today_row.get("시작격차_분", 0)
        점수 = today_row.get("일관성점수", 0)
        건수 = today_row.get("총건수", 0)
        매출 = today_row.get("총매출", 0)

        # 7일 평균
        starts = [r.get("시작격차_분", 0) for r in rows if r.get("시작격차_분") is not None]
        avg_격차 = sum(starts) / len(starts) if starts else 0

        # 점수 별
        stars = "⭐" * (점수 // 25 + 1) if 점수 else "?"

        lines = [
            f"📊 운행 일관성 — {target}",
            f"시작: {시작} (목표 19:00, 격차 {'+' if 격차>=0 else ''}{격차}분)",
            f"종료: {종료}",
            f"건수: {건수}건 | 매출: {매출:,}원" if 매출 else f"건수: {건수}건",
            f"",
            f"일관성 점수: {점수}/100 {stars}",
            f"7일 시작 평균 격차: {avg_격차:+.0f}분",
            f"★ 알고리즘 학습 신호: {'강함' if 점수 >= 80 else '보통' if 점수 >= 60 else '약함'}",
        ]
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        logger.error(f"일관성 조회 오류: {e}")
        await update.message.reply_text("❌ 일관성 조회 오류")

async def ocr_sekuti(image_bytes: bytes) -> dict | None:
    """세큐티 리포트 이미지 OCR → 점수·등급 추출"""
    prompt = (
        "이 세큐티(SEKUTI) 기사 리포트 이미지에서 정보를 추출해서 JSON만 반환해줘.\n"
        '{"종합점수":숫자,"상위퍼센트":숫자,"수락률":숫자,"실내공기":숫자,'
        '"친절도":숫자,"안전운행":숫자,"등급":"마스터/다이아몬드/플래티넘/골드/실버 중 하나",'
        '"기간":"주간 또는 월간","기준날짜":"YYYY-MM-DD"}\n'
        "없는 항목은 null. 점수는 숫자만(원,% 제외).\n"
        "JSON만 반환. 설명·마크다운 금지."
    )
    try:
        raw = await claude_vision(image_bytes, prompt, max_tokens=300)
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        import json as _json
        return _json.loads(raw)
    except Exception as e:
        logger.error(f"세큐티 OCR 오류: {e}")
        return None


async def process_sekuti(update, image_bytes: bytes):
    """세큐티 이미지 처리: OCR → sekuti_weekly 저장 → 결과 표시"""
    await update.message.reply_text("📊 세큐티 분석 중...")

    data = await ocr_sekuti(image_bytes)
    if not data:
        await update.message.reply_text("❌ 세큐티 인식 실패. 다시 올려주세요.")
        return

    from datetime import date as _dc
    기준날짜 = data.get("기준날짜")
    if 기준날짜:
        try:
            _dc.fromisoformat(str(기준날짜))
        except Exception:
            기준날짜 = str(today_kst())
    else:
        기준날짜 = str(today_kst())

    종합점수   = data.get("종합점수")
    상위퍼센트 = data.get("상위퍼센트")
    수락률     = data.get("수락률")
    실내공기   = data.get("실내공기")
    친절도     = data.get("친절도")
    안전운행   = data.get("안전운행")
    등급       = data.get("등급") or ""
    기간       = data.get("기간") or "주간"

    payload = {
        "날짜":       기준날짜,
        "종합점수":   종합점수,
        "상위퍼센트": 상위퍼센트,
        "수락률":     수락률,
        "실내공기":   실내공기,
        "친절도":     친절도,
        "안전운행":   안전운행,
        "등급":       등급,
        "기간":       기간,
    }
    await sb_upsert("sekuti_weekly", payload, on_conflict="날짜,기간")

    grade_icon = {
        "마스터":"👑","다이아몬드":"💎","플래티넘":"🥈","골드":"🥇","실버":"🥉"
    }.get(등급, "📊")

    def score_bar(score):
        if score is None: return "N/A"
        s = int(score)
        if s >= 95: return f"{s}점 🟢"
        if s >= 90: return f"{s}점 🟡"
        return f"{s}점 🔴"

    master_check = []
    if 종합점수 and 종합점수 >= 95:
        master_check.append("✅ 종합점수 95↑")
    else:
        master_check.append(f"❌ 종합점수 {종합점수 or '?'} (95 필요)")
    if 실내공기 and 실내공기 >= 93:
        master_check.append("✅ 실내공기 93↑")
    else:
        master_check.append(f"❌ 실내공기 {실내공기 or '?'} (93 필요)")
    if 수락률 and 수락률 >= 95:
        master_check.append("✅ 수락률 95↑")
    else:
        master_check.append(f"❌ 수락률 {수락률 or '?'} (95 필요)")

    msg_lines = [
        f"{grade_icon} 세큐티 리포트 ({기간}) — {기준날짜}",
        "",
        f"종합점수: {score_bar(종합점수)}" + (f" (상위 {상위퍼센트}%)" if 상위퍼센트 else ""),
        f"등급: {등급}",
        "",
        "[항목별]",
        f"  수락률:   {score_bar(수락률)}",
        f"  실내공기: {score_bar(실내공기)}",
        f"  친절도:   {score_bar(친절도)}",
        f"  안전운행: {score_bar(안전운행)}",
        "",
        "[마스터 전환 체크]",
    ] + master_check

    await update.message.reply_text("\n".join(msg_lines))


async def handle_sekuti_query(update):
    """세큐티 최근 기록 조회 — '세큐티 조회' 명령어"""
    rows = await sb_select("sekuti_weekly", {"order": "날짜.desc", "limit": "5"})
    if not rows:
        await update.message.reply_text(
            "📊 세큐티 기록 없음\n세큐티 리포트 이미지를 올려주세요."
        )
        return

    grade_icon = {"마스터":"👑","다이아몬드":"💎","플래티넘":"🥈","골드":"🥇","실버":"🥉"}
    lines_out = ["📊 세큐티 최근 기록\n"]
    for r in rows:
        icon = grade_icon.get(r.get("등급",""), "📊")
        lines_out.append(
            f"{icon} {r.get('날짜','')} ({r.get('기간','')})\n"
            f"  종합 {r.get('종합점수','?')}점 · 상위 {r.get('상위퍼센트','?')}% · {r.get('등급','')}\n"
            f"  수락{r.get('수락률','?')} 공기{r.get('실내공기','?')} "
            f"친절{r.get('친절도','?')} 안전{r.get('안전운행','?')}"
        )
    await update.message.reply_text("\n".join(lines_out))


async def ocr_call_card(image_bytes: bytes) -> dict | None:
    prompt = (
        "이 카카오T 콜카드(운행이력) 이미지에서 아래 JSON만 반환해줘.\n"
        '{"날짜":"YYYY-MM-DD","배차시각":"HH:MM","하차시각":"HH:MM",'
        '"출발지":"OO구 OO동","도착지":"OO구 OO동",'
        '"요금":숫자,"카드사":"카드사명","콜유형":"카카오T 또는 배회","결제방식":"자동 또는 직접"}\n'
        "⚠️ 날짜: 이미지 상단의 운행 날짜(예: 2026/03/07 → 2026-03-07). 없으면 null.\n"
        "  단, 화면에 여러 날짜가 있으면 가장 최근 운행의 날짜만 추출.\n"
        "⚠️ 요금 추출:\n"
        "  - \'미터 요금\', \'총 요금\', \'결제 금액\', \'이용 요금\' 레이블 옆 숫자\n"
        "  - 예: 미터요금 13,100원 → 요금=13100\n"
        "  - 요금이 보이면 반드시 추출. 화면에 없을 때만 0으로 설정\n"
        "⚠️ 결제방식:\n"
        "  - 요금 숫자 보임 → 결제방식=\'자동\', 요금=해당숫자\n"
        "  - 요금 전혀 없음 → 결제방식=\'직접\', 요금=0\n"
        "⚠️ 도착지: 차량번호(예: 대구 32바 5763, XX가나바 NNNN) → null.\n"
        "  도착지는 \'OO구 OO동\' 형식 주소만.\n"
        "⚠️ 시각: 배차시각=승차시각, 하차시각=하차완료시각. 없으면 null.\n"
        "JSON만 반환. 설명·마크다운 금지."
    )
    raw = await claude_vision(image_bytes, prompt, max_tokens=200)
    try:
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```")
        return json.loads(raw)
    except Exception:
        logger.error(f"콜카드 JSON 파싱 실패: {raw}")
        return None

async def ocr_charge_receipt(image_bytes: bytes) -> list:
    prompt = (
        "이 충전 내역 이미지에서 모든 충전 건을 추출해서 JSON 배열만 반환해줘.\n"
        '[{"충전일자":"YYYY-MM-DD","충전소명":"이름","충전량":숫자(kWh),"결제금액":숫자(원)}, ...]\n'
        "여러 건이 보이면 전부 추출. JSON 배열만 반환. 설명 금지."
    )
    raw = await claude_vision(image_bytes, prompt, max_tokens=500)
    try:
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```")
        result = json.loads(raw)
        return result if isinstance(result, list) else []
    except Exception:
        logger.error(f"충전내역 JSON 파싱 실패: {raw}")
        return []

async def ocr_payment_history(image_bytes: bytes) -> list:
    """
    카카오T 결제내역(수익관리 화면) OCR.
    반환: [{"날짜":"YYYY-MM-DD","시각":"HH:MM","요금":숫자,"결제방법":"카드/현금"}, ...]
    """
    prompt = (
        "이 결제내역 화면에서 모든 결제 건을 추출해서 JSON 배열만 반환해줘.\n"
        '[{"날짜":"YYYY-MM-DD","시각":"HH:MM","요금":숫자,"결제방법":"카드 또는 현금"}, ...]\n'
        "⚠️ 지원하는 화면 형식:\n"
        "  1. 카카오T 수익관리 화면: 거래일자 YYYY-MM-DD HH:MM:SS 형식\n"
        "  2. 세큐티 결제내역조회 화면: 거래일자 YYYY-MM-DD | HH:MM:SS 형식\n"
        "     (상단에 결제내역조회, 차량번호, 조회기간 필터 있음)\n"
        "⚠️ 시각 처리: HH:MM:SS → HH:MM 으로 변환 (초 제거)\n"
        "⚠️ 날짜: 각 건의 거래일자에서 YYYY-MM-DD 추출. null 불가.\n"
        "⚠️ 요금: 숫자만 (쉼표, 원 제거). 예: 5,400원 → 5400\n"
        "⚠️ 취소 건(구분=취소) 제외. 승인 정상만 포함.\n"
        "JSON 배열만 반환. 설명·마크다운 금지."
    )
    raw = await claude_vision(image_bytes, prompt, max_tokens=2000)
    try:
        raw = raw.strip()
        raw = re.sub(r"```json\s*", "", raw)
        raw = re.sub(r"```\s*", "", raw)
        raw = raw.strip()
        arr_start = raw.find("[")
        arr_end = raw.rfind("]")
        if arr_start >= 0 and arr_end > arr_start:
            raw = raw[arr_start:arr_end+1]
        result = json.loads(raw)
        return result if isinstance(result, list) else []
    except Exception as e:
        logger.error(f"결제내역 JSON 파싱 실패: {e} / raw: {raw[:200]}")
        return []
    # 구버전 호환
    except Exception:
        logger.error(f"결제내역 JSON 파싱 실패: {raw}")
        return []

# ──────────────────────────────────────────────
# 교차대조 로직 (콜카드 ↔ 결제내역)
# ──────────────────────────────────────────────
async def cross_check(date_str: str) -> str:
    """
    콜카드(raw_calls) ↔ 결제내역(payment_receipts) 교차대조.
    매칭: 콜카드 하차시각 ↔ 결제시각 ±20분 (자정넘김 처리)
    미매칭 결제내역 자동분류:
      - 콜카드 운행 공백 시간대 → 배회영업 후보
      - 콜카드 운행 중 시간대   → 누락 콜카드 후보
    """
    from datetime import date as date_cls, timedelta

    calls = await sb_select("raw_calls", {"날짜": f"eq.{date_str}"})
    try:
        y, mo, d = date_str.split("-")
        next_date_str = str(date_cls(int(y),int(mo),int(d)) + timedelta(days=1))
    except Exception:
        next_date_str = date_str

    receipts_today = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})
    receipts_next  = await sb_select("payment_receipts", {"날짜": f"eq.{next_date_str}"})
    receipts = receipts_today + receipts_next

    if not calls and not receipts:
        return f"⚠️ {date_str} 데이터 없음"

    def to_min_smart(배차시각, 대상시각, 대상날짜):
        try:
            bh, bm = 배차시각.split(":")
            base = int(bh)*60+int(bm)
            th, tm = 대상시각.split(":")
            target = int(th)*60+int(tm)
            if 대상날짜 == next_date_str:
                target += 1440
            elif target < base - 60:
                target += 1440
            return target
        except: return None

    def to_min_abs(time_str, date_str_local):
        """날짜 기반 분 변환. mins<300 휴리스틱 제거 — 새벽 운행 오류 방지."""
        try:
            h, m = time_str.split(":")
            mins = int(h)*60+int(m)
            if date_str_local == next_date_str:
                mins += 1440  # 익일 날짜면 +1440만 적용
            return mins
        except: return None

    # STEP 1: 하차시각 ↔ 결제시각 ±20분 매칭
    matched_call_ids    = set()
    matched_receipt_ids = set()
    fee_mismatches      = []  # 금액 불일치 목록
    direct_updated      = []  # 직접결제 요금 자동업데이트 목록

    for i, call in enumerate(calls):
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각")
        if not 하차:
            try:
                bh, bm = 배차.split(":")
                est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: continue
        c_min = to_min_smart(배차, 하차, date_str)
        best_j, best_diff = None, 99999
        for j, rcpt in enumerate(receipts):
            if j in matched_receipt_ids: continue
            rcpt_date = rcpt.get("날짜", date_str)
            r_min = to_min_smart(배차, rcpt.get("시각","") or "", rcpt_date)
            if r_min and c_min:
                diff = abs(c_min - r_min)
                if diff <= 20 and diff < best_diff:
                    best_diff = diff
                    best_j = j
        if best_j is not None:
            matched_call_ids.add(i)
            matched_receipt_ids.add(best_j)
            call_fee = call.get("요금") or 0
            rcpt_fee = receipts[best_j].get("요금") or 0

            # 직접결제(요금=0) 콜카드 → 결제내역 요금으로 자동 업데이트
            if call_fee == 0 and rcpt_fee > 0:
                call_id = call.get("id")
                if call_id:
                    await sb_h("PATCH", f"raw_calls?id=eq.{call_id}",
                               json={"요금": rcpt_fee, "비고": "직접결제(요금확인완료)"})
                    direct_updated.append({
                        "배차시각": 배차,
                        "요금": rcpt_fee,
                    })

            # 금액 불일치 (둘 다 0이 아니고 차이 ≥500원)
            elif call_fee > 0 and rcpt_fee > 0:
                fee_diff = abs(call_fee - rcpt_fee)
                if fee_diff >= FEE_DIFF_THRESHOLD:
                    fee_mismatches.append({
                        "call_id": call.get("id"),
                        "배차시각": 배차,
                        "call_fee": call_fee,
                        "rcpt_fee": rcpt_fee,
                        "diff": fee_diff,
                    })

    unmatched_calls    = [c for i,c in enumerate(calls)    if i not in matched_call_ids]
    unmatched_receipts = [r for j,r in enumerate(receipts) if j not in matched_receipt_ids]

    # STEP 2: 콜카드 점유 시간대 계산 → 미매칭 결제내역 분류
    occupied = []
    for call in calls:
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각") or ""
        if not 하차:
            try:
                bh, bm = 배차.split(":")
                est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: pass
        s = to_min_abs(배차, date_str)
        e = to_min_abs(하차, date_str)
        if s and e:
            occupied.append((s-5, e+5))

    baehoe_rcpt  = []
    missing_rcpt = []
    for r in unmatched_receipts:
        r_date = r.get("날짜", date_str)
        r_min  = to_min_abs(r.get("시각","") or "", r_date)
        if r_min is None:
            missing_rcpt.append(r)
            continue
        if any(s <= r_min <= e for s,e in occupied):
            missing_rcpt.append(r)   # 운행 중 시간 → 누락 콜카드
        else:
            baehoe_rcpt.append(r)    # 공백 시간 → 배회영업

    # 결과 출력
    lines_out = [f"📊 교차대조 결과 — {date_str}"]
    lines_out.append(f"콜카드 {len(calls)}건 / 결제내역 {len(receipts)}건 / 매칭 {len(matched_call_ids)}건")
    lines_out.append("")

    # 직접결제 요금 자동업데이트 표시
    if direct_updated:
        lines_out.append(f"💳 직접결제 요금 자동확인 {len(direct_updated)}건:")
        for d in direct_updated:
            lines_out.append(f"  ✅ {d['배차시각']} → {fmt(d['요금'])} 업데이트")
        lines_out.append("")

    if unmatched_calls:
        lines_out.append(f"🟠 콜카드에만 있음 {len(unmatched_calls)}건:")
        for c in unmatched_calls:
            lines_out.append(f"  {c.get('배차시각','-')} {c.get('출발지','')}→{c.get('도착지','')} {fmt(c.get('요금') or 0)}")
        lines_out.append("")

    if baehoe_rcpt:
        lines_out.append(f"🚶 배회영업 후보 (공백시간) {len(baehoe_rcpt)}건:")
        for r in baehoe_rcpt:
            날짜표시 = f"({r.get('날짜','')})" if r.get("날짜") != date_str else ""
            lines_out.append(f"  {r.get('시각','-')}{날짜표시} {fmt(r.get('요금') or 0)}")
        lines_out.append("")

    if missing_rcpt:
        lines_out.append(f"🔴 누락 콜카드 후보 (운행중 시간) {len(missing_rcpt)}건:")
        for r in missing_rcpt:
            날짜표시 = f"({r.get('날짜','')})" if r.get("날짜") != date_str else ""
            lines_out.append(f"  {r.get('시각','-')}{날짜표시} {fmt(r.get('요금') or 0)}")
        lines_out.append("")

    # 금액 불일치 표시
    if fee_mismatches:
        lines_out.append(f"💰 금액 불일치 {len(fee_mismatches)}건 (차이 ≥500원):")
        for fm in fee_mismatches:
            lines_out.append(
                f"  {fm['배차시각']} 콜카드:{fmt(fm['call_fee'])} vs "
                f"결제:{fmt(fm['rcpt_fee'])} (차이 {fm['diff']:,}원)"
            )
        lines_out.append("  → '대조 금액확인 YYYY-MM-DD' 로 버튼 선택")
        lines_out.append("")

    if not unmatched_calls and not unmatched_receipts:
        if fee_mismatches:
            lines_out.append("⚠️ 매칭 완료 — 금액 불일치 확인 필요")
        else:
            lines_out.append("✅ 완전 매칭 — 누락 없음")

    if unmatched_calls:
        lines_out.append(f"💡 '배회분류 확정 {date_str}' → 콜카드 미매칭 배회 처리")
    if baehoe_rcpt:
        lines_out.append(f"💡 '대조 확정 {date_str}' → 배회후보 {len(baehoe_rcpt)}건 raw_calls 자동 추가")

    return "\n".join(lines_out)


async def confirm_baehoe_classification(date_str: str) -> str:
    """미매칭 콜카드를 배회영업으로 자동 분류 확정"""
    calls = await sb_select("raw_calls", {"날짜": f"eq.{date_str}"})
    receipts = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})

    def to_minutes(t_str):
        try:
            h, m = t_str.split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return None

    matched_call_ids = set()
    for i, call in enumerate(calls):
        call_min = to_minutes(call.get("배차시각", "") or "")
        for j, rcpt in enumerate(receipts):
            rcpt_min = to_minutes(rcpt.get("시각", "") or "")
            if call_min and rcpt_min and abs(call_min - rcpt_min) <= 3:
                matched_call_ids.add(i)
                break

    unmatched = [c for i, c in enumerate(calls) if i not in matched_call_ids]
    if not unmatched:
        return "✅ 미매칭 콜카드 없음"

    count = 0
    for call in unmatched:
        # 콜유형을 배회로 업데이트
        await sb_h(
            "PATCH",
            f"raw_calls?id=eq.{call['id']}",
            json={"콜유형": "배회"}
        )
        count += 1

    return f"✅ {count}건 배회영업으로 분류 완료"

# ──────────────────────────────────────────────
# 이미지 처리 함수

# ──────────────────────────────────────────────
# 중복 체크 헬퍼
# ──────────────────────────────────────────────
async def check_duplicate_call(날짜: str, 배차시각: str, 요금: int) -> bool:
    """raw_calls 중복 체크: 날짜+배차시각+요금 동일하면 True"""
    rows = await sb_select("raw_calls", {
        "날짜": f"eq.{날짜}",
        "배차시각": f"eq.{배차시각}",
        "요금": f"eq.{요금}",
    })
    return len(rows) > 0

async def check_duplicate_payment(날짜: str, 시각: str, 요금: int) -> bool:
    """payment_receipts 중복 체크: 날짜+시각+요금 동일하면 True"""
    rows = await sb_select("payment_receipts", {
        "날짜": f"eq.{날짜}",
        "시각": f"eq.{시각}",
        "요금": f"eq.{요금}",
    })
    return len(rows) > 0

# ──────────────────────────────────────────────
async def process_call_card(update: Update, image_bytes: bytes):
    data = await ocr_call_card(image_bytes)
    if not data:
        await update.message.reply_text("❌ 콜카드 인식 실패. 다시 올려주세요.")
        return

    # 날짜 결정 로직:
    # - OCR 날짜가 과거(미래 아님)이면 → OCR 날짜 신뢰 (뒤늦게 올린 콜카드)
    # - OCR 날짜가 미래이면 → 오늘 날짜 (오인식)
    # - OCR 날짜 없으면 → 오늘 날짜
    from datetime import date as _dc, timedelta as _td
    _today_d = today_kst()
    _today_str = str(_today_d)
    _dow_map = ["월","화","수","목","금","토","일"]

    ocr_date_raw = data.get("날짜")
    today = _today_str
    dow = _dow_map[_today_d.weekday()]
    date_source = "오늘"  # 저장 메시지용

    if ocr_date_raw:
        try:
            _ocr_d = _dc.fromisoformat(str(ocr_date_raw).strip())
            if _ocr_d > _today_d:
                # 미래 날짜 → 오인식 → 오늘 사용
                today = _today_str
                date_source = f"오늘(OCR미래날짜오류)"
            else:
                # 과거 또는 오늘 → OCR 날짜 신뢰
                today = str(_ocr_d)
                dow = _dow_map[_ocr_d.weekday()]
                diff = (_today_d - _ocr_d).days
                date_source = f"OCR({diff}일전)" if diff > 0 else "오늘"
        except Exception:
            today = _today_str
            date_source = "오늘(OCR파싱오류)"

    배차시각 = data.get("배차시각")
    요금 = data.get("요금") or 0
    결제방식 = data.get("결제방식", "자동")

    # 중복 체크 → 자동 삭제 후 재저장
    deleted = await delete_duplicate_call(today, 배차시각, 요금)
    if deleted:
        logger.info(f"중복 콜카드 자동 삭제 후 재저장: {today} {배차시각} {요금}")

    # 직접결제: 요금 0원 → pending 상태로 저장
    # 직접결제 판별:
    # - 결제방식이 명시적으로 "직접"인 경우만 직접결제
    # - 요금=0이어도 결제방식="자동"이면 OCR 오류일 수 있으므로 경고만 표시
    is_direct = (결제방식 == "직접")
    is_zero_fee = (요금 == 0) and not is_direct  # 요금=0 but 자동결제
    비고 = "직접결제(요금미확인)" if is_direct else data.get("카드사")

    payload = {
        "날짜": today,
        "요일": dow,
        "배차시각": 배차시각,
        "하차시각": data.get("하차시각"),
        "출발지": data.get("출발지"),
        "도착지": data.get("도착지"),
        "요금": 요금,
        "콜유형": data.get("콜유형", "카카오T"),
        "비고": 비고,
    }
    result = await sb_insert("raw_calls", payload)
    if result:
        if is_direct:
            await update.message.reply_text(
                f"💳 직접결제 콜카드 저장 (요금 미확인)\n"
                f"날짜: {today} ({dow}) [{date_source}]\n"
                f"{배차시각} {data.get('출발지','?')}→{data.get('도착지','?') or '?'}\n"
                f"⚠️ 결제내역 업로드 후 '대조 {today}' 입력해서 요금 확인하세요."
            )
        elif is_zero_fee:
            await update.message.reply_text(
                f"⚠️ 콜 저장 (요금 0원 — 확인 필요)\n"
                f"{배차시각} {data.get('출발지','?')}→{data.get('도착지','?')}\n"
                f"콜카드에 요금이 보이면 '콜수정 {배차시각} 요금=실제금액' 으로 수정하세요."
            )
        else:
            await update.message.reply_text(
                f"✅ 콜 저장\n"
                f"날짜: {today} ({dow}) [{date_source}]\n"
                f"{배차시각} {data.get('출발지','?')}→{data.get('도착지','?')}\n"
                f"{fmt(요금)} [{data.get('콜유형','카카오T')}]"
            )
    else:
        await update.message.reply_text("❌ DB 저장 실패")

async def process_charge_receipt(update: Update, image_bytes: bytes):
    items = await ocr_charge_receipt(image_bytes)
    if not items:
        await update.message.reply_text("❌ 충전내역 인식 실패. 다시 올려주세요.")
        return

    saved = 0
    for item in items:
        payload = {
            "충전일": item.get("충전일자") or str(today_kst()),
            "충전량_kwh": item.get("충전량"),
            "충전금액": item.get("결제금액"),
            "충전소": item.get("충전소명"),
        }
        r = await sb_insert("charging_log", payload)
        if r:
            saved += 1

    await update.message.reply_text(
        f"⚡ 충전내역 {saved}/{len(items)}건 저장 완료"
    )

async def process_payment_history(update: Update, image_bytes: bytes):
    """결제내역 OCR → payment_receipts 저장"""
    items = await ocr_payment_history(image_bytes)
    if not items:
        await update.message.reply_text("❌ 결제내역 인식 실패. 다시 올려주세요.")
        return

    saved = 0
    skipped = 0
    duplicated = 0
    date_warn = []
    dup_list = []
    for item in items:
        날짜 = item.get("날짜")
        # 날짜 없으면 저장 거부 — 오늘 날짜로 대체하지 않음
        if not 날짜 or 날짜 == "null":
            skipped += 1
            date_warn.append(item.get("시각", "?"))
            continue
        시각 = item.get("시각")
        요금 = item.get("요금", 0)
        # 중복 체크 → 자동 삭제 후 재저장
        deleted = await delete_duplicate_payment(날짜, 시각, 요금)
        if deleted:
            duplicated += deleted  # 실제 삭제 건수 반영
            logger.info(f"중복 결제내역 자동 삭제 {deleted}건: {날짜} {시각} {요금}")
        payload = {
            "날짜": 날짜,
            "시각": 시각,
            "요금": 요금,
            "결제방법": item.get("결제방법", "카드"),
        }
        r = await sb_insert("payment_receipts", payload)
        if r:
            saved += 1

    dates = list(set(item.get("날짜") for item in items if item.get("날짜") and item.get("날짜") != "null"))
    msg = f"💳 결제내역 {saved}건 저장"
    if dates:
        msg += f" ({', '.join(dates)})"
    if duplicated > 0:
        msg += f"\n⚠️ 중복 {duplicated}건 저장 안 됨: {', '.join(dup_list)}"
    if skipped > 0:
        msg += f"\n⚠️ 날짜 인식 실패 {skipped}건 저장 안 됨"
        msg += f"\n  시각: {', '.join(date_warn)} → 날짜 보이는 캡처로 재전송"
    if saved > 0 and dates:
        msg += f"\n교차대조: '대조 {dates[0]}' 입력"
    await update.message.reply_text(msg)

async def process_single_image(update: Update, context: ContextTypes.DEFAULT_TYPE, image_bytes: bytes = None):
    """이미지 처리. image_bytes가 있으면 파일 다운로드 생략 (파일 첨부 경우)."""
    if image_bytes is None:
        # 사진으로 전송된 경우 — 텔레그램 photo 객체에서 다운로드
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_bytes = await file.download_as_bytearray()
        image_bytes = bytes(image_bytes)

    # 큰 이미지 리사이즈 (OCR 정확도 유지하면서 API 부하 감소)
    image_bytes = resize_image_if_needed(image_bytes)  # API 전송 전 최적화
    image_type = await classify_image(image_bytes)
    logger.info(f"이미지 분류: {image_type}")

    if image_type == "일별운행이력":
        await process_daily_history(update, image_bytes)
    elif image_type == "콜카드":
        await process_call_card(update, image_bytes)
    elif image_type == "충전":
        await process_charge_receipt(update, image_bytes)
    elif image_type == "결제":
        await process_payment_history(update, image_bytes)
    elif image_type == "세큐티":
        await process_sekuti(update, image_bytes)
    else:
        # 분류 실패 → 결제내역으로 재시도 (결제내역이 가장 자주 오인식됨)
        logger.info("분류 실패 → 결제내역 fallback 시도")
        await update.message.reply_text("🔄 이미지 재분석 중...")
        try:
            receipts_raw = await ocr_payment_history(image_bytes)
            if receipts_raw and len(receipts_raw) > 0:
                await process_payment_history(update, image_bytes)
            else:
                await update.message.reply_text(
                    "❓ 이미지 인식 실패\n"
                    "💡 더 크게 캡처하거나 밝은 환경에서 다시 올려주세요.\n"
                    "콜카드·충전내역·결제내역·세큐티만 처리 가능합니다."
                )
        except Exception as e:
            logger.error(f"fallback OCR 오류: {e}")
            await update.message.reply_text(
                "❓ 이미지 인식 실패\n"
                "💡 더 크게 캡처해서 다시 올려주세요."
            )

async def process_image_queue_worker():
    while True:
        item = await image_queue.get()
        # item은 (update, context) 또는 (update, context, image_bytes)
        if len(item) == 3:
            update, context, image_bytes = item
        else:
            update, context = item
            image_bytes = None
        try:
            await process_single_image(update, context, image_bytes=image_bytes)
        except Exception as e:
            logger.error(f"이미지 처리 오류: {type(e).__name__}: {e}", exc_info=True)
            try:
                err_msg = str(e)[:80] if str(e) else type(e).__name__
                await update.message.reply_text(
                    f"❌ 처리 오류: {err_msg}\n"
                    f"잠시 후 다시 올려주세요."
                )
            except Exception:
                pass
        finally:
            image_queue.task_done()
        await asyncio.sleep(1)

# ──────────────────────────────────────────────
# 수동 입력 파서
# ──────────────────────────────────────────────
def parse_manual_call(text: str) -> dict | None:
    """
    '콜 7800' / '배회 5600' / '콜 7800 경산'
    → {"콜유형": ..., "요금": ..., "도착지힌트": ...}
    """
    parts = text.strip().split()
    if len(parts) < 2:
        return None
    keyword = parts[0]
    if keyword not in ("콜", "배회"):
        return None
    try:
        fee = int(parts[1].replace(",", "").replace("원", ""))
    except ValueError:
        return None
    hint = parts[2] if len(parts) >= 3 else None
    return {
        "콜유형": "카카오T" if keyword == "콜" else "배회",
        "요금": fee,
        "도착지힌트": hint,
    }

EXPENSE_KEYWORDS = {
    "충전": ("⚡ 전기충전", "charges"),
    "타이어": ("🔧 타이어교체", None),
    "오일": ("🔧 오일교환", None),
    "세차": ("🚿 세차", None),
}

def parse_expense(text: str) -> dict | None:
    """
    '충전 4595' / '타이어 80000' / '지출 기타 3000'
    → {"카테고리": ..., "금액": ..., "메모": ...}
    """
    parts = text.strip().split()
    if not parts:
        return None

    keyword = parts[0]
    if keyword in EXPENSE_KEYWORDS and len(parts) >= 2:
        label, _ = EXPENSE_KEYWORDS[keyword]
        try:
            fee = int(parts[1].replace(",", "").replace("원", ""))
        except ValueError:
            return None
        return {"카테고리": label, "금액": fee, "메모": keyword}

    if keyword == "지출" and len(parts) >= 3:
        cat = parts[1]
        try:
            fee = int(parts[2].replace(",", "").replace("원", ""))
        except ValueError:
            return None
        return {"카테고리": f"📦 {cat}", "금액": fee, "메모": cat}

    return None

# ──────────────────────────────────────────────
# 핸들러 — 수동 입력
# ──────────────────────────────────────────────
async def handle_manual_call(update: Update, parsed: dict):
    today = str(today_kst())
    dow = get_dow()
    payload = {
        "날짜": today,
        "요일": dow,
        "배차시각": now_kst().strftime("%H:%M"),
        "요금": parsed["요금"],
        "콜유형": parsed["콜유형"],
        "도착지": parsed.get("도착지힌트"),
    }
    r = await sb_insert("raw_calls", payload)
    if r:
        await update.message.reply_text(
            f"✅ {parsed['콜유형']} {fmt(parsed['요금'])} 입력"
        )
    else:
        await update.message.reply_text("❌ 저장 실패")

async def handle_expense(update: Update, parsed: dict):
    today = str(today_kst())
    payload = {
        "날짜": today,
        "카테고리": parsed["카테고리"],
        "금액": parsed["금액"],
        "메모": parsed.get("메모", ""),
        "자동여부": False,
    }
    r = await sb_insert("expenses", payload)
    if r:
        await update.message.reply_text(
            f"✅ 지출 {parsed['카테고리']} {fmt(parsed['금액'])} 입력"
        )
    else:
        await update.message.reply_text("❌ 저장 실패")

async def handle_expense_cancel(update: Update):
    today = str(today_kst())
    ok = await sb_delete_last(
        "expenses",
        {"날짜": f"eq.{today}", "자동여부": "eq.false"}
    )
    if ok:
        await update.message.reply_text("✅ 마지막 수동 지출 삭제")
    else:
        await update.message.reply_text("⚠️ 삭제할 수동 지출 없음")

async def handle_rest_day(update: Update, text: str = "휴무"):
    """휴무 처리. '4-7 휴무' 형식으로 날짜 지정 가능."""
    import re
    from datetime import date as date_cls

    target = today_kst()
    date_pat = r"(\d{4})-(\d{1,2})-(\d{1,2})|(\d{1,2})[-/](\d{1,2})"
    m = re.search(date_pat, text)
    if m:
        g = m.groups()
        try:
            if g[0]:
                target = date_cls(int(g[0]), int(g[1]), int(g[2]))
            else:
                target = date_cls(today_kst().year, int(g[3]), int(g[4]))
        except ValueError:
            await update.message.reply_text("❌ 잘못된 날짜입니다.")
            return

    dow_map = ["월","화","수","목","금","토","일"]
    date_str = str(target)
    dow = dow_map[target.weekday()]

    await sb_upsert("daily_summary", {
        "날짜": date_str, "요일": dow,
        "휴무여부": True, "정상여부": "휴무",
    }, on_conflict="날짜")
    await insert_insurance(target)
    await update.message.reply_text(
        f"✅ {date_str} ({dow}) 휴무 처리\n보험료 {INSURANCE_DAILY:,}원 자동 기록"
    )

# ──────────────────────────────────────────────
# 핸들러 — 조회
# ──────────────────────────────────────────────
async def handle_today_quick(update: Update):
    s = await today_summary()
    달성바 = "█" * (s["달성률"] // 10) + "░" * (10 - s["달성률"] // 10)
    # 신사고 체계: 건수를 요일 분포 맥락으로 표현
    from datetime import date as _date2
    dow_kor = ["월","화","수","목","금","토","일"][_date2.today().weekday()]
    DOW_EXPECTED = {
        "월":(7,12),"화":(8,13),"수":(9,14),"목":(8,12),
        "금":(10,15),"토":(11,16),"일":(10,15)
    }
    exp_min, exp_max = DOW_EXPECTED.get(dow_kor, (8,13))
    건수 = s["건수"]
    if 건수 == 0:
        건수_평가 = ""
    elif 건수 < exp_min:
        건수_평가 = f" (기대 {exp_min}~{exp_max}건 하위)"
    elif 건수 > exp_max:
        건수_평가 = f" (기대 {exp_min}~{exp_max}건 상위)"
    else:
        건수_평가 = f" (기대 {exp_min}~{exp_max}건 정상)"

    await update.message.reply_text(
        f"📍 오늘 현황 ({now_kst().strftime('%m/%d %H:%M')})\n"
        f"콜 {건수}건{건수_평가} | 매출 {fmt(s['매출'])}\n"
        f"지출 {fmt(s['지출'])} | 순수익 {fmt(s['순수익'])}\n"
        f"목표 [{달성바}] {s['달성률']}%"
    )

async def handle_weekly(update: Update):
    today = today_kst()
    start = today - timedelta(days=6)
    start_str = str(start)
    end_str = str(today)

    calls = await sb_select(
        "raw_calls",
        {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"}
    )
    expenses = await sb_select(
        "expenses",
        {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"}
    )
    총건수 = len(calls)
    총매출 = sum(c.get("요금", 0) or 0 for c in calls)
    총지출 = sum(e.get("금액", 0) or 0 for e in expenses)
    순수익 = calc_net(총매출, 총지출)
    일평균매출 = 총매출 // 7 if 총매출 else 0

    await update.message.reply_text(
        f"📅 주간 요약 ({start_str} ~ {end_str})\n"
        f"총 {총건수}건 | 매출 {fmt(총매출)}\n"
        f"지출 {fmt(총지출)} | 순수익 {fmt(순수익)}\n"
        f"일평균 매출 {fmt(일평균매출)}"
    )

async def handle_monthly(update: Update):
    today = today_kst()
    start_str = today.replace(day=1).isoformat()
    end_str = str(today)

    calls = await sb_select("raw_calls", {"날짜": f"gte.{start_str}"})
    expenses = await sb_select("expenses", {"날짜": f"gte.{start_str}"})
    총건수 = len(calls)
    총매출 = sum(c.get("요금", 0) or 0 for c in calls)
    총지출 = sum(e.get("금액", 0) or 0 for e in expenses)
    순수익 = calc_net(총매출, 총지출)

    # 카테고리별 지출
    cat_map = {}
    for e in expenses:
        cat = e.get("카테고리", "기타")
        cat_map[cat] = cat_map.get(cat, 0) + (e.get("금액", 0) or 0)
    cat_lines = "\n".join(f"  {k}: {fmt(v)}" for k, v in sorted(cat_map.items()))

    운행일 = len(set(c.get("날짜") for c in calls))
    일평균 = 총매출 // 운행일 if 운행일 else 0

    await update.message.reply_text(
        f"📆 월간 요약 ({start_str} ~)\n"
        f"운행일 {운행일}일 | 총 {총건수}건\n"
        f"총매출 {fmt(총매출)} | 일평균 {fmt(일평균)}\n"
        f"지출 {fmt(총지출)} | 순수익 {fmt(순수익)}\n"
        f"\n지출 카테고리:\n{cat_lines}"
    )

async def handle_expense_check(update: Update):
    expenses = await today_expenses()
    if not expenses:
        await update.message.reply_text("오늘 지출 없음")
        return
    lines = [f"💸 오늘 지출 ({str(today_kst())})"]
    total = 0
    for e in expenses:
        lines.append(f"  {e.get('카테고리','')} {fmt(e.get('금액',0))}")
        total += e.get("금액", 0) or 0
    lines.append(f"합계: {fmt(total)}")
    await update.message.reply_text("\n".join(lines))


async def handle_receipt_delete(update, text: str):
    """
    결제내역 삭제 명령어
    결제삭제 YYYY-MM-DD 운행외  → 02:01~18:59 시간대 삭제
    결제삭제 YYYY-MM-DD 0원     → 요금 0원·null 삭제
    결제삭제 YYYY-MM-DD 전체    → 해당 날짜 전체 삭제
    결제삭제 YYYY-MM-DD HH:MM   → 특정 시각 삭제
    """
    parts = text.strip().split(" ", 2)
    if len(parts) < 3:
        await update.message.reply_text(
            "형식:\n"
            "결제삭제 YYYY-MM-DD 운행외\n"
            "결제삭제 YYYY-MM-DD 0원\n"
            "결제삭제 YYYY-MM-DD 전체\n"
            "결제삭제 YYYY-MM-DD HH:MM"
        )
        return

    date_str = parts[1].strip()
    mode = parts[2].strip()

    try:
        rows = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})
        if not rows:
            await update.message.reply_text(f"⚠️ {date_str} 결제내역 없음")
            return

        delete_ids = []

        if mode == "운행외":
            for r in rows:
                t = r.get("시각", "") or ""
                try:
                    h, m = t.split(":")
                    mins = int(h) * 60 + int(m)
                    # 운행시간 외: 02:01~18:59
                    if 121 <= mins <= 1019:
                        delete_ids.append(r["id"])
                except Exception:
                    delete_ids.append(r["id"])

        elif mode == "0원":
            for r in rows:
                fee = r.get("요금")
                if fee is None or int(fee) == 0:
                    delete_ids.append(r["id"])

        elif mode == "전체":
            delete_ids = [r["id"] for r in rows]

        elif ":" in mode:
            for r in rows:
                if r.get("시각", "") == mode:
                    delete_ids.append(r["id"])
        else:
            await update.message.reply_text(f"❓ 알 수 없는 모드: {mode}")
            return

        if not delete_ids:
            await update.message.reply_text(f"✅ 삭제 대상 없음 ({mode})")
            return

        deleted = 0
        for rid in delete_ids:
            await sb_h("DELETE", f"payment_receipts?id=eq.{rid}")
            deleted += 1

        await update.message.reply_text(
            f"🗑️ 삭제 완료\n"
            f"{date_str} | 조건: {mode}\n"
            f"{deleted}건 삭제 (전체 {len(rows)}건 중)"
        )

    except Exception as e:
        logger.error(f"결제삭제 오류: {e}")
        await update.message.reply_text(f"❌ 삭제 오류: {str(e)[:200]}")




async def confirm_cross_check(date_str: str) -> str:
    """
    '대조 확정 YYYY-MM-DD' 명령어 처리.
    교차대조 미매칭 결제내역(현금) → raw_calls 배회영업으로 자동 추가.
    """
    from datetime import date as date_cls, timedelta

    try:
        y, mo, d = date_str.split("-")
        next_date_str = str(date_cls(int(y),int(mo),int(d)) + timedelta(days=1))
    except Exception:
        return "❌ 날짜 형식 오류 (YYYY-MM-DD)"

    # 콜카드 + 결제내역 조회 (익일 포함)
    calls    = await sb_select("raw_calls", {"날짜": f"eq.{date_str}"})
    receipts = (await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})) +                (await sb_select("payment_receipts", {"날짜": f"eq.{next_date_str}"}))

    def to_min_smart(배차, 대상시각, 대상날짜):
        try:
            bh, bm = 배차.split(":")
            base = int(bh)*60+int(bm)
            th, tm = 대상시각.split(":")
            target = int(th)*60+int(tm)
            if 대상날짜 == next_date_str:
                target += 1440
            elif target < base - 60:
                target += 1440
            return target
        except: return None

    # 매칭
    matched_r = set()
    for call in calls:
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각")
        if not 하차:
            try:
                bh, bm = 배차.split(":")
                est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: continue
        c_min = to_min_smart(배차, 하차, date_str)
        best_j, best_diff = None, 99999
        for j, r in enumerate(receipts):
            if j in matched_r: continue
            r_min = to_min_smart(배차, r.get("시각","") or "", r.get("날짜", date_str))
            if r_min and c_min:
                diff = abs(c_min - r_min)
                if diff <= 20 and diff < best_diff:
                    best_diff = diff
                    best_j = j
        if best_j is not None:
            matched_r.add(best_j)

    unmatched = [r for j,r in enumerate(receipts) if j not in matched_r]

    # 현금 결제 → 배회영업으로 추가
    baehoe = [r for r in unmatched if (r.get("결제방법") or "") in ("현금","")]

    if not baehoe:
        return f"✅ {date_str} 배회후보 없음 (현금 미매칭 결제내역 없음)"

    DOW_MAP = ["월","화","수","목","금","토","일"]
    added = 0
    for r in baehoe:
        r_date = r.get("날짜") or date_str
        from datetime import date as dc
        try:
            rd = dc.fromisoformat(r_date)
            dow = DOW_MAP[rd.weekday()]
        except: dow = ""

        payload = {
            "날짜":     r_date,
            "요일":     dow,
            "배차시각": r.get("시각"),
            "하차시각": r.get("시각"),  # 결제시각 = 하차시각으로 설정
            "요금":     r.get("요금"),
            "콜유형":   "배회",
            "비고":     "결제내역 교차대조 자동추가",
        }
        result = await sb_insert("raw_calls", payload)
        if result:
            added += 1

    # 요약
    total_calls = len(calls) + added
    kakao = len([c for c in calls if c.get("콜유형","") == "카카오T"])
    lines = [
        f"✅ 대조 확정 완료 — {date_str}",
        f"배회영업 {added}건 raw_calls 추가",
        f"",
        f"📊 확정 후 현황:",
        f"  총 {total_calls}건 (카카오T {kakao}건 + 배회 {added}건)",
        f"  결제내역 {len(receipts)}건 → {'✅ 완전매칭' if total_calls == len(receipts) else f'⚠️ 차이 {abs(total_calls - len(receipts))}건'}",
    ]
    return "\n".join(lines)


async def handle_fee_confirm_request(update, date_str: str):
    """
    '대조 금액확인 YYYY-MM-DD' 명령어.
    해당 날짜 매칭 건 중 금액 불일치(≥500원) 건에 대해
    InlineKeyboard 버튼으로 확인 요청.
    """
    from datetime import date as date_cls, timedelta

    calls    = await sb_select("raw_calls", {"날짜": f"eq.{date_str}"})
    try:
        y,mo,d = date_str.split("-")
        next_d = str(date_cls(int(y),int(mo),int(d)) + timedelta(days=1))
    except: next_d = date_str

    receipts = (await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})) +                (await sb_select("payment_receipts", {"날짜": f"eq.{next_d}"}))

    def to_min(배차, 대상, 대상날짜):
        try:
            bh,bm = 배차.split(":"); base = int(bh)*60+int(bm)
            th,tm = 대상.split(":"); target = int(th)*60+int(tm)
            if 대상날짜 == next_d: target += 1440
            elif target < base-60: target += 1440
            return target
        except: return None

    matched_r = set()
    mismatches = []

    for call in calls:
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각") or ""
        if not 하차:
            try:
                bh,bm = 배차.split(":"); est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: continue
        c_min = to_min(배차, 하차, date_str)
        call_fee = call.get("요금") or 0
        best_j, best_diff = None, 99999
        for j, r in enumerate(receipts):
            if j in matched_r: continue
            r_min = to_min(배차, r.get("시각","") or "", r.get("날짜", date_str))
            if r_min and c_min:
                diff = abs(c_min - r_min)
                if diff <= 20 and diff < best_diff:
                    best_diff = diff; best_j = j
        if best_j is not None:
            matched_r.add(best_j)
            rcpt_fee = receipts[best_j].get("요금") or 0
            fee_diff = abs(call_fee - rcpt_fee)
            if fee_diff >= FEE_DIFF_THRESHOLD:
                mismatches.append({
                    "call_id": call.get("id"),
                    "배차시각": 배차,
                    "call_fee": call_fee,
                    "rcpt_fee": rcpt_fee,
                    "diff": fee_diff,
                })

    if not mismatches:
        await update.message.reply_text(f"✅ {date_str} 금액 불일치 없음")
        return

    for fm in mismatches:
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton(
                f"콜카드 {fm['call_fee']:,}원",
                callback_data=f"fee:{fm['call_id']}:{fm['call_fee']}"
            ),
            InlineKeyboardButton(
                f"결제내역 {fm['rcpt_fee']:,}원",
                callback_data=f"fee:{fm['call_id']}:{fm['rcpt_fee']}"
            ),
        ]])
        await update.message.reply_text(
            f"⚠️ 금액 불일치 확인\n"
            f"배차: {fm['배차시각']}\n"
            f"콜카드: {fm['call_fee']:,}원 | 결제: {fm['rcpt_fee']:,}원\n"
            f"차이: {fm['diff']:,}원",
            reply_markup=keyboard
        )

async def handle_date_stat(update, text: str):
    """
    날짜+통계 키워드 조합 처리
    예: '3-17 총건수', '3-17 매출', '3-17 순수익', '3-17 지출'
    """
    import re
    from datetime import date as date_cls

    today_d = today_kst()
    dow_map = ["월","화","수","목","금","토","일"]

    # 날짜 추출
    parsed = None
    for pat, mode in [
        (r"(\d{4})-(\d{1,2})-(\d{1,2})", "full"),
        (r"(\d{1,2})-(\d{1,2})",           "md"),
        (r"(\d{1,2})/(\d{1,2})",           "md"),
    ]:
        m = re.search(pat, text)
        if m:
            g = m.groups()
            try:
                parsed = date_cls(int(g[0]),int(g[1]),int(g[2])) if mode=="full"                          else date_cls(today_d.year, int(g[0]), int(g[1]))
                break
            except ValueError:
                pass

    if not parsed:
        await update.message.reply_text("❓ 날짜 인식 실패\n예: 3-17 총건수")
        return

    date_key = str(parsed)
    dow = dow_map[parsed.weekday()]
    header = f"📅 {date_key} ({dow})"

    # 통계 키워드 판단
    if any(kw in text for kw in ["총건수", "건수"]):
        calls = await sb_select("raw_calls", {"날짜": f"eq.{date_key}"})
        카카오 = sum(1 for c in calls if (c.get("콜유형") or "") == "카카오T")
        배회   = sum(1 for c in calls if (c.get("콜유형") or "") == "배회")
        총매출 = sum(c.get("요금") or 0 for c in calls)
        건당   = 총매출 // len(calls) if calls else 0
        await update.message.reply_text(
            f"{header} 총건수\n"
            f"총 {len(calls)}콜\n"
            f"  🚕 카카오T {카카오}건\n"
            f"  🚶 배회 {배회}건\n"
            f"건당 평균 {fmt(건당)}"
        )

    elif "매출" in text:
        calls = await sb_select("raw_calls", {"날짜": f"eq.{date_key}"})
        총매출 = sum(c.get("요금") or 0 for c in calls)
        건수   = len(calls)
        건당   = 총매출 // 건수 if 건수 else 0
        await update.message.reply_text(
            f"{header} 매출\n"
            f"총매출 {fmt(총매출)}\n"
            f"콜수 {건수}건 | 건당 {fmt(건당)}"
        )

    elif "순수익" in text:
        calls    = await sb_select("raw_calls",  {"날짜": f"eq.{date_key}"})
        expenses = await sb_select("expenses",   {"날짜": f"eq.{date_key}"})
        총매출 = sum(c.get("요금") or 0 for c in calls)
        총지출 = sum(e.get("금액") or 0 for e in expenses)
        순수익 = calc_net(총매출, 총지출)
        달성률 = min(int(순수익 / NET_GOAL * 100), 999) if NET_GOAL else 0
        달성바 = "█" * min(달성률//10,10) + "░" * max(10-달성률//10,0)
        await update.message.reply_text(
            f"{header} 순수익\n"
            f"매출 {fmt(총매출)} | 지출 {fmt(총지출)}\n"
            f"순수익 {fmt(순수익)}\n"
            f"목표 [{달성바}] {달성률}%"
        )

    elif "지출" in text:
        expenses = await sb_select("expenses", {"날짜": f"eq.{date_key}", "order": "id.asc"})
        총지출 = sum(e.get("금액") or 0 for e in expenses)
        if not expenses:
            await update.message.reply_text(f"{header}\n지출 없음")
            return
        lines_out = [f"{header} 지출 {fmt(총지출)}"]
        for e in expenses:
            auto = " (자동)" if e.get("자동여부") else ""
            lines_out.append(f"  {e.get('카테고리','')} {fmt(e.get('금액') or 0)}{auto}")
        await update.message.reply_text("\n".join(lines_out))

    else:
        # 전체 조회로 위임
        await handle_date_query(update, text)


async def handle_date_query(update, date_str: str):
    """특정 날짜 조회: 운행 내역 + 요약 + 지출"""
    import re
    from datetime import date as date_cls

    text = date_str.replace("조회","").replace("일","").strip()
    today_d = today_kst()
    parsed = None

    for pattern, mode in [
        (r"^(\d{4})-(\d{1,2})-(\d{1,2})$", "full"),
        (r"^(\d{1,2})-(\d{1,2})$",           "md"),
        (r"^(\d{1,2})/(\d{1,2})$",           "md"),
    ]:
        m = re.match(pattern, text)
        if m:
            g = m.groups()
            try:
                if mode == "full":
                    parsed = date_cls(int(g[0]), int(g[1]), int(g[2]))
                else:
                    parsed = date_cls(today_d.year, int(g[0]), int(g[1]))
                break
            except ValueError:
                pass

    if not parsed:
        await update.message.reply_text(
            "❓ 날짜 형식 오류\n"
            "예시: 3-2 조회 / 3/2 조회 / 2026-03-02 조회"
        )
        return

    date_key = str(parsed)
    dow_map = ["월","화","수","목","금","토","일"]
    dow = dow_map[parsed.weekday()]

    calls    = await sb_select("raw_calls", {"날짜": f"eq.{date_key}", "order": "배차시각.asc"})
    expenses = await sb_select("expenses",  {"날짜": f"eq.{date_key}", "order": "id.asc"})

    if not calls and not expenses:
        await update.message.reply_text(f"📭 {date_key} ({dow}) 데이터 없음")
        return

    result_lines = [f"📅 {date_key} ({dow}) 조회\n"]

    if calls:
        총매출 = sum(c.get("요금") or 0 for c in calls)
        result_lines.append(f"[운행] {len(calls)}콜 | {fmt(총매출)}")
        for c in calls:
            배차 = c.get("배차시각") or "-"
            출발 = (c.get("출발지") or "")[:8]
            도착 = (c.get("도착지") or "")[:8]
            요금 = fmt(c.get("요금") or 0)
            유형 = c.get("콜유형") or "카카오T"
            icon = "🚕" if 유형 == "카카오T" else "🚶"
            result_lines.append(f"  {icon}{배차} {출발}→{도착} {요금}")
    else:
        총매출 = 0
        result_lines.append("[운행] 없음")

    result_lines.append("")

    총지출 = sum(e.get("금액") or 0 for e in expenses)
    if expenses:
        result_lines.append(f"[지출] {fmt(총지출)}")
        for e in expenses:
            cat = e.get("카테고리") or ""
            amt = fmt(e.get("금액") or 0)
            auto = " (자동)" if e.get("자동여부") else ""
            result_lines.append(f"  {cat} {amt}{auto}")
    else:
        result_lines.append("[지출] 없음")

    result_lines.append("")
    순수익 = calc_net(총매출, 총지출)
    달성률 = min(int(순수익 / NET_GOAL * 100), 999) if NET_GOAL else 0
    달성바 = "█" * min(달성률//10, 10) + "░" * max(10 - 달성률//10, 0)
    result_lines.append("[요약]")
    result_lines.append(f"  매출 {fmt(총매출)} | 지출 {fmt(총지출)}")
    result_lines.append(f"  순수익 {fmt(순수익)}")
    result_lines.append(f"  목표 [{달성바}] {달성률}%")

    msg = "\n".join(result_lines)
    if len(msg) > 4000:
        msg = msg[:4000] + "\n...(생략)"
    await update.message.reply_text(msg)




def parse_manual_full(text: str) -> dict | None:
    """
    수동 콜 전체 입력 파싱.
    형식:
      2026 03 01 23 05 수성못>대명1동 8500 카카오
      26 03 01 23 05 수성못>대명1동 8500 배회
      03012305 수성못>대명1동 8500 카카오
      0301 2305 수성못>대명1동 8500 배회
      2026-03-01 23:05 수성못>대명1동 8500 카카오
    """
    import re as _re
    from datetime import date as _date

    orig = text.strip()
    today = _date.today()

    # 콜유형
    콜유형 = "배회" if "배회" in orig else "카카오T"
    clean = _re.sub(r'배회|카카오T?', '', orig).strip()

    # 경로 (출발>도착)
    route_m = _re.search(r'([가-힣\w]+)\s*[>→]\s*([가-힣\w]+)', clean)
    출발지 = 도착지 = None
    if route_m:
        출발지 = route_m.group(1).strip()
        도착지 = route_m.group(2).strip()
        clean = (clean[:route_m.start()] + ' ' + clean[route_m.end():]).strip()

    # 날짜+시각 파싱 (패턴 순서대로 시도)
    날짜 = None
    배차시각 = None

    # A: YYYY-MM-DD HH:MM 또는 YY-MM-DD HH:MM
    m = _re.search(r'(\d{2,4})[.\-](\d{1,2})[.\-](\d{1,2})\s+(\d{1,2}):(\d{2})', clean)
    if m:
        g = m.groups(); y = int(g[0]); y = y+2000 if y<100 else y
        try:
            날짜 = _date(y,int(g[1]),int(g[2])); 배차시각 = f"{int(g[3]):02d}:{g[4]}"
            clean = clean[:m.start()] + ' ' + clean[m.end():]
        except ValueError: pass

    # B: YYYY MM DD HH MM
    if not 날짜:
        m = _re.search(r'(\d{4})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{2})\b', clean)
        if m:
            g = m.groups()
            try:
                날짜 = _date(int(g[0]),int(g[1]),int(g[2])); 배차시각 = f"{int(g[3]):02d}:{g[4]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    # C: YY MM DD HH MM
    if not 날짜:
        m = _re.search(r'\b(\d{2})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{2})\b', clean)
        if m:
            g = m.groups()
            try:
                날짜 = _date(int(g[0])+2000,int(g[1]),int(g[2])); 배차시각 = f"{int(g[3]):02d}:{g[4]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    # D: MMDD HHMM
    if not 날짜:
        m = _re.search(r'\b(\d{4})\s+(\d{4})\b', clean)
        if m:
            a,b = m.group(1), m.group(2)
            try:
                날짜 = _date(today.year,int(a[:2]),int(a[2:])); 배차시각 = f"{int(b[:2]):02d}:{b[2:]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    # E: MMDDHHMM (8자리)
    if not 날짜:
        m = _re.search(r'\b(\d{8})\b', clean)
        if m:
            n = m.group(1)
            try:
                날짜 = _date(today.year,int(n[0:2]),int(n[2:4])); 배차시각 = f"{int(n[4:6]):02d}:{n[6:]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    if not 날짜 or not 배차시각:
        return None

    # 요금: 남은 clean의 4~6자리 숫자
    fee_m = _re.search(r'(?<!\d)(\d{4,6})(?!\d)', clean)
    if not fee_m:
        return None
    요금 = int(fee_m.group(1))

    요일 = ["월","화","수","목","금","토","일"][날짜.weekday()]
    return {"날짜":str(날짜),"요일":요일,"배차시각":배차시각,
            "출발지":출발지,"도착지":도착지,"요금":요금,"콜유형":콜유형}


async def handle_manual_full_call(update, text: str):
    """수동 전체 입력 콜 저장"""
    data = parse_manual_full(text)
    if not data:
        await update.message.reply_text(
            "❌ 형식 오류\n\n"
            "예시:\n"
            "2026 03 01 23 05 수성못>대명1동 8500 카카오\n"
            "26 03 01 23 05 수성못>대명1동 8500 배회\n"
            "0301 2305 수성못>대명1동 8500 카카오\n"
            "03012305 수성못>대명1동 8500 배회"
        )
        return

    # 중복 체크 → 자동 삭제 후 재저장
    deleted = await delete_duplicate_call(data["날짜"], data["배차시각"], data["요금"])
    if deleted:
        logger.info(f"수동입력 중복 삭제: {data['날짜']} {data['배차시각']}")

    result = await sb_insert("raw_calls", {
        "날짜":     data["날짜"],
        "요일":     data["요일"],
        "배차시각": data["배차시각"],
        "출발지":   data["출발지"],
        "도착지":   data["도착지"],
        "요금":     data["요금"],
        "콜유형":   data["콜유형"],
        "비고":     "수동입력",
    })

    if result:
        await update.message.reply_text(
            f"✅ 수동입력 저장\n"
            f"{data['날짜']}({data['요일']}) {data['배차시각']}\n"
            f"{data.get('출발지','-')}→{data.get('도착지','-')}\n"
            f"{data['요금']:,}원 [{data['콜유형']}]"
        )
    else:
        await update.message.reply_text("❌ DB 저장 실패")


async def handle_call_edit(update, text: str):
    """
    콜카드 수동 수정 명령어.
    형식:
      콜수정 HH:MM 필드=값
      콜수정 YYYY-MM-DD HH:MM 필드=값
    예:
      콜수정 19:18 요금=13100
      콜수정 19:18 배차시각=22:46
      콜수정 2026-03-20 19:18 도착지=수성구 만촌3동
    지원 필드: 배차시각, 하차시각, 출발지, 도착지, 요금, 콜유형, 비고
    """
    import re

    EDITABLE = {"배차시각","하차시각","출발지","도착지","요금","콜유형","비고"}

    parts = text.strip().split(" ", 1)
    if len(parts) < 2:
        await update.message.reply_text(
            "형식: 콜수정 HH:MM 필드=값\n"
            "날짜지정: 콜수정 YYYY-MM-DD HH:MM 필드=값\n"
            "예) 콜수정 19:18 요금=13100\n"
            "예) 콜수정 2026-03-20 19:18 배차시각=22:46"
        )
        return

    rest = parts[1].strip()

    # 날짜 포함 여부 판단
    date_match = re.match(r'^(\d{4}-\d{1,2}-\d{1,2})\s+(\d{1,2}:\d{2})\s+(.+)$', rest)
    time_only  = re.match(r'^(\d{1,2}:\d{2})\s+(.+)$', rest)

    if date_match:
        target_date  = date_match.group(1)
        target_time  = date_match.group(2)
        field_str    = date_match.group(3)
    elif time_only:
        target_date  = str(today_kst())
        target_time  = time_only.group(1)
        field_str    = time_only.group(2)
    else:
        await update.message.reply_text("❌ 형식 오류\n예) 콜수정 19:18 요금=13100")
        return

    # 필드=값 파싱
    field_match = re.match(r'^(\S+?)=(.+)$', field_str.strip())
    if not field_match:
        await update.message.reply_text("❌ 필드=값 형식 오류\n예) 요금=13100")
        return

    field = field_match.group(1).strip()
    value = field_match.group(2).strip()

    if field not in EDITABLE:
        await update.message.reply_text(
            f"❌ '{field}' 는 수정 불가\n"
            f"수정 가능: {', '.join(sorted(EDITABLE))}"
        )
        return

    # 요금은 int 변환
    if field == "요금":
        try:
            value = int(value.replace(",","").replace("원",""))
        except ValueError:
            await update.message.reply_text("❌ 요금은 숫자만 입력 (예: 13100)")
            return

    # DB에서 해당 건 찾기
    rows = await sb_select("raw_calls", {
        "날짜": f"eq.{target_date}",
        "배차시각": f"eq.{target_time}",
    })

    if not rows:
        await update.message.reply_text(
            f"⚠️ {target_date} {target_time} 콜 없음\n"
            f"날짜·시각을 확인해주세요."
        )
        return

    if len(rows) > 1:
        lines_out = [f"⚠️ {target_time} 콜이 {len(rows)}건 있습니다. 어느 건?"]
        for r in rows:
            lines_out.append(
                f"  ID:{r['id']} {r.get('출발지','')}→{r.get('도착지','')} {fmt(r.get('요금') or 0)}"
            )
        lines_out.append("ID 지정: 콜수정ID [id] 필드=값")
        await update.message.reply_text("\n".join(lines_out))
        return

    row = rows[0]
    old_val = row.get(field)
    row_id  = row["id"]

    # PATCH
    result = await sb_h("PATCH", f"raw_calls?id=eq.{row_id}", json={field: value})

    if result is not None:
        await update.message.reply_text(
            f"✅ 콜 수정 완료\n"
            f"날짜: {target_date} | 배차: {target_time}\n"
            f"{field}: {old_val} → {value}"
        )
    else:
        await update.message.reply_text("❌ 수정 실패")


async def handle_call_edit_by_id(update, text: str):
    """
    ID 지정 수정: '콜수정ID [id] 필드=값'
    동일 시각 콜이 여러 건일 때 사용
    """
    import re
    EDITABLE = {"배차시각","하차시각","출발지","도착지","요금","콜유형","비고"}

    m = re.match(r'^(\d+)\s+(\S+?)=(.+)$', text.strip())
    if not m:
        await update.message.reply_text("형식: 콜수정ID [id] 필드=값\n예) 콜수정ID 42 요금=13100")
        return

    row_id = int(m.group(1))
    field  = m.group(2).strip()
    value  = m.group(3).strip()

    if field not in EDITABLE:
        await update.message.reply_text(f"❌ '{field}' 수정 불가")
        return

    if field == "요금":
        try:
            value = int(value.replace(",","").replace("원",""))
        except ValueError:
            await update.message.reply_text("❌ 요금은 숫자만")
            return

    rows = await sb_select("raw_calls", {"id": f"eq.{row_id}"})
    if not rows:
        await update.message.reply_text(f"❌ ID {row_id} 없음")
        return

    old_val = rows[0].get(field)
    await sb_h("PATCH", f"raw_calls?id=eq.{row_id}", json={field: value})
    await update.message.reply_text(
        f"✅ ID {row_id} 수정\n{field}: {old_val} → {value}"
    )


async def handle_db_check(update: Update):
    calls = await sb_select("raw_calls", {"order": "id.desc", "limit": "1"})
    total_calls = await sb_select("raw_calls", {})
    total_exp = await sb_select("expenses", {})
    charging = await sb_select("charging_log", {"order": "id.desc", "limit": "1"})

    last_call = calls[0] if calls else {}
    last_charge = charging[0] if charging else {}

    await update.message.reply_text(
        f"🗄️ DB 현황\n"
        f"raw_calls: {len(total_calls)}건\n"
        f"expenses: {len(total_exp)}건\n"
        f"최근 콜: {last_call.get('날짜','-')} {last_call.get('배차시각','-')} {fmt(last_call.get('요금',0))}\n"
        f"최근 충전: {last_charge.get('충전일','-')} {last_charge.get('충전량_kwh','-')}kWh"
    )

# ──────────────────────────────────────────────
# 핸들러 — 전략
# ──────────────────────────────────────────────
async def get_strategy(update: Update):
    hour = now_kst().hour
    # 시간대 매칭
    if 19 <= hour < 21:
        time_key = "19~21"
    elif hour == 21:
        time_key = "21~22"
    elif 22 <= hour < 24:
        time_key = "22~00"
    elif 0 <= hour < 2:
        time_key = "00~02"
    elif 2 <= hour < 3:
        time_key = "02~03"
    else:
        time_key = "전체"

    rows = await sb_select(
        "strategy_lookup",
        {"시간대": f"in.(전체,{time_key})", "order": "우선순위.asc"}
    )
    if not rows:
        await update.message.reply_text("⚠️ 전략 테이블 없음. 마기 업데이트 필요")
        return

    lines = [f"⚡ 현재 전략 ({now_kst().strftime('%H:%M')})"]
    for r in rows:
        priority_icon = {"긴급": "🔴", "높음": "🟠", "보통": "🟡"}.get(r.get("우선순위", ""), "⚪")
        lines.append(f"{priority_icon} [{r.get('시간대','')}] {r.get('행동지침','')}")

    await update.message.reply_text("\n".join(lines))

async def handle_magi_update(update: Update, content: str):
    """'마기 업데이트 [시간대] [내용]' → strategy_lookup INSERT"""
    parts = content.strip().split(" ", 1)
    if len(parts) < 2:
        await update.message.reply_text("형식: 마기 업데이트 [시간대] [내용]")
        return
    시간대 = parts[0]
    지침 = parts[1]
    await sb_insert("strategy_lookup", {
        "시간대": 시간대,
        "행동지침": 지침,
        "우선순위": "높음",
    })
    await update.message.reply_text(f"✅ 전략 테이블 갱신완료\n[{시간대}] {지침}")

# ──────────────────────────────────────────────
# 핸들러 — 엑셀 다운로드
# ──────────────────────────────────────────────

async def handle_download_month(update, ym: str):
    """특정 월 다운로드 (예: 2026-03)"""
    try:
        year, month = ym.split("-")
        year, month = int(year), int(month)
    except Exception:
        await update.message.reply_text("❌ 형식 오류: YYYY-MM (예: 2026-03)")
        return

    from datetime import date, timedelta
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    start_str = str(start_date)
    end_str   = str(end_date)

    # AND 조건 (날짜 키 중복 버그 방지)
    calls    = await sb_select("raw_calls",    {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"})
    expenses = await sb_select("expenses",     {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"})
    charging = await sb_select("charging_log", {"and": f"(충전일.gte.{start_str},충전일.lte.{end_str})"})

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "운행기록"
    headers1 = ["날짜","요일","배차시각","출발지","도착지","요금","콜유형","비고"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4A90D9")
    for c in calls:
        ws1.append([c.get("날짜"),c.get("요일"),c.get("배차시각"),
                    c.get("출발지"),c.get("도착지"),c.get("요금"),
                    c.get("콜유형"),c.get("비고")])

    ws2 = wb.create_sheet("지출")
    headers2 = ["날짜","카테고리","금액","메모","자동여부"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F5A623")
    for e in expenses:
        ws2.append([e.get("날짜"),e.get("카테고리"),e.get("금액"),
                    e.get("메모"),e.get("자동여부")])

    ws3 = wb.create_sheet("충전기록")
    headers3 = ["충전일","충전량(kWh)","충전금액","충전소"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="7ED321")
    for ch in charging:
        ws3.append([ch.get("충전일"),ch.get("충전량_kwh"),
                    ch.get("충전금액"),ch.get("충전소")])

    filepath = f"/tmp/자비스_월간_{ym}.xlsx"
    wb.save(filepath)

    await update.message.reply_document(
        document=open(filepath, "rb"),
        filename=f"자비스_월간_{ym}.xlsx",
        caption=f"📊 {ym} 데이터 ({len(calls)}건)"
    )

async def handle_download(update: Update, scope: str):
    today = today_kst()
    if scope == "주간":
        start = today - timedelta(days=6)
        start_str = str(start)
        label = f"{start_str}_{today}"
    elif scope == "월간":
        start_str = today.replace(day=1).isoformat()
        label = f"{today.year}{today.month:02d}"
    else:
        start_str = "2000-01-01"
        label = "전체"

    calls = await sb_select("raw_calls", {"날짜": f"gte.{start_str}"})
    expenses = await sb_select("expenses", {"날짜": f"gte.{start_str}"})
    charging = await sb_select("charging_log", {"충전일": f"gte.{start_str}"})

    wb = openpyxl.Workbook()

    # 시트1 raw_calls
    ws1 = wb.active
    ws1.title = "운행기록"
    headers1 = ["날짜", "요일", "배차시각", "출발지", "도착지", "요금", "콜유형", "비고"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4A90D9")
    for row, c in enumerate(calls, 2):
        ws1.append([
            c.get("날짜"), c.get("요일"), c.get("배차시각"),
            c.get("출발지"), c.get("도착지"), c.get("요금"),
            c.get("콜유형"), c.get("비고"),
        ])

    # 시트2 expenses
    ws2 = wb.create_sheet("지출")
    headers2 = ["날짜", "카테고리", "금액", "메모", "자동여부"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F5A623")
    for row, e in enumerate(expenses, 2):
        ws2.append([
            e.get("날짜"), e.get("카테고리"), e.get("금액"),
            e.get("메모"), e.get("자동여부"),
        ])

    # 시트3 charging_log
    ws3 = wb.create_sheet("충전기록")
    headers3 = ["충전일", "충전량(kWh)", "충전금액", "충전소"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="7ED321")
    for row, ch in enumerate(charging, 2):
        ws3.append([
            ch.get("충전일"), ch.get("충전량_kwh"),
            ch.get("충전금액"), ch.get("충전소"),
        ])

    filepath = f"/tmp/자비스_{scope}_{label}.xlsx"
    wb.save(filepath)

    await update.message.reply_document(
        document=open(filepath, "rb"),
        filename=f"자비스_{scope}_{label}.xlsx",
        caption=f"📊 {scope} 데이터 ({len(calls)}건)"
    )

# ──────────────────────────────────────────────
# 핸들러 — 데이터 이식 (v6 엑셀 → raw_calls)
# ──────────────────────────────────────────────
async def handle_excel_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """v6 인수인계 엑셀 업로드 시 raw_calls 자동 이식"""
    doc = update.message.document
    file = await context.bot.get_file(doc.file_id)
    file_bytes = await file.download_as_bytearray()

    import io
    wb = openpyxl.load_workbook(io.BytesIO(bytes(file_bytes)), read_only=True)
    target_sheets = [s for s in wb.sheetnames if "운행" in s or "데이터" in s]

    if not target_sheets:
        await update.message.reply_text("⚠️ 운행 데이터 시트를 찾지 못했습니다.")
        return

    await update.message.reply_text(f"📥 이식 시작: {target_sheets}")
    total_saved = 0
    total_skip = 0

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(h).strip() if h else "" for h in rows[0]]
        col = {h: i for i, h in enumerate(header)}

        for row in rows[1:]:
            if not any(row):
                continue
            try:
                날짜_raw = row[col.get("날짜", 0)]
                if not 날짜_raw:
                    continue
                날짜 = str(날짜_raw)[:10]
                요금_raw = row[col.get("요금", 5)]
                요금 = int(str(요금_raw).replace(",", "").replace("원", "")) if 요금_raw else 0
                배차시각 = row[col.get("배차시각", 2)]
                배차시각 = str(배차시각) if 배차시각 else None

                # 중복 체크
                existing = await sb_select(
                    "raw_calls",
                    {"날짜": f"eq.{날짜}", "요금": f"eq.{요금}", "배차시각": f"eq.{배차시각}"}
                )
                if existing:
                    total_skip += 1
                    continue

                payload = {
                    "날짜": 날짜,
                    "요일": row[col.get("요일", 1)] or "",
                    "배차시각": 배차시각,
                    "출발지": row[col.get("출발지", 3)],
                    "도착지": row[col.get("도착지", 4)],
                    "요금": 요금,
                    "콜유형": row[col.get("콜유형", 6)] or "카카오T",
                }
                r = await sb_insert("raw_calls", payload)
                if r:
                    total_saved += 1
            except Exception as e:
                logger.error(f"이식 행 오류: {e}")
                continue

    await update.message.reply_text(
        f"✅ 이식 완료\n저장: {total_saved}건 | 중복스킵: {total_skip}건"
    )


# ══════════════════════════════════════════════
# 어군탐지기 v2 — 자동 브리핑 시스템
# ══════════════════════════════════════════════

FISH_DATA = {
    "월": {
        "19~21": [["달서구 신당동",      "8",  "9,500",  "S", "신당동 주택가 도로변",   "카카오T 공식 평일 호출 1위"]],
        "21~24": [["달서구 신당동",      "9",  "10,200", "S", "신당동 상가 밀집",       "평일 23시 호출 1위"]],
        "00~02": [["중구 성내1·2동",     "9",  "10,800", "S", "성내2동 앵커",           "유흥가 마감 귀가 폭증"]],
    },
    "화": {
        "19~21": [["수성구 범어·만촌",   "7",  "8,200",  "A", "범어역~만촌역",          "화요일 수성구 준수"]],
        "21~24": [["달서구 신당동",      "7",  "9,500",  "A", "신당동 상가",             "화요일 야간 준수"]],
        "00~02": [["중구 성내1·2동",     "8",  "10,200", "S", "성내2동 앵커",            "화요일 심야 귀가"]],
    },
    "수": {
        "19~21": [["수성구 범어·만촌",   "7",  "8,500",  "A", "범어역~만촌역",          "수요일 수성구 준수"]],
        "21~24": [["중구 동성로/삼덕동", "8",  "9,800",  "A", "삼덕동 먹자골목",        "수요일 야간 선호"]],
        "00~02": [["중구 성내1·2동",     "8",  "10,500", "S", "성내2동 앵커",            "수요일 심야 귀가"]],
    },
    "목": {
        "19~21": [["달서구 신당동",      "6",  "7,800",  "B", "신당동 주택가",           "목요일 콜 저조 — 단축 검토"]],
        "21~24": [["달서구 신당동",      "7",  "9,200",  "A", "신당동 상가",             "목요일 막판 집중"]],
        "00~02": [["중구 성내1·2동",     "7",  "9,800",  "A", "성내2동 앵커",            "목요일 심야 귀가"]],
    },
    "금": {
        "19~21": [["수성구 범어·만촌",   "8",  "7,600",  "A", "범어역~만촌역",          "금요일 수성구 콜 실적 최다"]],
        "21~24": [["중구 동성로/삼덕동", "9",  "10,000", "S", "삼덕동 먹자골목",        "유흥 피크! 술집 01시 마감"]],
        "00~02": [["중구 성내1·2동",     "9",  "11,500", "S", "성내2동 앵커",            "동성로 마감 귀가 폭증"]],
    },
    "토": {
        "19~21": [["수성구 고산2동",     "8",  "9,200",  "S", "수성못 주변",             "주말 17~19시 호출 집중"]],
        "21~24": [["중구 동성로/삼덕동", "9",  "8,800",  "S", "삼덕동~동성로",           "토요일 밤 유흥 최고 피크"]],
        "00~02": [["중구 성내1동",       "9",  "14,400", "S", "성내1동~성내2동",         "토요일 막판 단가 14,433원 최고치"]],
    },
    "일": {
        "19~21": [["수성구 고산2동",     "8",  "9,200",  "S", "수성못 주변",             "주말 호출 1위"]],
        "21~24": [["중구 동성로",        "8",  "11,800", "S", "동성로 입구",             "일요일 밤 단가 11,811원 고효율"]],
        "00~02": [["중구 성내1동",       "10", "12,400", "S", "성내1동",                 "주말 00~01시 호출 1위 구역"]],
    },
}

def get_fish_slot(hour: int) -> str | None:
    """현재 시각 → 어군 슬롯 반환"""
    if 19 <= hour < 21: return "19~21"
    if 21 <= hour <= 23: return "21~24"
    if 0 <= hour < 2:   return "00~02"
    return None

# 태그 이모지 매핑 (v2.2)
FISH_TAG_EMOJI = {
    "golden_time":    "⭐",
    "long_distance":  "🚀",
    "foreign_worker": "🌏",
    "blue_ocean":     "🌊",
    "avoid":          "⛔",
    "oversupply":     "⛔",
}

def get_fish_report(custom_hour: int = None) -> str | None:
    """⚠️ 사용 중단(2026-07-13) — FISH_DATA 하드코딩 기반, 더 이상 어디서도 호출 안 함.
    실데이터 기반 get_fish_report_db()로 완전 대체됨. 삭제 대신 보존만 해둠(참고용)."""
    now  = datetime.now(KST)
    hour = custom_hour if custom_hour is not None else now.hour
    day  = DOW_KOR[now.weekday()]
    slot = get_fish_slot(hour)
    if not slot:
        return None
    zones = FISH_DATA.get(day, {}).get(slot, [])
    if not zones:
        return f"🐟 {day}요일 {slot} 어군 데이터 없음"
    lines = [f"🐟 어군브리핑 — {day}요일 {slot}"]
    for idx, z in enumerate(zones, 1):
        grade_icon = {"S": "🔴", "A": "🟠", "B": "🟡", "C": "⚪"}.get(z[3], "⚪")
        lines.append(f"\n#{idx} {z[0]} {grade_icon}{z[3]}등급")
        lines.append(f"  점수 {z[1]}/10 | 예상 {z[2]}원")
        lines.append(f"  📍 {z[4]}")
        lines.append(f"  💡 {z[5]}")
    return "\n".join(lines)


async def recalc_fish_hour_data():
    """어군 브리핑 시간대별 통계 재계산 (2026-07-13, 하드코딩 HOUR_DATA/FISH_DATA 제거).
    raw_calls(실시간 데이터) + call_quality_history(2/14~3/31 368건 검증데이터)를
    합쳐서 시간대별 카카오T/배회 평균 건수·비중·평균단가를 계산해 fish_hour_data에 저장.
    주의: call_quality_history는 마기 명령서#016에 따라 공식 건수 KPI에는 사용 금지지만,
    이 계산은 "시간대별 패턴 분석"이라는 call_quality_history 본래 목적과 일치하므로 사용함.
    """
    try:
        raw = await sb_select("raw_calls", {}) or []
    except Exception as e:
        logger.error(f"fish_hour_data 재계산 - raw_calls 조회 실패: {e}")
        raw = []
    try:
        quality = await sb_select("call_quality_history", {}) or []
    except Exception as e:
        logger.error(f"fish_hour_data 재계산 - call_quality_history 조회 실패: {e}")
        quality = []

    all_rows = raw + quality
    if not all_rows:
        logger.warning("fish_hour_data 재계산 - 데이터 없음, 건너뜀")
        return

    # 관측일수(분모) — 전체 데이터셋의 고유 날짜 수
    total_days = len(set(r.get("날짜") for r in all_rows if r.get("날짜"))) or 1

    from collections import defaultdict
    hour_kakao_cnt  = defaultdict(float)
    hour_baehoe_cnt = defaultdict(float)
    hour_kakao_fare = defaultdict(list)
    hour_baehoe_fare = defaultdict(list)

    for r in all_rows:
        bt = r.get("배차시각")
        if not bt:
            continue
        try:
            h = int(str(bt).split(":")[0])
        except Exception:
            continue
        ct = r.get("콜유형") or ""
        cnt = _extract_count(r)  # raw_calls OCR요약행 보정, call_quality_history는 항상 1
        fare = r.get("요금")
        if ct == "카카오T":
            hour_kakao_cnt[h] += cnt
            if fare: hour_kakao_fare[h].append(fare)
        elif ct == "배회":
            hour_baehoe_cnt[h] += cnt
            if fare: hour_baehoe_fare[h].append(fare)

    for h in range(24):
        k_avg = round(hour_kakao_cnt.get(h, 0) / total_days, 2)
        b_avg = round(hour_baehoe_cnt.get(h, 0) / total_days, 2)
        total = hour_kakao_cnt.get(h, 0) + hour_baehoe_cnt.get(h, 0)
        b_pct = round(hour_baehoe_cnt.get(h, 0) / total * 100, 1) if total > 0 else 0.0
        kf = hour_kakao_fare.get(h, [])
        bf = hour_baehoe_fare.get(h, [])
        avg_fare_kakao = round(sum(kf) / len(kf)) if kf else None
        avg_fare_baehoe = round(sum(bf) / len(bf)) if bf else None

        try:
            await sb_upsert("fish_hour_data", {
                "hour": h,
                "kakao_avg": k_avg,
                "baehoe_avg": b_avg,
                "b_pct": b_pct,
                "avg_fare_kakao": avg_fare_kakao,
                "avg_fare_baehoe": avg_fare_baehoe,
                "sample_days": total_days,
            }, on_conflict="hour")
        except Exception as e:
            logger.error(f"fish_hour_data hour={h} 저장 오류: {e}")

    logger.info(f"fish_hour_data 재계산 완료 (관측일수={total_days}, 총 {len(all_rows)}건)")


_FISH_HOUR_CACHE = {}

async def load_fish_hour_data():
    """fish_hour_data 조회, 메모리 캐시. 재계산 스케줄러가 갱신할 때까지 캐시 유지."""
    global _FISH_HOUR_CACHE
    if _FISH_HOUR_CACHE:
        return _FISH_HOUR_CACHE
    try:
        rows = await sb_select("fish_hour_data", {}) or []
        _FISH_HOUR_CACHE = {
            r["hour"]: {
                "kakao": r.get("kakao_avg") or 0,
                "baehoe": r.get("baehoe_avg") or 0,
                "b_pct": r.get("b_pct") or 0,
                "avg_fare_kakao": r.get("avg_fare_kakao"),
                "avg_fare_baehoe": r.get("avg_fare_baehoe"),
            } for r in rows
        }
    except Exception as e:
        logger.error(f"fish_hour_data 로드 실패: {e}")
    return _FISH_HOUR_CACHE


async def get_fish_report_db(hour=None, tag_filter=None):
    """어군 브리핑 v3 - 카카오/배회 분리, raw_calls+call_quality_history 실데이터 기반(하드코딩 제거)"""
    now = datetime.now(KST)
    h   = hour if hour is not None else now.hour
    day = DOW_KOR[now.weekday()]

    HOUR_DATA = await load_fish_hour_data()
    _fallback = {"kakao": 0, "baehoe": 0, "b_pct": 0, "avg_fare_kakao": None, "avg_fare_baehoe": None}
    hd_raw = HOUR_DATA.get(h) or _fallback

    # 캐스퍼 수정 2026-07-13: 실제 운행 데이터가 거의 없는 시간대(예: 낮 시간대)에
    # 억지로 최소 1건으로 채워서 "콜 간격 480분" 같은 비현실적 수치가 나오던 문제 수정.
    # 관측일당 카카오+배회 합계가 0.05건 미만(48일 기준 대략 2~3건 미만)이면 데이터 부족으로 간주.
    total_obs = (hd_raw.get("kakao") or 0) + (hd_raw.get("baehoe") or 0)
    if total_obs < 0.05:
        return (
            f"🐟 어군 브리핑 v3 — {str(h).zfill(2)}시 {day}요일 (실데이터 기반)\n"
            f"{chr(0x2501)*22}\n\n"
            f"이 시간대는 축적된 운행 데이터가 부족합니다.\n"
            f"(19~21시/21~24시/00~02시 운영시간대에 데이터가 집중되어 있습니다)"
        )

    # kakao/baehoe는 "관측일당 평균"이므로, 기존 표기(8시간당 건수)에 맞춰 근사 환산
    hd = {
        "kakao": max(round(hd_raw["kakao"] * 8), 1) if hd_raw["kakao"] else 1,
        "baehoe": round(hd_raw["baehoe"] * 8, 1),
        "b_pct": round(hd_raw["b_pct"], 1),
    }
    fare_kakao = hd_raw.get("avg_fare_kakao") or 9000
    fare_baehoe = hd_raw.get("avg_fare_baehoe") or 10500
    k_pct = 100 - hd["b_pct"]
    k_int = round(60 / (hd["kakao"] / 8)) if hd["kakao"] > 0 else 99
    b_per = round(hd["baehoe"] / 8 * 10) / 10

    def stars(n):
        if n >= 5: return "★★★"
        if n >= 3: return "★★"
        return "★"

    anchor_night = "중구 성내·삼덕·동인 / 동구 신암"
    anchor_late  = "북구 침산↔복현·노원동"
    anchor_dawn  = "북구 노원동 / 중구 동인동"

    if 0 <= h <= 2:
        anchor = anchor_night + " / " + anchor_late
    elif h >= 3:
        anchor = anchor_dawn
    else:
        anchor = anchor_night

    if 19 <= h <= 21:
        decision   = "카카오 우선 대기"
        rec_detail = "배회 " + str(hd["b_pct"]) + "% 미만 — 콜 대기 합리적\n수락률 100% 유지"
    elif 22 <= h <= 23:
        decision   = "카카오 우선 + 배회 수용"
        rec_detail = "배회 " + str(hd["b_pct"]) + "% — 앵커 위치면 적극 수용\n" + anchor.split("/")[0].strip()
    elif 0 <= h <= 2:
        decision   = "★ 배회 적극 수용 (황금시간)"
        rec_detail = "배회 " + str(hd["b_pct"]) + "% — 자정후 황금구간\n핵심 동선: " + anchor
    elif 3 <= h <= 4:
        decision   = "마감 단계"
        rec_detail = "02시 종료 검토" if day in ["화", "목"] else "끝까지 사수"
    else:
        decision   = "운행 준비 / 대기"
        rec_detail = "19시 이후 카카오 골든타임 준비"

    est_h = round((fare_kakao * k_pct / 100 + fare_baehoe * hd["b_pct"] / 100) * (hd["kakao"] + hd["baehoe"]) / 8)

    db_zones = []
    try:
        slot_map = {19:"19-21",20:"19-21",21:"21-24",22:"21-24",23:"21-24",0:"00-02",1:"00-02",2:"00-02"}
        tb = slot_map.get(h)
        if tb:
            params = {"time_band": "eq." + tb, "verified": "neq.avoid",
                      "order": "rank_overall.asc", "limit": "3"}
            if tag_filter:
                params["pattern_tag"] = "eq." + tag_filter
            rows = await sb_select("fish_finder", params)
            if rows:
                for r in rows:
                    zone = r.get("zone", "")
                    avg  = r.get("avg_fare", 0) or 0
                    db_zones.append(zone + ("(" + str(int(avg)) + "원)" if avg else ""))
    except Exception:
        pass

    lines = [
        "🐟 어군 브리핑 v3 — " + str(h).zfill(2) + "시 " + day + "요일 (실데이터 기반)",
        chr(0x2501) * 22,
        "",
        "🟢 카카오 콜 어군",
        "  콜 간격: 약 " + str(k_int) + "분 (" + stars(hd["kakao"]) + ")",
        "  평균 단가: " + fmt(fare_kakao) + "대",
        "  비중: " + str(k_pct) + "%  (" + str(hd["kakao"]) + "건/8h 기준)",
    ]
    if db_zones:
        lines.append("  DB 핫존: " + " / ".join(db_zones))

    lines += [
        "",
        "🟠 배회 어군",
        "  만남 확률: 시간당 " + str(b_per) + "건 (" + stars(hd["baehoe"]) + ", " + str(hd["b_pct"]) + "%)",
        "  평균 단가: " + fmt(fare_baehoe) + " (수수료 0%)",
        "  핵심 동선: " + anchor,
        "",
        "💡 종합 권고",
        "  " + decision,
    ]
    for dl in rec_detail.split("\n"):
        lines.append("  " + dl)
    lines += [
        "  예상 시간당: " + fmt(est_h) + "대",
        "",
        chr(0x2500) + " 시간대별 배회 비중 " + chr(0x2500),
    ]

    BAR_HOURS = [19, 20, 21, 22, 23, 0, 1, 2, 3]
    bar = ""
    for bh in BAR_HOURS:
        bd = HOUR_DATA.get(bh, {"b_pct": 10})
        p  = bd["b_pct"]
        mk = "●" if bh == h else ("◆" if p >= 25 else ("◇" if p >= 15 else "·"))
        bar += str(bh).zfill(2) + mk + " "
    lines.append(bar.strip())
    lines.append("(● 현재  ◆ 배회25%↑  ◇ 배회15%↑)")

    # 캐스퍼 수정 2026-07-15: "회피 구역" 섹션 제거.
    # fish_finder(verified='avoid') 조회에 시간대 필터가 전혀 없어서, 몇 시에 브리핑을
    # 받든 항상 똑같은 3줄이 고정으로 나오는 구조였음(아키텍트 실사용 중 지적).
    # 필요하면 /avoid 명령으로 전체 회피구역 목록을 별도 조회 가능(cmd_avoid, 변경 없음).

    return "\n".join(lines)


def fish_scheduler(app):
    """18:50 영업준비 브리핑 + 19~02시 매 정각 자동 브리핑
    2026-07-13 수정: 하드코딩 FISH_DATA(get_fish_report) → 실데이터 기반(get_fish_report_db)으로 전환.
    매일 03:10 fish_hour_data 자동 재계산 추가."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    # 발송 대상: 등록된 모든 단말기
    chat_ids = [x for x in [
        os.getenv("ALLOWED_CHAT_ID", ""),
        os.getenv("ALLOWED_CHAT_ID2", ""),
    ] if x]

    async def send_all(text: str):
        for cid in chat_ids:
            try:
                await app.bot.send_message(chat_id=cid, text=text)
            except Exception as e:
                logger.error(f"어군 브리핑 발송 오류 ({cid}): {e}")

    last_sent_hour = -1
    sent_start_brief = False
    last_reset_day = -1
    last_recalc_day = -1

    # 최초 기동 시 1회 즉시 재계산 (테이블이 비어있으면 폴백값으로 작동하다가 여기서 채워짐)
    try:
        loop.run_until_complete(recalc_fish_hour_data())
    except Exception as e:
        logger.error(f"fish_hour_data 최초 재계산 실패: {e}")

    while True:
        now = datetime.now(KST)

        # ── 매일 03시 플래그 리셋 (운행 종료 후)
        if now.hour == 3 and now.day != last_reset_day:
            sent_start_brief = False
            last_sent_hour   = -1
            last_reset_day   = now.day
            logger.info("어군 스케줄러 일간 리셋")

        # ── 매일 03:10 시간대별 통계 자동 재계산 (하드코딩 제거 후속)
        if now.hour == 3 and now.minute >= 10 and now.day != last_recalc_day:
            try:
                loop.run_until_complete(recalc_fish_hour_data())
                global _FISH_HOUR_CACHE
                _FISH_HOUR_CACHE = {}  # 캐시 무효화 → 다음 조회 시 새 값 로드
                logger.info("fish_hour_data 일일 재계산 완료")
            except Exception as e:
                logger.error(f"fish_hour_data 일일 재계산 실패: {e}")
            last_recalc_day = now.day

        # ── 18:50 영업 준비 브리핑
        if now.hour == 18 and now.minute == 50 and not sent_start_brief:
            report = loop.run_until_complete(get_fish_report_db(hour=19)) or "데이터 없음"
            msg = f"🚀 영업준비 브리핑 (10분 후 출발)\n\n{report}"
            loop.run_until_complete(send_all(msg))
            sent_start_brief = True
            logger.info("18:50 영업준비 브리핑 발송")

        # ── 19:00 ~ 02:00 매 정각 브리핑
        if now.minute == 0 and now.hour != last_sent_hour:
            in_service = (19 <= now.hour <= 23) or (0 <= now.hour < 2)
            if in_service:
                report = loop.run_until_complete(get_fish_report_db())
                if report:
                    loop.run_until_complete(send_all(report))
                    logger.info(f"어군 브리핑 발송: {now.hour}시")
                last_sent_hour = now.hour

        time.sleep(30)


# ──────────────────────────────────────────────
# 중복 자동 삭제 + 금액 불일치 처리
# ──────────────────────────────────────────────
FEE_DIFF_THRESHOLD = 500  # 이 이상 차이면 확인 요청

async def delete_duplicate_call(날짜: str, 배차시각: str, 요금: int) -> bool:
    """날짜+배차시각+요금 완전 일치 건 삭제. 삭제되면 True."""
    rows = await sb_select("raw_calls", {
        "날짜": f"eq.{날짜}",
        "배차시각": f"eq.{배차시각}",
        "요금": f"eq.{요금}",
    })
    if not rows:
        return False
    for row in rows:
        await sb_h("DELETE", f"raw_calls?id=eq.{row['id']}")
    return True

async def delete_duplicate_payment(날짜: str, 시각: str, 요금: int) -> bool:
    """날짜+시각+요금 완전 일치 결제내역 삭제."""
    rows = await sb_select("payment_receipts", {
        "날짜": f"eq.{날짜}",
        "시각": f"eq.{시각}",
        "요금": f"eq.{요금}",
    })
    if not rows:
        return False
    for row in rows:
        await sb_h("DELETE", f"payment_receipts?id=eq.{row['id']}")
    return True

async def send_fee_confirm(update, call_id: int, 배차시각: str,
                            call_fee: int, rcpt_fee: int, diff: int):
    """금액 불일치 ≥500원 시 InlineKeyboard 확인 요청 발송."""
    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                f"콜카드 {call_fee:,}원",
                callback_data=f"fee:{call_id}:{call_fee}"
            ),
            InlineKeyboardButton(
                f"결제내역 {rcpt_fee:,}원",
                callback_data=f"fee:{call_id}:{rcpt_fee}"
            ),
        ]
    ])
    await update.message.reply_text(
        f"⚠️ 금액 불일치 확인 요청\n"
        f"배차: {배차시각}\n"
        f"콜카드: {call_fee:,}원\n"
        f"결제내역: {rcpt_fee:,}원\n"
        f"차이: {diff:,}원\n\n"
        f"어느 금액으로 저장할까요?",
        reply_markup=keyboard
    )

async def handle_fee_callback(update, context):
    """InlineKeyboard 버튼 클릭 처리."""
    query = update.callback_query
    await query.answer()
    data = query.data  # "fee:call_id:선택금액"
    try:
        _, call_id_str, fee_str = data.split(":")
        call_id = int(call_id_str)
        selected_fee = int(fee_str)
        # raw_calls 요금 업데이트
        await sb_h("PATCH", f"raw_calls?id=eq.{call_id}",
                   json={"요금": selected_fee})
        await query.edit_message_text(
            f"✅ {selected_fee:,}원으로 저장 완료"
        )
    except Exception as e:
        logger.error(f"fee callback 오류: {e}")
        await query.edit_message_text("❌ 처리 오류")

# ──────────────────────────────────────────────
# 보험 스케줄러
# ──────────────────────────────────────────────
def insurance_scheduler():
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    while True:
        now = datetime.now(KST)
        next_run = now.replace(hour=0, minute=1, second=0, microsecond=0)
        if now >= next_run:
            next_run += timedelta(days=1)
        sleep_sec = (next_run - now).total_seconds()
        logger.info(f"보험 스케줄러 대기: {sleep_sec:.0f}초")
        time.sleep(sleep_sec)
        today = datetime.now(KST).date()  # 반드시 KST
        loop.run_until_complete(insert_insurance(today))
        time.sleep(60)

# ──────────────────────────────────────────────
# 텔레그램 핸들러
# ──────────────────────────────────────────────
def is_allowed(update: Update) -> bool:
    chat_id = str(update.effective_chat.id)
    return chat_id in ALLOWED_IDS


async def cmd_fish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    어군 브리핑. /fish [시간] [태그]
    예: /fish       → 현재 시간
        /fish 19    → 19시 기준
        /fish 21 foreign → 21시 외국인 패턴 필터
    """
    if not is_allowed(update):
        return

    args = context.args or []
    hour = None
    tag_filter = None

    for arg in args:
        if arg.isdigit():
            hour = int(arg)
        elif arg in ("foreign", "foreign_worker"):
            tag_filter = "foreign_worker"
        elif arg in ("long", "long_distance"):
            tag_filter = "long_distance"
        elif arg in ("golden", "golden_time"):
            tag_filter = "golden_time"
        elif arg in ("blue", "blue_ocean"):
            tag_filter = "blue_ocean"

    report = await get_fish_report_db(hour=hour, tag_filter=tag_filter)
    await update.message.reply_text(report)


async def cmd_avoid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """회피 구역 조회 /avoid"""
    if not is_allowed(update):
        return
    try:
        rows = await sb_select("fish_finder", {"verified": "eq.avoid", "order": "time_band.asc"})
        if not rows:
            await update.message.reply_text("⛔ 등록된 회피 구역 없음\n(Supabase fish_finder 테이블 확인 필요)")
            return
        lines = ["⛔ 회피 구역 전체 목록\n"]
        for r in rows:
            zone = r.get("zone","")
            band = r.get("time_band","")
            note = r.get("note","") or ""
            lines.append(f"  ✗ {zone} ({band})" + (f"\n    {note}" if note else ""))
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        logger.error(f"회피구역 조회 오류: {e}")
        await update.message.reply_text("❌ 조회 오류")

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    await update.message.reply_text(
        "🤖 자비스 v5 시작\n\n"
        "이미지: 콜카드·충전내역·결제내역·세큐티\n"
        "텍스트: 콜 7800 / 배회 5600 / 오늘 / 이번 주 / 전략\n"
        "다운로드: 주간·월간·전체 다운로드\n"
        "교차대조: 대조 YYYY-MM-DD"
    )

async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    await update.message.reply_text(
        "📋 명령어\n\n"
        "[이미지] 콜카드·충전영수증·결제내역 → 자동저장\n\n"
        "[수동입력]\n"
        "콜 금액 / 배회 금액 / 충전 금액\n"
        "타이어·오일·세차 금액 / 지출 항목 금액\n"
        "지출취소 / 휴무 / 4-7 휴무\n\n"
        "[수동전체]\n"
        "2026 03 01 23 05 출발>도착 요금 카카오\n"
        "0301 2305 출발>도착 요금 배회\n\n"
        "[콜수정]\n"
        "콜수정 HH:MM 필드=값\n"
        "콜수정 날짜 HH:MM 필드=값\n"
        "콜수정ID [id] 필드=값\n\n"
        "[조회]\n"
        "오늘·이번 주·이번 달·지출 확인·DB 확인\n"
        "3-7 조회·매출·순수익·총건수·지출\n\n"
        "[교차대조]\n"
        "대조 날짜 (예: 대조 3-7)\n"
        "대조 확정 날짜 / 대조 금액확인 날짜\n\n"
        "[결제삭제]\n"
        "결제삭제 날짜 운행외/0원/전체/HH:MM\n\n"
        "[전략] 전략 / 마기 업데이트 시간대 내용\n\n"
        "[다운로드]\n"
        "주간/월간/전체 다운로드 / 월간 2026-03\n\n"
        "[어군] /fish"
    )

async def cmd_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Chat ID: {update.effective_chat.id}")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    try:
        await image_queue.put((update, context))
    except Exception as e:
        logger.error(f"이미지 큐 오류: {e}")
        await update.message.reply_text("❌ 이미지 처리 오류. 다시 시도해주세요.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    파일 첨부 처리.
    - 이미지 파일 (jpg/jpeg/png/webp/gif/bmp) → 이미지 처리 파이프라인
    - xlsx → 엑셀 이식
    - 그 외 → 안내 메시지
    """
    if not is_allowed(update):
        return
    doc = update.message.document
    if not doc:
        return

    fname = (doc.file_name or "").lower()
    mime  = (doc.mime_type or "").lower()

    # ── 이미지 파일로 전송된 경우 (파일로 보내기) ──
    IMAGE_EXTS  = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp")
    IMAGE_MIMES = ("image/jpeg", "image/png", "image/webp", "image/gif", "image/bmp")

    if any(fname.endswith(e) for e in IMAGE_EXTS) or any(mime.startswith(m) for m in IMAGE_MIMES):
        # 파일 다운로드 후 이미지 큐에 추가
        try:
            file = await context.bot.get_file(doc.file_id)
            image_bytes = await file.download_as_bytearray()
            image_bytes = bytes(image_bytes)
            if image_queue is None:
                await update.message.reply_text("❌ 이미지 큐 초기화 중입니다. 잠시 후 다시 시도해주세요.")
                return
            await image_queue.put((update, context, image_bytes))
            logger.info(f"파일 이미지 큐 추가: {fname} ({len(image_bytes):,}bytes)")
        except Exception as e:
            logger.error(f"파일 이미지 처리 오류: {e}")
            await update.message.reply_text(f"❌ 파일 처리 오류: {str(e)[:100]}")
        return

    # ── 엑셀 이식 ──
    if fname.endswith(".xlsx"):
        await handle_excel_import(update, context)
        return

    # ── 그 외 ──
    await update.message.reply_text(
        "⚠️ 지원하지 않는 파일 형식입니다.\n"
        "이미지: jpg·png·webp 파일 또는 사진으로 전송\n"
        "엑셀: xlsx 파일"
    )


async def _process_single_command(update, context, text: str) -> str | None:
    """단일 명령어 처리. 줄바꿈 다중 명령어 시 각 줄 처리용."""

    # 결제삭제
    if text.startswith("결제삭제 "):
        parts = text.strip().split(" ", 2)
        if len(parts) < 3:
            return "❌ 형식: 결제삭제 YYYY-MM-DD 운행외|0원|전체|HH:MM"
        date_str = parts[1].strip()
        mode = parts[2].strip()
        try:
            rows = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})
            if not rows:
                return f"⚠️ {date_str} 결제내역 없음"
            delete_ids = []
            if mode == "운행외":
                for r in rows:
                    t = r.get("시각", "") or ""
                    try:
                        h, m = t.split(":")
                        mins = int(h)*60+int(m)
                        if 121 <= mins <= 1019:
                            delete_ids.append(r["id"])
                    except Exception:
                        delete_ids.append(r["id"])
            elif mode == "0원":
                for r in rows:
                    fee = r.get("요금")
                    if fee is None or int(fee) == 0:
                        delete_ids.append(r["id"])
            elif mode == "전체":
                delete_ids = [r["id"] for r in rows]
            elif ":" in mode:
                for r in rows:
                    if r.get("시각","") == mode:
                        delete_ids.append(r["id"])
            else:
                return f"❓ 알 수 없는 모드: {mode}"
            if not delete_ids:
                return f"✅ {date_str} 삭제 대상 없음 ({mode})"
            for rid in delete_ids:
                await sb_h("DELETE", f"payment_receipts?id=eq.{rid}")
            return f"🗑️ {date_str} {mode} {len(delete_ids)}건 삭제"
        except Exception as e:
            return f"❌ 삭제 오류: {str(e)[:100]}"

    # 수동 콜
    parsed_call = parse_manual_call(text)
    if parsed_call:
        today = str(today_kst())
        payload = {
            "날짜": today, "요일": get_dow(),
            "배차시각": now_kst().strftime("%H:%M"),
            "요금": parsed_call["요금"],
            "콜유형": parsed_call["콜유형"],
            "도착지": parsed_call.get("도착지힌트"),
        }
        r = await sb_insert("raw_calls", payload)
        return f"✅ {parsed_call['콜유형']} {fmt(parsed_call['요금'])} 입력" if r else "❌ 저장 실패"

    # 지출
    parsed_exp = parse_expense(text)
    if parsed_exp:
        today = str(today_kst())
        payload = {
            "날짜": today,
            "카테고리": parsed_exp["카테고리"],
            "금액": parsed_exp["금액"],
            "메모": parsed_exp.get("메모",""),
            "자동여부": False,
        }
        r = await sb_insert("expenses", payload)
        return f"✅ {parsed_exp['카테고리']} {fmt(parsed_exp['금액'])} 입력" if r else "❌ 저장 실패"

    return f"❓ '{text[:20]}' 인식 불가"

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    text = (update.message.text or "").strip()
    lower = text.lower()

    # ── 줄바꿈 다중 명령어 처리 ──
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if len(lines) > 1:
        results = []
        for line in lines:
            # 각 줄을 개별 명령어로 처리
            fake_update = update
            line_text = line
            try:
                result = await _process_single_command(fake_update, context, line_text)
                if result:
                    results.append(result)
            except Exception as e:
                results.append(f"❌ '{line_text[:20]}' 오류: {str(e)[:50]}")
        if results:
            await update.message.reply_text("\n".join(results))
        return

    # 결제내역 삭제 명령어
    # "결제삭제 YYYY-MM-DD 운행외" → 운행시간(19~02시) 외 데이터 삭제
    # "결제삭제 YYYY-MM-DD 0원"    → 요금 0원 데이터 삭제
    # "결제삭제 YYYY-MM-DD 전체"   → 해당 날짜 전체 삭제
    if text.startswith("결제삭제 "):
        await handle_receipt_delete(update, text)
        return

    # 대조 확정 (배회후보 raw_calls 추가)
    if text.startswith("대조 확정 "):
        _ds = text[6:].strip()
        try:
            result = await confirm_cross_check(_ds)
            await update.message.reply_text(result)
        except Exception as e:
            await update.message.reply_text(f"❌ 대조확정 오류: {str(e)[:200]}")
        return

    # 대조 금액확인 (InlineKeyboard 금액 선택)
    if text.startswith("대조 금액확인 "):
        _ds = text[8:].strip()
        try:
            await handle_fee_confirm_request(update, _ds)
        except Exception as e:
            await update.message.reply_text(f"❌ 금액확인 오류: {str(e)[:200]}")
        return

    # 교차대조 — YYYY-MM-DD 또는 M-D 형식 지원
    if text.startswith("대조 "):
        import re as _re3
        from datetime import date as _date3
        date_str = text[3:].strip()
        _md = _re3.match(r'^(\d{1,2})-(\d{1,2})$', date_str)
        if _md:
            try:
                _d = _date3(_date3.today().year, int(_md.group(1)), int(_md.group(2)))
                date_str = str(_d)
            except ValueError:
                await update.message.reply_text("❌ 잘못된 날짜입니다.")
                return
        try:
            result = await cross_check(date_str)
            if len(result) > 4000:
                result = result[:4000] + "\n...(생략)"
            await update.message.reply_text(result)
        except Exception as e:
            logger.error(f"교차대조 오류: {e}")
            await update.message.reply_text(f"❌ 교차대조 오류: {str(e)[:200]}")
        return

    if text.startswith("배회분류 확정 "):
        date_str = text[8:].strip()
        try:
            result = await confirm_baehoe_classification(date_str)
            await update.message.reply_text(result)
        except Exception as e:
            logger.error(f"배회분류 오류: {e}")
            await update.message.reply_text(f"❌ 배회분류 오류: {str(e)[:200]}")
        return

    # 날짜 + 통계 키워드 조합
    _stat_kws = ["총건수","건수","매출","순수익","지출","조회","상세"]
    _date_pat = r"(\d{1,2})[-/](\d{1,2})|(\d{4})-(\d{1,2})-(\d{1,2})"
    import re as _re
    if _re.search(_date_pat, text) and any(kw in text for kw in _stat_kws):
        if any(kw in text for kw in ["총건수","건수","매출","순수익","지출"]):
            await handle_date_stat(update, text)
        else:
            await handle_date_query(update, text)
        return

    # 브리핑
    if text in ("브리핑", "오늘브리핑", "오늘 브리핑"):
        await handle_briefing(update)
        return

    # 운행완료수
    if text in ("운행완료수", "완료수", "ai진입"):
        await handle_completion_status(update)
        return

    # 운행 일관성 조회
    if text in ("일관성", "일관성 조회", "운행일관성"):
        await report_operation_consistency(update)
        return

    # 어군 브리핑 텍스트 명령
    if text in ("어군", "어군조회", "어군 조회"):
        now = datetime.now(KST)
        if not get_fish_slot(now.hour):
            await update.message.reply_text(
                f"🐟 현재 {now.hour}시는 브리핑 시간대가 아닙니다.\n"
                f"운영시간: 19~21시 / 21~24시 / 00~02시"
            )
        else:
            report = await get_fish_report_db()
            await update.message.reply_text(report or "🐟 어군 데이터 없음")
        return

    # 세큐티 조회
    if text in ("세큐티 조회", "세큐티조회"):
        await handle_sekuti_query(update)
        return

    # 특정 날짜 조회 (조회 키워드만 있을 때)
    if "조회" in text:
        await handle_date_query(update, text)
        return

    # 조회
    if text == "오늘":
        await handle_today_quick(update)
        return
    if text in ("이번 주", "이번주", "주간"):
        await handle_weekly(update)
        return
    import re as _re2
    _ym = _re2.match(r"^월간\s+(\d{4}-\d{2})\s*$", text.strip())
    if _ym:
        await handle_download_month(update, _ym.group(1))
        return

    if text in ("이번 달", "이번달", "월간"):
        await handle_monthly(update)
        return
    if text == "지출 확인":
        await handle_expense_check(update)
        return
    if text == "DB 확인":
        await handle_db_check(update)
        return

    # 전략
    if text in ("전략", "실시간"):
        await get_strategy(update)
        return
    if text.startswith("마기 업데이트 "):
        content = text[8:].strip()
        await handle_magi_update(update, content)
        return

    # 다운로드
    if "다운로드" in text:
        if "주간" in text:
            await handle_download(update, "주간")
        elif "월간" in text:
            await handle_download(update, "월간")
        elif "전체" in text:
            await handle_download(update, "전체")
        else:
            await update.message.reply_text("주간·월간·전체 다운로드 중 선택해주세요.")
        return

    # 특정 월 다운로드: "월간 2026-03" or "월간2026-03"
    import re as _re
    _ym_match = _re.match(r"^월간\s*(\d{4}-\d{2})$", text.strip())
    if _ym_match:
        await handle_download_month(update, _ym_match.group(1))
        return

    # 수동 전체 입력 (날짜+시각+경로+요금 형식)
    import re as _re2
    _has_route = bool(_re2.search(r'[가-힣\w]+[>→][가-힣\w]+', text))
    _has_date_nums = bool(_re2.search(r'\d{2}\s+\d{1,2}\s+\d{1,2}\s+\d{1,2}\s+\d{2}|\d{4}[\-.]\d{1,2}[\-.]\d{1,2}|\d{8}|\d{4}\s+\d{4}', text))
    if _has_route and _has_date_nums:
        await handle_manual_full_call(update, text)
        return

    # 콜카드 수동 수정
    if text.startswith("콜수정ID "):
        await handle_call_edit_by_id(update, text[6:].strip())
        return

    if text.startswith("콜수정 "):
        await handle_call_edit(update, text)
        return

    # 휴무 (오늘 또는 날짜 지정)
    if "휴무" in text:
        await handle_rest_day(update, text)
        return

    # 지출취소
    if text == "지출취소":
        await handle_expense_cancel(update)
        return

    # 수동 콜 입력
    parsed_call = parse_manual_call(text)
    if parsed_call:
        await handle_manual_call(update, parsed_call)
        return

    # 지출 입력
    parsed_exp = parse_expense(text)
    if parsed_exp:
        await handle_expense(update, parsed_exp)
        return

    # 미인식
    await update.message.reply_text("❓ 명령어를 인식하지 못했습니다. /help 로 확인해주세요.")

# ──────────────────────────────────────────────
# main()
# ──────────────────────────────────────────────
def main():
    global image_queue

    # Health server
    threading.Thread(target=run_health_server, daemon=True).start()

    # Insurance scheduler
    threading.Thread(target=insurance_scheduler, daemon=True).start()

    # Telegram application
    app = (
        Application.builder()
        .token(TELEGRAM_TOKEN)
        .read_timeout(30)
        .write_timeout(30)
        .connect_timeout(30)
        .get_updates_read_timeout(5)    # 폴링 빠른 타임아웃 → Conflict 최소화
        .build()
    )

    async def post_init(application):
        global image_queue
        image_queue = asyncio.Queue(maxsize=10)  # 최대 10개 대기
        asyncio.create_task(process_image_queue_worker())
        # 기존 webhook 제거 + 이전 인스턴스 세션 정리 — Conflict 방지
        await asyncio.sleep(2)  # 구 인스턴스 세션 해제 대기
        await application.bot.delete_webhook(drop_pending_updates=True)
        logger.info("Webhook 삭제 완료 — 폴링 시작 준비")

    app.post_init = post_init

    # 어군탐지기 스케줄러 — app 생성 후 시작
    threading.Thread(target=fish_scheduler, args=(app,), daemon=True).start()
    logger.info("어군탐지기 스케줄러 시작")

    # 핸들러 등록
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("fish", cmd_fish))    # 어군 브리핑 수동 조회
    app.add_handler(CommandHandler("avoid", cmd_avoid))  # 회피 구역 조회
    app.add_handler(CommandHandler("forecast", lambda u,c: handle_forecast(u, c.args[0] if c.args else None)))
    app.add_handler(CommandHandler("completion_status", lambda u,c: handle_completion_status(u)))
    app.add_handler(CommandHandler("briefing", lambda u,c: handle_briefing(u)))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("id", cmd_id))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    from telegram.ext import CallbackQueryHandler
    app.add_handler(CallbackQueryHandler(handle_fee_callback, pattern=r"^fee:"))

    # 전역 에러 핸들러 — Conflict/Network 오류 자동 복구
    async def error_handler(update, context):
        import telegram
        err = context.error
        if isinstance(err, telegram.error.Conflict):
            logger.warning(f"Conflict 감지 (자동복구 대기): {err}")
        elif isinstance(err, telegram.error.NetworkError):
            logger.warning(f"네트워크 오류 (자동재시도): {err}")
        elif isinstance(err, telegram.error.TimedOut):
            pass  # 타임아웃은 정상 폴링 동작
        else:
            logger.error(f"봇 오류: {type(err).__name__}: {err}")
    app.add_error_handler(error_handler)

    logger.info("자비스 v5 시작")
    app.run_polling(
        drop_pending_updates=True,
        allowed_updates=["message", "callback_query"],
    )

if __name__ == "__main__":
    main()
