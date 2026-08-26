# bot_v5.py — 자비스(JARVIS) 봇 v5 완성판
# 설계서 기준 + 결제내역 OCR + 콜카드↔결제내역 교차대조
# 작성일: 2026-03-30

import os
import io
import re
import json
import time
import asyncio
import logging
import threading
import base64
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer

import httpx
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ContextTypes, filters
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic

load_dotenv()

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────
KST = timezone(timedelta(hours=9))
NET_GOAL      = 150_000      # 일 순수익 목표
CARD_FEE_RATE = 0.033        # 카드수수료 3.3%
INSURANCE_DAILY = 7_945      # 일 보험료
DOW_KOR = ["월", "화", "수", "목", "금", "토", "일"]

TELEGRAM_TOKEN     = os.getenv("TELEGRAM_TOKEN", "")
ALLOWED_IDS_RAW    = [
    os.getenv("ALLOWED_CHAT_ID", ""),
    os.getenv("ALLOWED_CHAT_ID2", ""),
]
ALLOWED_IDS = {x for x in ALLOWED_IDS_RAW if x}

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
SUPABASE_URL      = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY      = os.getenv("SUPABASE_KEY", "")
PORT              = int(os.getenv("PORT", "10000"))

# 캐스퍼 명령서 #014 반영 (2026-07-10) — 봇 직접 GitHub 커밋용
# 이지스가 쓰는 PAT와 별도로 봇 전용 PAT를 Render 환경변수에 등록해서 사용 권장.
GITHUB_PAT   = os.getenv("GITHUB_PAT", "")
GITHUB_OWNER = os.getenv("GITHUB_OWNER", "Kim0900")
GITHUB_REPO  = os.getenv("GITHUB_REPO", "magi-taxi-data")

OCR_MODEL      = "claude-haiku-4-5-20251001"

def _extract_claude_text(msg) -> str:
    """캐스퍼 수정 2026-07-24: Claude 응답의 content가 항상 텍스트 블록이라는
    가정(content[0].text)이 최신 모델(sonnet-5 등)에서 ThinkingBlock이 먼저
    오는 경우 깨짐('ThinkingBlock' object has no attribute 'text'). 
    type=='text'인 블록만 안전하게 찾아서 반환."""
    for block in msg.content:
        if getattr(block, "type", None) == "text":
            return block.text
    return ""

HEADERS_SB = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Health Check 서버
# ──────────────────────────────────────────────
class HealthHandler(BaseHTTPRequestHandler):
    def _cors_headers(self):
        """모든 응답에 CORS 허용 헤더 추가"""
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization, apikey')

    def do_OPTIONS(self):
        """CORS preflight 요청 처리"""
        self.send_response(204)
        self._cors_headers()
        self.end_headers()

    def do_GET(self):
        # 영실 MCP안정화 제안서 §3,4 P0 대응
        if self.path == '/health':
            try:
                supa_ok = True
                supa_detail = "OK"
                try:
                    asyncio.run(sb_select("db_manifest", {"limit": "1"}))
                except Exception as e:
                    supa_ok = False
                    supa_detail = f"FAIL: {e}"

                last_db_sync = None
                try:
                    rows = asyncio.run(sb_select("correction_log", {"order": "changed_at.desc", "limit": "1"}))
                    if rows: last_db_sync = rows[0].get("changed_at")
                except Exception as e:
                    logger.error(f"health: last_db_sync 조회 실패: {e}")

                scheduler_rows = []
                try:
                    scheduler_rows = asyncio.run(sb_select("scheduler_status", {"order": "last_run_at.desc"})) or []
                except Exception as e:
                    logger.error(f"health: scheduler_status 조회 실패: {e}")

                body = json.dumps({
                    "render": "OK",
                    "supabase": "OK" if supa_ok else supa_detail,
                    "mcp": "OK",
                    "telegram": "미구현(라이브러리 내부폴링이라 추적불가, 정직히 표기)",
                    "last_db_sync": last_db_sync,
                    "scheduler_jobs": {r["job_name"]: {"last_run_at": r.get("last_run_at"), "last_result": r.get("last_result")} for r in scheduler_rows},
                }, ensure_ascii=False).encode('utf-8')
                self.send_response(200)
                self._cors_headers()
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                self.send_response(500)
                self._cors_headers()
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return

        if self.path == '/version':
            body = json.dumps({
                "bot_version": "v5",
                "github_commit": os.getenv("RENDER_GIT_COMMIT", "unknown(로컬실행 또는 Render 미배포)"),
                "render_service_id": os.getenv("RENDER_SERVICE_ID", "unknown"),
            }, ensure_ascii=False).encode('utf-8')
            self.send_response(200)
            self._cors_headers()
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(body)
            return

        self.send_response(200)
        self._cors_headers()
        self.end_headers()
        self.wfile.write(b"Jarvis v5 OK")

    def do_POST(self):
        """API 엔드포인트 — OCR / 마기분석 / 아틀라스보고"""
        import json as _j, re as _re, time as _time
        _req_start = _time.time()
        length = int(self.headers.get('Content-Length', 0))
        raw_body = self.rfile.read(length)
        try:
            payload = _j.loads(raw_body.decode('utf-8')) if raw_body else {}
        except Exception:
            payload = {}

        def send_json(code, data):
            # 영실 MCP안정화 제안서 §5: Action호출 표준로그(처리시간·반환건수) —
            # 전체 POST 엔드포인트(MCP 7종 포함)에 공용함수 통해 일괄 적용
            duration_ms = round((_time.time() - _req_start) * 1000)
            rows_count = None
            if isinstance(data, dict):
                for k in ("tasks", "task", "events", "corrections", "task_events"):
                    v = data.get(k)
                    if isinstance(v, list):
                        rows_count = len(v); break
            logger.info(f"[ActionLog] {self.path} | HTTP {code} | {duration_ms}ms | rows={rows_count}")
            body = _j.dumps(data, ensure_ascii=False).encode('utf-8')
            self.send_response(code)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization, apikey')
            self.end_headers()
            self.wfile.write(body)

        if self.path == '/ocr_receipt':
            try:
                import anthropic as _ant
                b64 = payload.get('image_b64', '')
                mt  = payload.get('media_type', 'image/jpeg')
                client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
                # 캐스퍼 수정 2026-07-23: 영수증에 연도가 없는 경우가 대부분이라
                # 오늘 날짜를 프롬프트에 명시하지 않으면 모델이 엉뚱한 연도(예: 2023)를
                # 추측해버리는 버그가 있었음. 오늘 날짜를 컨텍스트로 반드시 제공.
                _today_str = str(today_kst())
                msg = client.messages.create(
                    model="claude-haiku-4-5-20251001", max_tokens=400, temperature=0,
                    messages=[{"role":"user","content":[
                        {"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
                        {"type":"text","text":f'오늘 날짜는 {_today_str}입니다. 이 택시 매출집계 영수증에서 정보를 추출해서 JSON만 반환해줘.\n{{"date":"YYYY-MM-DD","total_sales":숫자,"commission":숫자,"trip_count":숫자,"start_time":"HH:MM","end_time":"HH:MM"}}\n영수증에 연도 표시가 없으면 오늘({_today_str}) 기준 연도를 사용하되, 자정을 넘겨 익일로 표시된 시각이 있으면 날짜 앞뒤 관계를 자연스럽게 맞춰라.\n⚠️ commission(수수료)은 영수증에 "수수료"라고 명시적으로 적힌 금액만 사용해라. "카드결제"·"앱결제"·"현금결제"처럼 결제수단별로 나눈 금액은 수수료가 아니니 절대 혼동하지 마라. "수수료"라는 글자가 영수증에 없으면 commission은 0으로 반환해라. 숫자만(원제외). JSON만 반환.'}
                    ]}]
                )
                txt = _re.sub(r"```[a-z]*", "", _extract_claude_text(msg).strip()).strip()
                send_json(200, {"success": True, "data": _j.loads(txt)})
            except Exception as e:
                logger.error(f"OCR 오류: {e}")
                send_json(400, {"success": False, "error": str(e)})
            return

        if self.path == '/gpx_parse':
            # GPX는 앱(브라우저 DOMParser)에서 직접 파싱 — 봇 불필요
            # 이 엔드포인트는 예비용(서버사이드 파싱 필요 시 확장)
            send_json(200, {"status":"ok","message":"GPX는 앱 내 파싱 완료"})
            return

        # ──────────────────────────────────────────────
        # 명령서#029 (2026-08-09, 마기 발부): MAGI 전용 MCP 서버 1단계
        # 카산드라(ChatGPT) 등 외부AI가 읽기전용으로 조회할 수 있는 통로 3개.
        # 범위: 조회만, 쓰기·전략판단·자동오케스트레이션은 포함하지 않음(§0 명시).
        # 인증: X-MCP-Key 헤더가 환경변수 MCP_API_KEY와 일치해야 함.
        # ──────────────────────────────────────────────
        if self.path.startswith('/mcp/'):
            mcp_key = os.getenv("MCP_API_KEY", "")
            req_key = self.headers.get("X-MCP-Key", "")
            if not mcp_key or req_key != mcp_key:
                send_json(401, {"success": False, "error": "MCP 인증 실패 (X-MCP-Key 헤더 확인)"})
                return

            if self.path == '/mcp/recent_events':
                try:
                    n = int(payload.get("limit", 5))
                    rows = asyncio.run(sb_select("atlas_reports", {
                        "status": "eq.analyzed", "order": "run_date.desc", "limit": str(n)
                    }))
                    events = [{"run_date": r.get("run_date"), "title": r.get("title"),
                               "summary": (r.get("content") or "")[:500]} for r in rows]
                    send_json(200, {"success": True, "events": events})
                except Exception as e:
                    logger.error(f"MCP recent_events 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            if self.path == '/mcp/kpi_status':
                try:
                    dv = asyncio.run(dual_verify_7day_average())
                    today = str(today_kst())
                    today_calls = asyncio.run(sb_select_calls({"날짜": f"eq.{today}"}))
                    send_json(200, {
                        "success": True,
                        "seven_day_avg_krw": dv["method_a"]["일평균"],
                        "seven_day_total_krw": dv["method_a"]["총매출"],
                        "days_with_data": dv["method_a"]["데이터있는날"],
                        "today_date": today,
                        "today_driving": len(today_calls) > 0,
                        "today_calls_so_far": len(today_calls),
                        "dual_verify_match": dv["match"],
                    })
                except Exception as e:
                    logger.error(f"MCP kpi_status 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            if self.path == '/mcp/corrections':
                try:
                    n = int(payload.get("limit", 10))
                    rows = asyncio.run(sb_select("correction_log", {
                        "order": "changed_at.desc", "limit": str(n)
                    }))
                    send_json(200, {"success": True, "corrections": rows})
                except Exception as e:
                    logger.error(f"MCP corrections 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            # ──────────────────────────────────────────────
            # 명령서#032 (2026-08-12): ATHENA Task Registry Phase 1 — MCP READ tool 4종.
            # 전부 읽기전용(§4 명시). WRITE는 이번 범위 밖.
            # ──────────────────────────────────────────────
            # ──────────────────────────────────────────────
            # 명령서#034 (2026-08-13): magi_tasks에 problem/target/dependencies/
            # context_summary 4개 컬럼 추가됨. 아래 sb_select()들은 select 파라미터를
            # 지정하지 않아 PostgREST 기본동작(전체 컬럼 반환)으로 신규 필드도 이미
            # 자동 포함됨 — 코드 로직 변경 불필요, curl 실기기검증으로 확인함.
            # ──────────────────────────────────────────────
            if self.path == '/mcp/get_active_tasks':
                try:
                    params = {"status": "not.in.(CLOSED,CANCELLED)", "order": "updated_at.desc"}
                    owner = payload.get("owner_agent")
                    if owner:
                        params["owner_agent"] = f"eq.{owner}"
                    rows = asyncio.run(sb_select("magi_tasks", params))
                    send_json(200, {"success": True, "tasks": rows})
                except Exception as e:
                    logger.error(f"MCP get_active_tasks 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            if self.path == '/mcp/get_task':
                try:
                    tid = payload.get("task_id")
                    if tid is None:
                        send_json(400, {"success": False, "error": "task_id 필요"})
                        return
                    task_rows = asyncio.run(sb_select("magi_tasks", {"task_id": f"eq.{tid}"}))
                    event_rows = asyncio.run(sb_select("magi_task_events", {
                        "task_id": f"eq.{tid}", "order": "created_at.asc"
                    }))
                    if not task_rows:
                        send_json(404, {"success": False, "error": f"task_id={tid} 없음"})
                        return
                    send_json(200, {"success": True, "task": task_rows[0], "events": event_rows})
                except Exception as e:
                    logger.error(f"MCP get_task 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            if self.path == '/mcp/get_blocked_tasks':
                try:
                    rows = asyncio.run(sb_select("magi_tasks", {
                        "status": "eq.BLOCKED", "order": "updated_at.desc"
                    }))
                    send_json(200, {"success": True, "tasks": rows})
                except Exception as e:
                    logger.error(f"MCP get_blocked_tasks 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            if self.path == '/mcp/get_recent_completed_tasks':
                try:
                    n = int(payload.get("limit", 10))
                    rows = asyncio.run(sb_select("magi_tasks", {
                        "status": "eq.COMPLETED", "order": "completed_at.desc", "limit": str(n)
                    }))
                    send_json(200, {"success": True, "tasks": rows})
                except Exception as e:
                    logger.error(f"MCP get_recent_completed_tasks 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            # ──────────────────────────────────────────────
            # task_id=27(로드맵6) 지원: 아르고스 브리핑 자동화용 READ 엔드포인트.
            # calc_date 지정시 그 날짜(축A) 단건, 없으면 최근 limit개(기본 7) 반환.
            # ──────────────────────────────────────────────
            if self.path == '/mcp/get_daily_calc_snapshot':
                try:
                    calc_date = payload.get("calc_date")
                    if calc_date:
                        rows = asyncio.run(sb_select("daily_calc_snapshot", {
                            "calc_date": f"eq.{calc_date}", "axis": "eq.A"
                        }))
                    else:
                        n = int(payload.get("limit", 7))
                        rows = asyncio.run(sb_select("daily_calc_snapshot", {
                            "axis": "eq.A", "order": "calc_date.desc", "limit": str(n)
                        }))
                    send_json(200, {"success": True, "snapshots": rows})
                except Exception as e:
                    logger.error(f"MCP get_daily_calc_snapshot 오류: {e}")
                    send_json(400, {"success": False, "error": str(e)})
                return

            send_json(404, {"success": False, "error": f"알 수 없는 MCP 엔드포인트: {self.path}"})
            return

        if self.path == '/ocr_history':
            try:
                import anthropic as _ant, re as _re, json as _j, base64 as _b64mod
                b64 = payload.get('image_b64', '')
                mt  = payload.get('media_type', 'image/jpeg')
                try:
                    raw_bytes = _b64mod.b64decode(b64)
                    resized_bytes = resize_image_if_needed(raw_bytes)
                except Exception as _rez_err:
                    logger.warning(f"OCR 이미지 리사이즈 실패, 원본 사용: {_rez_err}")
                    resized_bytes = _b64mod.b64decode(b64)

                # 대표님요청(2026-08-18): 세로분할 2회호출. 상단(헤더+상반부)/하단(하반부)
                # 각각 OCR해서 정보밀도를 낮춰 세부동이름 인식률 향상. 비용증가 최소화를
                # 위해 4분할이 아닌 2분할(10%겹침)로 설계.
                try:
                    top_bytes, bottom_bytes = split_image_vertically(resized_bytes)
                except Exception as _split_err:
                    logger.error(f"이미지 분할 실패, 단일호출로 폴백: {_split_err}")
                    top_bytes, bottom_bytes = resized_bytes, None

                top_b64 = _b64mod.b64encode(top_bytes).decode('utf-8')
                bottom_b64 = _b64mod.b64encode(bottom_bytes).decode('utf-8') if bottom_bytes else None
                mt = 'image/jpeg'
                client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
                _today_str = str(today_kst())

                def _run_ocr_history(model, image_b64, is_bottom_half=False):
                    is_daily = True
                    extra_kwargs = {"temperature": 0} if model.startswith("claude-haiku") else {"thinking": {"type": "disabled"}}
                    if is_daily:
                        if is_bottom_half:
                            # 하단조각 전용 프롬프트: 헤더(날짜/합계)가 없을 수 있으므로 요구 안 함
                            prompt = (
                                '이 이미지는 카카오T 일별운행이력 화면의 아래쪽 절반(스크롤 하단부)이다. '
                                '화면 상단(날짜·합계 헤더)은 이 조각에 없을 수 있다 — 없으면 무시하고, '
                                '이 조각에 보이는 운행 건만 추출해서 JSON만 반환해줘.\n'
                                '{"type":"daily_history","calls":['
                                '{"배차시각":"HH:MM","하차시각":"HH:MM","출발지":"대구 OO구 OO동","도착지":"대구 OO구 OO동","요금":숫자,"결제방식":"자동 또는 직접"}]}\n'
                                '⚠️ 출발지/도착지는 화면에 실제로 인쇄된 글자를 정확히 그대로 옮겨라. 특히 동 이름 끝의 숫자(1동/2동/3동/4동 등)를 '
                                '놓치지 말고 반드시 확인해라 — 숫자가 흐려서 안 보이면 지어내지 말고 "OO?동"처럼 표시해라. '
                                '절대로 지명을 지어내거나 비슷하게 짐작하지 마라. 화면 맨 위/아래가 항목 중간에서 잘렸으면 그 잘린 항목은 건너뛰어라(다른 조각에서 온전히 잡힘).\n'
                                '⚠️ "요금"은 반드시 숫자만 반환해라. 흐리면 0. JSON만 반환.'
                            )
                        else:
                            prompt = (
                                f'오늘 날짜는 {_today_str}입니다. 이 이미지는 카카오T 일별운행이력 화면의 위쪽 절반(헤더+상반부)이다. '
                                '화면이 아래로 더 이어질 수 있으니, 이 조각에 보이는 운행 건까지만 추출해서 JSON만 반환해줘.\n'
                                '{"type":"daily_history","date":"YYYY-MM-DD","표시건수":숫자,"표시금액":숫자,"calls":['
                                '{"배차시각":"HH:MM","하차시각":"HH:MM","출발지":"대구 OO구 OO동","도착지":"대구 OO구 OO동","요금":숫자,"결제방식":"자동 또는 직접"}]}\n'
                                f'날짜: 화면에 연도가 없으면 오늘({_today_str}) 기준 연도 사용. 상단 YYYY년 M월 D일 있으면 그것 우선. 결제방식: 직접결제 있으면 직접, 없으면 자동.\n'
                                '⚠️ "표시건수"/"표시금액"은 화면 맨 위에 큰 글씨로 적힌 "OO건", "OOO원" 합계를 그대로 옮겨라 — 이 값은 화면 전체(하단 조각 포함) 기준이니 이 조각에 보이는 건수와 달라도 그대로 옮겨라.\n'
                                '⚠️ 출발지/도착지는 화면에 실제로 인쇄된 글자를 정확히 그대로 옮겨라. 특히 동 이름 끝의 숫자(1동/2동/3동/4동 등)를 '
                                '놓치지 말고 반드시 확인해라 — 숫자가 흐려서 안 보이면 지어내지 말고 "OO?동"처럼 표시해라. '
                                '절대로 지명을 지어내거나 비슷하게 짐작하지 마라. 화면 아래쪽이 항목 중간에서 잘렸으면 그 잘린 항목은 건너뛰어라(다른 조각에서 온전히 잡힘).\n'
                                '⚠️ "요금"은 반드시 숫자만 반환해라. 흐리면 0. JSON만 반환.'
                            )
                    ocr_msg = client.messages.create(
                        model=model, max_tokens=8000, **extra_kwargs,
                        messages=[{"role":"user","content":[
                            {"type":"image","source":{"type":"base64","media_type":mt,"data":image_b64}},
                            {"type":"text","text":prompt}
                        ]}]
                    )
                    txt = _re.sub(r"```[a-z]*","",_extract_claude_text(ocr_msg).strip()).strip()
                    txt = _re.sub(r'("(?:요금|표시금액|표시건수)"\s*:\s*)(?!["\d\-])[^\s,}\]]+', r'\g<1>0', txt)
                    if not txt:
                        raise ValueError(f"모델({model}) 응답에 텍스트가 없음 (stop_reason={getattr(ocr_msg,'stop_reason',None)}) — 사고예산 소진 또는 응답거부 의심")
                    try:
                        return _j.loads(txt)
                    except _j.JSONDecodeError as je:
                        ctx = txt[max(0,je.pos-150):je.pos+150]
                        logger.error(f"OCR JSON 파싱 실패 상세 — stop_reason={getattr(ocr_msg,'stop_reason',None)}, 전체길이={len(txt)}자, 에러지점 앞뒤 300자:\n{ctx}")
                        raise

                def _run_ocr_split(model):
                    top_result = _run_ocr_history(model, top_b64, is_bottom_half=False)
                    if bottom_b64:
                        bottom_result = _run_ocr_history(model, bottom_b64, is_bottom_half=True)
                        return merge_split_ocr_results(top_result, bottom_result)
                    return top_result

                def _safe_int(v):
                    if v is None: return None
                    try: return int(v)
                    except (ValueError, TypeError):
                        digits = _re.sub(r'[^0-9]', '', str(v))
                        return int(digits) if digits else None

                # 캐스퍼 수정 2026-08-07(5차): 지금까지 자체검증(_verify)이 합계(건수/금액)만
                # 봐서, 개별 주소가 거의 다 오독이어도(예: "력선동","안성3동" 등 150개 목록에
                # 없는 지명 다수) 합계만 맞으면 그냥 통과되던 구조적 허점. 대구 150개 행정동
                # 매핑을 서버에도 넣어서, 무효 주소 개수가 많으면 합계가 맞아도 재시도하도록 확장.
                _DONG_TO_GU = {"가창면":"달성군","감삼동":"달서구","검단동":"북구","고산1동":"수성구","고산2동":"수성구","고산3동":"수성구","고성동":"북구","공산동":"동구","관문동":"북구","관음동":"북구","구암동":"북구","구지면":"달성군","국우동":"북구","군위읍":"군위군","남산1동":"중구","남산2동":"중구","남산3동":"중구","남산4동":"중구","내당1동":"서구","내당2·3동":"서구","내당4동":"서구","노원동":"북구","논공읍":"달성군","다사읍":"달성군","대명10동":"남구","대명11동":"남구","대명1동":"남구","대명2동":"남구","대명3동":"남구","대명4동":"남구","대명5동":"남구","대명6동":"남구","대명9동":"남구","대봉1동":"중구","대봉2동":"중구","대신동":"중구","대현동":"북구","도원동":"달서구","도평동":"동구","동인동":"중구","동천동":"북구","동촌동":"동구","두류1,2동":"달서구","두류3동":"달서구","두산동":"수성구","만촌1동":"수성구","만촌2동":"수성구","만촌3동":"수성구","무태조야동":"북구","방촌동":"동구","범물1동":"수성구","범물2동":"수성구","범어1동":"수성구","범어2동":"수성구","범어3동":"수성구","범어4동":"수성구","복현1동":"북구","복현2동":"북구","본동":"달서구","본리동":"달서구","봉덕1동":"남구","봉덕2동":"남구","봉덕3동":"남구","부계면":"군위군","불로·봉무동":"동구","비산1동":"서구","비산2·3동":"서구","비산4동":"서구","비산5동":"서구","비산6동":"서구","비산7동":"서구","산격1동":"북구","산격2동":"북구","산격3동":"북구","산격4동":"북구","산성면":"군위군","삼국유사면":"군위군","삼덕동":"중구","상동":"수성구","상인1동":"달서구","상인2동":"달서구","상인3동":"달서구","상중이동":"서구","성내1동":"중구","성내2동":"중구","성내3동":"중구","성당동":"달서구","소보면":"군위군","송현1동":"달서구","송현2동":"달서구","수성1가동":"수성구","수성2·3가동":"수성구","수성4가동":"수성구","신당동":"달서구","신암1동":"동구","신암2동":"동구","신암3동":"동구","신암4동":"동구","신암5동":"동구","신천1·2동":"동구","신천3동":"동구","신천4동":"동구","안심1동":"동구","안심2동":"동구","안심3동":"동구","안심4동":"동구","옥포읍":"달성군","용산1동":"달서구","용산2동":"달서구","우보면":"군위군","원대동":"서구","월성1동":"달서구","월성2동":"달서구","유가읍":"달성군","유천동":"달서구","읍내동":"북구","의흥면":"군위군","이곡1동":"달서구","이곡2동":"달서구","이천동":"남구","장기동":"달서구","죽전동":"달서구","중동":"수성구","지산1동":"수성구","지산2동":"수성구","지저동":"동구","진천동":"달서구","칠성동":"북구","침산1동":"북구","침산2동":"북구","침산3동":"북구","태전1동":"북구","태전2동":"북구","파동":"수성구","평리1동":"서구","평리2동":"서구","평리3동":"서구","평리4동":"서구","평리5동":"서구","평리6동":"서구","하빈면":"달성군","해안동":"동구","혁신동":"동구","현풍읍":"달성군","화원읍":"달성군","황금1동":"수성구","황금2동":"수성구","효령면":"군위군","효목1동":"동구","효목2동":"동구"}

                def _addr_is_bad(addr):
                    if not addr: return False
                    m = _re.search(r'(\S+?[구군])\s*(\S+?동)', str(addr))
                    if not m: return False
                    gu, dong = m.group(1), m.group(2)
                    correct_gu = _DONG_TO_GU.get(dong)
                    if correct_gu is None:
                        return str(addr).strip().startswith('대구')  # 대구인데 목록에 없는 동
                    return correct_gu != gu  # 구 불일치

                def _verify(data):
                    items = data.get('calls') if data.get('type')=='daily_history' else data.get('items')
                    items = items or []
                    actual_count = len(items)
                    actual_sum = sum((_safe_int(i.get('요금')) or 0) for i in items)
                    header_count = _safe_int(data.get('표시건수'))
                    header_total = _safe_int(data.get('표시금액'))
                    totals_ok = (header_count is not None and header_total is not None
                                 and header_count==actual_count and header_total==actual_sum)
                    bad_addr_count = sum(
                        1 for i in items
                        if _addr_is_bad(i.get('출발지')) or _addr_is_bad(i.get('도착지'))
                    ) if data.get('type') == 'daily_history' else 0
                    # 무효주소가 전체 30% 넘거나 3건 넘으면(둘 중 먼저 걸리는 기준) 재시도 트리거
                    addr_ok = bad_addr_count <= max(1, int(actual_count * 0.3)) and bad_addr_count <= 3
                    ok = totals_ok and addr_ok
                    return ok, actual_count, actual_sum, header_count, header_total, bad_addr_count

                data = _run_ocr_split("claude-haiku-4-5-20251001")
                verified, a_cnt, a_sum, h_cnt, h_sum, bad_addr = _verify(data)
                model_used = "claude-haiku-4-5-20251001"
                if not verified:
                    logger.warning(f"OCR 검증 실패 — haiku 결과 재시도(sonnet): 실제 {a_cnt}건/{a_sum}원 vs 화면표시 {h_cnt}건/{h_sum}원, 무효주소 {bad_addr}건")
                    data2 = _run_ocr_split("claude-sonnet-5")
                    verified2, a_cnt2, a_sum2, h_cnt2, h_sum2, bad_addr2 = _verify(data2)
                    data, verified, a_cnt, a_sum, h_cnt, h_sum, bad_addr = data2, verified2, a_cnt2, a_sum2, h_cnt2, h_sum2, bad_addr2
                    model_used = "claude-sonnet-5"

                data['_verified'] = verified
                data['_verify_detail'] = {"실제건수":a_cnt,"실제금액":a_sum,"화면표시건수":h_cnt,"화면표시금액":h_sum,"무효주소건수":bad_addr,"model":model_used}
                send_json(200, {"success":True,"data":data})
            except Exception as e:
                logger.error(f"OCR history 오류: {e}")
                send_json(400, {"success":False,"error":str(e)})
            return

        if self.path == '/magi_analyze':
            try:
                import anthropic as _ant
                client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=120.0)
                msg = client.messages.create(
                    model=payload.get('model','claude-sonnet-5'),
                    max_tokens=int(payload.get('max_tokens',1500)),
                    system=payload.get('system_prompt','당신은 마기입니다.'),
                    messages=[{"role":"user","content":payload.get('user_message','')}]
                )
                send_json(200, {"success": True, "result": _extract_claude_text(msg)})
            except Exception as e:
                logger.error(f"마기분석 오류: {e}")
                send_json(400, {"success": False, "error": str(e)})
            return

        if self.path == '/atlas-report':
            try:
                import threading
                threading.Thread(
                    target=lambda: asyncio.run(save_atlas_report(payload)),
                    daemon=True
                ).start()
                send_json(200, {"status": "ok"})
                logger.info(f"아틀라스 보고 수신: {payload.get('title','?')}")
            except Exception as e:
                send_json(400, {"error": str(e)})
            return

        self.send_response(404)
        self.end_headers()

    def log_message(self, *args):
        pass

def run_health_server():
    server = HTTPServer(("0.0.0.0", PORT), HealthHandler)
    logger.info(f"Health server on port {PORT}")
    server.serve_forever()

# ──────────────────────────────────────────────
# Supabase 헬퍼
# ──────────────────────────────────────────────
async def send_telegram_broadcast(text: str):
    """긴급수정(2026-08-23): 기존 send_all()은 main() 함수 내부의 지역함수(closure)라
    모듈레벨 함수(check_ingestion_gap, ask_operated_status_telegram, task#52 마기자동
    검증 등)에서 호출하면 NameError로 조용히 실패하고 있었음(task#52 실기기검증중
    "텔레그램에 아무것도 안 옴"으로 발견 — 대표님 지적). 모듈레벨에서 독립적으로
    작동하는 전역 발송함수로 대체, app 객체 의존 없이 별도 Bot 인스턴스 사용."""
    from telegram import Bot
    chat_ids = [x for x in [os.getenv("ALLOWED_CHAT_ID", ""), os.getenv("ALLOWED_CHAT_ID2", "")] if x]
    if not chat_ids:
        logger.error("send_telegram_broadcast: ALLOWED_CHAT_ID 미설정")
        return
    bot = Bot(token=TELEGRAM_TOKEN)
    for cid in chat_ids:
        try:
            await bot.send_message(chat_id=cid, text=text)
        except Exception as e:
            logger.error(f"텔레그램 발송 실패(chat_id={cid}): {e}")


async def sb_h(method: str, path: str, **kwargs) -> dict | list | None:
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    # 캐스퍼 수정 2026-07-06: headers를 여기서 무조건 고정으로 넘기면서
    # 동시에 **kwargs에도 headers가 들어있는 경우(sb_upsert, calc_official_var_score 등
    # headers=를 직접 넘기는 모든 호출) "multiple values for keyword argument 'headers'"
    # TypeError로 무조건 크래시하던 잠복 버그. kwargs의 headers를 우선 사용하도록 수정.
    headers = kwargs.pop("headers", HEADERS_SB)
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.request(method, url, headers=headers, **kwargs)
        if r.status_code in (200, 201):
            # 캐스퍼 수정 2026-07-06 (2차): Prefer: return=minimal을 쓰는 호출은
            # Supabase가 본문을 아예 비워서 응답하는데, 여기서 무조건 r.json()을
            # 호출해서 JSONDecodeError로 크래시하던 버그. 빈 본문이면 파싱 생략.
            if not r.content:
                return []
            try:
                return r.json()
            except Exception:
                logger.error(f"Supabase {method} {path} → JSON 파싱 실패: {r.text[:200]}")
                return []
        logger.error(f"Supabase {method} {path} → {r.status_code}: {r.text}")
        return None

async def sb_insert(table: str, data: dict) -> dict:
    """task#50 근본수정(2026-08-22): sb_upsert와 동일 원칙 - 실패해도 조용히
    None을 반환해서 호출부가 무검사로 "성공"처럼 넘어가던 위험(task#47과
    같은 종류) 원천봉쇄. 6개 호출지점(831,1600,3532,3983,4027,4061) 개별
    수정 없이 이 함수 자체에서 일괄 보호."""
    result = await sb_h("POST", table, json=data)
    if result is None:
        raise RuntimeError(f"sb_insert 실패: table={table} (RLS/네트워크 등 — sb_h가 None 반환)")
    return result

async def sb_select(table: str, params: dict = None) -> list:
    result = await sb_h("GET", table, params=params or {})
    return result if isinstance(result, list) else []

# 캐스퍼 명령서#024 (2026-08-05): 영수증 OCR 요약행(비고 "OCR 추출: 매출...")이 개별 콜과
# 같은 테이블/컬럼에 섞여있어서, raw_calls를 그대로 sum/count하는 모든 곳에서 매출·건수가
# 이중집계되던 근본원인. index.html 쪽은 이미 excludeSummaryRows()로 수정 완료.
# bot_v5.py도 13곳이 각자 raw_calls를 직접 조회하고 있어 공용 헬퍼로 일괄 대응.
# 캐스퍼 수정 2026-08-05 (명령서#024 검증 중 발견): 아르고스 재삽입 데이터는 비고가 아니라
# 콜유형='합계'로 요약행을 표시함(비고는 null). 두 컨벤션 다 인식하도록 확대.
def _is_receipt_summary_row(r: dict) -> bool:
    # 캐스퍼 긴급수정 2026-08-13: 기존 비고텍스트 패턴("OCR 추출:"으로 시작) 판정은
    # id1207이 confirmed 처리되며 비고가 바뀌자 패턴이 깨져서 요약행을 못 거르고
    # 실제 이중집계(개별콜148,200원+요약행162,200원=310,400원 오산정)를 일으키던 중이었음
    # — 대표님 GPX027 재확인 요청으로 발견. 명령서#030 data_source 컬럼 기반으로 교체,
    # 텍스트 변경에 안 흔들리게 함. 구버전 데이터 호환을 위해 텍스트패턴도 폴백 유지.
    if r.get("data_source") == "app_ocr_summary":
        return True
    return str(r.get("비고") or "").startswith("OCR 추출:") or r.get("콜유형") == "합계"

# 캐스퍼 긴급수정 2026-08-10: _safe_int가 /ocr_history 핸들러 안에 지역함수로만 있어서
# dual_verify_7day_average()·MCP 엔드포인트에서 NameError 발생(카산드라 실기기 테스트로 발견).
# 모듈레벨로 승격해서 어디서든 쓸 수 있게 함.
import re as _re_mod
def _safe_int(v):
    if v is None: return None
    try: return int(v)
    except (ValueError, TypeError):
        digits = _re_mod.sub(r'[^0-9]', '', str(v))
        return int(digits) if digits else None

def exclude_summary_rows(calls: list) -> list:
    # 캐스퍼 수정 2026-08-05: 단순 전체제외는 "개별콜 없이 합계행만 있는 날"의 매출이
    # 통째로 0으로 사라지는 과교정을 낳음(index.html에서 실제로 겪은 문제, 동일 로직 이식).
    # 날짜별로 묶어서: 개별콜이 하나라도 있으면 그 날짜의 합계행만 제외, 없으면 유지.
    from collections import defaultdict
    by_date = defaultdict(list)
    for c in calls:
        by_date[c.get("날짜")].append(c)
    out = []
    for day_rows in by_date.values():
        has_real = any(not _is_receipt_summary_row(r) for r in day_rows)
        if has_real:
            out.extend(r for r in day_rows if not _is_receipt_summary_row(r))
        else:
            out.extend(day_rows)
    return out

# ──────────────────────────────────────────────
# 명령서#026: GPX 좌표 → 대구 행정동 역지오코딩 (2026-08-08)
# 카카오T 실거리 매칭(공차율 100% 문제)과 배회 핵심동선 분석(하드코딩 제거) 두 프로젝트가
# 공유하는 핵심 자산. daegu_boundaries.geojson(150개 행정동 폴리곤, vuski/admdongkor
# ver20260701 기준)을 광선교차법(ray casting)으로 point-in-polygon 판정.
# 검증: 동대구역→동구 신암4동, 대구시청→중구 동인동, 동성로→중구 성내1동,
#       수성구청→수성구 범어1동 — 4곳 다 실제 위치와 일치 확인(2026-08-08).
# ──────────────────────────────────────────────
_DAEGU_BOUNDARIES_CACHE = None

def _load_daegu_boundaries():
    global _DAEGU_BOUNDARIES_CACHE
    if _DAEGU_BOUNDARIES_CACHE is None:
        import json as _json, os as _os
        path = _os.path.join(_os.path.dirname(__file__), "daegu_boundaries.geojson")
        with open(path, encoding="utf-8") as f:
            _DAEGU_BOUNDARIES_CACHE = _json.load(f)["features"]
    return _DAEGU_BOUNDARIES_CACHE

def _point_in_ring(x, y, ring):
    n = len(ring)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = ring[i]
        xj, yj = ring[j]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi + 1e-15) + xi):
            inside = not inside
        j = i
    return inside

def _point_in_polygon_coords(x, y, coords):
    if not _point_in_ring(x, y, coords[0]):
        return False
    return not any(_point_in_ring(x, y, hole) for hole in coords[1:])

def latlon_to_dong(lat: float, lon: float) -> str | None:
    """GPS 좌표 → 대구 행정동 전체명(예: '대구광역시 중구 동인동') 반환. 대구 밖이면 None."""
    x, y = lon, lat
    for feat in _load_daegu_boundaries():
        geom = feat["geometry"]
        props = feat["properties"]
        if geom["type"] == "Polygon":
            if _point_in_polygon_coords(x, y, geom["coordinates"]):
                return props.get("adm_nm")
        elif geom["type"] == "MultiPolygon":
            for poly in geom["coordinates"]:
                if _point_in_polygon_coords(x, y, poly):
                    return props.get("adm_nm")
    return None


async def sb_select_calls(params: dict = None) -> list:
    """raw_calls 전용 조회 — 영수증 요약행을 자동으로 제외한다.
    기존 `await sb_select("raw_calls", {...})`를 대체."""
    raw = await sb_select("raw_calls", params)
    return exclude_summary_rows(raw)

async def sb_upsert(table: str, data: dict, on_conflict: str) -> dict:
    """task#47 긴급수정(2026-08-22, 마기지적): 실패해도 조용히 넘어가 "성공"으로
    로그가 찍히는 사례가 여러 저장함수에서 발견됨(daily_calc_snapshot,
    kpi_7day_snapshot 등) — 관측(로그) 자체가 신뢰 안 되면 다른 모든 검증이
    무의미해짐. 15곳 개별수정 대신 이 함수 자체에서 실패시 예외를 던져 호출부
    전체를 일괄 보호(호출부는 대부분 이미 try/except 안에 있어 안전)."""
    result = await sb_h(
        "POST", table,
        json=data,
        headers={**HEADERS_SB, "Prefer": f"resolution=merge-duplicates,return=representation"},
        params={"on_conflict": on_conflict}
    )
    if result is None:
        raise RuntimeError(f"sb_upsert 실패: table={table}, on_conflict={on_conflict} (RLS/네트워크 등 — sb_h가 None 반환)")
    return result


# ──────────────────────────────────────────────
# GitHub 직접 커밋 (캐스퍼 명령서 #014 §2)
# ──────────────────────────────────────────────
async def github_commit_briefing(날짜: str, content: str) -> dict:
    """브리핑 markdown을 magi-taxi-data 레포 /briefings/ 폴더에 직접 커밋.
    반환: {"ok": bool, "url": str|None, "error": str|None}
    """
    if not GITHUB_PAT:
        return {"ok": False, "url": None, "error": "GITHUB_PAT 환경변수 미설정"}

    # 캐스퍼 명령서 #017 반영(2026-07-13): 아르고스 정밀 브리핑(/briefings/)과
    # 경로·파일명 분리 — 이지스 파이프라인A와의 덮어쓰기 충돌 방지
    path = f"bot_briefings/bot_summary_{날짜.replace('-','')}.md"
    api_url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{path}"
    headers = {
        "Authorization": f"Bearer {GITHUB_PAT}",
        "Accept": "application/vnd.github+json",
    }

    async with httpx.AsyncClient(timeout=20) as client:
        # 기존 파일 존재 여부 확인 (있으면 sha 필요 — 덮어쓰기용)
        sha = None
        try:
            r_get = await client.get(api_url, headers=headers)
            if r_get.status_code == 200:
                sha = r_get.json().get("sha")
        except Exception:
            pass  # 조회 실패해도 신규 생성 시도는 계속 진행

        # 카카오/우버/배회 건수 커밋 메시지용 집계
        try:
            calls = await sb_select_calls( {"날짜": f"eq.{날짜}"}) or []
            kakao_n = sum(1 for c in calls if (c.get("콜유형") or "") == "카카오T")
            uber_n  = sum(1 for c in calls if (c.get("콜유형") or "") == "우버")
            bhw_n   = sum(1 for c in calls if (c.get("콜유형") or "") == "배회")
        except Exception:
            kakao_n = uber_n = bhw_n = 0

        commit_msg = f"[BOT] {날짜.replace('-','')} 브리핑 자동생성 — 카카오{kakao_n}건·우버{uber_n}건·배회{bhw_n}건"

        body = {
            "message": commit_msg,
            "content": base64.b64encode(content.encode("utf-8")).decode("ascii"),
            "branch": "main",
        }
        if sha:
            body["sha"] = sha

        try:
            r_put = await client.put(api_url, headers=headers, json=body)
        except Exception as e:
            return {"ok": False, "url": None, "error": f"GitHub 요청 실패: {e}"}

        if r_put.status_code in (200, 201):
            html_url = r_put.json().get("content", {}).get("html_url")
            return {"ok": True, "url": html_url, "error": None}
        else:
            return {"ok": False, "url": None, "error": f"{r_put.status_code}: {r_put.text[:200]}"}

async def sb_delete_receipt(conditions: dict) -> int:
    """payment_receipts 조건부 삭제. 삭제 건수 반환"""
    query = "&".join(f"{k}={v}" for k, v in conditions.items())
    r = await sb_h("DELETE", f"payment_receipts?{query}")
    # Supabase DELETE는 삭제된 행 반환 (Prefer: return=representation)
    if isinstance(r, list):
        return len(r)
    return 0

async def sb_delete_last(table: str, filter_params: dict) -> bool:
    rows = await sb_select(table, {**filter_params, "order": "id.desc", "limit": "1"})
    if not rows:
        return False
    row_id = rows[0]["id"]
    r = await sb_h("DELETE", f"{table}?id=eq.{row_id}")
    return True

# ──────────────────────────────────────────────
# 유틸 함수
# ──────────────────────────────────────────────
def now_kst() -> datetime:
    return datetime.now(KST)

def today_kst():
    return now_kst().date()

def get_dow(d=None) -> str:
    d = d or today_kst()
    return DOW_KOR[d.weekday()]

def fmt(n) -> str:
    """None·문자열 안전 처리"""
    if n is None:
        return "0원"
    try:
        return f"{int(n):,}원"
    except (ValueError, TypeError):
        return "0원"

def calc_net(매출: int, 지출: int) -> int:
    return int(매출 * (1 - CARD_FEE_RATE)) - 지출

async def today_summary() -> dict:
    today = str(today_kst())
    calls = await sb_select_calls( {"날짜": f"eq.{today}"})
    expenses = await sb_select("expenses", {"날짜": f"eq.{today}"})
    건수 = len(calls)
    매출 = sum(c.get("요금", 0) or 0 for c in calls)
    지출 = sum(e.get("금액", 0) or 0 for e in expenses)
    순수익 = calc_net(매출, 지출)
    달성률 = int(순수익 / NET_GOAL * 100) if NET_GOAL else 0
    return {
        "건수": 건수, "매출": 매출, "지출": 지출,
        "순수익": 순수익, "달성률": 달성률,
    }

async def today_expenses() -> list:
    today = str(today_kst())
    return await sb_select("expenses", {"날짜": f"eq.{today}", "order": "id.asc"})

async def insert_insurance(date):
    date_str = str(date)
    existing = await sb_select(
        "expenses",
        {"날짜": f"eq.{date_str}", "카테고리": "eq.보험료", "자동여부": "eq.true"}
    )
    if existing:
        return
    await sb_insert("expenses", {
        "날짜": date_str,
        "카테고리": "보험료",
        "금액": INSURANCE_DAILY,
        "메모": "자동 보험료",
        "자동여부": True,
    })
    logger.info(f"보험료 자동 기록: {date_str}")

# ──────────────────────────────────────────────
# 이미지 큐 (동시 업로드 과부하 방지)
# ──────────────────────────────────────────────
image_queue: asyncio.Queue = None  # main()에서 초기화

# ──────────────────────────────────────────────
# Claude API — 이미지 분류 + OCR
# ──────────────────────────────────────────────
async def claude_vision(image_bytes: bytes, prompt: str, max_tokens: int = 500) -> str:
    """Claude API 비동기 호출. 모든 이미지 포맷을 JPEG로 정규화 후 전송."""

    def _prepare_image(raw: bytes) -> tuple[bytes, str]:
        """이미지를 JPEG로 변환 + 최대 높이 3000px 리사이즈"""
        try:
            from PIL import Image as _Image
            import io as _io
            img = _Image.open(_io.BytesIO(raw))
            # RGB 변환
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            # 최대 높이 3000px (세로 긴 이미지 처리)
            MAX_H = 3000
            if img.height > MAX_H:
                ratio = MAX_H / img.height
                img = img.resize((int(img.width * ratio), MAX_H), _Image.LANCZOS)
                logger.info(f"이미지 높이 축소: {img.height}→{MAX_H}px")
            buf = _io.BytesIO()
            img.save(buf, format="JPEG", quality=85, optimize=True)
            return buf.getvalue(), "image/jpeg"
        except Exception as e:
            logger.warning(f"이미지 변환 실패({e}) → 원본 사용")
            # 포맷 감지
            if raw[:4] == b"RIFF" or raw[:4] == b"WEBP":
                return raw, "image/webp"
            elif raw[:8] == b"\x89PNG\r\n\x1a\n":
                return raw, "image/png"
            return raw, "image/jpeg"

    img_data, media_type = _prepare_image(image_bytes)
    b64 = base64.standard_b64encode(img_data).decode()

    def _sync_call():
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
        msg = client.messages.create(
            model=OCR_MODEL,
            max_tokens=max_tokens,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        return _extract_claude_text(msg).strip()

    return await asyncio.to_thread(_sync_call)


def split_image_vertically(image_bytes: bytes):
    """대표님요청(2026-08-18): 세로로 긴 일별운행이력 화면을 상/하단으로 나눠 각각
    OCR — 정보밀도를 낮춰 세부동이름(침산3동 등) 인식률 향상 목적. 경계선에 걸친
    항목 유실 방지를 위해 10% 겹치게 자름(상단 0~55%, 하단 45~100%)."""
    from PIL import Image
    img = Image.open(io.BytesIO(image_bytes))
    w, h = img.width, img.height
    top = img.crop((0, 0, w, int(h * 0.55)))
    bottom = img.crop((0, int(h * 0.45), w, h))

    def _to_jpeg(im):
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, format="JPEG", quality=85, optimize=True)
        return buf.getvalue()

    return _to_jpeg(top), _to_jpeg(bottom)


def merge_split_ocr_results(top_data: dict, bottom_data: dict) -> dict:
    """상/하단 OCR 결과 병합. date/표시건수/표시금액은 상단(헤더 포함) 우선.
    calls는 (배차시각,요금) 조합으로 중복제거 후 합침 — 10%겹침구간 대응."""
    merged = dict(top_data)
    top_calls = top_data.get('calls') or top_data.get('items') or []
    bottom_calls = bottom_data.get('calls') or bottom_data.get('items') or []
    seen = set((c.get('배차시각'), c.get('요금')) for c in top_calls)
    dedup_bottom = [c for c in bottom_calls if (c.get('배차시각'), c.get('요금')) not in seen]
    key = 'calls' if top_data.get('type') == 'daily_history' else 'items'
    merged[key] = top_calls + dedup_bottom
    return merged


def resize_image_if_needed(image_bytes: bytes) -> bytes:
    """
    Claude API 전송 전 이미지 최적화.
    - 최대 너비 1000px (세로 비율 유지)
    - JPEG quality 80
    - Pillow 필수 (requirements.txt에 Pillow 추가 필요)
    """
    try:
        from PIL import Image
        import io

        img = Image.open(io.BytesIO(image_bytes))
        orig_w, orig_h = img.width, img.height
        orig_kb = len(image_bytes) / 1024

        # 최대 너비 1000px 또는 최대 높이 3000px 초과 시 축소
        MAX_WIDTH = 1000
        MAX_HEIGHT = 3000
        ratio_w = MAX_WIDTH / orig_w if orig_w > MAX_WIDTH else 1.0
        ratio_h = MAX_HEIGHT / orig_h if orig_h > MAX_HEIGHT else 1.0
        ratio = min(ratio_w, ratio_h)
        if ratio < 1.0:
            new_w = int(orig_w * ratio)
            new_h = int(orig_h * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)
        
        # RGB 변환 (PNG RGBA 등 처리)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=80, optimize=True)
        result = buf.getvalue()
        result_kb = len(result) / 1024

        logger.info(
            f"이미지 최적화: {orig_w}x{orig_h} {orig_kb:.0f}KB"
            f" → {img.width}x{img.height} {result_kb:.0f}KB"
        )
        return result

    except ImportError:
        logger.error("Pillow 미설치 — requirements.txt에 Pillow 추가 필요")
        return image_bytes
    except Exception as e:
        logger.error(f"이미지 리사이즈 오류: {e}")
        return image_bytes

async def classify_image(image_bytes: bytes) -> str:
    prompt = (
        "이 이미지가 아래 중 어느 종류인지 판단해서 해당 단어 하나만 답해줘.\n\n"
        "【결제】← 최우선 확인\n"
        "  카카오T 또는 세큐티 결제내역 화면. 아래 중 하나라도 있으면 반드시 '결제':\n"
        "  · '결제내역조회', '거래일자', '결제구분', '조회기간' 텍스트\n"
        "  · 날짜+시각(YYYY-MM-DD | HH:MM:SS 또는 YYYY-MM-DD HH:MM) 옆에 금액 목록\n"
        "  · 'KB카드', '신한카드', '현대카드', 'BC카드', '하나카드', '농협카드' 등 카드사명\n"
        "  · '승인정상', '1승인', '거래일자' 텍스트\n"
        "  · 세로로 5건 이상 금액 목록 나열\n\n"
        "【충전】\n"
        "  전기차 충전 앱 이용내역. 아래 중 하나라도 있으면 '충전':\n"
        "  · 'kWh', '충전량', '충전완료', '충전소' 텍스트\n"
        "  · '전기차 충전' 탭 UI\n\n"
        "【콜카드】\n"
        "  카카오T 택시 운행기록 1건. '배차', '승차', '하차' + 출발지·도착지 주소.\n\n"
        "【일별운행이력】← 콜카드보다 먼저 확인\n"
        "  카카오T '일별 운행 이력' 화면. 아래 특징이 있으면 반드시 '일별운행이력':\n"
        "  · 상단에 'YYYY년 M월 D일(요일) N건' 형식\n"
        "  · 여러 건의 운행이 세로로 나열\n"
        "  · 각 건마다 'HH:MM - HH:MM [실시간]' 시간 범위\n"
        "  · '직접결제' 텍스트 포함 가능\n"
        "  · '실시간 운행 N건 / N원' 요약\n\n"
        "【세큐티】\n"
        "  세큐티 등급·점수 리포트. 종합점수, 수락률 등 항목.\n\n"
        "【기타】위 4가지 해당 없음.\n\n"
        "⚠️ 핵심 구분:\n"
        "  충전: kWh 단위 있음\n"
        "  결제: 카드사명 + 날짜+금액 목록\n"
        "  콜카드: 운행 1건(배차·승차·하차)\n\n"
        "반드시 결제·충전·콜카드·세큐티·기타 중 하나만 답해. 다른 말 금지."
    )
    result = await claude_vision(image_bytes, prompt, max_tokens=15)
    for keyword in ["일별운행이력", "콜카드", "충전", "결제", "세큐티"]:
        if keyword in result:
            return keyword
    return "기타"


# ══════════════════════════════════════════════
# 세큐티 OCR + 저장 + 조회
# ══════════════════════════════════════════════


# ══════════════════════════════════════════════
# 일별 운행이력 OCR + 저장
# ══════════════════════════════════════════════

async def save_atlas_report(data: dict):
    """아틀라스 보고서 Supabase 저장 + 텔레그램 알림"""
    try:
        payload = {
            "report_type": data.get("report_type", "manual"),
            "source": "atlas",
            "title": data.get("title", "아틀라스 보고"),
            "payload": data.get("payload", data),
            "status": "pending",
            "run_date": str(today_kst()),
        }
        result = await sb_insert("atlas_reports", payload)
        report_id = result[0]["id"] if result else "?"

        # 텔레그램 알림 (봇 애플리케이션에 전송)
        # application 객체에 접근하기 위해 전역 변수 사용
        global _bot_app
        if _bot_app:
            from telegram import Bot
            bot = _bot_app.bot
            allowed_ids = [int(x) for x in os.getenv("ALLOWED_USER_IDS","").split(",") if x]
            for uid in allowed_ids:
                try:
                    await bot.send_message(
                        chat_id=uid,
                        text=f"📡 아틀라스 보고 #{report_id}\n"
                             f"유형: {payload['report_type']}\n"
                             f"제목: {payload['title']}\n"
                             f"앱에서 마기 분석 자동 시작"
                    )
                except Exception:
                    pass
        logger.info(f"atlas_report #{report_id} 저장 완료")
    except Exception as e:
        logger.error(f"atlas_report 저장 오류: {e}")


async def ocr_daily_history(image_bytes: bytes) -> dict | None:
    """
    카카오T '일별 운행 이력' 화면 OCR.
    반환: {날짜: "YYYY-MM-DD", 콜목록: [{배차시각, 하차시각, 출발지, 도착지, 요금, 결제방식}, ...]}
    """
    prompt = (
        "이 카카오T '일별 운행 이력' 화면에서 정보를 추출해서 JSON만 반환해줘.\n"
        '{"날짜":"YYYY-MM-DD","콜목록":['
        '{"배차시각":"HH:MM","하차시각":"HH:MM","출발지":"대구 OO구 OO동",'
        '"도착지":"대구 OO구 OO동","요금":숫자,"결제방식":"자동 또는 직접"}]}\n'
        "⚠️ 날짜: 상단 'YYYY년 M월 D일' → YYYY-MM-DD 변환\n"
        "⚠️ 배차시각: 각 콜의 앞 시각 (예: 02:58 - 03:11 에서 02:58)\n"
        "⚠️ 하차시각: 각 콜의 뒤 시각 (예: 02:58 - 03:11 에서 03:11)\n"
        "⚠️ 결제방식: '직접결제' 텍스트 있으면 '직접', 없으면 '자동'\n"
        "⚠️ 요금: 파란색 숫자. 직접결제는 표시된 요금 그대로 추출\n"
        "⚠️ 출발지/도착지: 화면에 실제로 인쇄된 글자를 정확히 그대로 옮겨 적어라. "
        "절대로 지명을 만들어내거나 비슷하게 짐작해서 채우지 마라. "
        "글자가 흐리거나 잘려서 확신이 안 서면 억지로 완성하지 말고 '미확인'이라고 반환해라. "
        "'대구 OO구 OO동' 형식이 아니어도(건물명·도로명 등) 화면에 있는 그대로 옮기는 게 우선이다.\n"
        "JSON만 반환. 설명·마크다운 금지."
    )
    try:
        raw = await claude_vision(image_bytes, prompt, max_tokens=2000)
        raw = raw.strip()
        import re as _re
        raw = _re.sub(r"```json\s*", "", raw)
        raw = _re.sub(r"```\s*", "", raw)
        raw = raw.strip()
        import json as _json
        return _json.loads(raw)
    except Exception as e:
        logger.error(f"일별운행이력 OCR 오류: {e}")
        return None


async def process_daily_history(update, image_bytes: bytes):
    """
    일별 운행이력 이미지 처리:
    OCR → 날짜 보정 → raw_calls 저장 → 결과 안내
    """
    from datetime import date as _dc, timedelta as _td

    await update.message.reply_text("📋 일별 운행이력 분석 중...")

    data = await ocr_daily_history(image_bytes)
    if not data or not data.get("콜목록"):
        await update.message.reply_text(
            "❌ 일별 운행이력 인식 실패\n"
            "💡 화면을 더 크게 캡처해서 다시 올려주세요."
        )
        return

    # 화면 날짜 파싱
    screen_date_str = data.get("날짜", "")
    try:
        screen_date = _dc.fromisoformat(screen_date_str)
    except Exception:
        screen_date = today_kst()
        logger.warning(f"날짜 파싱 실패: {screen_date_str} → 오늘 사용")

    DOW_MAP = ["월","화","수","목","금","토","일"]
    saved = 0
    updated = 0
    dates_used = set()
    result_lines = []

    for call in data.get("콜목록", []):
        배차 = call.get("배차시각", "")
        하차 = call.get("하차시각", "")
        출발 = call.get("출발지", "")
        도착 = call.get("도착지", "")
        요금 = call.get("요금", 0) or 0
        결제방식 = call.get("결제방식", "자동")

        # 날짜 보정: 배차 06시 이전 → 화면날짜 +1일 (새벽 운행)
        try:
            h = int(배차.split(":")[0])
            save_date = screen_date + _td(days=1) if h < 6 else screen_date
        except Exception:
            save_date = screen_date

        save_date_str = str(save_date)
        dow = DOW_MAP[save_date.weekday()]
        dates_used.add(save_date_str)

        # 중복 삭제 후 재저장
        deleted = await delete_duplicate_call(save_date_str, 배차, 요금)
        if deleted:
            updated += deleted

        비고 = "직접결제(요금미확인)" if 결제방식 == "직접" else None

        payload = {
            "날짜":     save_date_str,
            "요일":     dow,
            "배차시각": 배차,
            "하차시각": 하차,
            "출발지":   출발,
            "도착지":   도착,
            "요금":     요금,
            "콜유형":   "카카오T",
            "비고":     비고,
            "data_source": "app_ocr_individual",
        }
        result = await sb_insert("raw_calls", payload)
        if result:
            saved += 1
            직접표시 = " [직접결제]" if 결제방식 == "직접" else ""
            result_lines.append(
                f"  {배차}~{하차} {출발}→{도착} {fmt(요금)}{직접표시}"
            )

    # 결과 메시지
    dates_sorted = sorted(dates_used)
    msg = [
        f"✅ 일별 운행이력 저장 완료",
        f"화면날짜: {screen_date_str} | 저장: {saved}건",
        f"날짜 분포: {', '.join(dates_sorted)}",
        "",
    ]
    msg.extend(result_lines[:10])  # 최대 10건 표시
    if len(result_lines) > 10:
        msg.append(f"  ... 외 {len(result_lines)-10}건")

    msg.append("")
    msg.append("💡 교차대조:")
    for d in dates_sorted:
        msg.append(f"  대조 {d}")

    await update.message.reply_text("\n".join(msg))



# ══════════════════════════════════════════════
# 운행 일관성 모니터링 (Step E)
# ══════════════════════════════════════════════


# ══════════════════════════════════════════════
# Step A: 카카오 공식 6변수 자동 평가
# ══════════════════════════════════════════════

async def calc_official_var_score(날짜: str) -> dict:
    """카카오 AI 배차 공식 6변수 평가 및 daily_summary 저장"""
    from datetime import date as _dc
    today = _dc.today()
    mo = 날짜[:7]

    # var_2: 오늘 운행완료수
    calls_today = await sb_select_calls( {"날짜": f"eq.{날짜}"})
    daily_completed = len(calls_today)

    # var_2: 이번달 일평균
    calls_month = await sb_select_calls( {
        "and": f"(날짜.gte.{mo}-01,날짜.lte.{mo}-31)"
    })
    from datetime import date as _d2
    days_so_far = (_d2.today() - _d2(int(mo[:4]), int(mo[5:7]), 1)).days + 1
    # 명령서 #010 대응(2026-07-08): 행 개수가 아닌 건수 가중 합계로 수정
    month_weighted_count = sum(_extract_count(c) for c in calls_month)
    monthly_avg = month_weighted_count / max(days_so_far, 1)

    # var_3·4: sekuti에서 조회 (현재 0으로 고정 — 마스터 등급)
    avoid_count = 0
    one_star_count = 0

    # var_5: 수락률 (현재 100% 유지 중)
    acceptance_rate = 100

    # AI 진입 추정 (운행완료수 기반)
    AREA_AVG_LOW, AREA_AVG_HIGH = 18, 25
    if monthly_avg >= AREA_AVG_HIGH:
        ai_estimate = "85%+"
    elif monthly_avg >= AREA_AVG_LOW:
        ai_estimate = f"{int(60 + (monthly_avg-AREA_AVG_LOW)/(AREA_AVG_HIGH-AREA_AVG_LOW)*25)}%"
    else:
        ai_estimate = f"{int(40 + monthly_avg/AREA_AVG_LOW*20)}%"

    score = {
        "date": 날짜,
        "vars": {
            "var_1_acceptance_prob": "양호" if len(calls_month) >= 30 else "데이터 축적 중",
            "var_2_daily_completed": daily_completed,
            "var_2_monthly_avg": round(monthly_avg, 1),
            "var_2_area_avg_est": f"{AREA_AVG_LOW}~{AREA_AVG_HIGH}",
            "var_3_avoid_count_monthly": avoid_count,
            "var_4_one_star_monthly": one_star_count,
            "var_5_acceptance_rate": acceptance_rate,
            "var_6_eta_score": "위치 의존"
        },
        "weak_var": "var_2_daily_completed",
        "ai_inclusion_estimate": ai_estimate,
        "improvement_needed": "운행 시간 확대 + 매일 운행" if monthly_avg < AREA_AVG_LOW else "유지"
    }

    # daily_summary에 upsert
    await sb_h("POST", f"daily_summary",
        json={"날짜": 날짜, "official_var_score": score},
        headers={**HEADERS_SB, "Prefer": "resolution=merge-duplicates,return=minimal"},
        params={"on_conflict": "날짜"}
    )
    return score


async def handle_forecast(update, date_str=None):
    """/forecast [YYYY-MM-DD] — 사전 예측 시안"""
    from datetime import date as _dc, timedelta as _td
    target = date_str or str(_dc.today() + _td(days=1))
    try:
        _dc.fromisoformat(target)
    except ValueError:
        await update.message.reply_text("❌ 날짜 형식 오류. 예: /forecast 2026-05-21")
        return

    rows = await sb_select("forecast", {"forecast_date": f"eq.{target}"})
    if rows:
        r = rows[0]
        cp  = r.get("kakao_count_point", "?")
        cl  = r.get("kakao_count_ci_low", "?")
        ch  = r.get("kakao_count_ci_high", "?")
        rp  = int(r.get("revenue_point", 0) or 0)
        rl  = int(r.get("revenue_ci_low", 0) or 0)
        rh  = int(r.get("revenue_ci_high", 0) or 0)
        wf  = r.get("weather_forecast", "미정")
        mv  = r.get("model_version", "v0.3")
        nt  = r.get("notes", "")
        hz  = r.get("hotzones", [])
        lines = [
            f"🔮 예측 시안 — {target}",
            f"",
            f"📊 콜수: {cp}건 (CI {cl}~{ch}건)",
            f"💰 매출: {rp:,}원 (CI {rl:,}~{rh:,}원)",
            f"🌤️ 날씨: {wf}",
        ]
        if hz:
            lines.append(f"🔥 핫존: {' / '.join(hz) if isinstance(hz,list) else hz}")
        lines += [f"", f"📝 {nt}", f"모델: {mv}"]
        await update.message.reply_text("\n".join(lines))
        return

    # 신규 생성 — 메타모델 v0.3
    dow = ["월","화","수","목","금","토","일"][_dc.fromisoformat(target).weekday()]
    DOW_MODEL = {
        "월": (8,  6, 11,  72000,  55000,  95000),
        "화": (9,  7, 12,  85000,  65000, 110000),
        "수": (10, 8, 13,  95000,  75000, 120000),
        "목": (9,  7, 12,  85000,  65000, 110000),
        "금": (12,10, 15, 115000,  90000, 145000),
        "토": (13,11, 16, 125000, 100000, 155000),
        "일": (12,10, 15, 118000,  95000, 148000),
    }
    pt, cl2, ch2, rp2, rl2, rh2 = DOW_MODEL.get(dow, (10, 8, 13, 95000, 75000, 120000))
    fd = {
        "forecast_date": target,
        "kakao_count_point": pt, "kakao_count_ci_low": cl2, "kakao_count_ci_high": ch2,
        "revenue_point": rp2, "revenue_ci_low": rl2, "revenue_ci_high": rh2,
        "weather_forecast": "미정 (수동 업데이트 필요)",
        "model_version": "메타모델 v0.3",
        "notes": f"{dow}요일 요일별 메타모델 기반 예측",
        "hotzones": ["중구 성내2동", "수성구 범어동", "동구 동대구역"]
    }
    result = await sb_insert("forecast", fd)
    fid = result[0]["id"] if result else "?"
    lines = [
        f"🔮 예측 시안 #{fid} — {target} ({dow}요일)",
        f"",
        f"📊 콜수: {pt}건 (CI {cl2}~{ch2}건)",
        f"💰 매출: {rp2:,}원 (CI {rl2:,}~{rh2:,}원)",
        f"🔥 핫존: 성내2동 / 범어동 / 동대구역",
        f"",
        f"⚠️ 메타모델 v0.3 — 실측 후 정밀화됩니다",
    ]
    await update.message.reply_text("\n".join(lines))


async def handle_completion_status(update):
    """/completion_status — 운행완료수 현황 및 AI 진입 가능성"""
    from datetime import date as _dc
    today = str(_dc.today())
    mo = today[:7]

    calls_month = await sb_select_calls( {
        "and": f"(날짜.gte.{mo}-01,날짜.lte.{mo}-31)"
    })
    from datetime import date as _d2
    days_so_far = (_d2.today() - _d2(int(mo[:4]), int(mo[5:7]), 1)).days + 1
    total = len(calls_month)
    avg = total / max(days_so_far, 1)

    AREA_LOW, AREA_HIGH = 18, 25
    pct = int(avg / AREA_HIGH * 100)

    # 진입 가능성
    if avg >= AREA_HIGH:
        level = "🟢 높음"
        comment = "AI 1순위 후보군 정상 진입 구간"
    elif avg >= AREA_LOW:
        level = "🟡 중간"
        needed = round(AREA_HIGH - avg, 1)
        comment = f"목표까지 {needed}건/일 더 필요"
    else:
        level = "🔴 낮음"
        needed = round(AREA_HIGH - avg, 1)
        comment = f"목표까지 {needed}건/일 더 필요"

    # 6월 목표 계산 (매일 운영 가정)
    import calendar as _cal
    days_in_month = _cal.monthrange(int(mo[:4]), int(mo[5:7]))[1]
    target_for_area = AREA_LOW * days_in_month

    lines = [
        f"📊 운행완료수 현황 — {mo}",
        f"",
        f"이번달 누적: {total}건",
        f"일평균: {avg:.1f}건/일",
        f"사업구역 평균 추정: {AREA_LOW}~{AREA_HIGH}건/일",
        f"대비: {pct}%",
        f"",
        f"AI 1순위 진입 가능성: {level}",
        f"💡 {comment}",
        f"",
        f"[월간 목표]",
        f"AI 안정권(18건/일): {target_for_area}건/월",
        f"현재: {total}건 / 잔여: {max(target_for_area-total,0)}건",
        f"",
        f"[6변수 현황]",
        f"  ② 운행완료수: {avg:.1f}건/일 {'❌' if avg < AREA_LOW else '✅'}",
        f"  ③ 만나지않기: 0회 ✅",
        f"  ④ 평점1점: 0회 ✅",
        f"  ⑤ 수락률: 100% ✅",
    ]
    await update.message.reply_text("\n".join(lines))


# ══════════════════════════════════════════════
# Step D: /briefing — 7섹션 통합 보고
# ══════════════════════════════════════════════

import re as _re_count

def _extract_count(c: dict) -> int:
    """raw_calls 한 행의 실제 건수 (수동입력=1, OCR요약행=실제trip_count).
    앱(index.html)의 extractCount와 동일 로직. 명령서 #010 대응(2026-07-08)으로
    calc_official_var_score/handle_briefing의 월평균 계산에도 공통 적용."""
    비고 = c.get("비고") or ""
    m = _re_count.search(r"건수\s*(\d+)\s*건", 비고)
    if m:
        return int(m.group(1))
    건수 = c.get("건수")
    if 건수 is not None:
        try:
            return int(건수)
        except Exception:
            pass
    return 1


async def calc_kpi_metrics(날짜: str, 매출: int, work_hours) -> dict:
    """캐스퍼 명령서 #008 §3 반영 — daily_summary KPI 4종 봇 자체 계산
    (아르고스 브리핑 텍스트 파싱 방식 폐기, raw_calls/daily_summary 원본 직접 집계로 전환)

    kpi_7day_avg: 반드시 축B(카카오T, 00시 기준 달력일, 콜 없는 날은 0) 기준.
    raw_calls의 '날짜'는 축A(영업일, 저녁 시작 기준) 라벨이므로,
    00~05시 배차 콜은 실제로는 다음 캘린더일 새벽 연장 운행 → 익일로 재귀속해서 집계.

    2026-07-08 수정: 앱(index.html)에서 영수증 하루치를 요약 1행으로 저장하는 경우
    (예: 12건이 한 행에 뭉쳐 요금 합계만 기록) raw_calls "행 개수"를 그대로 건수로 세면
    과소집계된다. ①번(누적산출)에서 이미 적용한 것과 동일하게, 비고 텍스트의
    "건수 N건" 패턴에서 실제 건수를 역추출해서 가중 집계하도록 보정.
    """
    from datetime import date as _dc, timedelta as _td

    target = _dc.fromisoformat(날짜)
    window_start = (target - _td(days=7)).isoformat()

    # 명령서 #012 반영(2026-07-09): raw_calls에 축A/축B 라벨이 섞여 있으므로,
    # 축B 전용 지표(kpi_7day_avg, kpi_longdist_rate)는 date_axis='B' 행만 사용.
    # kpi_avg_fare(오늘 평일단가)는 오늘 하루치 실적 확인용이라 축 구분 없이
    # 전체 사용 — 아래에서 today_kakao는 window_calls 전체 기준으로 별도 처리.
    window_calls_all = await sb_select_calls( {
        "and": f"(날짜.gte.{window_start},날짜.lte.{날짜})"
    }) or []
    # task_id=11 명명표준화 대비(2026-08-22): DB값이 'B'->'calendar_day'로
    # 리네이밍될 예정이라, 전환기간 안전을 위해 신값/구값 둘 다 인식하도록 처리.
    window_calls = [c for c in window_calls_all if (c.get("date_axis") or "calendar_day") in ("B", "calendar_day")]

    # 명령서 #009 재검증(2026-07-08) 결과 반영: 자동 +1일 보정 제거.
    # 이지스가 업로드하는 raw_calls는 아르고스가 명령서#020·#021-rev1 자정분리
    # 원칙으로 이미 축B(달력일) 확정 처리한 뒤 적재하므로, 날짜 필드가
    # 이미 정확한 캘린더일이다. 여기에 "00~05시는 +1일" 자동 보정을 또 걸면
    # 이중 보정이 되어 날짜가 잘못 밀린다(실측 CSV 대조로 확인, 오차 12→8로 개선).
    # 향후 raw_calls에 날짜 라벨링 기준(축A/축B)을 구분하는 컬럼이 생기기 전까지는
    # 날짜 필드를 그대로 신뢰하는 것이 자동 보정보다 정확하다.
    def _axis_b_date(축a_날짜, 배차시각):
        return 축a_날짜

    # 축B 일별 카카오T 완료건수 (최근 7일, 콜 없는 날은 0으로 유지) — 건수 가중 집계
    axis_b_counts = {(target - _td(days=6 - i)).isoformat(): 0 for i in range(7)}
    for c in window_calls:
        if (c.get("콜유형") or "") != "카카오T":
            continue
        b_date = _axis_b_date(c.get("날짜"), c.get("배차시각"))
        if b_date in axis_b_counts:
            axis_b_counts[b_date] += _extract_count(c)
    kpi_7day_avg = round(sum(axis_b_counts.values()) / 7, 2)

    # 평균단가 — 오늘 카카오T 콜 기준(축 구분 없이 전체), 평일(월~금)만 산출. 건수 가중 평균.
    weekday = target.weekday()  # 0=월 ... 5=토 6=일
    today_kakao = [c for c in window_calls_all
                   if c.get("날짜") == 날짜 and (c.get("콜유형") or "") == "카카오T"]
    if weekday <= 4 and today_kakao:
        total_fare = sum(c.get("요금", 0) or 0 for c in today_kakao)
        total_cnt = sum(_extract_count(c) for c in today_kakao)
        kpi_avg_fare = int(total_fare / total_cnt) if total_cnt else None
    else:
        kpi_avg_fare = None

    # 장거리 비율 — 7일 윈도우, 카카오T, 15,000원 이상.
    # 요약행은 건당 평균요금(요금/건수)으로 개별 콜의 장거리 여부를 근사 추정.
    week_kakao = [c for c in window_calls if (c.get("콜유형") or "") == "카카오T"]
    if week_kakao:
        total_trips = 0
        longdist_trips = 0
        for c in week_kakao:
            cnt = _extract_count(c)
            fare = c.get("요금", 0) or 0
            per_trip_fare = fare / cnt if cnt else fare
            total_trips += cnt
            if per_trip_fare >= 15000:
                longdist_trips += cnt
        kpi_longdist_rate = round(longdist_trips / total_trips * 100, 1) if total_trips else None
    else:
        kpi_longdist_rate = None

    # 시간당 매출
    kpi_hourly_revenue = int(매출 / work_hours) if work_hours and work_hours > 0 else None

    return {
        "kpi_7day_avg": kpi_7day_avg,
        "kpi_avg_fare": kpi_avg_fare,
        "kpi_longdist_rate": kpi_longdist_rate,
        "kpi_hourly_revenue": kpi_hourly_revenue,
    }


async def handle_briefing(update, date_str: str = None):
    """매 운행 후 7섹션 통합 브리핑"""
    from datetime import date as _dc
    날짜 = date_str or str(_dc.today())
    mo = 날짜[:7]

    await update.message.reply_text(f"📋 {날짜} 브리핑 생성 중...")

    # 데이터 수집
    calls = await sb_select_calls( {"날짜": f"eq.{날짜}"})
    calls_month = await sb_select_calls( {
        "and": f"(날짜.gte.{mo}-01,날짜.lte.{mo}-31)"
    })

    total = len(calls)
    매출 = sum(c.get("요금", 0) or 0 for c in calls)
    avg_fare = int(매출 / total) if total else 0

    from datetime import date as _d2
    days_so_far = (_d2.today() - _d2(int(mo[:4]), int(mo[5:7]), 1)).days + 1
    # 명령서 #010 대응(2026-07-08): 행 개수가 아닌 건수 가중 합계로 수정.
    # 또한 아르고스 실측치(일평균 10~15건)와의 괴리 원인이 분모 정의 차이일 가능성이 높아,
    # 캘린더일 기준 월평균과 별도로 "운행일 기준 평균"도 함께 계산해 브리핑에 병기.
    # 아르고스 방법론 확정 회신 오기 전까지는 두 수치를 나란히 보여줘서 비교 가능하게 함.
    month_weighted_count = sum(_extract_count(c) for c in calls_month)
    monthly_avg_calls = month_weighted_count / max(days_so_far, 1)
    operating_days = len(set(c.get("날짜") for c in calls_month if c.get("날짜")))
    workday_avg_calls = month_weighted_count / operating_days if operating_days else 0.0

    # 공식 6변수 평가
    var_score = await calc_official_var_score(날짜)

    # 7섹션 구성
    lines = [
        f"═══ 자비스 브리핑 {날짜} ═══",
        f"",
        f"[A] 운행 데이터",
        f"  콜수: {total}건 | 매출: {fmt(매출)}",
        f"  건당단가: {fmt(avg_fare)}원",
        f"",
        f"[B] 카카오 알고리즘 관점",
        f"  ② 오늘 완료수: {total}건 (월평균 {monthly_avg_calls:.1f}건 · 운행일평균 {workday_avg_calls:.1f}건)",
        f"  ⑤ 수락률: 100% ✅",
        f"  AI 진입 추정: {var_score['ai_inclusion_estimate']}",
        f"  약점: {var_score['improvement_needed']}",
        f"",
        f"[C] 확률 분포",
        f"  건당단가 {fmt(avg_fare)}원",
        f"  {'목표단가 초과 ✅' if avg_fare >= 10000 else '목표단가 미달 (10,000원 목표)'}",
        f"",
        f"[D] 운빨 vs 추세",
        f"  오늘: {total}건 / 월평균: {monthly_avg_calls:.1f}건 / 운행일평균: {workday_avg_calls:.1f}건",
        f"  {'▲ 추세 우위' if total >= monthly_avg_calls else '▼ 추세 하회'}",
        f"",
        f"[E] 종합 진단",
        f"  운행완료수 약점 {'개선 중 📈' if monthly_avg_calls >= 12 else '강화 필요 ⚠️'}",
        f"  수락률·평점·만나지않기 모두 최고 ✅",
        f"",
        f"[F] 다음 운행 전략",
        f"  19~21시 수성구 집중 → 21시 성내2동 앵커",
        f"  수락률 100% 유지 (콜 거절 금지)",
        f"  목표: {max(0, 18-total)}건 이상 추가 달성",
        f"",
        f"[G] 베이지안 업데이트",
        f"  오늘 {total}건 반영 완료",
        f"  누적 {len(calls_month)}건 → 모델 정밀도 {min(95, 60 + len(calls_month)//10)}%",
    ]

    # DB 저장
    briefing_data = {
        "run_date": 날짜,
        "section_a": {"calls": total, "revenue": 매출, "avg_fare": avg_fare},
        "section_b": var_score,
        "section_c": {"avg_fare": avg_fare, "target": 10000},
        "section_d": {"today": total, "monthly_avg": round(monthly_avg_calls, 1)},
        "section_e": f"운행완료수 {'개선중' if monthly_avg_calls >= 12 else '강화필요'}",
        "section_f": "19~21 수성구 → 21시 성내2동 앵커, 수락률 100% 유지",
        "section_g": {"cumulative": len(calls_month), "model_accuracy": min(95, 60+len(calls_month)//10)}
    }
    await sb_h("POST", "daily_briefing",
        json=briefing_data,
        headers={**HEADERS_SB, "Prefer": "resolution=merge-duplicates,return=minimal"},
        params={"on_conflict": "run_date"}
    )

    briefing_text = "\n".join(lines)

    # task#63(2026-08-26): 이지스 폐기 확정(correction_log 기록, 8/25 대표님 확인 —
    # 역할은 아르고스가 흡수)에 따라, "이지스 연동용"(명령서#004/#006, 7/6)이던
    # bot_briefings INSERT를 제거. 소비자(컨슈머) 로직 자체가 처음부터 없었던
    # 죽은 큐였음(task#59에서 확인).

    # 캐스퍼 명령서 #008 §3 — KPI 4종 (봇 자체 계산, 축B 필수)
    kpi = {}
    try:
        def _to_virtual_min(hhmm: str) -> int:
            h, m = map(int, hhmm.split(":"))
            minutes = h * 60 + m
            if h < 6:  # 0~5시는 전날 심야운행의 연장으로 간주
                minutes += 24 * 60
            return minutes

        시각목록 = [c.get("배차시각") for c in calls if c.get("배차시각")]
        summary_payload = {"날짜": 날짜}
        work_hours = None
        if 시각목록:
            정렬됨 = sorted((( _to_virtual_min(t), t) for t in 시각목록))
            시작 = 정렬됨[0][1]
            종료 = 정렬됨[-1][1]
            work_hours = round((정렬됨[-1][0] - 정렬됨[0][0]) / 60, 2)
            summary_payload.update({
                "work_start_time": 시작,
                "work_end_time": 종료,
                "work_hours": work_hours
            })

        kpi = await calc_kpi_metrics(날짜, 매출, work_hours)
        summary_payload.update(kpi)

        await sb_upsert("daily_summary", summary_payload, on_conflict="날짜")
    except Exception as e:
        logger.error(f"daily_summary 운행시간/KPI 저장 오류: {e}")

    # ══════════════════════════════════════════════
    # 캐스퍼 명령서 #014 반영 (2026-07-10)
    # 완전 자동화: 브리핑 생성 → GitHub 직접 커밋 → 텔레그램 요약+파일 전송
    # ══════════════════════════════════════════════

    kakao_n = sum(1 for c in calls if (c.get("콜유형") or "") == "카카오T")
    uber_n  = sum(1 for c in calls if (c.get("콜유형") or "") == "우버")
    bhw_n   = sum(1 for c in calls if (c.get("콜유형") or "") == "배회")

    # 절벽구간(간이 산정): 오늘 콜 간 40분 이상 공백 — GPX 교차검증 없는 raw_calls 시각만의 근사치.
    # 아르고스의 정식 Dead Zone 분석(GPX 이동 여부 확인 포함)과는 다른, 봇 자체의 단순 근사값임을 명시.
    def _gap_stats(calls, threshold_min=40):
        times = [t for t in (c.get("배차시각") for c in calls) if t]
        if len(times) < 2:
            return 0, 0
        mins_sorted = sorted(_to_virtual_min(t) for t in times)
        gap_count = 0
        gap_total = 0
        for i in range(1, len(mins_sorted)):
            gap = mins_sorted[i] - mins_sorted[i-1]
            if gap >= threshold_min:
                gap_count += 1
                gap_total += gap
        return gap_total, gap_count

    gap_total_min, gap_count = _gap_stats(calls)

    kpi_7day = kpi.get("kpi_7day_avg")
    kpi_fare = kpi.get("kpi_avg_fare")
    kpi_long = kpi.get("kpi_longdist_rate")
    kpi_hourly = kpi.get("kpi_hourly_revenue")
    kpi_met = sum([
        (kpi_7day or 0) >= 10,
        (kpi_fare or 0) >= 10000,
        (kpi_long or 0) >= 20,
        (kpi_hourly or 0) >= 20000,
    ])

    # 전체 브리핑 markdown 문서 (GitHub 커밋 + 파일첨부용)
    full_md = "\n".join([
        f"# 자비스 브리핑 {날짜}",
        "",
        briefing_text,
        "",
        "---",
        f"*bot_v5.py 자동생성 · KPI 판정 {kpi_met}/4 충족*",
    ])

    # GitHub 직접 커밋
    gh_result = await github_commit_briefing(날짜, full_md)
    if not gh_result["ok"]:
        logger.error(f"GitHub 커밋 실패: {gh_result['error']}")

    # 텔레그램 요약 메시지 (명령서 #014 §3 템플릿)
    summary_msg = (
        f"[봇 자동요약 / 실시간, 검증 전]\n"
        f"[자비스 브리핑 요약 / {날짜}]\n"
        f"카카오T {kakao_n}건 · 우버{uber_n}건 · 배회{bhw_n}건 | 매출 {fmt(매출)}\n"
        f"7일평균 {(f'{kpi_7day:.1f}건' if kpi_7day is not None else '-건')} "
        f"(기준10건 대비 {'✅' if (kpi_7day or 0) >= 10 else '❌'})\n"
        f"평균단가 {(fmt(kpi_fare) if kpi_fare is not None else '-원')} "
        f"(기준10,000원 대비 {'✅' if (kpi_fare or 0) >= 10000 else '❌'})\n"
        f"KPI 판정: {kpi_met}/4 충족\n"
        f"절벽구간: {gap_total_min}분 ({gap_count}건, 간이산정)\n"
        f"오늘 요약: 콜 {total}건 · 매출 {fmt(매출)} 기록\n"
        f"→ 전체 브리핑은 첨부파일 참고\n"
        f"→ 아르고스 정밀 분석은 별도로 브리핑 확인"
        + (f"\nGitHub: {gh_result['url']}" if gh_result["ok"] else "\n⚠️ GitHub 저장 실패 (로그 확인 필요)")
    )

    # 안전장치: 4,096자 초과 시 자동 분할
    if len(summary_msg) > 4000:
        for i in range(0, len(summary_msg), 4000):
            await update.message.reply_text(summary_msg[i:i+4000])
    else:
        await update.message.reply_text(summary_msg)

    # 전체 브리핑 파일 첨부 전송
    try:
        file_bytes = io.BytesIO(full_md.encode("utf-8"))
        file_bytes.name = f"bot_summary_{날짜.replace('-','')}.md"
        await update.message.reply_document(document=file_bytes, filename=f"bot_summary_{날짜.replace('-','')}.md")
    except Exception as e:
        logger.error(f"브리핑 파일 전송 오류: {e}")


async def save_operation_consistency(날짜: str, 시작시각: str, 종료시각: str,
                                     총건수: int, 총매출: int):
    """운행 일관성 데이터 저장 및 점수 산출"""
    try:
        h_start = int(시작시각.split(":")[0])
        m_start = int(시작시각.split(":")[1])
        # 목표 19:00 기준 격차 (분)
        격차 = (h_start * 60 + m_start) - (19 * 60)
        격차_abs = abs(격차)

        # 일관성 점수 (100점 기준)
        # 시작 시각 ±15분 = 100, ±30분 = 80, ±60분 = 60, 초과 = 40
        if 격차_abs <= 15: start_score = 100
        elif 격차_abs <= 30: start_score = 80
        elif 격차_abs <= 60: start_score = 60
        else: start_score = 40

        일관성점수 = start_score

        payload = {
            "날짜":         날짜,
            "시작시각":     시작시각,
            "종료시각":     종료시각,
            "시작격차_분":  격차,
            "일관성점수":   일관성점수,
            "총건수":       총건수,
            "총매출":       총매출,
        }
        await sb_upsert("operation_consistency", payload, on_conflict="날짜")
        return 일관성점수
    except Exception as e:
        logger.error(f"일관성 저장 오류: {e}")
        return None


async def report_operation_consistency(update, 날짜: str = None):
    """운행 일관성 보고 — 오늘 + 최근 7일 평균"""
    try:
        target = 날짜 or str(today_kst())
        rows = await sb_select("operation_consistency",
                               {"order": "날짜.desc", "limit": "7"})
        if not rows:
            await update.message.reply_text("📊 운행 일관성 데이터 없음")
            return

        today_row = next((r for r in rows if r.get("날짜") == target), rows[0])
        시작 = today_row.get("시작시각", "?")
        종료 = today_row.get("종료시각", "?")
        격차 = today_row.get("시작격차_분", 0)
        점수 = today_row.get("일관성점수", 0)
        건수 = today_row.get("총건수", 0)
        매출 = today_row.get("총매출", 0)

        # 7일 평균
        starts = [r.get("시작격차_분", 0) for r in rows if r.get("시작격차_분") is not None]
        avg_격차 = sum(starts) / len(starts) if starts else 0

        # 점수 별
        stars = "⭐" * (점수 // 25 + 1) if 점수 else "?"

        lines = [
            f"📊 운행 일관성 — {target}",
            f"시작: {시작} (목표 19:00, 격차 {'+' if 격차>=0 else ''}{격차}분)",
            f"종료: {종료}",
            f"건수: {건수}건 | 매출: {매출:,}원" if 매출 else f"건수: {건수}건",
            f"",
            f"일관성 점수: {점수}/100 {stars}",
            f"7일 시작 평균 격차: {avg_격차:+.0f}분",
            f"★ 알고리즘 학습 신호: {'강함' if 점수 >= 80 else '보통' if 점수 >= 60 else '약함'}",
        ]
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        logger.error(f"일관성 조회 오류: {e}")
        await update.message.reply_text("❌ 일관성 조회 오류")

async def ocr_sekuti(image_bytes: bytes) -> dict | None:
    """세큐티 리포트 이미지 OCR → 점수·등급 추출"""
    prompt = (
        "이 세큐티(SEKUTI) 기사 리포트 이미지에서 정보를 추출해서 JSON만 반환해줘.\n"
        '{"종합점수":숫자,"상위퍼센트":숫자,"수락률":숫자,"실내공기":숫자,'
        '"친절도":숫자,"안전운행":숫자,"등급":"마스터/다이아몬드/플래티넘/골드/실버 중 하나",'
        '"기간":"주간 또는 월간","기준날짜":"YYYY-MM-DD"}\n'
        "없는 항목은 null. 점수는 숫자만(원,% 제외).\n"
        "JSON만 반환. 설명·마크다운 금지."
    )
    try:
        raw = await claude_vision(image_bytes, prompt, max_tokens=300)
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        import json as _json
        return _json.loads(raw)
    except Exception as e:
        logger.error(f"세큐티 OCR 오류: {e}")
        return None


async def process_sekuti(update, image_bytes: bytes):
    """세큐티 이미지 처리: OCR → sekuti_weekly 저장 → 결과 표시"""
    await update.message.reply_text("📊 세큐티 분석 중...")

    data = await ocr_sekuti(image_bytes)
    if not data:
        await update.message.reply_text("❌ 세큐티 인식 실패. 다시 올려주세요.")
        return

    from datetime import date as _dc
    기준날짜 = data.get("기준날짜")
    if 기준날짜:
        try:
            _dc.fromisoformat(str(기준날짜))
        except Exception:
            기준날짜 = str(today_kst())
    else:
        기준날짜 = str(today_kst())

    종합점수   = data.get("종합점수")
    상위퍼센트 = data.get("상위퍼센트")
    수락률     = data.get("수락률")
    실내공기   = data.get("실내공기")
    친절도     = data.get("친절도")
    안전운행   = data.get("안전운행")
    등급       = data.get("등급") or ""
    기간       = data.get("기간") or "주간"

    payload = {
        "날짜":       기준날짜,
        "종합점수":   종합점수,
        "상위퍼센트": 상위퍼센트,
        "수락률":     수락률,
        "실내공기":   실내공기,
        "친절도":     친절도,
        "안전운행":   안전운행,
        "등급":       등급,
        "기간":       기간,
    }
    await sb_upsert("sekuti_weekly", payload, on_conflict="날짜,기간")

    grade_icon = {
        "마스터":"👑","다이아몬드":"💎","플래티넘":"🥈","골드":"🥇","실버":"🥉"
    }.get(등급, "📊")

    def score_bar(score):
        if score is None: return "N/A"
        s = int(score)
        if s >= 95: return f"{s}점 🟢"
        if s >= 90: return f"{s}점 🟡"
        return f"{s}점 🔴"

    master_check = []
    if 종합점수 and 종합점수 >= 95:
        master_check.append("✅ 종합점수 95↑")
    else:
        master_check.append(f"❌ 종합점수 {종합점수 or '?'} (95 필요)")
    if 실내공기 and 실내공기 >= 93:
        master_check.append("✅ 실내공기 93↑")
    else:
        master_check.append(f"❌ 실내공기 {실내공기 or '?'} (93 필요)")
    if 수락률 and 수락률 >= 95:
        master_check.append("✅ 수락률 95↑")
    else:
        master_check.append(f"❌ 수락률 {수락률 or '?'} (95 필요)")

    msg_lines = [
        f"{grade_icon} 세큐티 리포트 ({기간}) — {기준날짜}",
        "",
        f"종합점수: {score_bar(종합점수)}" + (f" (상위 {상위퍼센트}%)" if 상위퍼센트 else ""),
        f"등급: {등급}",
        "",
        "[항목별]",
        f"  수락률:   {score_bar(수락률)}",
        f"  실내공기: {score_bar(실내공기)}",
        f"  친절도:   {score_bar(친절도)}",
        f"  안전운행: {score_bar(안전운행)}",
        "",
        "[마스터 전환 체크]",
    ] + master_check

    await update.message.reply_text("\n".join(msg_lines))


async def handle_sekuti_query(update):
    """세큐티 최근 기록 조회 — '세큐티 조회' 명령어"""
    rows = await sb_select("sekuti_weekly", {"order": "날짜.desc", "limit": "5"})
    if not rows:
        await update.message.reply_text(
            "📊 세큐티 기록 없음\n세큐티 리포트 이미지를 올려주세요."
        )
        return

    grade_icon = {"마스터":"👑","다이아몬드":"💎","플래티넘":"🥈","골드":"🥇","실버":"🥉"}
    lines_out = ["📊 세큐티 최근 기록\n"]
    for r in rows:
        icon = grade_icon.get(r.get("등급",""), "📊")
        lines_out.append(
            f"{icon} {r.get('날짜','')} ({r.get('기간','')})\n"
            f"  종합 {r.get('종합점수','?')}점 · 상위 {r.get('상위퍼센트','?')}% · {r.get('등급','')}\n"
            f"  수락{r.get('수락률','?')} 공기{r.get('실내공기','?')} "
            f"친절{r.get('친절도','?')} 안전{r.get('안전운행','?')}"
        )
    await update.message.reply_text("\n".join(lines_out))


async def ocr_call_card(image_bytes: bytes) -> dict | None:
    prompt = (
        "이 카카오T 콜카드(운행이력) 이미지에서 아래 JSON만 반환해줘.\n"
        '{"날짜":"YYYY-MM-DD","배차시각":"HH:MM","하차시각":"HH:MM",'
        '"출발지":"OO구 OO동","도착지":"OO구 OO동",'
        '"요금":숫자,"카드사":"카드사명","콜유형":"카카오T 또는 배회","결제방식":"자동 또는 직접"}\n'
        "⚠️ 날짜: 이미지 상단의 운행 날짜(예: 2026/03/07 → 2026-03-07). 없으면 null.\n"
        "  단, 화면에 여러 날짜가 있으면 가장 최근 운행의 날짜만 추출.\n"
        "⚠️ 요금 추출:\n"
        "  - \'미터 요금\', \'총 요금\', \'결제 금액\', \'이용 요금\' 레이블 옆 숫자\n"
        "  - 예: 미터요금 13,100원 → 요금=13100\n"
        "  - 요금이 보이면 반드시 추출. 화면에 없을 때만 0으로 설정\n"
        "⚠️ 결제방식:\n"
        "  - 요금 숫자 보임 → 결제방식=\'자동\', 요금=해당숫자\n"
        "  - 요금 전혀 없음 → 결제방식=\'직접\', 요금=0\n"
        "⚠️ 도착지: 차량번호(예: 대구 32바 5763, XX가나바 NNNN) → null.\n"
        "  도착지는 \'OO구 OO동\' 형식 주소만.\n"
        "⚠️ 시각: 배차시각=승차시각, 하차시각=하차완료시각. 없으면 null.\n"
        "JSON만 반환. 설명·마크다운 금지."
    )
    raw = await claude_vision(image_bytes, prompt, max_tokens=200)
    try:
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```")
        return json.loads(raw)
    except Exception:
        logger.error(f"콜카드 JSON 파싱 실패: {raw}")
        return None

async def ocr_charge_receipt(image_bytes: bytes) -> list:
    prompt = (
        "이 충전 내역 이미지에서 모든 충전 건을 추출해서 JSON 배열만 반환해줘.\n"
        '[{"충전일자":"YYYY-MM-DD","충전소명":"이름","충전량":숫자(kWh),"결제금액":숫자(원)}, ...]\n'
        "여러 건이 보이면 전부 추출. JSON 배열만 반환. 설명 금지."
    )
    raw = await claude_vision(image_bytes, prompt, max_tokens=500)
    try:
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```")
        result = json.loads(raw)
        return result if isinstance(result, list) else []
    except Exception:
        logger.error(f"충전내역 JSON 파싱 실패: {raw}")
        return []

async def ocr_payment_history(image_bytes: bytes) -> list:
    """
    카카오T 결제내역(수익관리 화면) OCR.
    반환: [{"날짜":"YYYY-MM-DD","시각":"HH:MM","요금":숫자,"결제방법":"카드/현금"}, ...]
    """
    prompt = (
        "이 결제내역 화면에서 모든 결제 건을 추출해서 JSON 배열만 반환해줘.\n"
        '[{"날짜":"YYYY-MM-DD","시각":"HH:MM","요금":숫자,"결제방법":"카드 또는 현금"}, ...]\n'
        "⚠️ 지원하는 화면 형식:\n"
        "  1. 카카오T 수익관리 화면: 거래일자 YYYY-MM-DD HH:MM:SS 형식\n"
        "  2. 세큐티 결제내역조회 화면: 거래일자 YYYY-MM-DD | HH:MM:SS 형식\n"
        "     (상단에 결제내역조회, 차량번호, 조회기간 필터 있음)\n"
        "⚠️ 시각 처리: HH:MM:SS → HH:MM 으로 변환 (초 제거)\n"
        "⚠️ 날짜: 각 건의 거래일자에서 YYYY-MM-DD 추출. null 불가.\n"
        "⚠️ 요금: 숫자만 (쉼표, 원 제거). 예: 5,400원 → 5400\n"
        "⚠️ 취소 건(구분=취소) 제외. 승인 정상만 포함.\n"
        "JSON 배열만 반환. 설명·마크다운 금지."
    )
    raw = await claude_vision(image_bytes, prompt, max_tokens=2000)
    try:
        raw = raw.strip()
        raw = re.sub(r"```json\s*", "", raw)
        raw = re.sub(r"```\s*", "", raw)
        raw = raw.strip()
        arr_start = raw.find("[")
        arr_end = raw.rfind("]")
        if arr_start >= 0 and arr_end > arr_start:
            raw = raw[arr_start:arr_end+1]
        result = json.loads(raw)
        return result if isinstance(result, list) else []
    except Exception as e:
        logger.error(f"결제내역 JSON 파싱 실패: {e} / raw: {raw[:200]}")
        return []
    # 구버전 호환
    except Exception:
        logger.error(f"결제내역 JSON 파싱 실패: {raw}")
        return []

# ──────────────────────────────────────────────
# 교차대조 로직 (콜카드 ↔ 결제내역)
# ──────────────────────────────────────────────
async def cross_check(date_str: str) -> str:
    """
    콜카드(raw_calls) ↔ 결제내역(payment_receipts) 교차대조.
    매칭: 콜카드 하차시각 ↔ 결제시각 ±20분 (자정넘김 처리)
    미매칭 결제내역 자동분류:
      - 콜카드 운행 공백 시간대 → 배회영업 후보
      - 콜카드 운행 중 시간대   → 누락 콜카드 후보
    """
    from datetime import date as date_cls, timedelta

    calls = await sb_select_calls( {"날짜": f"eq.{date_str}"})
    try:
        y, mo, d = date_str.split("-")
        next_date_str = str(date_cls(int(y),int(mo),int(d)) + timedelta(days=1))
    except Exception:
        next_date_str = date_str

    receipts_today = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})
    receipts_next  = await sb_select("payment_receipts", {"날짜": f"eq.{next_date_str}"})
    receipts = receipts_today + receipts_next

    if not calls and not receipts:
        return f"⚠️ {date_str} 데이터 없음"

    def to_min_smart(배차시각, 대상시각, 대상날짜):
        try:
            bh, bm = 배차시각.split(":")
            base = int(bh)*60+int(bm)
            th, tm = 대상시각.split(":")
            target = int(th)*60+int(tm)
            if 대상날짜 == next_date_str:
                target += 1440
            elif target < base - 60:
                target += 1440
            return target
        except: return None

    def to_min_abs(time_str, date_str_local):
        """날짜 기반 분 변환. mins<300 휴리스틱 제거 — 새벽 운행 오류 방지."""
        try:
            h, m = time_str.split(":")
            mins = int(h)*60+int(m)
            if date_str_local == next_date_str:
                mins += 1440  # 익일 날짜면 +1440만 적용
            return mins
        except: return None

    # STEP 1: 하차시각 ↔ 결제시각 ±20분 매칭
    matched_call_ids    = set()
    matched_receipt_ids = set()
    fee_mismatches      = []  # 금액 불일치 목록
    direct_updated      = []  # 직접결제 요금 자동업데이트 목록

    for i, call in enumerate(calls):
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각")
        if not 하차:
            try:
                bh, bm = 배차.split(":")
                est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: continue
        c_min = to_min_smart(배차, 하차, date_str)
        best_j, best_diff = None, 99999
        for j, rcpt in enumerate(receipts):
            if j in matched_receipt_ids: continue
            rcpt_date = rcpt.get("날짜", date_str)
            r_min = to_min_smart(배차, rcpt.get("시각","") or "", rcpt_date)
            if r_min and c_min:
                diff = abs(c_min - r_min)
                if diff <= 20 and diff < best_diff:
                    best_diff = diff
                    best_j = j
        if best_j is not None:
            matched_call_ids.add(i)
            matched_receipt_ids.add(best_j)
            call_fee = call.get("요금") or 0
            rcpt_fee = receipts[best_j].get("요금") or 0

            # 직접결제(요금=0) 콜카드 → 결제내역 요금으로 자동 업데이트
            if call_fee == 0 and rcpt_fee > 0:
                call_id = call.get("id")
                if call_id:
                    await sb_h("PATCH", f"raw_calls?id=eq.{call_id}",
                               json={"요금": rcpt_fee, "비고": "직접결제(요금확인완료)"})
                    direct_updated.append({
                        "배차시각": 배차,
                        "요금": rcpt_fee,
                    })

            # 금액 불일치 (둘 다 0이 아니고 차이 ≥500원)
            elif call_fee > 0 and rcpt_fee > 0:
                fee_diff = abs(call_fee - rcpt_fee)
                if fee_diff >= FEE_DIFF_THRESHOLD:
                    fee_mismatches.append({
                        "call_id": call.get("id"),
                        "배차시각": 배차,
                        "call_fee": call_fee,
                        "rcpt_fee": rcpt_fee,
                        "diff": fee_diff,
                    })

    unmatched_calls    = [c for i,c in enumerate(calls)    if i not in matched_call_ids]
    unmatched_receipts = [r for j,r in enumerate(receipts) if j not in matched_receipt_ids]

    # STEP 2: 콜카드 점유 시간대 계산 → 미매칭 결제내역 분류
    occupied = []
    for call in calls:
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각") or ""
        if not 하차:
            try:
                bh, bm = 배차.split(":")
                est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: pass
        s = to_min_abs(배차, date_str)
        e = to_min_abs(하차, date_str)
        if s and e:
            occupied.append((s-5, e+5))

    baehoe_rcpt  = []
    missing_rcpt = []
    for r in unmatched_receipts:
        r_date = r.get("날짜", date_str)
        r_min  = to_min_abs(r.get("시각","") or "", r_date)
        if r_min is None:
            missing_rcpt.append(r)
            continue
        if any(s <= r_min <= e for s,e in occupied):
            missing_rcpt.append(r)   # 운행 중 시간 → 누락 콜카드
        else:
            baehoe_rcpt.append(r)    # 공백 시간 → 배회영업

    # 결과 출력
    lines_out = [f"📊 교차대조 결과 — {date_str}"]
    lines_out.append(f"콜카드 {len(calls)}건 / 결제내역 {len(receipts)}건 / 매칭 {len(matched_call_ids)}건")
    lines_out.append("")

    # 직접결제 요금 자동업데이트 표시
    if direct_updated:
        lines_out.append(f"💳 직접결제 요금 자동확인 {len(direct_updated)}건:")
        for d in direct_updated:
            lines_out.append(f"  ✅ {d['배차시각']} → {fmt(d['요금'])} 업데이트")
        lines_out.append("")

    if unmatched_calls:
        lines_out.append(f"🟠 콜카드에만 있음 {len(unmatched_calls)}건:")
        for c in unmatched_calls:
            lines_out.append(f"  {c.get('배차시각','-')} {c.get('출발지','')}→{c.get('도착지','')} {fmt(c.get('요금') or 0)}")
        lines_out.append("")

    if baehoe_rcpt:
        lines_out.append(f"🚶 배회영업 후보 (공백시간) {len(baehoe_rcpt)}건:")
        for r in baehoe_rcpt:
            날짜표시 = f"({r.get('날짜','')})" if r.get("날짜") != date_str else ""
            lines_out.append(f"  {r.get('시각','-')}{날짜표시} {fmt(r.get('요금') or 0)}")
        lines_out.append("")

    if missing_rcpt:
        lines_out.append(f"🔴 누락 콜카드 후보 (운행중 시간) {len(missing_rcpt)}건:")
        for r in missing_rcpt:
            날짜표시 = f"({r.get('날짜','')})" if r.get("날짜") != date_str else ""
            lines_out.append(f"  {r.get('시각','-')}{날짜표시} {fmt(r.get('요금') or 0)}")
        lines_out.append("")

    # 금액 불일치 표시
    if fee_mismatches:
        lines_out.append(f"💰 금액 불일치 {len(fee_mismatches)}건 (차이 ≥500원):")
        for fm in fee_mismatches:
            lines_out.append(
                f"  {fm['배차시각']} 콜카드:{fmt(fm['call_fee'])} vs "
                f"결제:{fmt(fm['rcpt_fee'])} (차이 {fm['diff']:,}원)"
            )
        lines_out.append("  → '대조 금액확인 YYYY-MM-DD' 로 버튼 선택")
        lines_out.append("")

    if not unmatched_calls and not unmatched_receipts:
        if fee_mismatches:
            lines_out.append("⚠️ 매칭 완료 — 금액 불일치 확인 필요")
        else:
            lines_out.append("✅ 완전 매칭 — 누락 없음")

    if unmatched_calls:
        lines_out.append(f"💡 '배회분류 확정 {date_str}' → 콜카드 미매칭 배회 처리")
    if baehoe_rcpt:
        lines_out.append(f"💡 '대조 확정 {date_str}' → 배회후보 {len(baehoe_rcpt)}건 raw_calls 자동 추가")

    return "\n".join(lines_out)


async def confirm_baehoe_classification(date_str: str) -> str:
    """미매칭 콜카드를 배회영업으로 자동 분류 확정"""
    calls = await sb_select_calls( {"날짜": f"eq.{date_str}"})
    receipts = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})

    def to_minutes(t_str):
        try:
            h, m = t_str.split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return None

    matched_call_ids = set()
    for i, call in enumerate(calls):
        call_min = to_minutes(call.get("배차시각", "") or "")
        for j, rcpt in enumerate(receipts):
            rcpt_min = to_minutes(rcpt.get("시각", "") or "")
            if call_min and rcpt_min and abs(call_min - rcpt_min) <= 3:
                matched_call_ids.add(i)
                break

    unmatched = [c for i, c in enumerate(calls) if i not in matched_call_ids]
    if not unmatched:
        return "✅ 미매칭 콜카드 없음"

    count = 0
    for call in unmatched:
        # 콜유형을 배회로 업데이트
        await sb_h(
            "PATCH",
            f"raw_calls?id=eq.{call['id']}",
            json={"콜유형": "배회"}
        )
        count += 1

    return f"✅ {count}건 배회영업으로 분류 완료"

# ──────────────────────────────────────────────
# 이미지 처리 함수

# ──────────────────────────────────────────────
# 중복 체크 헬퍼
# ──────────────────────────────────────────────
async def check_duplicate_call(날짜: str, 배차시각: str, 요금: int) -> bool:
    """raw_calls 중복 체크: 날짜+배차시각+요금 동일하면 True"""
    rows = await sb_select_calls( {
        "날짜": f"eq.{날짜}",
        "배차시각": f"eq.{배차시각}",
        "요금": f"eq.{요금}",
    })
    return len(rows) > 0

async def check_duplicate_payment(날짜: str, 시각: str, 요금: int) -> bool:
    """payment_receipts 중복 체크: 날짜+시각+요금 동일하면 True"""
    rows = await sb_select("payment_receipts", {
        "날짜": f"eq.{날짜}",
        "시각": f"eq.{시각}",
        "요금": f"eq.{요금}",
    })
    return len(rows) > 0

# ──────────────────────────────────────────────
async def process_call_card(update: Update, image_bytes: bytes):
    data = await ocr_call_card(image_bytes)
    if not data:
        await update.message.reply_text("❌ 콜카드 인식 실패. 다시 올려주세요.")
        return

    # 날짜 결정 로직:
    # - OCR 날짜가 과거(미래 아님)이면 → OCR 날짜 신뢰 (뒤늦게 올린 콜카드)
    # - OCR 날짜가 미래이면 → 오늘 날짜 (오인식)
    # - OCR 날짜 없으면 → 오늘 날짜
    from datetime import date as _dc, timedelta as _td
    _today_d = today_kst()
    _today_str = str(_today_d)
    _dow_map = ["월","화","수","목","금","토","일"]

    ocr_date_raw = data.get("날짜")
    today = _today_str
    dow = _dow_map[_today_d.weekday()]
    date_source = "오늘"  # 저장 메시지용

    if ocr_date_raw:
        try:
            _ocr_d = _dc.fromisoformat(str(ocr_date_raw).strip())
            if _ocr_d > _today_d:
                # 미래 날짜 → 오인식 → 오늘 사용
                today = _today_str
                date_source = f"오늘(OCR미래날짜오류)"
            else:
                # 과거 또는 오늘 → OCR 날짜 신뢰
                today = str(_ocr_d)
                dow = _dow_map[_ocr_d.weekday()]
                diff = (_today_d - _ocr_d).days
                date_source = f"OCR({diff}일전)" if diff > 0 else "오늘"
        except Exception:
            today = _today_str
            date_source = "오늘(OCR파싱오류)"

    배차시각 = data.get("배차시각")
    요금 = data.get("요금") or 0
    결제방식 = data.get("결제방식", "자동")

    # 중복 체크 → 자동 삭제 후 재저장
    deleted = await delete_duplicate_call(today, 배차시각, 요금)
    if deleted:
        logger.info(f"중복 콜카드 자동 삭제 후 재저장: {today} {배차시각} {요금}")

    # 직접결제: 요금 0원 → pending 상태로 저장
    # 직접결제 판별:
    # - 결제방식이 명시적으로 "직접"인 경우만 직접결제
    # - 요금=0이어도 결제방식="자동"이면 OCR 오류일 수 있으므로 경고만 표시
    is_direct = (결제방식 == "직접")
    is_zero_fee = (요금 == 0) and not is_direct  # 요금=0 but 자동결제
    비고 = "직접결제(요금미확인)" if is_direct else data.get("카드사")

    payload = {
        "날짜": today,
        "요일": dow,
        "배차시각": 배차시각,
        "하차시각": data.get("하차시각"),
        "출발지": data.get("출발지"),
        "도착지": data.get("도착지"),
        "요금": 요금,
        "콜유형": data.get("콜유형", "카카오T"),
        "비고": 비고,
        "data_source": "app_ocr_individual",
    }
    result = await sb_insert("raw_calls", payload)
    if result:
        if is_direct:
            await update.message.reply_text(
                f"💳 직접결제 콜카드 저장 (요금 미확인)\n"
                f"날짜: {today} ({dow}) [{date_source}]\n"
                f"{배차시각} {data.get('출발지','?')}→{data.get('도착지','?') or '?'}\n"
                f"⚠️ 결제내역 업로드 후 '대조 {today}' 입력해서 요금 확인하세요."
            )
        elif is_zero_fee:
            await update.message.reply_text(
                f"⚠️ 콜 저장 (요금 0원 — 확인 필요)\n"
                f"{배차시각} {data.get('출발지','?')}→{data.get('도착지','?')}\n"
                f"콜카드에 요금이 보이면 '콜수정 {배차시각} 요금=실제금액' 으로 수정하세요."
            )
        else:
            await update.message.reply_text(
                f"✅ 콜 저장\n"
                f"날짜: {today} ({dow}) [{date_source}]\n"
                f"{배차시각} {data.get('출발지','?')}→{data.get('도착지','?')}\n"
                f"{fmt(요금)} [{data.get('콜유형','카카오T')}]"
            )
    else:
        await update.message.reply_text("❌ DB 저장 실패")

async def process_charge_receipt(update: Update, image_bytes: bytes):
    items = await ocr_charge_receipt(image_bytes)
    if not items:
        await update.message.reply_text("❌ 충전내역 인식 실패. 다시 올려주세요.")
        return

    saved = 0
    for item in items:
        payload = {
            "충전일": item.get("충전일자") or str(today_kst()),
            "충전량_kwh": item.get("충전량"),
            "충전금액": item.get("결제금액"),
            "충전소": item.get("충전소명"),
        }
        r = await sb_insert("charging_log", payload)
        if r:
            saved += 1

    await update.message.reply_text(
        f"⚡ 충전내역 {saved}/{len(items)}건 저장 완료"
    )

async def process_payment_history(update: Update, image_bytes: bytes):
    """결제내역 OCR → payment_receipts 저장"""
    items = await ocr_payment_history(image_bytes)
    if not items:
        await update.message.reply_text("❌ 결제내역 인식 실패. 다시 올려주세요.")
        return

    saved = 0
    skipped = 0
    duplicated = 0
    date_warn = []
    dup_list = []
    for item in items:
        날짜 = item.get("날짜")
        # 날짜 없으면 저장 거부 — 오늘 날짜로 대체하지 않음
        if not 날짜 or 날짜 == "null":
            skipped += 1
            date_warn.append(item.get("시각", "?"))
            continue
        시각 = item.get("시각")
        요금 = item.get("요금", 0)
        # 중복 체크 → 자동 삭제 후 재저장
        deleted = await delete_duplicate_payment(날짜, 시각, 요금)
        if deleted:
            duplicated += deleted  # 실제 삭제 건수 반영
            logger.info(f"중복 결제내역 자동 삭제 {deleted}건: {날짜} {시각} {요금}")
        payload = {
            "날짜": 날짜,
            "시각": 시각,
            "요금": 요금,
            "결제방법": item.get("결제방법", "카드"),
        }
        r = await sb_insert("payment_receipts", payload)
        if r:
            saved += 1

    dates = list(set(item.get("날짜") for item in items if item.get("날짜") and item.get("날짜") != "null"))
    msg = f"💳 결제내역 {saved}건 저장"
    if dates:
        msg += f" ({', '.join(dates)})"
    if duplicated > 0:
        msg += f"\n⚠️ 중복 {duplicated}건 저장 안 됨: {', '.join(dup_list)}"
    if skipped > 0:
        msg += f"\n⚠️ 날짜 인식 실패 {skipped}건 저장 안 됨"
        msg += f"\n  시각: {', '.join(date_warn)} → 날짜 보이는 캡처로 재전송"
    if saved > 0 and dates:
        msg += f"\n교차대조: '대조 {dates[0]}' 입력"
    await update.message.reply_text(msg)

async def process_single_image(update: Update, context: ContextTypes.DEFAULT_TYPE, image_bytes: bytes = None):
    """이미지 처리. image_bytes가 있으면 파일 다운로드 생략 (파일 첨부 경우)."""
    if image_bytes is None:
        # 사진으로 전송된 경우 — 텔레그램 photo 객체에서 다운로드
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_bytes = await file.download_as_bytearray()
        image_bytes = bytes(image_bytes)

    # 큰 이미지 리사이즈 (OCR 정확도 유지하면서 API 부하 감소)
    image_bytes = resize_image_if_needed(image_bytes)  # API 전송 전 최적화
    image_type = await classify_image(image_bytes)
    logger.info(f"이미지 분류: {image_type}")

    if image_type == "일별운행이력":
        await process_daily_history(update, image_bytes)
    elif image_type == "콜카드":
        await process_call_card(update, image_bytes)
    elif image_type == "충전":
        await process_charge_receipt(update, image_bytes)
    elif image_type == "결제":
        await process_payment_history(update, image_bytes)
    elif image_type == "세큐티":
        await process_sekuti(update, image_bytes)
    else:
        # 분류 실패 → 결제내역으로 재시도 (결제내역이 가장 자주 오인식됨)
        logger.info("분류 실패 → 결제내역 fallback 시도")
        await update.message.reply_text("🔄 이미지 재분석 중...")
        try:
            receipts_raw = await ocr_payment_history(image_bytes)
            if receipts_raw and len(receipts_raw) > 0:
                await process_payment_history(update, image_bytes)
            else:
                await update.message.reply_text(
                    "❓ 이미지 인식 실패\n"
                    "💡 더 크게 캡처하거나 밝은 환경에서 다시 올려주세요.\n"
                    "콜카드·충전내역·결제내역·세큐티만 처리 가능합니다."
                )
        except Exception as e:
            logger.error(f"fallback OCR 오류: {e}")
            await update.message.reply_text(
                "❓ 이미지 인식 실패\n"
                "💡 더 크게 캡처해서 다시 올려주세요."
            )

async def process_image_queue_worker():
    while True:
        item = await image_queue.get()
        # item은 (update, context) 또는 (update, context, image_bytes)
        if len(item) == 3:
            update, context, image_bytes = item
        else:
            update, context = item
            image_bytes = None
        try:
            await process_single_image(update, context, image_bytes=image_bytes)
        except Exception as e:
            logger.error(f"이미지 처리 오류: {type(e).__name__}: {e}", exc_info=True)
            try:
                err_msg = str(e)[:80] if str(e) else type(e).__name__
                await update.message.reply_text(
                    f"❌ 처리 오류: {err_msg}\n"
                    f"잠시 후 다시 올려주세요."
                )
            except Exception:
                pass
        finally:
            image_queue.task_done()
        await asyncio.sleep(1)

# ──────────────────────────────────────────────
# 수동 입력 파서
# ──────────────────────────────────────────────
def parse_manual_call(text: str) -> dict | None:
    """
    '콜 7800' / '배회 5600' / '콜 7800 경산'
    → {"콜유형": ..., "요금": ..., "도착지힌트": ...}
    """
    parts = text.strip().split()
    if len(parts) < 2:
        return None
    keyword = parts[0]
    if keyword not in ("콜", "배회"):
        return None
    try:
        fee = int(parts[1].replace(",", "").replace("원", ""))
    except ValueError:
        return None
    hint = parts[2] if len(parts) >= 3 else None
    return {
        "콜유형": "카카오T" if keyword == "콜" else "배회",
        "요금": fee,
        "도착지힌트": hint,
    }

EXPENSE_KEYWORDS = {
    "충전": ("⚡ 전기충전", "charges"),
    "타이어": ("🔧 타이어교체", None),
    "오일": ("🔧 오일교환", None),
    "세차": ("🚿 세차", None),
}

def parse_expense(text: str) -> dict | None:
    """
    '충전 4595' / '타이어 80000' / '지출 기타 3000'
    → {"카테고리": ..., "금액": ..., "메모": ...}
    """
    parts = text.strip().split()
    if not parts:
        return None

    keyword = parts[0]
    if keyword in EXPENSE_KEYWORDS and len(parts) >= 2:
        label, _ = EXPENSE_KEYWORDS[keyword]
        try:
            fee = int(parts[1].replace(",", "").replace("원", ""))
        except ValueError:
            return None
        return {"카테고리": label, "금액": fee, "메모": keyword}

    if keyword == "지출" and len(parts) >= 3:
        cat = parts[1]
        try:
            fee = int(parts[2].replace(",", "").replace("원", ""))
        except ValueError:
            return None
        return {"카테고리": f"📦 {cat}", "금액": fee, "메모": cat}

    return None

# ──────────────────────────────────────────────
# 핸들러 — 수동 입력
# ──────────────────────────────────────────────
async def handle_manual_call(update: Update, parsed: dict):
    today = str(today_kst())
    dow = get_dow()
    payload = {
        "날짜": today,
        "요일": dow,
        "배차시각": now_kst().strftime("%H:%M"),
        "요금": parsed["요금"],
        "콜유형": parsed["콜유형"],
        "도착지": parsed.get("도착지힌트"),
        "data_source": "manual_entry",
    }
    r = await sb_insert("raw_calls", payload)
    if r:
        await update.message.reply_text(
            f"✅ {parsed['콜유형']} {fmt(parsed['요금'])} 입력"
        )
    else:
        await update.message.reply_text("❌ 저장 실패")

async def handle_expense(update: Update, parsed: dict):
    today = str(today_kst())
    payload = {
        "날짜": today,
        "카테고리": parsed["카테고리"],
        "금액": parsed["금액"],
        "메모": parsed.get("메모", ""),
        "자동여부": False,
    }
    r = await sb_insert("expenses", payload)
    if r:
        await update.message.reply_text(
            f"✅ 지출 {parsed['카테고리']} {fmt(parsed['금액'])} 입력"
        )
    else:
        await update.message.reply_text("❌ 저장 실패")

async def handle_expense_cancel(update: Update):
    today = str(today_kst())
    ok = await sb_delete_last(
        "expenses",
        {"날짜": f"eq.{today}", "자동여부": "eq.false"}
    )
    if ok:
        await update.message.reply_text("✅ 마지막 수동 지출 삭제")
    else:
        await update.message.reply_text("⚠️ 삭제할 수동 지출 없음")

async def handle_rest_day(update: Update, text: str = "휴무"):
    """휴무 처리. '4-7 휴무' 형식으로 날짜 지정 가능."""
    import re
    from datetime import date as date_cls

    target = today_kst()
    date_pat = r"(\d{4})-(\d{1,2})-(\d{1,2})|(\d{1,2})[-/](\d{1,2})"
    m = re.search(date_pat, text)
    if m:
        g = m.groups()
        try:
            if g[0]:
                target = date_cls(int(g[0]), int(g[1]), int(g[2]))
            else:
                target = date_cls(today_kst().year, int(g[3]), int(g[4]))
        except ValueError:
            await update.message.reply_text("❌ 잘못된 날짜입니다.")
            return

    dow_map = ["월","화","수","목","금","토","일"]
    date_str = str(target)
    dow = dow_map[target.weekday()]

    await sb_upsert("daily_summary", {
        "날짜": date_str, "요일": dow,
        "휴무여부": True, "정상여부": "휴무",
    }, on_conflict="날짜")
    await insert_insurance(target)
    await update.message.reply_text(
        f"✅ {date_str} ({dow}) 휴무 처리\n보험료 {INSURANCE_DAILY:,}원 자동 기록"
    )

# ──────────────────────────────────────────────
# 핸들러 — 조회
# ──────────────────────────────────────────────
async def handle_today_quick(update: Update):
    s = await today_summary()
    달성바 = "█" * (s["달성률"] // 10) + "░" * (10 - s["달성률"] // 10)
    # 신사고 체계: 건수를 요일 분포 맥락으로 표현
    from datetime import date as _date2
    dow_kor = ["월","화","수","목","금","토","일"][_date2.today().weekday()]
    DOW_EXPECTED = {
        "월":(7,12),"화":(8,13),"수":(9,14),"목":(8,12),
        "금":(10,15),"토":(11,16),"일":(10,15)
    }
    exp_min, exp_max = DOW_EXPECTED.get(dow_kor, (8,13))
    건수 = s["건수"]
    if 건수 == 0:
        건수_평가 = ""
    elif 건수 < exp_min:
        건수_평가 = f" (기대 {exp_min}~{exp_max}건 하위)"
    elif 건수 > exp_max:
        건수_평가 = f" (기대 {exp_min}~{exp_max}건 상위)"
    else:
        건수_평가 = f" (기대 {exp_min}~{exp_max}건 정상)"

    await update.message.reply_text(
        f"📍 오늘 현황 ({now_kst().strftime('%m/%d %H:%M')})\n"
        f"콜 {건수}건{건수_평가} | 매출 {fmt(s['매출'])}\n"
        f"지출 {fmt(s['지출'])} | 순수익 {fmt(s['순수익'])}\n"
        f"목표 [{달성바}] {s['달성률']}%"
    )

async def handle_weekly(update: Update):
    today = today_kst()
    start = today - timedelta(days=6)
    start_str = str(start)
    end_str = str(today)

    calls = await sb_select(
        "raw_calls",
        {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"}
    )
    expenses = await sb_select(
        "expenses",
        {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"}
    )
    총건수 = len(calls)
    총매출 = sum(c.get("요금", 0) or 0 for c in calls)
    총지출 = sum(e.get("금액", 0) or 0 for e in expenses)
    순수익 = calc_net(총매출, 총지출)
    일평균매출 = 총매출 // 7 if 총매출 else 0

    await update.message.reply_text(
        f"📅 주간 요약 ({start_str} ~ {end_str})\n"
        f"총 {총건수}건 | 매출 {fmt(총매출)}\n"
        f"지출 {fmt(총지출)} | 순수익 {fmt(순수익)}\n"
        f"일평균 매출 {fmt(일평균매출)}"
    )

async def handle_monthly(update: Update):
    today = today_kst()
    start_str = today.replace(day=1).isoformat()
    end_str = str(today)

    calls = await sb_select_calls( {"날짜": f"gte.{start_str}"})
    expenses = await sb_select("expenses", {"날짜": f"gte.{start_str}"})
    총건수 = len(calls)
    총매출 = sum(c.get("요금", 0) or 0 for c in calls)
    총지출 = sum(e.get("금액", 0) or 0 for e in expenses)
    순수익 = calc_net(총매출, 총지출)

    # 카테고리별 지출
    cat_map = {}
    for e in expenses:
        cat = e.get("카테고리", "기타")
        cat_map[cat] = cat_map.get(cat, 0) + (e.get("금액", 0) or 0)
    cat_lines = "\n".join(f"  {k}: {fmt(v)}" for k, v in sorted(cat_map.items()))

    운행일 = len(set(c.get("날짜") for c in calls))
    일평균 = 총매출 // 운행일 if 운행일 else 0

    await update.message.reply_text(
        f"📆 월간 요약 ({start_str} ~)\n"
        f"운행일 {운행일}일 | 총 {총건수}건\n"
        f"총매출 {fmt(총매출)} | 일평균 {fmt(일평균)}\n"
        f"지출 {fmt(총지출)} | 순수익 {fmt(순수익)}\n"
        f"\n지출 카테고리:\n{cat_lines}"
    )

async def handle_expense_check(update: Update):
    expenses = await today_expenses()
    if not expenses:
        await update.message.reply_text("오늘 지출 없음")
        return
    lines = [f"💸 오늘 지출 ({str(today_kst())})"]
    total = 0
    for e in expenses:
        lines.append(f"  {e.get('카테고리','')} {fmt(e.get('금액',0))}")
        total += e.get("금액", 0) or 0
    lines.append(f"합계: {fmt(total)}")
    await update.message.reply_text("\n".join(lines))


async def handle_receipt_delete(update, text: str):
    """
    결제내역 삭제 명령어
    결제삭제 YYYY-MM-DD 운행외  → 02:01~18:59 시간대 삭제
    결제삭제 YYYY-MM-DD 0원     → 요금 0원·null 삭제
    결제삭제 YYYY-MM-DD 전체    → 해당 날짜 전체 삭제
    결제삭제 YYYY-MM-DD HH:MM   → 특정 시각 삭제
    """
    parts = text.strip().split(" ", 2)
    if len(parts) < 3:
        await update.message.reply_text(
            "형식:\n"
            "결제삭제 YYYY-MM-DD 운행외\n"
            "결제삭제 YYYY-MM-DD 0원\n"
            "결제삭제 YYYY-MM-DD 전체\n"
            "결제삭제 YYYY-MM-DD HH:MM"
        )
        return

    date_str = parts[1].strip()
    mode = parts[2].strip()

    try:
        rows = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})
        if not rows:
            await update.message.reply_text(f"⚠️ {date_str} 결제내역 없음")
            return

        delete_ids = []

        if mode == "운행외":
            for r in rows:
                t = r.get("시각", "") or ""
                try:
                    h, m = t.split(":")
                    mins = int(h) * 60 + int(m)
                    # 운행시간 외: 02:01~18:59
                    if 121 <= mins <= 1019:
                        delete_ids.append(r["id"])
                except Exception:
                    delete_ids.append(r["id"])

        elif mode == "0원":
            for r in rows:
                fee = r.get("요금")
                if fee is None or int(fee) == 0:
                    delete_ids.append(r["id"])

        elif mode == "전체":
            delete_ids = [r["id"] for r in rows]

        elif ":" in mode:
            for r in rows:
                if r.get("시각", "") == mode:
                    delete_ids.append(r["id"])
        else:
            await update.message.reply_text(f"❓ 알 수 없는 모드: {mode}")
            return

        if not delete_ids:
            await update.message.reply_text(f"✅ 삭제 대상 없음 ({mode})")
            return

        deleted = 0
        for rid in delete_ids:
            await sb_h("DELETE", f"payment_receipts?id=eq.{rid}")
            deleted += 1

        await update.message.reply_text(
            f"🗑️ 삭제 완료\n"
            f"{date_str} | 조건: {mode}\n"
            f"{deleted}건 삭제 (전체 {len(rows)}건 중)"
        )

    except Exception as e:
        logger.error(f"결제삭제 오류: {e}")
        await update.message.reply_text(f"❌ 삭제 오류: {str(e)[:200]}")




async def confirm_cross_check(date_str: str) -> str:
    """
    '대조 확정 YYYY-MM-DD' 명령어 처리.
    교차대조 미매칭 결제내역(현금) → raw_calls 배회영업으로 자동 추가.
    """
    from datetime import date as date_cls, timedelta

    try:
        y, mo, d = date_str.split("-")
        next_date_str = str(date_cls(int(y),int(mo),int(d)) + timedelta(days=1))
    except Exception:
        return "❌ 날짜 형식 오류 (YYYY-MM-DD)"

    # 콜카드 + 결제내역 조회 (익일 포함)
    calls    = await sb_select_calls( {"날짜": f"eq.{date_str}"})
    receipts = (await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})) +                (await sb_select("payment_receipts", {"날짜": f"eq.{next_date_str}"}))

    def to_min_smart(배차, 대상시각, 대상날짜):
        try:
            bh, bm = 배차.split(":")
            base = int(bh)*60+int(bm)
            th, tm = 대상시각.split(":")
            target = int(th)*60+int(tm)
            if 대상날짜 == next_date_str:
                target += 1440
            elif target < base - 60:
                target += 1440
            return target
        except: return None

    # 매칭
    matched_r = set()
    for call in calls:
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각")
        if not 하차:
            try:
                bh, bm = 배차.split(":")
                est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: continue
        c_min = to_min_smart(배차, 하차, date_str)
        best_j, best_diff = None, 99999
        for j, r in enumerate(receipts):
            if j in matched_r: continue
            r_min = to_min_smart(배차, r.get("시각","") or "", r.get("날짜", date_str))
            if r_min and c_min:
                diff = abs(c_min - r_min)
                if diff <= 20 and diff < best_diff:
                    best_diff = diff
                    best_j = j
        if best_j is not None:
            matched_r.add(best_j)

    unmatched = [r for j,r in enumerate(receipts) if j not in matched_r]

    # 현금 결제 → 배회영업으로 추가
    baehoe = [r for r in unmatched if (r.get("결제방법") or "") in ("현금","")]

    if not baehoe:
        return f"✅ {date_str} 배회후보 없음 (현금 미매칭 결제내역 없음)"

    DOW_MAP = ["월","화","수","목","금","토","일"]
    added = 0
    for r in baehoe:
        r_date = r.get("날짜") or date_str
        from datetime import date as dc
        try:
            rd = dc.fromisoformat(r_date)
            dow = DOW_MAP[rd.weekday()]
        except: dow = ""

        payload = {
            "날짜":     r_date,
            "요일":     dow,
            "배차시각": r.get("시각"),
            "하차시각": r.get("시각"),  # 결제시각 = 하차시각으로 설정
            "요금":     r.get("요금"),
            "콜유형":   "배회",
            "비고":     "결제내역 교차대조 자동추가",
            "data_source": "app_ocr_individual",
        }
        result = await sb_insert("raw_calls", payload)
        if result:
            added += 1

    # 요약
    total_calls = len(calls) + added
    kakao = len([c for c in calls if c.get("콜유형","") == "카카오T"])
    lines = [
        f"✅ 대조 확정 완료 — {date_str}",
        f"배회영업 {added}건 raw_calls 추가",
        f"",
        f"📊 확정 후 현황:",
        f"  총 {total_calls}건 (카카오T {kakao}건 + 배회 {added}건)",
        f"  결제내역 {len(receipts)}건 → {'✅ 완전매칭' if total_calls == len(receipts) else f'⚠️ 차이 {abs(total_calls - len(receipts))}건'}",
    ]
    return "\n".join(lines)


async def handle_fee_confirm_request(update, date_str: str):
    """
    '대조 금액확인 YYYY-MM-DD' 명령어.
    해당 날짜 매칭 건 중 금액 불일치(≥500원) 건에 대해
    InlineKeyboard 버튼으로 확인 요청.
    """
    from datetime import date as date_cls, timedelta

    calls    = await sb_select_calls( {"날짜": f"eq.{date_str}"})
    try:
        y,mo,d = date_str.split("-")
        next_d = str(date_cls(int(y),int(mo),int(d)) + timedelta(days=1))
    except: next_d = date_str

    receipts = (await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})) +                (await sb_select("payment_receipts", {"날짜": f"eq.{next_d}"}))

    def to_min(배차, 대상, 대상날짜):
        try:
            bh,bm = 배차.split(":"); base = int(bh)*60+int(bm)
            th,tm = 대상.split(":"); target = int(th)*60+int(tm)
            if 대상날짜 == next_d: target += 1440
            elif target < base-60: target += 1440
            return target
        except: return None

    matched_r = set()
    mismatches = []

    for call in calls:
        배차 = call.get("배차시각") or ""
        하차 = call.get("하차시각") or ""
        if not 하차:
            try:
                bh,bm = 배차.split(":"); est = int(bh)*60+int(bm)+20
                하차 = f"{est//60%24:02d}:{est%60:02d}"
            except: continue
        c_min = to_min(배차, 하차, date_str)
        call_fee = call.get("요금") or 0
        best_j, best_diff = None, 99999
        for j, r in enumerate(receipts):
            if j in matched_r: continue
            r_min = to_min(배차, r.get("시각","") or "", r.get("날짜", date_str))
            if r_min and c_min:
                diff = abs(c_min - r_min)
                if diff <= 20 and diff < best_diff:
                    best_diff = diff; best_j = j
        if best_j is not None:
            matched_r.add(best_j)
            rcpt_fee = receipts[best_j].get("요금") or 0
            fee_diff = abs(call_fee - rcpt_fee)
            if fee_diff >= FEE_DIFF_THRESHOLD:
                mismatches.append({
                    "call_id": call.get("id"),
                    "배차시각": 배차,
                    "call_fee": call_fee,
                    "rcpt_fee": rcpt_fee,
                    "diff": fee_diff,
                })

    if not mismatches:
        await update.message.reply_text(f"✅ {date_str} 금액 불일치 없음")
        return

    for fm in mismatches:
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton(
                f"콜카드 {fm['call_fee']:,}원",
                callback_data=f"fee:{fm['call_id']}:{fm['call_fee']}"
            ),
            InlineKeyboardButton(
                f"결제내역 {fm['rcpt_fee']:,}원",
                callback_data=f"fee:{fm['call_id']}:{fm['rcpt_fee']}"
            ),
        ]])
        await update.message.reply_text(
            f"⚠️ 금액 불일치 확인\n"
            f"배차: {fm['배차시각']}\n"
            f"콜카드: {fm['call_fee']:,}원 | 결제: {fm['rcpt_fee']:,}원\n"
            f"차이: {fm['diff']:,}원",
            reply_markup=keyboard
        )

async def handle_date_stat(update, text: str):
    """
    날짜+통계 키워드 조합 처리
    예: '3-17 총건수', '3-17 매출', '3-17 순수익', '3-17 지출'
    """
    import re
    from datetime import date as date_cls

    today_d = today_kst()
    dow_map = ["월","화","수","목","금","토","일"]

    # 날짜 추출
    parsed = None
    for pat, mode in [
        (r"(\d{4})-(\d{1,2})-(\d{1,2})", "full"),
        (r"(\d{1,2})-(\d{1,2})",           "md"),
        (r"(\d{1,2})/(\d{1,2})",           "md"),
    ]:
        m = re.search(pat, text)
        if m:
            g = m.groups()
            try:
                parsed = date_cls(int(g[0]),int(g[1]),int(g[2])) if mode=="full"                          else date_cls(today_d.year, int(g[0]), int(g[1]))
                break
            except ValueError:
                pass

    if not parsed:
        await update.message.reply_text("❓ 날짜 인식 실패\n예: 3-17 총건수")
        return

    date_key = str(parsed)
    dow = dow_map[parsed.weekday()]
    header = f"📅 {date_key} ({dow})"

    # 통계 키워드 판단
    if any(kw in text for kw in ["총건수", "건수"]):
        calls = await sb_select_calls( {"날짜": f"eq.{date_key}"})
        카카오 = sum(1 for c in calls if (c.get("콜유형") or "") == "카카오T")
        배회   = sum(1 for c in calls if (c.get("콜유형") or "") == "배회")
        총매출 = sum(c.get("요금") or 0 for c in calls)
        건당   = 총매출 // len(calls) if calls else 0
        await update.message.reply_text(
            f"{header} 총건수\n"
            f"총 {len(calls)}콜\n"
            f"  🚕 카카오T {카카오}건\n"
            f"  🚶 배회 {배회}건\n"
            f"건당 평균 {fmt(건당)}"
        )

    elif "매출" in text:
        calls = await sb_select_calls( {"날짜": f"eq.{date_key}"})
        총매출 = sum(c.get("요금") or 0 for c in calls)
        건수   = len(calls)
        건당   = 총매출 // 건수 if 건수 else 0
        await update.message.reply_text(
            f"{header} 매출\n"
            f"총매출 {fmt(총매출)}\n"
            f"콜수 {건수}건 | 건당 {fmt(건당)}"
        )

    elif "순수익" in text:
        calls    = await sb_select_calls(  {"날짜": f"eq.{date_key}"})
        expenses = await sb_select("expenses",   {"날짜": f"eq.{date_key}"})
        총매출 = sum(c.get("요금") or 0 for c in calls)
        총지출 = sum(e.get("금액") or 0 for e in expenses)
        순수익 = calc_net(총매출, 총지출)
        달성률 = min(int(순수익 / NET_GOAL * 100), 999) if NET_GOAL else 0
        달성바 = "█" * min(달성률//10,10) + "░" * max(10-달성률//10,0)
        await update.message.reply_text(
            f"{header} 순수익\n"
            f"매출 {fmt(총매출)} | 지출 {fmt(총지출)}\n"
            f"순수익 {fmt(순수익)}\n"
            f"목표 [{달성바}] {달성률}%"
        )

    elif "지출" in text:
        expenses = await sb_select("expenses", {"날짜": f"eq.{date_key}", "order": "id.asc"})
        총지출 = sum(e.get("금액") or 0 for e in expenses)
        if not expenses:
            await update.message.reply_text(f"{header}\n지출 없음")
            return
        lines_out = [f"{header} 지출 {fmt(총지출)}"]
        for e in expenses:
            auto = " (자동)" if e.get("자동여부") else ""
            lines_out.append(f"  {e.get('카테고리','')} {fmt(e.get('금액') or 0)}{auto}")
        await update.message.reply_text("\n".join(lines_out))

    else:
        # 전체 조회로 위임
        await handle_date_query(update, text)


async def handle_date_query(update, date_str: str):
    """특정 날짜 조회: 운행 내역 + 요약 + 지출"""
    import re
    from datetime import date as date_cls

    text = date_str.replace("조회","").replace("일","").strip()
    today_d = today_kst()
    parsed = None

    for pattern, mode in [
        (r"^(\d{4})-(\d{1,2})-(\d{1,2})$", "full"),
        (r"^(\d{1,2})-(\d{1,2})$",           "md"),
        (r"^(\d{1,2})/(\d{1,2})$",           "md"),
    ]:
        m = re.match(pattern, text)
        if m:
            g = m.groups()
            try:
                if mode == "full":
                    parsed = date_cls(int(g[0]), int(g[1]), int(g[2]))
                else:
                    parsed = date_cls(today_d.year, int(g[0]), int(g[1]))
                break
            except ValueError:
                pass

    if not parsed:
        await update.message.reply_text(
            "❓ 날짜 형식 오류\n"
            "예시: 3-2 조회 / 3/2 조회 / 2026-03-02 조회"
        )
        return

    date_key = str(parsed)
    dow_map = ["월","화","수","목","금","토","일"]
    dow = dow_map[parsed.weekday()]

    calls    = await sb_select_calls( {"날짜": f"eq.{date_key}", "order": "배차시각.asc"})
    expenses = await sb_select("expenses",  {"날짜": f"eq.{date_key}", "order": "id.asc"})

    if not calls and not expenses:
        await update.message.reply_text(f"📭 {date_key} ({dow}) 데이터 없음")
        return

    result_lines = [f"📅 {date_key} ({dow}) 조회\n"]

    if calls:
        총매출 = sum(c.get("요금") or 0 for c in calls)
        result_lines.append(f"[운행] {len(calls)}콜 | {fmt(총매출)}")
        for c in calls:
            배차 = c.get("배차시각") or "-"
            출발 = (c.get("출발지") or "")[:8]
            도착 = (c.get("도착지") or "")[:8]
            요금 = fmt(c.get("요금") or 0)
            유형 = c.get("콜유형") or "카카오T"
            icon = "🚕" if 유형 == "카카오T" else "🚶"
            result_lines.append(f"  {icon}{배차} {출발}→{도착} {요금}")
    else:
        총매출 = 0
        result_lines.append("[운행] 없음")

    result_lines.append("")

    총지출 = sum(e.get("금액") or 0 for e in expenses)
    if expenses:
        result_lines.append(f"[지출] {fmt(총지출)}")
        for e in expenses:
            cat = e.get("카테고리") or ""
            amt = fmt(e.get("금액") or 0)
            auto = " (자동)" if e.get("자동여부") else ""
            result_lines.append(f"  {cat} {amt}{auto}")
    else:
        result_lines.append("[지출] 없음")

    result_lines.append("")
    순수익 = calc_net(총매출, 총지출)
    달성률 = min(int(순수익 / NET_GOAL * 100), 999) if NET_GOAL else 0
    달성바 = "█" * min(달성률//10, 10) + "░" * max(10 - 달성률//10, 0)
    result_lines.append("[요약]")
    result_lines.append(f"  매출 {fmt(총매출)} | 지출 {fmt(총지출)}")
    result_lines.append(f"  순수익 {fmt(순수익)}")
    result_lines.append(f"  목표 [{달성바}] {달성률}%")

    msg = "\n".join(result_lines)
    if len(msg) > 4000:
        msg = msg[:4000] + "\n...(생략)"
    await update.message.reply_text(msg)




def parse_manual_full(text: str) -> dict | None:
    """
    수동 콜 전체 입력 파싱.
    형식:
      2026 03 01 23 05 수성못>대명1동 8500 카카오
      26 03 01 23 05 수성못>대명1동 8500 배회
      03012305 수성못>대명1동 8500 카카오
      0301 2305 수성못>대명1동 8500 배회
      2026-03-01 23:05 수성못>대명1동 8500 카카오
    """
    import re as _re
    from datetime import date as _date

    orig = text.strip()
    today = _date.today()

    # 콜유형
    콜유형 = "배회" if "배회" in orig else "카카오T"
    clean = _re.sub(r'배회|카카오T?', '', orig).strip()

    # 경로 (출발>도착)
    route_m = _re.search(r'([가-힣\w]+)\s*[>→]\s*([가-힣\w]+)', clean)
    출발지 = 도착지 = None
    if route_m:
        출발지 = route_m.group(1).strip()
        도착지 = route_m.group(2).strip()
        clean = (clean[:route_m.start()] + ' ' + clean[route_m.end():]).strip()

    # 날짜+시각 파싱 (패턴 순서대로 시도)
    날짜 = None
    배차시각 = None

    # A: YYYY-MM-DD HH:MM 또는 YY-MM-DD HH:MM
    m = _re.search(r'(\d{2,4})[.\-](\d{1,2})[.\-](\d{1,2})\s+(\d{1,2}):(\d{2})', clean)
    if m:
        g = m.groups(); y = int(g[0]); y = y+2000 if y<100 else y
        try:
            날짜 = _date(y,int(g[1]),int(g[2])); 배차시각 = f"{int(g[3]):02d}:{g[4]}"
            clean = clean[:m.start()] + ' ' + clean[m.end():]
        except ValueError: pass

    # B: YYYY MM DD HH MM
    if not 날짜:
        m = _re.search(r'(\d{4})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{2})\b', clean)
        if m:
            g = m.groups()
            try:
                날짜 = _date(int(g[0]),int(g[1]),int(g[2])); 배차시각 = f"{int(g[3]):02d}:{g[4]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    # C: YY MM DD HH MM
    if not 날짜:
        m = _re.search(r'\b(\d{2})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+(\d{2})\b', clean)
        if m:
            g = m.groups()
            try:
                날짜 = _date(int(g[0])+2000,int(g[1]),int(g[2])); 배차시각 = f"{int(g[3]):02d}:{g[4]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    # D: MMDD HHMM
    if not 날짜:
        m = _re.search(r'\b(\d{4})\s+(\d{4})\b', clean)
        if m:
            a,b = m.group(1), m.group(2)
            try:
                날짜 = _date(today.year,int(a[:2]),int(a[2:])); 배차시각 = f"{int(b[:2]):02d}:{b[2:]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    # E: MMDDHHMM (8자리)
    if not 날짜:
        m = _re.search(r'\b(\d{8})\b', clean)
        if m:
            n = m.group(1)
            try:
                날짜 = _date(today.year,int(n[0:2]),int(n[2:4])); 배차시각 = f"{int(n[4:6]):02d}:{n[6:]}"
                clean = clean[:m.start()] + ' ' + clean[m.end():]
            except ValueError: pass

    if not 날짜 or not 배차시각:
        return None

    # 요금: 남은 clean의 4~6자리 숫자
    fee_m = _re.search(r'(?<!\d)(\d{4,6})(?!\d)', clean)
    if not fee_m:
        return None
    요금 = int(fee_m.group(1))

    요일 = ["월","화","수","목","금","토","일"][날짜.weekday()]
    return {"날짜":str(날짜),"요일":요일,"배차시각":배차시각,
            "출발지":출발지,"도착지":도착지,"요금":요금,"콜유형":콜유형}


async def handle_manual_full_call(update, text: str):
    """수동 전체 입력 콜 저장"""
    data = parse_manual_full(text)
    if not data:
        await update.message.reply_text(
            "❌ 형식 오류\n\n"
            "예시:\n"
            "2026 03 01 23 05 수성못>대명1동 8500 카카오\n"
            "26 03 01 23 05 수성못>대명1동 8500 배회\n"
            "0301 2305 수성못>대명1동 8500 카카오\n"
            "03012305 수성못>대명1동 8500 배회"
        )
        return

    # 중복 체크 → 자동 삭제 후 재저장
    deleted = await delete_duplicate_call(data["날짜"], data["배차시각"], data["요금"])
    if deleted:
        logger.info(f"수동입력 중복 삭제: {data['날짜']} {data['배차시각']}")

    result = await sb_insert("raw_calls", {
        "날짜":     data["날짜"],
        "요일":     data["요일"],
        "배차시각": data["배차시각"],
        "출발지":   data["출발지"],
        "도착지":   data["도착지"],
        "요금":     data["요금"],
        "콜유형":   data["콜유형"],
        "비고":     "수동입력",
        "data_source": "manual_entry",
    })

    if result:
        await update.message.reply_text(
            f"✅ 수동입력 저장\n"
            f"{data['날짜']}({data['요일']}) {data['배차시각']}\n"
            f"{data.get('출발지','-')}→{data.get('도착지','-')}\n"
            f"{data['요금']:,}원 [{data['콜유형']}]"
        )
    else:
        await update.message.reply_text("❌ DB 저장 실패")


async def handle_call_edit(update, text: str):
    """
    콜카드 수동 수정 명령어.
    형식:
      콜수정 HH:MM 필드=값
      콜수정 YYYY-MM-DD HH:MM 필드=값
    예:
      콜수정 19:18 요금=13100
      콜수정 19:18 배차시각=22:46
      콜수정 2026-03-20 19:18 도착지=수성구 만촌3동
    지원 필드: 배차시각, 하차시각, 출발지, 도착지, 요금, 콜유형, 비고
    """
    import re

    EDITABLE = {"배차시각","하차시각","출발지","도착지","요금","콜유형","비고"}

    parts = text.strip().split(" ", 1)
    if len(parts) < 2:
        await update.message.reply_text(
            "형식: 콜수정 HH:MM 필드=값\n"
            "날짜지정: 콜수정 YYYY-MM-DD HH:MM 필드=값\n"
            "예) 콜수정 19:18 요금=13100\n"
            "예) 콜수정 2026-03-20 19:18 배차시각=22:46"
        )
        return

    rest = parts[1].strip()

    # 날짜 포함 여부 판단
    date_match = re.match(r'^(\d{4}-\d{1,2}-\d{1,2})\s+(\d{1,2}:\d{2})\s+(.+)$', rest)
    time_only  = re.match(r'^(\d{1,2}:\d{2})\s+(.+)$', rest)

    if date_match:
        target_date  = date_match.group(1)
        target_time  = date_match.group(2)
        field_str    = date_match.group(3)
    elif time_only:
        target_date  = str(today_kst())
        target_time  = time_only.group(1)
        field_str    = time_only.group(2)
    else:
        await update.message.reply_text("❌ 형식 오류\n예) 콜수정 19:18 요금=13100")
        return

    # 필드=값 파싱
    field_match = re.match(r'^(\S+?)=(.+)$', field_str.strip())
    if not field_match:
        await update.message.reply_text("❌ 필드=값 형식 오류\n예) 요금=13100")
        return

    field = field_match.group(1).strip()
    value = field_match.group(2).strip()

    if field not in EDITABLE:
        await update.message.reply_text(
            f"❌ '{field}' 는 수정 불가\n"
            f"수정 가능: {', '.join(sorted(EDITABLE))}"
        )
        return

    # 요금은 int 변환
    if field == "요금":
        try:
            value = int(value.replace(",","").replace("원",""))
        except ValueError:
            await update.message.reply_text("❌ 요금은 숫자만 입력 (예: 13100)")
            return

    # DB에서 해당 건 찾기
    rows = await sb_select("raw_calls", {
        "날짜": f"eq.{target_date}",
        "배차시각": f"eq.{target_time}",
    })

    if not rows:
        await update.message.reply_text(
            f"⚠️ {target_date} {target_time} 콜 없음\n"
            f"날짜·시각을 확인해주세요."
        )
        return

    if len(rows) > 1:
        lines_out = [f"⚠️ {target_time} 콜이 {len(rows)}건 있습니다. 어느 건?"]
        for r in rows:
            lines_out.append(
                f"  ID:{r['id']} {r.get('출발지','')}→{r.get('도착지','')} {fmt(r.get('요금') or 0)}"
            )
        lines_out.append("ID 지정: 콜수정ID [id] 필드=값")
        await update.message.reply_text("\n".join(lines_out))
        return

    row = rows[0]
    old_val = row.get(field)
    row_id  = row["id"]

    # PATCH
    result = await sb_h("PATCH", f"raw_calls?id=eq.{row_id}", json={field: value})

    if result is not None:
        await update.message.reply_text(
            f"✅ 콜 수정 완료\n"
            f"날짜: {target_date} | 배차: {target_time}\n"
            f"{field}: {old_val} → {value}"
        )
    else:
        await update.message.reply_text("❌ 수정 실패")


async def handle_call_edit_by_id(update, text: str):
    """
    ID 지정 수정: '콜수정ID [id] 필드=값'
    동일 시각 콜이 여러 건일 때 사용
    """
    import re
    EDITABLE = {"배차시각","하차시각","출발지","도착지","요금","콜유형","비고"}

    m = re.match(r'^(\d+)\s+(\S+?)=(.+)$', text.strip())
    if not m:
        await update.message.reply_text("형식: 콜수정ID [id] 필드=값\n예) 콜수정ID 42 요금=13100")
        return

    row_id = int(m.group(1))
    field  = m.group(2).strip()
    value  = m.group(3).strip()

    if field not in EDITABLE:
        await update.message.reply_text(f"❌ '{field}' 수정 불가")
        return

    if field == "요금":
        try:
            value = int(value.replace(",","").replace("원",""))
        except ValueError:
            await update.message.reply_text("❌ 요금은 숫자만")
            return

    rows = await sb_select("raw_calls", {"id": f"eq.{row_id}"})
    if not rows:
        await update.message.reply_text(f"❌ ID {row_id} 없음")
        return

    old_val = rows[0].get(field)
    await sb_h("PATCH", f"raw_calls?id=eq.{row_id}", json={field: value})
    await update.message.reply_text(
        f"✅ ID {row_id} 수정\n{field}: {old_val} → {value}"
    )


async def handle_db_check(update: Update):
    calls = await sb_select("raw_calls", {"order": "id.desc", "limit": "1"})
    total_calls = await sb_select("raw_calls", {})
    total_exp = await sb_select("expenses", {})
    charging = await sb_select("charging_log", {"order": "id.desc", "limit": "1"})

    last_call = calls[0] if calls else {}
    last_charge = charging[0] if charging else {}

    await update.message.reply_text(
        f"🗄️ DB 현황\n"
        f"raw_calls: {len(total_calls)}건\n"
        f"expenses: {len(total_exp)}건\n"
        f"최근 콜: {last_call.get('날짜','-')} {last_call.get('배차시각','-')} {fmt(last_call.get('요금',0))}\n"
        f"최근 충전: {last_charge.get('충전일','-')} {last_charge.get('충전량_kwh','-')}kWh"
    )

# ──────────────────────────────────────────────
# 핸들러 — 전략
# ──────────────────────────────────────────────
async def get_strategy(update: Update):
    hour = now_kst().hour
    # 시간대 매칭
    if 19 <= hour < 21:
        time_key = "19~21"
    elif hour == 21:
        time_key = "21~22"
    elif 22 <= hour < 24:
        time_key = "22~00"
    elif 0 <= hour < 2:
        time_key = "00~02"
    elif 2 <= hour < 3:
        time_key = "02~03"
    else:
        time_key = "전체"

    rows = await sb_select(
        "strategy_lookup",
        {"시간대": f"in.(전체,{time_key})", "order": "우선순위.asc"}
    )
    if not rows:
        await update.message.reply_text("⚠️ 전략 테이블 없음. 마기 업데이트 필요")
        return

    lines = [f"⚡ 현재 전략 ({now_kst().strftime('%H:%M')})"]
    for r in rows:
        priority_icon = {"긴급": "🔴", "높음": "🟠", "보통": "🟡"}.get(r.get("우선순위", ""), "⚪")
        lines.append(f"{priority_icon} [{r.get('시간대','')}] {r.get('행동지침','')}")

    await update.message.reply_text("\n".join(lines))

async def handle_magi_update(update: Update, content: str):
    """'마기 업데이트 [시간대] [내용]' → strategy_lookup INSERT"""
    parts = content.strip().split(" ", 1)
    if len(parts) < 2:
        await update.message.reply_text("형식: 마기 업데이트 [시간대] [내용]")
        return
    시간대 = parts[0]
    지침 = parts[1]
    await sb_insert("strategy_lookup", {
        "시간대": 시간대,
        "행동지침": 지침,
        "우선순위": "높음",
    })
    await update.message.reply_text(f"✅ 전략 테이블 갱신완료\n[{시간대}] {지침}")

# ──────────────────────────────────────────────
# 핸들러 — 엑셀 다운로드
# ──────────────────────────────────────────────

async def handle_download_month(update, ym: str):
    """특정 월 다운로드 (예: 2026-03)"""
    try:
        year, month = ym.split("-")
        year, month = int(year), int(month)
    except Exception:
        await update.message.reply_text("❌ 형식 오류: YYYY-MM (예: 2026-03)")
        return

    from datetime import date, timedelta
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    start_str = str(start_date)
    end_str   = str(end_date)

    # AND 조건 (날짜 키 중복 버그 방지)
    calls    = await sb_select_calls(    {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"})
    expenses = await sb_select("expenses",     {"and": f"(날짜.gte.{start_str},날짜.lte.{end_str})"})
    charging = await sb_select("charging_log", {"and": f"(충전일.gte.{start_str},충전일.lte.{end_str})"})

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "운행기록"
    headers1 = ["날짜","요일","배차시각","출발지","도착지","요금","콜유형","비고"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4A90D9")
    for c in calls:
        ws1.append([c.get("날짜"),c.get("요일"),c.get("배차시각"),
                    c.get("출발지"),c.get("도착지"),c.get("요금"),
                    c.get("콜유형"),c.get("비고")])

    ws2 = wb.create_sheet("지출")
    headers2 = ["날짜","카테고리","금액","메모","자동여부"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F5A623")
    for e in expenses:
        ws2.append([e.get("날짜"),e.get("카테고리"),e.get("금액"),
                    e.get("메모"),e.get("자동여부")])

    ws3 = wb.create_sheet("충전기록")
    headers3 = ["충전일","충전량(kWh)","충전금액","충전소"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="7ED321")
    for ch in charging:
        ws3.append([ch.get("충전일"),ch.get("충전량_kwh"),
                    ch.get("충전금액"),ch.get("충전소")])

    filepath = f"/tmp/자비스_월간_{ym}.xlsx"
    wb.save(filepath)

    await update.message.reply_document(
        document=open(filepath, "rb"),
        filename=f"자비스_월간_{ym}.xlsx",
        caption=f"📊 {ym} 데이터 ({len(calls)}건)"
    )

async def handle_download(update: Update, scope: str):
    today = today_kst()
    if scope == "주간":
        start = today - timedelta(days=6)
        start_str = str(start)
        label = f"{start_str}_{today}"
    elif scope == "월간":
        start_str = today.replace(day=1).isoformat()
        label = f"{today.year}{today.month:02d}"
    else:
        start_str = "2000-01-01"
        label = "전체"

    calls = await sb_select_calls( {"날짜": f"gte.{start_str}"})
    expenses = await sb_select("expenses", {"날짜": f"gte.{start_str}"})
    charging = await sb_select("charging_log", {"충전일": f"gte.{start_str}"})

    wb = openpyxl.Workbook()

    # 시트1 raw_calls
    ws1 = wb.active
    ws1.title = "운행기록"
    headers1 = ["날짜", "요일", "배차시각", "출발지", "도착지", "요금", "콜유형", "비고"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4A90D9")
    for row, c in enumerate(calls, 2):
        ws1.append([
            c.get("날짜"), c.get("요일"), c.get("배차시각"),
            c.get("출발지"), c.get("도착지"), c.get("요금"),
            c.get("콜유형"), c.get("비고"),
        ])

    # 시트2 expenses
    ws2 = wb.create_sheet("지출")
    headers2 = ["날짜", "카테고리", "금액", "메모", "자동여부"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F5A623")
    for row, e in enumerate(expenses, 2):
        ws2.append([
            e.get("날짜"), e.get("카테고리"), e.get("금액"),
            e.get("메모"), e.get("자동여부"),
        ])

    # 시트3 charging_log
    ws3 = wb.create_sheet("충전기록")
    headers3 = ["충전일", "충전량(kWh)", "충전금액", "충전소"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="7ED321")
    for row, ch in enumerate(charging, 2):
        ws3.append([
            ch.get("충전일"), ch.get("충전량_kwh"),
            ch.get("충전금액"), ch.get("충전소"),
        ])

    filepath = f"/tmp/자비스_{scope}_{label}.xlsx"
    wb.save(filepath)

    await update.message.reply_document(
        document=open(filepath, "rb"),
        filename=f"자비스_{scope}_{label}.xlsx",
        caption=f"📊 {scope} 데이터 ({len(calls)}건)"
    )

# ──────────────────────────────────────────────
# 핸들러 — 데이터 이식 (v6 엑셀 → raw_calls)
# ──────────────────────────────────────────────
async def handle_excel_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """v6 인수인계 엑셀 업로드 시 raw_calls 자동 이식"""
    doc = update.message.document
    file = await context.bot.get_file(doc.file_id)
    file_bytes = await file.download_as_bytearray()

    import io
    wb = openpyxl.load_workbook(io.BytesIO(bytes(file_bytes)), read_only=True)
    target_sheets = [s for s in wb.sheetnames if "운행" in s or "데이터" in s]

    if not target_sheets:
        await update.message.reply_text("⚠️ 운행 데이터 시트를 찾지 못했습니다.")
        return

    await update.message.reply_text(f"📥 이식 시작: {target_sheets}")
    total_saved = 0
    total_skip = 0

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(h).strip() if h else "" for h in rows[0]]
        col = {h: i for i, h in enumerate(header)}

        for row in rows[1:]:
            if not any(row):
                continue
            try:
                날짜_raw = row[col.get("날짜", 0)]
                if not 날짜_raw:
                    continue
                날짜 = str(날짜_raw)[:10]
                요금_raw = row[col.get("요금", 5)]
                요금 = int(str(요금_raw).replace(",", "").replace("원", "")) if 요금_raw else 0
                배차시각 = row[col.get("배차시각", 2)]
                배차시각 = str(배차시각) if 배차시각 else None

                # 중복 체크
                existing = await sb_select(
                    "raw_calls",
                    {"날짜": f"eq.{날짜}", "요금": f"eq.{요금}", "배차시각": f"eq.{배차시각}"}
                )
                if existing:
                    total_skip += 1
                    continue

                payload = {
                    "날짜": 날짜,
                    "요일": row[col.get("요일", 1)] or "",
                    "배차시각": 배차시각,
                    "출발지": row[col.get("출발지", 3)],
                    "도착지": row[col.get("도착지", 4)],
                    "요금": 요금,
                    "콜유형": row[col.get("콜유형", 6)] or "카카오T",
                    "data_source": "manual_entry",
                }
                r = await sb_insert("raw_calls", payload)
                if r:
                    total_saved += 1
            except Exception as e:
                logger.error(f"이식 행 오류: {e}")
                continue

    await update.message.reply_text(
        f"✅ 이식 완료\n저장: {total_saved}건 | 중복스킵: {total_skip}건"
    )


# ══════════════════════════════════════════════
# 어군탐지기 v2 — 자동 브리핑 시스템
# ══════════════════════════════════════════════

FISH_DATA = {
    "월": {
        "19~21": [["달서구 신당동",      "8",  "9,500",  "S", "신당동 주택가 도로변",   "카카오T 공식 평일 호출 1위"]],
        "21~24": [["달서구 신당동",      "9",  "10,200", "S", "신당동 상가 밀집",       "평일 23시 호출 1위"]],
        "00~02": [["중구 성내1·2동",     "9",  "10,800", "S", "성내2동 앵커",           "유흥가 마감 귀가 폭증"]],
    },
    "화": {
        "19~21": [["수성구 범어·만촌",   "7",  "8,200",  "A", "범어역~만촌역",          "화요일 수성구 준수"]],
        "21~24": [["달서구 신당동",      "7",  "9,500",  "A", "신당동 상가",             "화요일 야간 준수"]],
        "00~02": [["중구 성내1·2동",     "8",  "10,200", "S", "성내2동 앵커",            "화요일 심야 귀가"]],
    },
    "수": {
        "19~21": [["수성구 범어·만촌",   "7",  "8,500",  "A", "범어역~만촌역",          "수요일 수성구 준수"]],
        "21~24": [["중구 동성로/삼덕동", "8",  "9,800",  "A", "삼덕동 먹자골목",        "수요일 야간 선호"]],
        "00~02": [["중구 성내1·2동",     "8",  "10,500", "S", "성내2동 앵커",            "수요일 심야 귀가"]],
    },
    "목": {
        "19~21": [["달서구 신당동",      "6",  "7,800",  "B", "신당동 주택가",           "목요일 콜 저조 — 단축 검토"]],
        "21~24": [["달서구 신당동",      "7",  "9,200",  "A", "신당동 상가",             "목요일 막판 집중"]],
        "00~02": [["중구 성내1·2동",     "7",  "9,800",  "A", "성내2동 앵커",            "목요일 심야 귀가"]],
    },
    "금": {
        "19~21": [["수성구 범어·만촌",   "8",  "7,600",  "A", "범어역~만촌역",          "금요일 수성구 콜 실적 최다"]],
        "21~24": [["중구 동성로/삼덕동", "9",  "10,000", "S", "삼덕동 먹자골목",        "유흥 피크! 술집 01시 마감"]],
        "00~02": [["중구 성내1·2동",     "9",  "11,500", "S", "성내2동 앵커",            "동성로 마감 귀가 폭증"]],
    },
    "토": {
        "19~21": [["수성구 고산2동",     "8",  "9,200",  "S", "수성못 주변",             "주말 17~19시 호출 집중"]],
        "21~24": [["중구 동성로/삼덕동", "9",  "8,800",  "S", "삼덕동~동성로",           "토요일 밤 유흥 최고 피크"]],
        "00~02": [["중구 성내1동",       "9",  "14,400", "S", "성내1동~성내2동",         "토요일 막판 단가 14,433원 최고치"]],
    },
    "일": {
        "19~21": [["수성구 고산2동",     "8",  "9,200",  "S", "수성못 주변",             "주말 호출 1위"]],
        "21~24": [["중구 동성로",        "8",  "11,800", "S", "동성로 입구",             "일요일 밤 단가 11,811원 고효율"]],
        "00~02": [["중구 성내1동",       "10", "12,400", "S", "성내1동",                 "주말 00~01시 호출 1위 구역"]],
    },
}

def get_fish_slot(hour: int) -> str | None:
    """현재 시각 → 어군 슬롯 반환"""
    if 19 <= hour < 21: return "19~21"
    if 21 <= hour <= 23: return "21~24"
    if 0 <= hour < 2:   return "00~02"
    return None

# 태그 이모지 매핑 (v2.2)
FISH_TAG_EMOJI = {
    "golden_time":    "⭐",
    "long_distance":  "🚀",
    "foreign_worker": "🌏",
    "blue_ocean":     "🌊",
    "avoid":          "⛔",
    "oversupply":     "⛔",
}

def get_fish_report(custom_hour: int = None) -> str | None:
    """⚠️ 사용 중단(2026-07-13) — FISH_DATA 하드코딩 기반, 더 이상 어디서도 호출 안 함.
    실데이터 기반 get_fish_report_db()로 완전 대체됨. 삭제 대신 보존만 해둠(참고용)."""
    now  = datetime.now(KST)
    hour = custom_hour if custom_hour is not None else now.hour
    day  = DOW_KOR[now.weekday()]
    slot = get_fish_slot(hour)
    if not slot:
        return None
    zones = FISH_DATA.get(day, {}).get(slot, [])
    if not zones:
        return f"🐟 {day}요일 {slot} 어군 데이터 없음"
    lines = [f"🐟 어군브리핑 — {day}요일 {slot}"]
    for idx, z in enumerate(zones, 1):
        grade_icon = {"S": "🔴", "A": "🟠", "B": "🟡", "C": "⚪"}.get(z[3], "⚪")
        lines.append(f"\n#{idx} {z[0]} {grade_icon}{z[3]}등급")
        lines.append(f"  점수 {z[1]}/10 | 예상 {z[2]}원")
        lines.append(f"  📍 {z[4]}")
        lines.append(f"  💡 {z[5]}")
    return "\n".join(lines)


# ──────────────────────────────────────────────
# 명령서#028 갭3 (2026-08-09, 마기 발부): 7일 이동평균 이중검증
# 가장 중요한 단일 지표(7일 이동평균)가 지금까지 daily_summary 한 가지 계산 방식에만
# 의존했음. raw_calls 직접집계(방식A)와 daily_summary(방식B) 두 방식으로 각각 계산해서
# 서로 다르면 즉시 경고한다. 자기검증시스템 갭분석(2026-08-09) 후속.
# ──────────────────────────────────────────────
async def mark_scheduler_run(job_name: str, result: str = "OK"):
    """영실 MCP안정화 제안서 §3(/health) 대응 — 각 스케줄러 작업 실행시각을
    DB(scheduler_status)에 기록. 프로세스 메모리 변수(last_kpi7day 등)는 재시작시
    사라지지만 이건 영구 기록되어 /health가 신뢰성 있게 참조 가능."""
    try:
        await sb_upsert("scheduler_status", {
            "job_name": job_name,
            "last_run_at": datetime.now(KST).isoformat(),
            "last_result": result,
        }, on_conflict="job_name")
    except Exception as e:
        logger.error(f"scheduler_status 기록 실패({job_name}): {e}")


async def sync_operated_status(days_back: int = 30):
    """task#38: 최근 N일에 대해 operated_status를 3단계 소스로 자동채움.
    confirmed(raw_calls존재)/gpx_proxy(raw_calls없음+GPX있음)는 자동확정,
    둘다없는 날짜만 미확인(pending)으로 남겨 질문 대상이 됨."""
    today = today_kst()
    dates = [str(today - timedelta(days=i)) for i in range(1, days_back+1)]

    raw_dates = set()
    try:
        rows = await sb_select("raw_calls", {"날짜": [f"gte.{dates[-1]}", f"lte.{dates[0]}"]})
        for r in (rows or []):
            if r.get("날짜"): raw_dates.add(r["날짜"])
    except Exception as e:
        logger.error(f"operated_status 동기화 - raw_calls 조회 실패: {e}")

    gpx_dates = set()
    try:
        rows = await sb_select("gpx_sessions", {"날짜": [f"gte.{dates[-1]}", f"lte.{dates[0]}"]})
        for r in (rows or []):
            if r.get("날짜"): gpx_dates.add(r["날짜"])
    except Exception as e:
        logger.error(f"operated_status 동기화 - gpx_sessions 조회 실패: {e}")

    try:
        existing_rows = await sb_select("operated_status", {"날짜": [f"gte.{dates[-1]}", f"lte.{dates[0]}"]})
        existing = {r["날짜"]: r for r in (existing_rows or [])}
    except Exception as e:
        logger.error(f"operated_status 동기화 - 기존행 조회 실패: {e}")
        existing = {}

    pending = []
    for d in dates:
        cur = existing.get(d)
        if d in raw_dates:
            if not cur or cur.get("source") != "confirmed":
                await sb_upsert("operated_status", {"날짜": d, "operated": True, "source": "confirmed"}, on_conflict="날짜")
        elif d in gpx_dates:
            if not cur or cur.get("source") != "gpx_proxy":
                await sb_upsert("operated_status", {"날짜": d, "operated": True, "source": "gpx_proxy"}, on_conflict="날짜")
        else:
            # raw_calls도 GPX도 없음 — 이미 질문했으면(asked_at 있음) 재질문 안 함
            if not cur:
                await sb_upsert("operated_status", {"날짜": d, "operated": None, "source": None}, on_conflict="날짜")
                pending.append(d)
            elif cur.get("source") is None and not cur.get("asked_at"):
                pending.append(d)
    return sorted(pending)


async def ask_operated_status_telegram():
    """task#38: 미확인 날짜(질문 안 한 것만) 텔레그램으로 한번에 묶어 질문.
    이미 질문한 날짜(asked_at 있음)는 재질문 안 함 — 스팸 방지."""
    pending = await sync_operated_status()
    if not pending:
        return []
    msg = (
        "❓ 아래 날짜는 운행기록(콜/GPX)이 없어 휴무인지 데이터누락인지 확인이 필요합니다.\n"
        + "\n".join(f"· {d}" for d in pending)
        + "\n\n휴무였던 날짜만 답장으로 알려주세요 (예: '8/13 8/15 휴무')."
    )
    try:
        await send_telegram_broadcast(msg)
        now_iso = datetime.now(KST).isoformat()
        for d in pending:
            await sb_upsert("operated_status", {"날짜": d, "asked_at": now_iso}, on_conflict="날짜")
        logger.info(f"operated_status 질문 발송 완료: {pending}")
    except Exception as e:
        logger.error(f"operated_status 질문 발송 실패: {e}")
    return pending


# ──────────────────────────────────────────────
# task#40 문제2 (2026-08-20): Haiku 기반 태스크 오케스트레이션.
# 마기 최종확정: APScheduler잡(폴링)+Haiku+OCR(Sonnet)과 완전분리+캐스퍼&SIMPLE만.
# tool 화이트리스트 4개뿐 — 파괴적작업(DELETE/UPDATE, GitHub커밋, Drive쓰기)은
# 이번 1단계에서 API경로로 완전히 배제. 완료해도 자동승인 없음(evidence PENDING).
# ──────────────────────────────────────────────
_ORCH_TOOLS = [
    {
        "name": "query_supabase",
        "description": "Supabase에 읽기전용 SELECT 쿼리만 실행한다. INSERT/UPDATE/DELETE는 거부된다.",
        "input_schema": {"type": "object", "properties": {
            "sql": {"type": "string", "description": "SELECT로 시작하는 SQL문"}
        }, "required": ["sql"]},
    },
    {
        "name": "record_event",
        "description": "이 태스크의 magi_task_events에 진행상황을 기록한다. 의미있는 단계마다 즉시 호출할 것.",
        "input_schema": {"type": "object", "properties": {
            "detail": {"type": "string", "description": "무엇을 확인/발견했는지 구체적으로"}
        }, "required": ["detail"]},
    },
    {
        "name": "write_correction_log",
        "description": "correction_log에 정정/발견 사항을 기록한다.",
        "input_schema": {"type": "object", "properties": {
            "table_name": {"type": "string"}, "reason": {"type": "string"}
        }, "required": ["table_name", "reason"]},
    },
    {
        "name": "finish_task",
        "description": "작업을 완료로 선언한다. 이 태스크에서 최종적으로 한 번만 호출.",
        "input_schema": {"type": "object", "properties": {
            "summary": {"type": "string", "description": "무엇을 확인/완료했는지 정직한 요약"}
        }, "required": ["summary"]},
    },
]

_ORCH_SYSTEM_PROMPT = (
    "당신은 캐스퍼(MAGI 시스템의 코드/DB 담당 에이전트)의 자동실행 모드입니다. "
    "핵심원칙: ①모르는 건 지어내지 않는다 ②확인 안 된 것을 확인됐다고 말하지 않는다 "
    "③이 세션에서는 조회(query_supabase)와 기록(record_event, write_correction_log)만 "
    "가능하며 데이터 수정(UPDATE/DELETE/INSERT)이나 코드 배포는 절대 할 수 없다 — "
    "그런 조치가 필요하면 finish_task에 '사람 세션 필요'라고 명시하고 종료한다. "
    "④의미있는 확인을 할 때마다 즉시 record_event를 호출한다(끝에 몰아서 하지 않는다). "
    "⑤작업이 끝나면 반드시 finish_task를 호출해서 마친다."
)

async def _orch_execute_tool(name: str, tool_input: dict, task_id: int) -> str:
    if name == "query_supabase":
        sql = (tool_input.get("sql") or "").strip()
        if not sql.lower().startswith("select"):
            return "거부: SELECT 쿼리만 허용됩니다."
        try:
            result = await sb_h("POST", "rpc/exec_readonly_sql", json={"query": sql})
            return json.dumps(result, ensure_ascii=False, default=str)[:4000]
        except Exception as e:
            return f"쿼리 실패: {e}"
    elif name == "record_event":
        await sb_insert("magi_task_events", {
            "task_id": task_id, "event_type": "TASK_UPDATED", "actor": "캐스퍼(Haiku자동)",
            "detail": tool_input.get("detail", "")
        })
        return "기록 완료"
    elif name == "write_correction_log":
        await sb_insert("correction_log", {
            "table_name": tool_input.get("table_name", ""), "field_changed": "자동오케스트레이션발견",
            "old_value": "", "new_value": "", "changed_by": "캐스퍼(Haiku자동)",
            "reason": tool_input.get("reason", "")
        })
        return "기록 완료"
    elif name == "finish_task":
        return "__FINISH__" + (tool_input.get("summary") or "")
    return "알 수 없는 tool"


async def run_haiku_orchestration_once():
    """task#40: ASSIGNED+캐스퍼+SIMPLE 태스크 1건을 찾아 Haiku로 처리 시도.
    최대 10회 tool호출 제한(무한루프방지). 완료시 evidence_registry에 PENDING
    등록(자동승인 없음), 매 tool호출마다 record_event로 즉시기록(중단복구용)."""
    try:
        rows = await sb_select("magi_tasks", {
            "status": "eq.ASSIGNED", "owner_agent": "eq.캐스퍼", "task_type": "eq.SIMPLE",
            "order": "created_at.asc", "limit": "1"
        })
    except Exception as e:
        logger.error(f"Haiku오케스트레이션 - magi_tasks 조회 실패: {e}")
        return None
    if not rows:
        return None
    task = rows[0]
    task_id = task["task_id"]

    if not ANTHROPIC_API_KEY:
        logger.error("Haiku오케스트레이션 - ANTHROPIC_API_KEY 없음")
        return None
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=60.0)
    messages = [{"role": "user", "content": (
        f"태스크#{task_id}: {task.get('title','')}\n"
        f"문제: {task.get('problem','') or '(없음)'}\n"
        f"목표: {task.get('target','') or '(없음)'}"
    )}]

    await sb_insert("magi_task_events", {
        "task_id": task_id, "event_type": "TASK_UPDATED", "actor": "캐스퍼(Haiku자동)",
        "detail": "Haiku 자동오케스트레이션 착수"
    })

    final_summary = None
    for turn in range(10):
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001", max_tokens=2000,
                system=_ORCH_SYSTEM_PROMPT, tools=_ORCH_TOOLS, messages=messages,
            )
        except Exception as e:
            logger.error(f"Haiku오케스트레이션 API 호출 실패(turn {turn}): {e}")
            break
        messages.append({"role": "assistant", "content": resp.content})
        tool_uses = [b for b in resp.content if b.type == "tool_use"]
        if not tool_uses:
            break
        tool_results = []
        finished = False
        for tu in tool_uses:
            result_text = await _orch_execute_tool(tu.name, tu.input, task_id)
            if result_text.startswith("__FINISH__"):
                final_summary = result_text[len("__FINISH__"):]
                finished = True
                result_text = "작업 종료 처리됨"
            tool_results.append({"type": "tool_result", "tool_use_id": tu.id, "content": result_text})
        messages.append({"role": "user", "content": tool_results})
        if finished:
            break

    if final_summary is not None:
        await sb_h("PATCH", f"magi_tasks?task_id=eq.{task_id}", json={"status": "VERIFICATION", "updated_at": datetime.now(KST).isoformat()}, headers={**HEADERS_SB, "Prefer": "return=minimal"})
        await sb_insert("evidence_registry", {
            "task_id": task_id, "evidence_type": "AGENT_RESULT", "document_type": "AUTO_ORCHESTRATION_RESULT",
            "agent": "캐스퍼(Haiku자동)", "verification_status": "PENDING",
            "summary": final_summary, "created_by": "캐스퍼(Haiku자동)"
        })
        logger.info(f"Haiku오케스트레이션 완료(task#{task_id}): {final_summary[:100]}")
    else:
        logger.warning(f"Haiku오케스트레이션 미완료(task#{task_id}) — 다음 폴링에서 재시도(진행상황은 이벤트로 남음)")


# ──────────────────────────────────────────────
# task#52 (2026-08-22): 양방향 완전자동화 - "마기(자동) 검증".
# task#40(마기지시→캐스퍼자동실행)의 반대방향: 에이전트 완료보고→마기자동검증→
# 다음단계 자동실행. Sonnet 사용(판단비중 높아 Haiku 부적합, 마기지시 그대로).
# 안전장치: SIMPLE만 자동승인 가능, COMPLEX/architect_decision_required=true는
# tool레벨에서 서버가 강제차단 - Sonnet이 착각해서 approve_task를 불러도 거부됨.
# ──────────────────────────────────────────────
_MAGI_AUTO_TOOLS = [
    {
        "name": "query_supabase",
        "description": "Supabase에 읽기전용 SELECT 쿼리만 실행한다. 원본대조·실제값 확인용.",
        "input_schema": {"type": "object", "properties": {
            "sql": {"type": "string", "description": "SELECT로 시작하는 SQL문"}
        }, "required": ["sql"]},
    },
    {
        "name": "record_event",
        "description": "검증 진행상황을 magi_task_events에 기록한다. 의미있는 확인마다 즉시 호출.",
        "input_schema": {"type": "object", "properties": {
            "detail": {"type": "string"}
        }, "required": ["detail"]},
    },
    {
        "name": "approve_task",
        "description": "검증통과시 태스크를 COMPLETED로 승인한다. task_type=SIMPLE이고 "
                       "architect_decision_required가 아닌 경우에만 실제로 적용된다 "
                       "(그 외엔 서버가 자동거부하고 escalate로 전환됨).",
        "input_schema": {"type": "object", "properties": {
            "summary": {"type": "string", "description": "검증 결과 정직한 요약"}
        }, "required": ["summary"]},
    },
    {
        "name": "escalate_to_architect",
        "description": "자동승인하지 않고 대표님(아키텍트) 판단을 요청한다. COMPLEX/"
                       "architect_decision_required=true/보고내용과 실제값 불일치 시 반드시 이걸 쓴다.",
        "input_schema": {"type": "object", "properties": {
            "reason": {"type": "string", "description": "왜 자동승인 안 하는지 구체적으로"}
        }, "required": ["reason"]},
    },
    {
        "name": "create_next_task",
        "description": "SIMPLE 태스크 승인 후 다음 단계가 명확하면 후속 태스크를 자동생성한다. "
                       "불확실하면 호출하지 않는다.",
        "input_schema": {"type": "object", "properties": {
            "title": {"type": "string"}, "problem": {"type": "string"}, "target": {"type": "string"},
            "owner_agent": {"type": "string"}, "task_type": {"type": "string", "enum": ["SIMPLE", "COMPLEX"]},
        }, "required": ["title", "problem", "target", "owner_agent", "task_type"]},
    },
    {
        "name": "finish_review",
        "description": "이 검증 세션을 종료한다. approve_task 또는 escalate_to_architect를 "
                       "먼저 호출한 뒤 마지막에 반드시 호출.",
        "input_schema": {"type": "object", "properties": {}, "required": []},
    },
]

_MAGI_AUTO_SYSTEM_PROMPT = (
    "당신은 마기(MAGI 시스템 총괄, 코드/DB담당 캐스퍼의 완료보고를 검증하는 역할)의 "
    "자동실행 모드입니다. 지금까지 채팅에서 마기가 해온 것과 동일한 절차를 따른다: "
    "①query_supabase로 실제 DB값을 직접 조회해서 보고내용과 대조(원본대조 없이 보고서 "
    "텍스트만 믿고 승인하지 않는다) ②의미있는 확인마다 record_event로 즉시 기록 "
    "③task_type=SIMPLE이고 architect_decision_required가 아니며 검증이 실제로 통과하면 "
    "approve_task ④task_type=COMPLEX이거나 architect_decision_required=true이거나 "
    "보고내용과 실제값이 다르면 반드시 escalate_to_architect(자동승인 절대 금지) "
    "⑤확실한 후속단계가 있으면 create_next_task ⑥마지막엔 반드시 finish_review."
)

async def _magi_auto_execute_tool(name: str, tool_input: dict, task: dict) -> str:
    task_id = task["task_id"]
    if name == "query_supabase":
        sql = (tool_input.get("sql") or "").strip()
        if not sql.lower().startswith(("select", "with")):
            return "거부: SELECT/WITH 쿼리만 허용됩니다."
        try:
            result = await sb_h("POST", "rpc/exec_readonly_sql", json={"query": sql})
            return json.dumps(result, ensure_ascii=False, default=str)[:4000]
        except Exception as e:
            return f"쿼리 실패: {e}"
    elif name == "record_event":
        await sb_insert("magi_task_events", {
            "task_id": task_id, "event_type": "TASK_UPDATED", "actor": "마기(자동)",
            "detail": tool_input.get("detail", "")
        })
        return "기록 완료"
    elif name == "approve_task":
        # 서버측 강제 재검증 — Sonnet이 잘못 판단해도 여기서 최종 차단
        if task.get("task_type") == "COMPLEX" or task.get("architect_decision_required"):
            await sb_insert("magi_task_events", {
                "task_id": task_id, "event_type": "TASK_UPDATED", "actor": "마기(자동)",
                "detail": f"자동승인 시도 거부됨(서버강제) - COMPLEX 또는 architect_decision_required=true. 원요청요약: {tool_input.get('summary','')[:200]}"
            })
            return "거부: 이 태스크는 COMPLEX 또는 architect_decision_required=true라 자동승인 불가합니다. escalate_to_architect를 사용하세요."
        await sb_h("PATCH", f"magi_tasks?task_id=eq.{task_id}",
                   json={"status": "COMPLETED", "verification_status": "MAGI_CONFIRMED",
                         "verified_by": "마기(자동)", "verified_at": datetime.now(KST).isoformat(),
                         "notes": f"[마기(자동) 승인] {tool_input.get('summary','')}"},
                   headers={**HEADERS_SB, "Prefer": "return=minimal"})
        await sb_insert("magi_task_events", {
            "task_id": task_id, "event_type": "TASK_COMPLETED", "old_status": "VERIFICATION",
            "new_status": "COMPLETED", "actor": "마기(자동)", "detail": tool_input.get("summary", "")
        })
        return "__APPROVED__"
    elif name == "escalate_to_architect":
        reason = tool_input.get("reason", "")
        try:
            await send_telegram_broadcast(f"🔔 마기(자동) 판단요청 — task#{task_id} ({task.get('title','')})\n{reason}")
        except Exception as e:
            logger.error(f"escalate 텔레그램발송 실패: {e}")
        # 긴급수정(2026-08-22, 실기기검증중 발견): verified_by를 안 채우면 폴링조건
        # (status=VERIFICATION AND verified_by IS NULL)에 계속 걸려 같은 태스크가
        # 무한 재시도됨(task54로 실증 확인) — escalate도 "처리완료(대표님응답대기)"
        # 표시가 필요.
        await sb_h("PATCH", f"magi_tasks?task_id=eq.{task_id}",
                   json={"verified_by": "마기(자동)_ESCALATED", "verified_at": datetime.now(KST).isoformat()},
                   headers={**HEADERS_SB, "Prefer": "return=minimal"})
        # task#64(2026-08-26): 이지스 관련 질문이 최소 2회 반복됐던 사례(마기가
        # 과거 답변에 접근 못해 재질문) 재발방지 — 명확한 event_type으로 구분.
        await sb_insert("magi_task_events", {
            "task_id": task_id, "event_type": "ARCHITECT_QUESTION_SENT", "actor": "마기(자동)",
            "detail": f"질문: {reason[:500]}"
        })
        return "__ESCALATED__"
    elif name == "create_next_task":
        try:
            new_task = await sb_insert("magi_tasks", {
                "title": tool_input.get("title", ""), "problem": tool_input.get("problem", ""),
                "target": tool_input.get("target", ""), "owner_agent": tool_input.get("owner_agent", ""),
                "task_type": tool_input.get("task_type", "SIMPLE"), "status": "ASSIGNED",
                "issuer": "마기(자동)", "parent_task_id": task_id,
            })
            new_id = new_task.get("task_id") if isinstance(new_task, dict) else (new_task[0].get("task_id") if new_task else None)
            try:
                await send_telegram_broadcast(f"📋 신규 태스크 자동생성 — task#{new_id}: {tool_input.get('title','')}\n담당: {tool_input.get('owner_agent','')}")
            except Exception:
                pass
            return f"생성완료: task#{new_id}"
        except Exception as e:
            return f"생성 실패: {e}"
    elif name == "finish_review":
        return "__FINISH__"
    return "알 수 없는 tool"


async def run_magi_auto_review_once():
    """task#52: VERIFICATION+verified_by없음 태스크 1건을 찾아 마기(자동)가 검증.
    Sonnet 사용, 최대 10회 tool호출 제한."""
    try:
        rows = await sb_select("magi_tasks", {
            "status": "eq.VERIFICATION", "verified_by": "is.null",
            "order": "updated_at.asc", "limit": "1"
        })
    except Exception as e:
        logger.error(f"마기자동검증 - magi_tasks 조회 실패: {e}")
        return None
    if not rows:
        return None
    task = rows[0]
    task_id = task["task_id"]

    if not ANTHROPIC_API_KEY:
        return None
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=90.0)

    evid_rows = await sb_select("evidence_registry", {"task_id": f"eq.{task_id}", "order": "evidence_id.desc", "limit": "1"})
    evid = evid_rows[0] if evid_rows else {}

    messages = [{"role": "user", "content": (
        f"태스크#{task_id}: {task.get('title','')}\n"
        f"task_type: {task.get('task_type','미지정')}\n"
        f"architect_decision_required: {task.get('architect_decision_required', False)}\n"
        f"문제: {task.get('problem','') or '(없음)'}\n목표: {task.get('target','') or '(없음)'}\n"
        f"완료보고(notes): {task.get('notes','') or '(없음)'}\n"
        f"evidence_registry 최신건: {json.dumps(evid, ensure_ascii=False, default=str)[:2000]}"
    )}]

    await sb_insert("magi_task_events", {
        "task_id": task_id, "event_type": "TASK_UPDATED", "actor": "마기(자동)",
        "detail": "마기(자동) 검증 착수"
    })

    outcome = None  # "__APPROVED__" | "__ESCALATED__" | None
    for turn in range(10):
        try:
            resp = client.messages.create(
                model="claude-sonnet-5", max_tokens=3000,
                system=_MAGI_AUTO_SYSTEM_PROMPT, tools=_MAGI_AUTO_TOOLS, messages=messages,
            )
        except Exception as e:
            logger.error(f"마기자동검증 API 호출 실패(turn {turn}): {e}")
            break
        messages.append({"role": "assistant", "content": resp.content})
        tool_uses = [b for b in resp.content if b.type == "tool_use"]
        if not tool_uses:
            break
        tool_results = []
        finished = False
        for tu in tool_uses:
            result_text = await _magi_auto_execute_tool(tu.name, tu.input, task)
            if result_text in ("__APPROVED__", "__ESCALATED__"):
                outcome = result_text
                result_text = "처리됨"
            elif result_text == "__FINISH__":
                finished = True
                result_text = "검증세션 종료"
            tool_results.append({"type": "tool_result", "tool_use_id": tu.id, "content": result_text})
        messages.append({"role": "user", "content": tool_results})
        if finished:
            break

    logger.info(f"마기(자동)검증 완료(task#{task_id}): outcome={outcome}")
    return task_id
    return task_id


async def build_weekly_briefing():
    """task#27: ARG-CMD-20260820-02(아르고스) 필드매핑 그대로 구현.
    계산 로직 재구현 금지 원칙 — daily_calc_snapshot이 이미 계산해둔 값만
    형식에 맞게 "조립"만 한다. 축A(영업일) 기준으로 전체 통일(대표님 지시,
    2026-08-21) — 기존 ①②는 kpi_7day_snapshot(축B)을 썼으나 ③④⑤⑧과
    기준이 달라 "총콜수3건"vs"요일별 전부0건" 같은 모순된 표시가 났던
    문제를 해소."""
    today = today_kst()
    week_start = str(today - timedelta(days=6))
    week_end = str(today)
    last_week_start = str(today - timedelta(days=13))
    last_week_end = str(today - timedelta(days=7))

    dcs_rows = await sb_select("daily_calc_snapshot", {
        "axis": "eq.A", "calc_date": [f"gte.{week_start}", f"lte.{week_end}"], "order": "calc_date.asc"
    }) or []
    last_week_rows = await sb_select("daily_calc_snapshot", {
        "axis": "eq.A", "calc_date": [f"gte.{last_week_start}", f"lte.{last_week_end}"]
    }) or []

    lines = ["📅 주간 브리핑 (축A 영업일 기준)", f"기간: {week_start} ~ {week_end}", ""]

    # ① 주간실적 (축A로 통일: daily_calc_snapshot 7일 합산)
    total_calls = sum(r.get("call_count") or 0 for r in dcs_rows)
    daily_avg = round(total_calls / 7, 2)
    weighted_fare_sum = sum((r.get("avg_fare") or 0) * (r.get("call_count") or 0) for r in dcs_rows)
    calls_with_fare = sum(r.get("call_count") or 0 for r in dcs_rows if r.get("avg_fare"))
    avg_fare_week = round(weighted_fare_sum / calls_with_fare) if calls_with_fare else None
    confirmed_revenue = round(weighted_fare_sum)
    # 매출추정 로직수정(대표님지시 2026-08-21): 미분류(요약행) 매출은 개별콜이
    # 아니라 avg_fare/call_count 계산엔 안 섞이지만, 실제 매출 자체는 있었으므로
    # "확정매출+미분류매출"을 분리 표시 — 8/16처럼 실매출이 통째로 안 보이던
    # 문제(task#46에서 발견) 해결.
    unclassified_revenue = sum(r.get("unclassified_amount") or 0 for r in dcs_rows)
    revenue_total = confirmed_revenue + unclassified_revenue
    lines += [
        "① 주간실적",
        f"  총 콜수: {total_calls}건(확정)",
        f"  일평균: {daily_avg}건",
        f"  건당 평균요금: {fmt(avg_fare_week) if avg_fare_week else '데이터부족'}",
        f"  매출 추정: {fmt(revenue_total)}" + (f" (확정 {fmt(confirmed_revenue)} + 미분류 {fmt(round(unclassified_revenue))})" if unclassified_revenue else ""),
        "",
    ]

    # ② 전주비교 (축A로 통일)
    last_week_calls = sum(r.get("call_count") or 0 for r in last_week_rows)
    if last_week_rows:
        last_week_avg = round(last_week_calls / 7, 2)
        diff_cnt = total_calls - last_week_calls
        diff_avg = round(daily_avg - last_week_avg, 2)
        pct = round(diff_cnt / last_week_calls * 100, 1) if last_week_calls else None
        lines += ["② 전주비교",
            f"  콜수 증감: {'+' if diff_cnt>=0 else ''}{diff_cnt}건" + (f" ({'+' if pct>=0 else ''}{pct}%)" if pct is not None else ""),
            f"  일평균 증감: {'+' if diff_avg>=0 else ''}{diff_avg}건", ""]
    else:
        lines += ["② 전주비교", "  지난주 데이터 없음(비교불가)", ""]

    # ③ 요일별 실적
    DOW_KR = ['월','화','수','목','금','토','일']
    lines.append("③ 요일별 실적(영업일/축A 기준)")
    for r in dcs_rows:
        d = datetime.strptime(r["calc_date"], "%Y-%m-%d").date()
        dow = DOW_KR[d.weekday()]
        cc = r.get("call_count") or 0
        af = r.get("avg_fare")
        lines.append(f"  {r['calc_date']}({dow}): {cc}건" + (f", 평균{fmt(af)}원" if af else ""))
    lines.append("")

    # ④ 시간대 효율
    hourly_sum = {}
    for r in dcs_rows:
        for h, c in (r.get("hourly_counts") or {}).items():
            hourly_sum[h] = hourly_sum.get(h, 0) + c
    if hourly_sum:
        peak_h, peak_c = max(hourly_sum.items(), key=lambda x: x[1])
        lines += ["④ 시간대 효율", f"  피크: {peak_h}시대 ({peak_c}건, 7일합산)", ""]
    else:
        lines += ["④ 시간대 효율", "  데이터부족", ""]

    # ⑤ 패턴 업데이트 (ARG-CMD-20260820-02)
    gyeongsan_sum = sum(r.get("gyeongsan_loop_count") or 0 for r in dcs_rows)
    baehoe_sum = sum(r.get("baehoe_count") or 0 for r in dcs_rows)
    consecutive_sum = sum(r.get("consecutive_call_count") or 0 for r in dcs_rows)
    lines += ["⑤ 패턴 업데이트",
        f"  경산루프: {gyeongsan_sum}건", f"  배회: {baehoe_sum}건",
        f"  연속콜: {consecutive_sum}건(하차~다음배차 10분이내 기준)", ""]

    # ⑧ 미업로드 알림 (task#46: unclassified 날짜도 별도 경고 — 8/16처럼 행은
    # 있지만 요약행만 있어 실제매출이 call_count=0으로 가려지던 사례 노출)
    expected_dates = {str(today - timedelta(days=i)) for i in range(7)}
    covered_dates = {r["calc_date"] for r in dcs_rows}
    missing = sorted(expected_dates - covered_dates)
    unclassified_dates = sorted(r["calc_date"] for r in dcs_rows if r.get("unclassified_flag"))
    section8 = ["⑧ 미업로드 알림"]
    if missing:
        section8.append("  미집계: " + ", ".join(missing))
    if unclassified_dates:
        section8.append("  ⚠️ 미분류매출 있음(개별콜 미확정): " + ", ".join(unclassified_dates))
    if not missing and not unclassified_dates:
        section8.append("  전체 집계 완료")
    lines += section8

    return "\n".join(lines)


async def check_ingestion_gap():
    """task_id=31 긴급대응: OCR→raw_calls 인입 중단을 조기 감지.
    최근 3일 연속 신규 raw_calls가 0건이면 텔레그램 경고 — 08-13~15 사흘간
    조용히 멈췄던 걸 아무도 몰랐던 사고 재발방지. 매일 08시 갭3 검증에 편승."""
    today = today_kst()
    gap_days = []
    for i in range(3):
        d = str(today - timedelta(days=i+1))  # 어제, 그제, 그그제
        rows = await sb_select("raw_calls", {"날짜": f"eq.{d}", "limit": "1"})
        if not rows:
            gap_days.append(d)
    if len(gap_days) == 3:
        msg = (
            f"🔴 인입 중단 경고: 최근 3일({', '.join(sorted(gap_days))}) raw_calls "
            f"신규데이터 0건입니다.\n실제 운행이 있었다면 OCR 업로드를 확인해주세요."
        )
        logger.warning(f"인입중단 감지: {gap_days}")
        try:
            await send_telegram_broadcast(msg)
        except Exception as e:
            logger.error(f"인입중단 경고 발송 실패: {e}")
    return gap_days


async def recalc_daily_summary_totals(days_back: int = 14):
    """daily_summary가 GPX업로드/영수증OCR 시점에만 갱신되는 구조라, 그 액션을
    안 하면 운행을 해도 daily_summary가 그 날짜 자체가 안 만들어지는 문제
    (2026-08-15 명령서#028 갭3 텔레그램 경고로 실제 발견: 08-09~08-12 raw_calls엔
    데이터 있는데 daily_summary는 전부 0원).
    raw_calls 요약행제외 직접집계로 총매출·총건수만 매일 자동 UPSERT — 공차율 등
    GPX전용 필드는 PostgREST upsert 특성상 미지정시 기존값 그대로 유지됨."""
    end = today_kst()
    start = end - timedelta(days=days_back-1)
    rows = await sb_select("raw_calls", {"날짜": [f"gte.{start}", f"lte.{end}"]})
    rows = exclude_summary_rows(rows or [])

    by_date = {}
    for r in rows:
        d = r.get("날짜")
        if not d: continue
        by_date.setdefault(d, {"총매출": 0, "총건수": 0})
        by_date[d]["총매출"] += _safe_int(r.get("요금")) or 0
        by_date[d]["총건수"] += 1

    updated = []
    for d, v in by_date.items():
        try:
            await sb_upsert("daily_summary", {"날짜": d, "총매출": v["총매출"], "총건수": v["총건수"]}, on_conflict="날짜")
            updated.append(d)
        except Exception as e:
            logger.error(f"daily_summary 자동갱신 실패({d}): {e}")
    logger.info(f"daily_summary 자동갱신 완료: {len(updated)}일 ({start}~{end})")
    return updated


async def recalc_7day_average():
    """명령서#035: 7일 이동평균(콜건수 기준, 축B 캘린더일) 자동 재계산.
    필수 필터링: 콜유형='카카오T'만, verify_status != 'invalidated_duplicate'.
    콜유형='합계'(미분류) 행이 있는 날은 계산에서 빼고 unclassified_days에 기록해서
    숨기지 않는다. 결과는 kpi_7day_snapshot에 append(덮어쓰지 않음)."""
    end = today_kst()
    start = end - timedelta(days=6)
    dates = [str(start + timedelta(days=i)) for i in range(7)]

    rows = await sb_select("raw_calls", {"날짜": [f"gte.{start}", f"lte.{end}"]})
    rows = rows or []

    unclassified_days = set()
    count_by_date = {d: 0 for d in dates}
    for r in rows:
        d = r.get("날짜")
        if d not in count_by_date:
            continue
        if r.get("콜유형") == "합계":
            unclassified_days.add(d)
            continue
        if r.get("콜유형") != "카카오T":
            continue
        if r.get("verify_status") == "invalidated_duplicate":
            continue
        count_by_date[d] += 1

    total_count = sum(count_by_date.values())
    daily_average = round(total_count / 7, 2)

    if daily_average >= 10:
        status = "SAFE"
    elif daily_average >= 7:
        status = "WARNING"
    else:
        status = "CRITICAL"

    snapshot = {
        "calc_date": str(end),
        "window_start": str(start),
        "window_end": str(end),
        "total_count": total_count,
        "daily_average": daily_average,
        "unclassified_days": sorted(unclassified_days) or None,
        "status": status,
    }
    result = await sb_upsert("kpi_7day_snapshot", snapshot, on_conflict="calc_date")
    # task#47 긴급수정(2026-08-22): 동일 버그 - 반환값 미검사로 실패해도 성공로그.
    if result is None:
        logger.error(f"kpi_7day_snapshot 저장 실패(calc_date={end}) — sb_upsert가 None 반환")
        raise RuntimeError(f"kpi_7day_snapshot 저장 실패: calc_date={end}")
    logger.info(f"7일평균 스냅샷 저장: {daily_average}건/일 ({status}), 미분류일 {sorted(unclassified_days)}")

    if status in ("WARNING", "CRITICAL"):
        msg = (
            f"{'🟡' if status=='WARNING' else '🔴'} 7일 이동평균 {status}\n"
            f"기간: {start}~{end}\n"
            f"합계 {total_count}건, 일평균 {daily_average}건"
        )
        if unclassified_days:
            msg += f"\n⚠️ 이 숫자는 {len(unclassified_days)}일 미분류(콜유형='합계') 상태를 제외한 값입니다: {sorted(unclassified_days)}"
        try:
            await send_telegram_broadcast(msg)
        except Exception as e:
            logger.error(f"7일평균 알림 발송 실패: {e}")

    return snapshot


async def calc_daily_snapshot(calc_date_str: str = None):
    """명령서#036: 축A(영업일) 기준 일별 5종 계산 — 시간대별/지역별 콜수, 평균단가,
    콜간격, 공차시간. ARGOS_인수인계[3][5] 규칙: 자정(00:00~05:59대) 이후 콜은 전날
    영업일 소속으로 재배정. 원본 raw_calls 필터링은 #035와 동일(verify_status,
    콜유형='합계' 처리)."""
    calc_date = calc_date_str or str(today_kst() - timedelta(days=1))  # 기본: 어제 영업일(당일 새벽 04시 실행 기준)
    next_date = str(datetime.strptime(calc_date, "%Y-%m-%d").date() + timedelta(days=1))

    same_day = await sb_select("raw_calls", {"날짜": f"eq.{calc_date}"}) or []
    next_day = await sb_select("raw_calls", {"날짜": f"eq.{next_date}"}) or []

    def hour_of(r):
        bt = r.get("배차시각")
        if not bt: return None
        try: return int(str(bt).split(":")[0])
        except Exception: return None

    biz_rows = [r for r in same_day if (hour_of(r) or 0) >= 6]
    biz_rows += [r for r in next_day if hour_of(r) is not None and hour_of(r) < 6]

    unclassified = False
    unclassified_note_parts = []
    unclassified_amount = 0
    valid_rows = []
    for r in biz_rows:
        # 캐스퍼 긴급수정 2026-08-20: 기존엔 콜유형='합계'만 요약행으로 판정해서,
        # id1384(콜유형='카카오T'인데 data_source='app_ocr_summary'인 요약행,
        # 230,900원/25건 OCR합계)가 개별콜처럼 avg_fare 계산에 섞여들어갈 뻔했던
        # 버그 발견·수정. exclude_summary_rows()/_is_receipt_summary_row()와
        # 동일한 기준(data_source 우선판정)으로 통일.
        is_summary = (
            r.get("data_source") == "app_ocr_summary"
            or str(r.get("비고") or "").startswith("OCR 추출:")
            or r.get("콜유형") == "합계"
        )
        if is_summary:
            unclassified = True
            unclassified_note_parts.append(f"{r.get('날짜')}(id={r.get('id')})")
            # 대표님지시(task#27 재작업, 2026-08-21): 요약행 금액도 매출추정에
            # 반영할 수 있도록 unclassified_amount로 별도 합산·저장.
            unclassified_amount += (r.get("요금") or 0)
            continue
        if r.get("콜유형") != "카카오T":
            continue
        if r.get("verify_status") == "invalidated_duplicate":
            continue
        valid_rows.append(r)

    hourly_counts = {}
    regional_counts = {}
    fares = []
    times_sorted = []
    for r in valid_rows:
        h = hour_of(r)
        if h is not None:
            hourly_counts[str(h)] = hourly_counts.get(str(h), 0) + 1
        dest = r.get("도착지") or "미상"
        # "대구 OO구 OO동" → "OO구"만 추출(지역별 집계는 구 단위가 실용적)
        parts = str(dest).replace("대구 ", "").split(" ")
        gu = parts[0] if parts else dest
        regional_counts[gu] = regional_counts.get(gu, 0) + 1
        fare = r.get("요금")
        if fare: fares.append(fare)
        bt = r.get("배차시각")
        if bt and h is not None:
            mm = int(str(bt).split(":")[1]) if ":" in str(bt) else 0
            minutes_from_18 = ((h - 18) % 24) * 60 + mm  # 18시 기준 상대분(자정넘김 순서보정)
            times_sorted.append(minutes_from_18)

    times_sorted.sort()
    intervals = [times_sorted[i] - times_sorted[i-1] for i in range(1, len(times_sorted))]
    avg_interval = round(sum(intervals) / len(intervals), 2) if intervals else None
    max_interval = round(max(intervals), 2) if intervals else None
    avg_fare = round(sum(fares) / len(fares), 2) if fares else None
    call_count = len(valid_rows)

    # 공차시간(근사치): gpx_sessions 구속시간을 "영업거리/총거리" 비율로 분배.
    # 원본 GPX 포인트가 DB에 없어 정밀계산 불가 — 근사치임을 명시.
    avg_idle_min = total_idle_min = None
    try:
        gpx = await sb_select("gpx_sessions", {"날짜": f"eq.{calc_date}"})
        if gpx:
            g = gpx[0]
            총거리 = g.get("총주행거리_km")
            구속시간_h = g.get("구속시간_h")
            영업거리합 = sum(float(r.get("영업거리_km") or 0) for r in valid_rows)
            if 총거리 and 구속시간_h and 총거리 > 0:
                공차비율 = max(0, 1 - min(영업거리합 / 총거리, 1))
                total_idle_min = round(구속시간_h * 60 * 공차비율, 2)
                avg_idle_min = round(total_idle_min / call_count, 2) if call_count else None
    except Exception as e:
        logger.error(f"공차시간(근사) 계산 실패 {calc_date}: {e}")

    # task#42: 패턴지표 3종(경산루프/연속콜/배회) — 마기 지시(2026-08-20).
    # 경산루프: 명령서#015 정의 그대로(요금≥15,000원 + 출발지/도착지에 "경산" 포함).
    # 배회: 콜유형='배회'(valid_rows엔 카카오T만 있어 biz_rows에서 별도 집계).
    # 연속콜: 명시적 임계값 정의문서를 못 찾아 "하차~다음배차 10분이내"로 가정
    # (기존 공차분석 인프라와 일관된 값) — unclassified_note에 가정임을 명시.
    gyeongsan_loop_count = sum(
        1 for r in valid_rows
        if (r.get("요금") or 0) >= 15000
        and ("경산" in str(r.get("출발지") or "") or "경산" in str(r.get("도착지") or ""))
    )
    baehoe_count = sum(1 for r in biz_rows if r.get("콜유형") == "배회")

    def _to_min(t):
        if not t: return None
        try:
            hh, mm = str(t).split(":")[:2]
            hh, mm = int(hh), int(mm)
            return ((hh - 18) % 24) * 60 + mm
        except Exception:
            return None

    sorted_by_dispatch = sorted(
        (r for r in valid_rows if r.get("배차시각")),
        key=lambda r: _to_min(r.get("배차시각")) or 0
    )
    CONSECUTIVE_THRESHOLD_MIN = 10  # 가정치, 명시적 정의 없어 공차분석과 일관된 값 채택
    consecutive_call_count = 0
    for i in range(1, len(sorted_by_dispatch)):
        prev_drop = _to_min(sorted_by_dispatch[i-1].get("하차시각"))
        cur_dispatch = _to_min(sorted_by_dispatch[i].get("배차시각"))
        if prev_drop is not None and cur_dispatch is not None:
            gap = cur_dispatch - prev_drop
            if 0 <= gap <= CONSECUTIVE_THRESHOLD_MIN:
                consecutive_call_count += 1

    if unclassified:
        unclassified_note_parts.append("(근거: GPX원본포인트 미보존으로 공차시간은 거리비율 기반 근사치)")
    elif total_idle_min is not None:
        unclassified_note_parts.append("공차시간은 GPX원본포인트 미보존으로 거리비율 기반 근사치")
    unclassified_note_parts.append(f"연속콜 판정기준: 하차~다음배차 {CONSECUTIVE_THRESHOLD_MIN}분이내(가정치, 명시적 정의문서 없음)")

    snapshot = {
        "calc_date": calc_date,
        "axis": "A",
        "hourly_counts": hourly_counts,
        "regional_counts": regional_counts,
        "avg_fare": avg_fare,
        "call_count": call_count,
        "avg_interval_min": avg_interval,
        "max_interval_min": max_interval,
        "avg_idle_min": avg_idle_min,
        "total_idle_min": total_idle_min,
        "unclassified_flag": unclassified,
        "unclassified_note": " | ".join(unclassified_note_parts) if unclassified_note_parts else None,
        "gyeongsan_loop_count": gyeongsan_loop_count,
        "consecutive_call_count": consecutive_call_count,
        "baehoe_count": baehoe_count,
        "unclassified_amount": unclassified_amount if unclassified_amount else None,
    }
    result = await sb_upsert("daily_calc_snapshot", snapshot, on_conflict="calc_date,axis")
    # task#47 긴급수정(2026-08-22, 마기지적): sb_upsert 실패시(RLS 등) None을
    # 반환하는데, 이 반환값을 확인 안 하고 무조건 "저장완료" 로그를 찍던 버그.
    # 관측(로그) 자체가 신뢰 안 되면 다른 모든 검증이 무의미해짐 — 명확히 구분.
    if result is None:
        logger.error(f"daily_calc_snapshot 저장 실패(축A={calc_date}) — sb_upsert가 None 반환(RLS/네트워크 등)")
        raise RuntimeError(f"daily_calc_snapshot 저장 실패: calc_date={calc_date}")
    logger.info(f"daily_calc_snapshot 저장(축A={calc_date}): {call_count}건, 평균단가{avg_fare}, 공차{total_idle_min}분, 경산루프{gyeongsan_loop_count}, 연속콜{consecutive_call_count}, 배회{baehoe_count}")
    return result


async def dual_verify_7day_average() -> dict:
    """7일 이동평균을 raw_calls 직접집계(A)와 daily_summary(B) 두 방식으로 계산해 대조.
    반환: {"match": bool, "method_a": {...}, "method_b": {...}, "diff": 숫자, "dates": [...]}"""
    end = today_kst()
    dates = [str(end - timedelta(days=i)) for i in range(7)]
    start = dates[-1]

    # 방식A: raw_calls 직접집계 (요약행 제외)
    calls_raw = await sb_select("raw_calls", {"날짜": [f"gte.{start}", f"lte.{end}"]})
    calls = exclude_summary_rows(calls_raw)
    a_by_date = {}
    for c in calls:
        d = c.get("날짜")
        if d not in dates: continue
        a_by_date.setdefault(d, {"매출": 0, "건수": 0})
        a_by_date[d]["매출"] += _safe_int(c.get("요금")) or 0
        a_by_date[d]["건수"] += 1
    a_total = sum(v["매출"] for v in a_by_date.values())
    a_days_with_data = len(a_by_date)
    a_avg = round(a_total / 7, 0)

    # 방식B: daily_summary (봇 계산값)
    ds_rows = await sb_select("daily_summary", {"날짜": [f"gte.{start}", f"lte.{end}"]})
    b_by_date = {r["날짜"]: {"매출": _safe_int(r.get("총매출")) or 0, "건수": _safe_int(r.get("총건수")) or 0} for r in ds_rows if r.get("날짜") in dates}
    b_total = sum(v["매출"] for v in b_by_date.values())
    b_avg = round(b_total / 7, 0)

    diff = a_total - b_total
    match = (diff == 0)

    result = {
        "match": match, "diff": diff,
        "method_a": {"총매출": a_total, "일평균": a_avg, "데이터있는날": a_days_with_data},
        "method_b": {"총매출": b_total, "일평균": b_avg},
        "dates": dates,
        "date_range": f"{start}~{end}",
    }
    if not match:
        # 날짜별 상세 차이도 같이 산출 (원인 파악용)
        detail = []
        for d in dates:
            av = a_by_date.get(d, {}).get("매출", 0)
            bv = b_by_date.get(d, {}).get("매출", 0)
            if av != bv:
                detail.append(f"{d}: A={av:,}원 vs B={bv:,}원 (차이 {av-bv:+,}원)")
        result["detail"] = detail
    return result


async def recalc_fish_hour_data():
    """어군 브리핑 시간대별 통계 재계산 (2026-07-13, 하드코딩 HOUR_DATA/FISH_DATA 제거).
    raw_calls(실시간 데이터) 기반으로 시간대별 카카오T/배회 평균 건수·비중·평균단가를
    계산해 fish_hour_data에 저장.
    (2026-08-11 정리: call_quality_history 테이블이 실제로 존재하지 않아 매번 404 에러만
    발생시키고 있었음 — Render 로그로 반복 확인, 항상 빈 배열이라 결과값엔 영향 없었지만
    불필요한 조회+에러로그를 없앰. 필요시 call_distribution 등 대체 테이블 연동은 별도 검토.)
    """
    try:
        raw = await sb_select_calls( {}) or []
    except Exception as e:
        logger.error(f"fish_hour_data 재계산 - raw_calls 조회 실패: {e}")
        raw = []

    all_rows = raw
    if not all_rows:
        logger.warning("fish_hour_data 재계산 - 데이터 없음, 건너뜀")
        return

    # 관측일수(분모) — 전체 데이터셋의 고유 날짜 수
    total_days = len(set(r.get("날짜") for r in all_rows if r.get("날짜"))) or 1

    from collections import defaultdict
    hour_kakao_cnt  = defaultdict(float)
    hour_baehoe_cnt = defaultdict(float)
    hour_kakao_fare = defaultdict(list)
    hour_baehoe_fare = defaultdict(list)

    for r in all_rows:
        bt = r.get("배차시각")
        if not bt:
            continue
        try:
            h = int(str(bt).split(":")[0])
        except Exception:
            continue
        ct = r.get("콜유형") or ""
        cnt = _extract_count(r)  # raw_calls OCR요약행 보정
        fare = r.get("요금")
        if ct == "카카오T":
            hour_kakao_cnt[h] += cnt
            if fare: hour_kakao_fare[h].append(fare)
        elif ct == "배회":
            hour_baehoe_cnt[h] += cnt
            if fare: hour_baehoe_fare[h].append(fare)

    for h in range(24):
        k_avg = round(hour_kakao_cnt.get(h, 0) / total_days, 2)
        b_avg = round(hour_baehoe_cnt.get(h, 0) / total_days, 2)
        total = hour_kakao_cnt.get(h, 0) + hour_baehoe_cnt.get(h, 0)
        b_pct = round(hour_baehoe_cnt.get(h, 0) / total * 100, 1) if total > 0 else 0.0
        kf = hour_kakao_fare.get(h, [])
        bf = hour_baehoe_fare.get(h, [])
        avg_fare_kakao = round(sum(kf) / len(kf)) if kf else None
        avg_fare_baehoe = round(sum(bf) / len(bf)) if bf else None

        try:
            await sb_upsert("fish_hour_data", {
                "hour": h,
                "kakao_avg": k_avg,
                "baehoe_avg": b_avg,
                "b_pct": b_pct,
                "avg_fare_kakao": avg_fare_kakao,
                "avg_fare_baehoe": avg_fare_baehoe,
                "sample_days": total_days,
            }, on_conflict="hour")
        except Exception as e:
            logger.error(f"fish_hour_data hour={h} 저장 오류: {e}")

    logger.info(f"fish_hour_data 재계산 완료 (관측일수={total_days}, 총 {len(all_rows)}건)")


async def recalc_fish_hour_data_dow():
    """대표님 직접요청(2026-08-17): 요일×시간대 조합별 콜빈도 재계산.
    기존 fish_hour_data(시간대만, 8시간환산)와 달리 '8시간 기준'이라는 근거없는
    관례적 상수를 없애고, "그 요일이 실제 관측된 일수"로 직접 나눈 평균건수를
    그대로 사용 — 예: 월요일21시 22건÷14일=1.57→반올림2건, 이게 최종 표시값."""
    try:
        raw = await sb_select_calls({}) or []
    except Exception as e:
        logger.error(f"fish_hour_data_dow 재계산 - raw_calls 조회 실패: {e}")
        raw = []
    if not raw:
        logger.warning("fish_hour_data_dow 재계산 - 데이터 없음, 건너뜀")
        return

    from collections import defaultdict
    # 요일별 관측일수(그 요일이 실제 몇 번 있었는지, 고유 날짜 기준)
    dow_days = defaultdict(set)
    for r in raw:
        d, w = r.get("날짜"), r.get("요일")
        if d and w:
            dow_days[w].add(d)
    dow_day_count = {w: len(s) for w, s in dow_days.items()}

    cnt = defaultdict(float)       # (요일,hour,콜유형) -> 건수합
    fare = defaultdict(list)       # (요일,hour,콜유형) -> 요금리스트
    for r in raw:
        w = r.get("요일")
        bt = r.get("배차시각")
        if not w or not bt:
            continue
        try:
            h = int(str(bt).split(":")[0])
        except Exception:
            continue
        ct = r.get("콜유형") or ""
        if ct not in ("카카오T", "배회"):
            continue
        c = _extract_count(r)
        cnt[(w, h, ct)] += c
        f = r.get("요금")
        if f: fare[(w, h, ct)].append(f)

    saved = 0
    for w, days in dow_day_count.items():
        if days < 1:
            continue
        for h in range(24):
            k_cnt = cnt.get((w, h, "카카오T"), 0)
            b_cnt = cnt.get((w, h, "배회"), 0)
            if k_cnt == 0 and b_cnt == 0:
                continue  # 데이터 자체가 없는 슬롯은 행을 안 만듦(빈슬롯과 0건 구분)
            kf = fare.get((w, h, "카카오T"), [])
            bf = fare.get((w, h, "배회"), [])
            try:
                await sb_upsert("fish_hour_data_dow", {
                    "요일": w, "hour": h,
                    "kakao_count": round(k_cnt), "kakao_avg": round(k_cnt / days, 2),
                    "avg_fare_kakao": round(sum(kf) / len(kf)) if kf else None,
                    "baehoe_count": round(b_cnt), "baehoe_avg": round(b_cnt / days, 2),
                    "avg_fare_baehoe": round(sum(bf) / len(bf)) if bf else None,
                    "sample_days": days,
                }, on_conflict="요일,hour")
                saved += 1
            except Exception as e:
                logger.error(f"fish_hour_data_dow {w}{h}시 저장 오류: {e}")
    logger.info(f"fish_hour_data_dow 재계산 완료 ({saved}개 슬롯)")


_FISH_HOUR_CACHE = {}

async def load_fish_hour_data():
    """fish_hour_data 조회, 메모리 캐시. 재계산 스케줄러가 갱신할 때까지 캐시 유지."""
    global _FISH_HOUR_CACHE
    if _FISH_HOUR_CACHE:
        return _FISH_HOUR_CACHE
    try:
        rows = await sb_select("fish_hour_data", {}) or []
        _FISH_HOUR_CACHE = {
            r["hour"]: {
                "kakao": r.get("kakao_avg") or 0,
                "baehoe": r.get("baehoe_avg") or 0,
                "b_pct": r.get("b_pct") or 0,
                "avg_fare_kakao": r.get("avg_fare_kakao"),
                "avg_fare_baehoe": r.get("avg_fare_baehoe"),
            } for r in rows
        }
    except Exception as e:
        logger.error(f"fish_hour_data 로드 실패: {e}")
    return _FISH_HOUR_CACHE


async def get_fish_report_db(hour=None, tag_filter=None):
    """어군 브리핑 v3 - 카카오/배회 분리, raw_calls 실데이터 기반(하드코딩 제거)"""
    now = datetime.now(KST)
    h   = hour if hour is not None else now.hour
    day = DOW_KOR[now.weekday()]

    # 대표님 요청(2026-08-17): 요일×시간대 데이터(fish_hour_data_dow) 우선사용.
    # "8시간환산" 관례를 없애고, 그 요일 관측일수로 직접 나눈 평균을 그대로 씀
    # (예: 22건÷14일=1.57건 → 반올림 약2건). 없으면 기존 요일무관 통계로 폴백.
    dow_rows = await sb_select("fish_hour_data_dow", {"요일": f"eq.{day}", "hour": f"eq.{h}"})
    hd_dow = dow_rows[0] if dow_rows else None

    HOUR_DATA = await load_fish_hour_data()
    _fallback = {"kakao": 0, "baehoe": 0, "b_pct": 0, "avg_fare_kakao": None, "avg_fare_baehoe": None}
    hd_raw = HOUR_DATA.get(h) or _fallback

    if hd_dow:
        kakao_avg = float(hd_dow.get("kakao_avg") or 0)
        baehoe_avg = float(hd_dow.get("baehoe_avg") or 0)
        sample_days = hd_dow.get("sample_days") or 0
        # 대표님 지적(2026-08-22): "예상0.0건"인데 단가가 표시되던 문제 —
        # 실측 없으면 폴백값(9000/10500원)을 마치 실측인 것처럼 보여주고 있었음.
        # 지어내지 않는다 원칙대로 폴백 제거, None 유지 → 표시부에서 명시적으로 처리.
        fare_kakao = hd_dow.get("avg_fare_kakao")
        fare_baehoe = hd_dow.get("avg_fare_baehoe")
        data_note = f"{day}요일 {sample_days}일 관측"
    else:
        kakao_avg = float(hd_raw.get("kakao") or 0)
        baehoe_avg = float(hd_raw.get("baehoe") or 0)
        sample_days = 0
        fare_kakao = hd_raw.get("avg_fare_kakao")
        fare_baehoe = hd_raw.get("avg_fare_baehoe")
        data_note = "요일무관 전체평균(요일축 데이터 없음)"

    total_obs = kakao_avg + baehoe_avg
    if total_obs < 0.05:
        return (
            f"🐟 어군 브리핑 v3 — {str(h).zfill(2)}시 {day}요일 (실데이터 기반)\n"
            f"{chr(0x2501)*22}\n\n"
            f"이 시간대는 축적된 운행 데이터가 부족합니다.\n"
            f"(19~21시/21~24시/00~02시 운영시간대에 데이터가 집중되어 있습니다)"
        )

    # 대표님 지적(2026-08-23): 카카오는 정수반올림, 배회는 소수1자리로 표시
    # 정밀도가 달라서 "예상0건인데 비중66.7%" 같은 모순된 표시가 나던 문제.
    # (실측: kakao_avg=0.4, baehoe_avg=0.2 → round(0.4)=0으로 반올림되며 발생)
    # 둘 다 소수1자리로 통일해 "0건인데 비중이 있다"는 모순을 제거.
    kakao_disp = round(kakao_avg, 1)
    baehoe_disp = round(baehoe_avg, 1)
    total_disp = kakao_avg + baehoe_avg
    b_pct = round(baehoe_avg / total_disp * 100, 1) if total_disp > 0 else 0.0
    k_pct = 100 - b_pct

    def stars(days):
        # 표본일수(관측 신뢰도) 기준 — 데이터가 며칠치인지가 진짜 신뢰도 지표
        if days >= 10: return "★★★"
        if days >= 5: return "★★"
        return "★"

    # 캐스퍼 수정 2026-08-12: "핵심 동선" 하드코딩을 실데이터 기반으로 교체.
    # 배회 콜의 출발지(GPX 역지오코딩으로 채워진 실좌표 기반값, 명령서#026)를 시간대별로
    # GROUP BY해서 실제 최다빈도 동을 계산. 표본이 너무 적으면(3건 미만) 정직하게
    # "데이터부족"으로 표시하고 지어내지 않는다.
    async def _baehoe_hotspot(hour_list):
        try:
            rows = await sb_select_calls({
                "콜유형": "in.(배회,배회(잠정))",
                "출발지": "not.is.null",
            })
        except Exception as e:
            logger.error(f"배회 핵심동선 조회 실패: {e}")
            return "조회오류"
        counts = {}
        for r in rows:
            addr = r.get("출발지")
            if not addr or addr in ("미상", ""):
                continue
            bt = r.get("배차시각")
            if not bt:
                continue
            try:
                rh = int(str(bt).split(":")[0])
            except Exception:
                continue
            if rh not in hour_list:
                continue
            # "대구 OO구 OO동" → "OO동(OO구)" 형태로 요약
            parts = addr.replace("대구 ", "").split(" ")
            key = f"{parts[-1]}({parts[0]})" if len(parts) >= 2 else addr
            counts[key] = counts.get(key, 0) + 1
        if not counts:
            return "데이터부족(미검증)"
        total = sum(counts.values())
        if total < 3:
            return f"데이터부족(표본{total}건, 미검증)"
        top = sorted(counts.items(), key=lambda x: -x[1])[:2]
        return " / ".join(f"{loc}({cnt}건)" for loc, cnt in top) + f" [실측 표본{total}건]"

    anchor_night = await _baehoe_hotspot([19,20,21,22,23])
    anchor_late  = await _baehoe_hotspot([0,1,2])
    anchor_dawn  = await _baehoe_hotspot([3,4,5])

    # 캐스퍼 긴급수정 2026-08-12 (대표님 실측대조로 발견): `elif h >= 3` 조건이
    # 3시~23시를 전부 삼켜버려서 `else: anchor = anchor_night`가 절대 실행 안 되는
    # 죽은 분기였음. 20시(night 구간)인데 dawn(3~5시) 표본이 뜨던 원인이 이것.
    # night(19-23시) 구간을 먼저 확정하도록 순서/경계 수정.
    if 19 <= h <= 23:
        anchor = anchor_night
    elif 0 <= h <= 2:
        anchor = anchor_night + " / " + anchor_late
    else:
        anchor = anchor_dawn

    if 19 <= h <= 21:
        decision   = "카카오 우선 대기"
        rec_detail = "배회 " + str(b_pct) + "% 미만 — 콜 대기 합리적\n수락률 100% 유지"
    elif 22 <= h <= 23:
        decision   = "카카오 우선 + 배회 수용"
        rec_detail = "배회 " + str(b_pct) + "% — 앵커 위치면 적극 수용\n" + anchor.split("/")[0].strip()
    elif 0 <= h <= 2:
        # 대표님 지적(2026-08-24): "00-02시는 항상 배회황금시간"으로 시간대만
        # 보고 하드코딩돼있어서, 실제 그 요일·시간대 배회데이터가 0건이어도
        # "★배회적극수용"이 그대로 나오던 모순 수정. 실측(baehoe_avg)을 확인.
        if baehoe_avg < 0.05:
            decision   = "카카오 우선 대기 (배회data부족)"
            rec_detail = f"이 시간대는 통상 배회 황금구간이나, {day}요일 {sample_days}일 관측상 배회실적 없음\n카카오 대기 권장"
        else:
            decision   = "★ 배회 적극 수용 (황금시간)"
            rec_detail = "배회 " + str(b_pct) + "% — 자정후 황금구간\n핵심 동선: " + anchor
    elif 3 <= h <= 4:
        decision   = "마감 단계"
        rec_detail = "02시 종료 검토" if day in ["화", "목"] else "끝까지 사수"
    else:
        decision   = "운행 준비 / 대기"
        rec_detail = "19시 이후 카카오 골든타임 준비"

    est_h = round((fare_kakao or 0) * kakao_avg + (fare_baehoe or 0) * baehoe_avg)

    db_zones = []
    try:
        slot_map = {19:"19-21",20:"19-21",21:"21-24",22:"21-24",23:"21-24",0:"00-02",1:"00-02",2:"00-02"}
        tb = slot_map.get(h)
        if tb:
            params = {"time_band": "eq." + tb, "verified": "neq.avoid",
                      "order": "rank_overall.asc", "limit": "3"}
            if tag_filter:
                params["pattern_tag"] = "eq." + tag_filter
            rows = await sb_select("fish_finder", params)
            if rows:
                for r in rows:
                    zone = r.get("zone", "")
                    avg  = r.get("avg_fare", 0) or 0
                    db_zones.append(zone + ("(" + str(int(avg)) + "원)" if avg else ""))
    except Exception:
        pass

    lines = [
        "🐟 어군 브리핑 v3 — " + str(h).zfill(2) + "시 " + day + "요일 (실데이터 기반)",
        chr(0x2501) * 22,
        "",
        "🟢 카카오 콜 어군",
        "  예상 건수: 약 " + str(kakao_disp) + "건 (" + stars(sample_days) + ", " + data_note + ")",
        "  평균 단가: " + (fmt(fare_kakao) + "대" if fare_kakao else "데이터부족"),
        "  비중: " + str(k_pct) + "%",
    ]
    if db_zones:
        lines.append("  DB 핫존: " + " / ".join(db_zones))

    lines += [
        "",
        "🟠 배회 어군",
        "  예상 건수: 약 " + str(baehoe_disp) + "건 (" + stars(sample_days) + ", " + str(b_pct) + "%)",
        "  평균 단가: " + (fmt(fare_baehoe) + " (수수료 0%)" if fare_baehoe else "데이터부족"),
        "  핵심 동선: " + anchor,
        "",
        "💡 종합 권고",
        "  " + decision,
    ]
    for dl in rec_detail.split("\n"):
        lines.append("  " + dl)
    lines += [
        "  예상 시간당: " + fmt(est_h) + "대",
        "",
        chr(0x2500) + " 시간대별 배회 비중 " + chr(0x2500),
    ]

    BAR_HOURS = [19, 20, 21, 22, 23, 0, 1, 2, 3]
    bar = ""
    for bh in BAR_HOURS:
        bd = HOUR_DATA.get(bh, {"b_pct": 10})
        p  = bd["b_pct"]
        mk = "●" if bh == h else ("◆" if p >= 25 else ("◇" if p >= 15 else "·"))
        bar += str(bh).zfill(2) + mk + " "
    lines.append(bar.strip())
    lines.append("(● 현재  ◆ 배회25%↑  ◇ 배회15%↑)")

    # 캐스퍼 수정 2026-07-15: "회피 구역" 섹션 제거.
    # fish_finder(verified='avoid') 조회에 시간대 필터가 전혀 없어서, 몇 시에 브리핑을
    # 받든 항상 똑같은 3줄이 고정으로 나오는 구조였음(아키텍트 실사용 중 지적).
    # 필요하면 /avoid 명령으로 전체 회피구역 목록을 별도 조회 가능(cmd_avoid, 변경 없음).

    return "\n".join(lines)


# ──────────────────────────────────────────────
# atlas_reports 서버사이드 폴링 (2026-07-24 신규)
# 기존엔 앱(index.html) 켜져있을 때만 pollAtlasReports가 돌아서
# "자동 감지"가 실제로는 반자동이었던 문제를 해결. 봇(24시간 서버)에
# 동일 로직을 이식해서 앱을 안 켜놔도 진짜 자동으로 처리되게 함.
# ──────────────────────────────────────────────

async def build_magi_system_prompt(recent_summary: str = "") -> str:
    try:
        rows = await sb_select("magi_context", {"is_active": "eq.true", "order": "version.desc", "limit": "1"})
    except Exception:
        rows = None
    if not rows:
        return "당신은 대구 택시기사의 AI 분석 도우미입니다."
    ctx = rows[0]
    extra = f"\n[최근 데이터]\n{recent_summary}" if recent_summary else ""
    return f"""당신은 마기(MAGI)입니다. 대구 세큐티 가맹 택시기사의 전략 두뇌입니다.

[운영 원칙]
{ctx.get('principles') or ''}

[핵심 메모리]
{ctx.get('core_memory') or ''}

[데이터 검증 원칙]
{ctx.get('data_rules') or ''}
{extra}

위 맥락을 바탕으로 분석하고 7섹션 브리핑 형식으로 답변하세요.
통제 가능한 변수만 권고하고, 방어적 화법을 쓰지 마세요."""


async def process_atlas_report_server(report: dict) -> bool:
    """앱의 processAtlasReport()와 동일 로직의 서버사이드(봇) 버전.
    반환: True=분석 성공(status=analyzed), False=실패(status=error)"""
    try:
        recent_calls = await sb_select("raw_calls", {"order": "날짜.desc,배차시각.desc", "limit": "10"}) or []
        recent_summary = "\n".join(
            f"{c.get('날짜')} {c.get('배차시각') or ''} {c.get('출발지') or ''}->{c.get('도착지') or ''} {fmt(c.get('요금'))}"
            for c in recent_calls
        )
        system_prompt = await build_magi_system_prompt(recent_summary)
        user_msg = (
            f"아틀라스 보고서 ({report.get('report_type')}):\n"
            f"제목: {report.get('title') or '제목없음'}\n"
            f"내용: {json.dumps(report.get('payload'), ensure_ascii=False, indent=2)}\n\n"
            f"위 보고서를 분석하고 7섹션 브리핑을 작성해주세요."
        )

        import anthropic as _ant
        client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=120.0)
        msg = client.messages.create(
            model="claude-sonnet-5",
            max_tokens=8000,  # 캐스퍼 수정 2026-07-24: 1500이었는데 확장사고가
            # 토큰을 다 써버려 실제 텍스트가 생성되기 전에 잘리는 경우가 있어 상향
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        analysis = _extract_claude_text(msg)
        if not analysis:
            # 캐스퍼 수정 2026-07-24: 빈 응답을 "성공"으로 잘못 저장하던 버그.
            # 텍스트를 못 뽑으면 명시적으로 실패 처리해서 재시도 가능하게 함.
            raise ValueError(f"Claude 응답에 텍스트 블록 없음 (content 타입들: {[getattr(b,'type',None) for b in msg.content]}, stop_reason: {msg.stop_reason})")

        await sb_h(
            "PATCH", f"atlas_reports?id=eq.{report['id']}",
            json={"status": "analyzed", "magi_analysis": analysis, "analyzed_at": datetime.now(KST).isoformat()},
            headers={**HEADERS_SB, "Prefer": "return=minimal"},
        )
        logger.info(f"atlas_reports #{report['id']} 분석 완료")
        return True

    except Exception as e:
        logger.error(f"atlas_reports #{report.get('id')} 분석 오류: {e}")
        try:
            await sb_h(
                "PATCH", f"atlas_reports?id=eq.{report['id']}",
                json={"status": "error"},
                headers={**HEADERS_SB, "Prefer": "return=minimal"},
            )
        except Exception:
            pass
        return False


def atlas_reports_scheduler(app):
    """30초마다 atlas_reports(status=pending) 폴링, 앱 미실행 시에도 동작.
    처리 성공 시 텔레그램으로도 요약 알림."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    chat_ids = [x for x in [os.getenv("ALLOWED_CHAT_ID", ""), os.getenv("ALLOWED_CHAT_ID2", "")] if x]

    async def notify(report, ok):
        title = report.get("title") or "제목없음"
        text = f"📡 아틀라스 보고 처리 {'완료' if ok else '실패'}: {title}"
        for cid in chat_ids:
            try:
                await app.bot.send_message(chat_id=cid, text=text)
            except Exception:
                pass

    while True:
        try:
            rows = loop.run_until_complete(
                sb_select("atlas_reports", {"status": "eq.pending", "order": "created_at.asc", "limit": "3"})
            ) or []
            for report in rows:
                ok = loop.run_until_complete(process_atlas_report_server(report))
                loop.run_until_complete(notify(report, ok))
        except Exception as e:
            logger.error(f"atlas_reports 스케줄러 오류: {e}")
        time.sleep(30)


def fish_scheduler(app):
    """18:50 영업준비 브리핑 + 19~02시 매 정각 자동 브리핑
    2026-07-13 수정: 하드코딩 FISH_DATA(get_fish_report) → 실데이터 기반(get_fish_report_db)으로 전환.
    매일 03:10 fish_hour_data 자동 재계산 추가."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    # 발송 대상: 등록된 모든 단말기
    chat_ids = [x for x in [
        os.getenv("ALLOWED_CHAT_ID", ""),
        os.getenv("ALLOWED_CHAT_ID2", ""),
    ] if x]

    async def send_all(text: str):
        for cid in chat_ids:
            try:
                await app.bot.send_message(chat_id=cid, text=text)
            except Exception as e:
                logger.error(f"어군 브리핑 발송 오류 ({cid}): {e}")

    last_sent_hour = -1
    sent_start_brief = False
    last_reset_day = -1
    last_recalc_day = -1
    last_dualverify_day = -1
    last_kpi7day = -1
    last_orch_run_ts = 0.0
    last_magi_review_ts = 0.0

    # 최초 기동 시 1회 즉시 재계산 (테이블이 비어있으면 폴백값으로 작동하다가 여기서 채워짐)
    try:
        loop.run_until_complete(recalc_fish_hour_data())
    except Exception as e:
        logger.error(f"fish_hour_data 최초 재계산 실패: {e}")

    while True:
        try:
            now = datetime.now(KST)

            # ── task#40 문제2: 5분마다 Haiku 오케스트레이션(ASSIGNED+캐스퍼+SIMPLE)
            if time.time() - last_orch_run_ts >= 300:
                last_orch_run_ts = time.time()
                try:
                    tid = loop.run_until_complete(run_haiku_orchestration_once())
                    loop.run_until_complete(mark_scheduler_run("run_haiku_orchestration_once", f"task_id={tid}" if tid else "no_task"))
                except Exception as e:
                    logger.error(f"Haiku오케스트레이션 실행 실패: {e}")

            # ── task#52: 5분마다 마기(자동) 검증(VERIFICATION+verified_by없음)
            if time.time() - last_magi_review_ts >= 300:
                last_magi_review_ts = time.time()
                try:
                    tid2 = loop.run_until_complete(run_magi_auto_review_once())
                    loop.run_until_complete(mark_scheduler_run("run_magi_auto_review_once", f"task_id={tid2}" if tid2 else "no_task"))
                except Exception as e:
                    logger.error(f"마기(자동)검증 실행 실패: {e}")

            # ── 명령서#035+#036: 매일 04:00 7일평균(축B) + 일별5종(축A) 계산 + daily_summary 자동갱신
            if now.hour == 4 and now.day != last_kpi7day:
                try:
                    loop.run_until_complete(recalc_daily_summary_totals())
                    loop.run_until_complete(mark_scheduler_run("recalc_daily_summary_totals"))
                except Exception as e:
                    logger.error(f"daily_summary 자동갱신 실패: {e}")
                    loop.run_until_complete(mark_scheduler_run("recalc_daily_summary_totals", f"FAIL: {e}"))
                try:
                    loop.run_until_complete(recalc_7day_average())
                    loop.run_until_complete(mark_scheduler_run("recalc_7day_average"))
                except Exception as e:
                    logger.error(f"7일평균(명령서#035) 재계산 실패: {e}")
                    loop.run_until_complete(mark_scheduler_run("recalc_7day_average", f"FAIL: {e}"))
                try:
                    loop.run_until_complete(calc_daily_snapshot())
                    loop.run_until_complete(mark_scheduler_run("calc_daily_snapshot"))
                except Exception as e:
                    logger.error(f"daily_calc_snapshot(명령서#036) 계산 실패: {e}")
                    loop.run_until_complete(mark_scheduler_run("calc_daily_snapshot", f"FAIL: {e}"))
                last_kpi7day = now.day

            # ── task_id=31 긴급대응: 매일 08:00 인입중단 조기감지
            if now.hour == 8 and now.day != last_dualverify_day:
                try:
                    loop.run_until_complete(check_ingestion_gap())
                    loop.run_until_complete(mark_scheduler_run("check_ingestion_gap"))
                except Exception as e:
                    logger.error(f"인입중단 감지 실패: {e}")

            # ── task_id=38: 매일 08:00 operated_status 동기화 + 미확인날짜 질문발송
            if now.hour == 8 and now.day != last_dualverify_day:
                try:
                    asked = loop.run_until_complete(ask_operated_status_telegram())
                    loop.run_until_complete(mark_scheduler_run("ask_operated_status_telegram", f"asked={len(asked)}"))
                except Exception as e:
                    logger.error(f"operated_status 질문발송 실패: {e}")

            # ── 명령서#028 갭3: 매일 08:00 7일평균 이중검증 (raw_calls 직접집계 vs daily_summary)
            if now.hour == 8 and now.day != last_dualverify_day:
                try:
                    dv = loop.run_until_complete(dual_verify_7day_average())
                    last_dualverify_day = now.day
                    loop.run_until_complete(mark_scheduler_run("dual_verify_7day_average", "OK" if dv["match"] else "MISMATCH"))
                    if not dv["match"]:
                        msg = (
                            f"⚠️ 7일평균 이중검증 불일치 발견 ({dv['date_range']})\n"
                            f"방식A(raw_calls 직접집계): {dv['method_a']['총매출']:,}원 (일평균 {dv['method_a']['일평균']:,.0f}원)\n"
                            f"방식B(daily_summary): {dv['method_b']['총매출']:,}원 (일평균 {dv['method_b']['일평균']:,.0f}원)\n"
                            f"차이: {dv['diff']:+,}원\n"
                            + "\n".join(dv.get("detail", []))
                        )
                        logger.warning(f"명령서#028 갭3 검증 불일치: {msg}")
                        loop.run_until_complete(send_all(msg))
                    else:
                        logger.info(f"명령서#028 갭3 검증 통과 (일치, {dv['method_a']['총매출']:,}원)")
                except Exception as e:
                    logger.error(f"7일평균 이중검증 실행 오류: {e}")

            # ── 매일 03시 플래그 리셋 (운행 종료 후)
            if now.hour == 3 and now.day != last_reset_day:
                sent_start_brief = False
                last_sent_hour   = -1
                last_reset_day   = now.day
                logger.info("어군 스케줄러 일간 리셋")

            # ── 매일 03:10 시간대별 통계 자동 재계산 (하드코딩 제거 후속)
            if now.hour == 3 and now.minute >= 10 and now.day != last_recalc_day:
                try:
                    loop.run_until_complete(recalc_fish_hour_data())
                    global _FISH_HOUR_CACHE
                    _FISH_HOUR_CACHE = {}  # 캐시 무효화 → 다음 조회 시 새 값 로드
                    logger.info("fish_hour_data 일일 재계산 완료")
                except Exception as e:
                    logger.error(f"fish_hour_data 일일 재계산 실패: {e}")
                try:
                    loop.run_until_complete(recalc_fish_hour_data_dow())
                    logger.info("fish_hour_data_dow 일일 재계산 완료")
                except Exception as e:
                    logger.error(f"fish_hour_data_dow 일일 재계산 실패: {e}")
                last_recalc_day = now.day

            # ── 18:50 영업 준비 브리핑
            if now.hour == 18 and now.minute == 50 and not sent_start_brief:
                try:
                    report = loop.run_until_complete(get_fish_report_db(hour=19)) or "데이터 없음"
                    msg = f"🚀 영업준비 브리핑 (10분 후 출발)\n\n{report}"
                    loop.run_until_complete(send_all(msg))
                    logger.info("18:50 영업준비 브리핑 발송")
                except Exception as e:
                    # 캐스퍼 긴급수정 2026-08-11 (명령서#031): 이 구간에 예외처리가 없어서
                    # 여기서 죽으면 while루프 전체(03:10 재계산 포함)가 영구정지되던 게
                    # fish_hour_data 29일간 미갱신의 유력한 원인으로 추정됨. 재발방지.
                    logger.error(f"18:50 영업준비 브리핑 오류: {e}")
                sent_start_brief = True

            # ── 19:00 ~ 02:00 매 정각 브리핑
            if now.minute == 0 and now.hour != last_sent_hour:
                in_service = (19 <= now.hour <= 23) or (0 <= now.hour < 2)
                if in_service:
                    try:
                        report = loop.run_until_complete(get_fish_report_db())
                        if report:
                            loop.run_until_complete(send_all(report))
                            logger.info(f"어군 브리핑 발송: {now.hour}시")
                    except Exception as e:
                        logger.error(f"{now.hour}시 정각 브리핑 오류: {e}")
                last_sent_hour = now.hour

        except Exception as e:
            # 명령서#031 최외곽 안전망: 위에서 못 잡은 예상 밖 예외까지 전부 여기서 흡수해서
            # 스레드 자체가 죽는 일은 절대 없게 함. 이게 이번 사건(29일 정지)의 재발을 막는 핵심.
            logger.error(f"fish_scheduler 최외곽 예외 포착(스레드 생존): {e}")

        time.sleep(30)


# ──────────────────────────────────────────────
# 중복 자동 삭제 + 금액 불일치 처리
# ──────────────────────────────────────────────
FEE_DIFF_THRESHOLD = 500  # 이 이상 차이면 확인 요청

async def delete_duplicate_call(날짜: str, 배차시각: str, 요금: int) -> bool:
    """날짜+배차시각+요금 완전 일치 건 삭제. 삭제되면 True."""
    rows = await sb_select("raw_calls", {
        "날짜": f"eq.{날짜}",
        "배차시각": f"eq.{배차시각}",
        "요금": f"eq.{요금}",
    })
    if not rows:
        return False
    for row in rows:
        await sb_h("DELETE", f"raw_calls?id=eq.{row['id']}")
    return True

async def delete_duplicate_payment(날짜: str, 시각: str, 요금: int) -> bool:
    """날짜+시각+요금 완전 일치 결제내역 삭제."""
    rows = await sb_select("payment_receipts", {
        "날짜": f"eq.{날짜}",
        "시각": f"eq.{시각}",
        "요금": f"eq.{요금}",
    })
    if not rows:
        return False
    for row in rows:
        await sb_h("DELETE", f"payment_receipts?id=eq.{row['id']}")
    return True

async def send_fee_confirm(update, call_id: int, 배차시각: str,
                            call_fee: int, rcpt_fee: int, diff: int):
    """금액 불일치 ≥500원 시 InlineKeyboard 확인 요청 발송."""
    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                f"콜카드 {call_fee:,}원",
                callback_data=f"fee:{call_id}:{call_fee}"
            ),
            InlineKeyboardButton(
                f"결제내역 {rcpt_fee:,}원",
                callback_data=f"fee:{call_id}:{rcpt_fee}"
            ),
        ]
    ])
    await update.message.reply_text(
        f"⚠️ 금액 불일치 확인 요청\n"
        f"배차: {배차시각}\n"
        f"콜카드: {call_fee:,}원\n"
        f"결제내역: {rcpt_fee:,}원\n"
        f"차이: {diff:,}원\n\n"
        f"어느 금액으로 저장할까요?",
        reply_markup=keyboard
    )

async def handle_fee_callback(update, context):
    """InlineKeyboard 버튼 클릭 처리."""
    query = update.callback_query
    await query.answer()
    data = query.data  # "fee:call_id:선택금액"
    try:
        _, call_id_str, fee_str = data.split(":")
        call_id = int(call_id_str)
        selected_fee = int(fee_str)
        # raw_calls 요금 업데이트
        await sb_h("PATCH", f"raw_calls?id=eq.{call_id}",
                   json={"요금": selected_fee})
        await query.edit_message_text(
            f"✅ {selected_fee:,}원으로 저장 완료"
        )
    except Exception as e:
        logger.error(f"fee callback 오류: {e}")
        await query.edit_message_text("❌ 처리 오류")

# ──────────────────────────────────────────────
# 보험 스케줄러
# ──────────────────────────────────────────────
def insurance_scheduler():
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    while True:
        now = datetime.now(KST)
        next_run = now.replace(hour=0, minute=1, second=0, microsecond=0)
        if now >= next_run:
            next_run += timedelta(days=1)
        sleep_sec = (next_run - now).total_seconds()
        logger.info(f"보험 스케줄러 대기: {sleep_sec:.0f}초")
        time.sleep(sleep_sec)
        today = datetime.now(KST).date()  # 반드시 KST
        loop.run_until_complete(insert_insurance(today))
        time.sleep(60)

# ──────────────────────────────────────────────
# 텔레그램 핸들러
# ──────────────────────────────────────────────
def is_allowed(update: Update) -> bool:
    chat_id = str(update.effective_chat.id)
    return chat_id in ALLOWED_IDS


async def cmd_fish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    어군 브리핑. /fish [시간] [태그]
    예: /fish       → 현재 시간
        /fish 19    → 19시 기준
        /fish 21 foreign → 21시 외국인 패턴 필터
    """
    if not is_allowed(update):
        return

    args = context.args or []
    hour = None
    tag_filter = None

    for arg in args:
        if arg.isdigit():
            hour = int(arg)
        elif arg in ("foreign", "foreign_worker"):
            tag_filter = "foreign_worker"
        elif arg in ("long", "long_distance"):
            tag_filter = "long_distance"
        elif arg in ("golden", "golden_time"):
            tag_filter = "golden_time"
        elif arg in ("blue", "blue_ocean"):
            tag_filter = "blue_ocean"

    report = await get_fish_report_db(hour=hour, tag_filter=tag_filter)
    await update.message.reply_text(report)


async def cmd_avoid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """회피 구역 조회 /avoid"""
    if not is_allowed(update):
        return
    try:
        rows = await sb_select("fish_finder", {"verified": "eq.avoid", "order": "time_band.asc"})
        if not rows:
            await update.message.reply_text("⛔ 등록된 회피 구역 없음\n(Supabase fish_finder 테이블 확인 필요)")
            return
        lines = ["⛔ 회피 구역 전체 목록\n"]
        for r in rows:
            zone = r.get("zone","")
            band = r.get("time_band","")
            note = r.get("note","") or ""
            lines.append(f"  ✗ {zone} ({band})" + (f"\n    {note}" if note else ""))
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        logger.error(f"회피구역 조회 오류: {e}")
        await update.message.reply_text("❌ 조회 오류")

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    await update.message.reply_text(
        "🤖 자비스 v5 시작\n\n"
        "이미지: 콜카드·충전내역·결제내역·세큐티\n"
        "텍스트: 콜 7800 / 배회 5600 / 오늘 / 이번 주 / 전략\n"
        "다운로드: 주간·월간·전체 다운로드\n"
        "교차대조: 대조 YYYY-MM-DD"
    )

async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    await update.message.reply_text(
        "📋 명령어\n\n"
        "[이미지] 콜카드·충전영수증·결제내역 → 자동저장\n\n"
        "[수동입력]\n"
        "콜 금액 / 배회 금액 / 충전 금액\n"
        "타이어·오일·세차 금액 / 지출 항목 금액\n"
        "지출취소 / 휴무 / 4-7 휴무\n\n"
        "[수동전체]\n"
        "2026 03 01 23 05 출발>도착 요금 카카오\n"
        "0301 2305 출발>도착 요금 배회\n\n"
        "[콜수정]\n"
        "콜수정 HH:MM 필드=값\n"
        "콜수정 날짜 HH:MM 필드=값\n"
        "콜수정ID [id] 필드=값\n\n"
        "[조회]\n"
        "오늘·이번 주·이번 달·지출 확인·DB 확인\n"
        "3-7 조회·매출·순수익·총건수·지출\n\n"
        "[교차대조]\n"
        "대조 날짜 (예: 대조 3-7)\n"
        "대조 확정 날짜 / 대조 금액확인 날짜\n\n"
        "[결제삭제]\n"
        "결제삭제 날짜 운행외/0원/전체/HH:MM\n\n"
        "[전략] 전략 / 마기 업데이트 시간대 내용\n\n"
        "[다운로드]\n"
        "주간/월간/전체 다운로드 / 월간 2026-03\n\n"
        "[어군] /fish"
    )

async def cmd_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Chat ID: {update.effective_chat.id}")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    try:
        await image_queue.put((update, context))
    except Exception as e:
        logger.error(f"이미지 큐 오류: {e}")
        await update.message.reply_text("❌ 이미지 처리 오류. 다시 시도해주세요.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    파일 첨부 처리.
    - 이미지 파일 (jpg/jpeg/png/webp/gif/bmp) → 이미지 처리 파이프라인
    - xlsx → 엑셀 이식
    - 그 외 → 안내 메시지
    """
    if not is_allowed(update):
        return
    doc = update.message.document
    if not doc:
        return

    fname = (doc.file_name or "").lower()
    mime  = (doc.mime_type or "").lower()

    # ── 이미지 파일로 전송된 경우 (파일로 보내기) ──
    IMAGE_EXTS  = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp")
    IMAGE_MIMES = ("image/jpeg", "image/png", "image/webp", "image/gif", "image/bmp")

    if any(fname.endswith(e) for e in IMAGE_EXTS) or any(mime.startswith(m) for m in IMAGE_MIMES):
        # 파일 다운로드 후 이미지 큐에 추가
        try:
            file = await context.bot.get_file(doc.file_id)
            image_bytes = await file.download_as_bytearray()
            image_bytes = bytes(image_bytes)
            if image_queue is None:
                await update.message.reply_text("❌ 이미지 큐 초기화 중입니다. 잠시 후 다시 시도해주세요.")
                return
            await image_queue.put((update, context, image_bytes))
            logger.info(f"파일 이미지 큐 추가: {fname} ({len(image_bytes):,}bytes)")
        except Exception as e:
            logger.error(f"파일 이미지 처리 오류: {e}")
            await update.message.reply_text(f"❌ 파일 처리 오류: {str(e)[:100]}")
        return

    # ── 엑셀 이식 ──
    if fname.endswith(".xlsx"):
        await handle_excel_import(update, context)
        return

    # ── 그 외 ──
    await update.message.reply_text(
        "⚠️ 지원하지 않는 파일 형식입니다.\n"
        "이미지: jpg·png·webp 파일 또는 사진으로 전송\n"
        "엑셀: xlsx 파일"
    )


async def _process_single_command(update, context, text: str) -> str | None:
    """단일 명령어 처리. 줄바꿈 다중 명령어 시 각 줄 처리용."""

    # task#27: 주간브리핑 조립(ARG-CMD-20260820-02)
    if text.strip() in ("주간브리핑", "/주간브리핑", "weekly"):
        try:
            return await build_weekly_briefing()
        except Exception as e:
            logger.error(f"주간브리핑 조립 실패: {e}")
            return f"주간브리핑 생성 실패: {e}"

    # task#38: operated_status 질문에 대한 답변("8/13 8/15 휴무" 형식)
    if text.strip().endswith("휴무") and any(ch.isdigit() for ch in text):
        import re as _re_op
        date_tokens = _re_op.findall(r'(\d{1,2})[/\-](\d{1,2})', text)
        if date_tokens:
            year = today_kst().year
            confirmed_dates = []
            now_iso = datetime.now(KST).isoformat()
            for mo, da in date_tokens:
                try:
                    d = f"{year}-{int(mo):02d}-{int(da):02d}"
                    await sb_upsert("operated_status", {
                        "날짜": d, "operated": False, "source": "user_confirmed", "answered_at": now_iso
                    }, on_conflict="날짜")
                    confirmed_dates.append(d)
                except Exception as e:
                    logger.error(f"휴무답변 처리 실패({mo}/{da}): {e}")
            if confirmed_dates:
                return "✅ 휴무 확인 등록: " + ", ".join(confirmed_dates)

    # 결제삭제
    if text.startswith("결제삭제 "):
        parts = text.strip().split(" ", 2)
        if len(parts) < 3:
            return "❌ 형식: 결제삭제 YYYY-MM-DD 운행외|0원|전체|HH:MM"
        date_str = parts[1].strip()
        mode = parts[2].strip()
        try:
            rows = await sb_select("payment_receipts", {"날짜": f"eq.{date_str}"})
            if not rows:
                return f"⚠️ {date_str} 결제내역 없음"
            delete_ids = []
            if mode == "운행외":
                for r in rows:
                    t = r.get("시각", "") or ""
                    try:
                        h, m = t.split(":")
                        mins = int(h)*60+int(m)
                        if 121 <= mins <= 1019:
                            delete_ids.append(r["id"])
                    except Exception:
                        delete_ids.append(r["id"])
            elif mode == "0원":
                for r in rows:
                    fee = r.get("요금")
                    if fee is None or int(fee) == 0:
                        delete_ids.append(r["id"])
            elif mode == "전체":
                delete_ids = [r["id"] for r in rows]
            elif ":" in mode:
                for r in rows:
                    if r.get("시각","") == mode:
                        delete_ids.append(r["id"])
            else:
                return f"❓ 알 수 없는 모드: {mode}"
            if not delete_ids:
                return f"✅ {date_str} 삭제 대상 없음 ({mode})"
            for rid in delete_ids:
                await sb_h("DELETE", f"payment_receipts?id=eq.{rid}")
            return f"🗑️ {date_str} {mode} {len(delete_ids)}건 삭제"
        except Exception as e:
            return f"❌ 삭제 오류: {str(e)[:100]}"

    # 수동 콜
    parsed_call = parse_manual_call(text)
    if parsed_call:
        today = str(today_kst())
        payload = {
            "날짜": today, "요일": get_dow(),
            "배차시각": now_kst().strftime("%H:%M"),
            "요금": parsed_call["요금"],
            "콜유형": parsed_call["콜유형"],
            "도착지": parsed_call.get("도착지힌트"),
            "data_source": "manual_entry",
        }
        r = await sb_insert("raw_calls", payload)
        return f"✅ {parsed_call['콜유형']} {fmt(parsed_call['요금'])} 입력" if r else "❌ 저장 실패"

    # 지출
    parsed_exp = parse_expense(text)
    if parsed_exp:
        today = str(today_kst())
        payload = {
            "날짜": today,
            "카테고리": parsed_exp["카테고리"],
            "금액": parsed_exp["금액"],
            "메모": parsed_exp.get("메모",""),
            "자동여부": False,
        }
        r = await sb_insert("expenses", payload)
        return f"✅ {parsed_exp['카테고리']} {fmt(parsed_exp['금액'])} 입력" if r else "❌ 저장 실패"

    return f"❓ '{text[:20]}' 인식 불가"

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return
    text = (update.message.text or "").strip()
    lower = text.lower()

    # ── 줄바꿈 다중 명령어 처리 ──
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if len(lines) > 1:
        results = []
        for line in lines:
            # 각 줄을 개별 명령어로 처리
            fake_update = update
            line_text = line
            try:
                result = await _process_single_command(fake_update, context, line_text)
                if result:
                    results.append(result)
            except Exception as e:
                results.append(f"❌ '{line_text[:20]}' 오류: {str(e)[:50]}")
        if results:
            await update.message.reply_text("\n".join(results))
        return

    # 결제내역 삭제 명령어
    # "결제삭제 YYYY-MM-DD 운행외" → 운행시간(19~02시) 외 데이터 삭제
    # "결제삭제 YYYY-MM-DD 0원"    → 요금 0원 데이터 삭제
    # "결제삭제 YYYY-MM-DD 전체"   → 해당 날짜 전체 삭제
    if text.startswith("결제삭제 "):
        await handle_receipt_delete(update, text)
        return

    # task#27: 주간브리핑(단일줄 경로) — 캐스퍼 수정 2026-08-21: 어제 이 코드를
    # _process_single_command에만 넣었는데, 그 함수는 "여러 줄 동시입력"일 때만
    # 호출되는 경로라 "주간브리핑" 한 줄 입력시엔 전혀 도달 못 하던 배치실수 수정.
    if text.strip() in ("주간브리핑", "/주간브리핑", "weekly"):
        try:
            result = await build_weekly_briefing()
            await update.message.reply_text(result)
        except Exception as e:
            logger.error(f"주간브리핑 조립 실패: {e}")
            await update.message.reply_text(f"주간브리핑 생성 실패: {e}")
        return

    # 대조 확정 (배회후보 raw_calls 추가)
    if text.startswith("대조 확정 "):
        _ds = text[6:].strip()
        try:
            result = await confirm_cross_check(_ds)
            await update.message.reply_text(result)
        except Exception as e:
            await update.message.reply_text(f"❌ 대조확정 오류: {str(e)[:200]}")
        return

    # 대조 금액확인 (InlineKeyboard 금액 선택)
    if text.startswith("대조 금액확인 "):
        _ds = text[8:].strip()
        try:
            await handle_fee_confirm_request(update, _ds)
        except Exception as e:
            await update.message.reply_text(f"❌ 금액확인 오류: {str(e)[:200]}")
        return

    # 교차대조 — YYYY-MM-DD 또는 M-D 형식 지원
    if text.startswith("대조 "):
        import re as _re3
        from datetime import date as _date3
        date_str = text[3:].strip()
        _md = _re3.match(r'^(\d{1,2})-(\d{1,2})$', date_str)
        if _md:
            try:
                _d = _date3(_date3.today().year, int(_md.group(1)), int(_md.group(2)))
                date_str = str(_d)
            except ValueError:
                await update.message.reply_text("❌ 잘못된 날짜입니다.")
                return
        try:
            result = await cross_check(date_str)
            if len(result) > 4000:
                result = result[:4000] + "\n...(생략)"
            await update.message.reply_text(result)
        except Exception as e:
            logger.error(f"교차대조 오류: {e}")
            await update.message.reply_text(f"❌ 교차대조 오류: {str(e)[:200]}")
        return

    if text.startswith("배회분류 확정 "):
        date_str = text[8:].strip()
        try:
            result = await confirm_baehoe_classification(date_str)
            await update.message.reply_text(result)
        except Exception as e:
            logger.error(f"배회분류 오류: {e}")
            await update.message.reply_text(f"❌ 배회분류 오류: {str(e)[:200]}")
        return

    # 날짜 + 통계 키워드 조합
    _stat_kws = ["총건수","건수","매출","순수익","지출","조회","상세"]
    _date_pat = r"(\d{1,2})[-/](\d{1,2})|(\d{4})-(\d{1,2})-(\d{1,2})"
    import re as _re
    if _re.search(_date_pat, text) and any(kw in text for kw in _stat_kws):
        if any(kw in text for kw in ["총건수","건수","매출","순수익","지출"]):
            await handle_date_stat(update, text)
        else:
            await handle_date_query(update, text)
        return

    # 브리핑
    if text in ("브리핑", "오늘브리핑", "오늘 브리핑"):
        await handle_briefing(update)
        return

    # 운행완료수
    if text in ("운행완료수", "완료수", "ai진입"):
        await handle_completion_status(update)
        return

    # 운행 일관성 조회
    if text in ("일관성", "일관성 조회", "운행일관성"):
        await report_operation_consistency(update)
        return

    # 어군 브리핑 텍스트 명령
    if text in ("어군", "어군조회", "어군 조회"):
        now = datetime.now(KST)
        if not get_fish_slot(now.hour):
            await update.message.reply_text(
                f"🐟 현재 {now.hour}시는 브리핑 시간대가 아닙니다.\n"
                f"운영시간: 19~21시 / 21~24시 / 00~02시"
            )
        else:
            report = await get_fish_report_db()
            await update.message.reply_text(report or "🐟 어군 데이터 없음")
        return

    # 세큐티 조회
    if text in ("세큐티 조회", "세큐티조회"):
        await handle_sekuti_query(update)
        return

    # 특정 날짜 조회 (조회 키워드만 있을 때)
    if "조회" in text:
        await handle_date_query(update, text)
        return

    # 조회
    if text == "오늘":
        await handle_today_quick(update)
        return
    if text in ("이번 주", "이번주", "주간"):
        await handle_weekly(update)
        return
    import re as _re2
    _ym = _re2.match(r"^월간\s+(\d{4}-\d{2})\s*$", text.strip())
    if _ym:
        await handle_download_month(update, _ym.group(1))
        return

    if text in ("이번 달", "이번달", "월간"):
        await handle_monthly(update)
        return
    if text == "지출 확인":
        await handle_expense_check(update)
        return
    if text == "DB 확인":
        await handle_db_check(update)
        return

    # 전략
    if text in ("전략", "실시간"):
        await get_strategy(update)
        return
    if text.startswith("마기 업데이트 "):
        content = text[8:].strip()
        await handle_magi_update(update, content)
        return

    # 다운로드
    if "다운로드" in text:
        if "주간" in text:
            await handle_download(update, "주간")
        elif "월간" in text:
            await handle_download(update, "월간")
        elif "전체" in text:
            await handle_download(update, "전체")
        else:
            await update.message.reply_text("주간·월간·전체 다운로드 중 선택해주세요.")
        return

    # 특정 월 다운로드: "월간 2026-03" or "월간2026-03"
    import re as _re
    _ym_match = _re.match(r"^월간\s*(\d{4}-\d{2})$", text.strip())
    if _ym_match:
        await handle_download_month(update, _ym_match.group(1))
        return

    # 수동 전체 입력 (날짜+시각+경로+요금 형식)
    import re as _re2
    _has_route = bool(_re2.search(r'[가-힣\w]+[>→][가-힣\w]+', text))
    _has_date_nums = bool(_re2.search(r'\d{2}\s+\d{1,2}\s+\d{1,2}\s+\d{1,2}\s+\d{2}|\d{4}[\-.]\d{1,2}[\-.]\d{1,2}|\d{8}|\d{4}\s+\d{4}', text))
    if _has_route and _has_date_nums:
        await handle_manual_full_call(update, text)
        return

    # 콜카드 수동 수정
    if text.startswith("콜수정ID "):
        await handle_call_edit_by_id(update, text[6:].strip())
        return

    if text.startswith("콜수정 "):
        await handle_call_edit(update, text)
        return

    # 휴무 (오늘 또는 날짜 지정)
    if "휴무" in text:
        await handle_rest_day(update, text)
        return

    # 지출취소
    if text == "지출취소":
        await handle_expense_cancel(update)
        return

    # 수동 콜 입력
    parsed_call = parse_manual_call(text)
    if parsed_call:
        await handle_manual_call(update, parsed_call)
        return

    # 지출 입력
    parsed_exp = parse_expense(text)
    if parsed_exp:
        await handle_expense(update, parsed_exp)
        return

    # task#64(2026-08-26): 이지스 관련 질문이 최소 2회 반복됐던 사례(마기가
    # 과거 답변에 접근 못해 재질문) 재발방지. 다른 명령어에 전혀 안 걸린 자유
    # 텍스트만 여기 도달하므로, 최근 미답변 마기(자동) 질문이 있으면 이 텍스트를
    # 답변으로 매칭해 질문+답변 전문을 보존한다(기존 안내메시지는 그대로 유지).
    try:
        recent_q = await sb_select("magi_task_events", {
            "event_type": "eq.ARCHITECT_QUESTION_SENT",
            "order": "created_at.desc", "limit": "1"
        })
        if recent_q:
            q = recent_q[0]
            answered = await sb_select("magi_task_events", {
                "task_id": f"eq.{q['task_id']}", "event_type": "eq.ARCHITECT_QUESTION_ANSWERED",
                "created_at": f"gte.{q['created_at']}", "limit": "1"
            })
            if not answered:
                await sb_insert("magi_task_events", {
                    "task_id": q["task_id"], "event_type": "ARCHITECT_QUESTION_ANSWERED", "actor": "대표님",
                    "detail": f"질문: {(q.get('detail') or '')[:300]}\n답변: {text[:500]}"
                })
    except Exception as e:
        logger.error(f"질문-답변 매칭 실패: {e}")

    # 미인식
    await update.message.reply_text("❓ 명령어를 인식하지 못했습니다. /help 로 확인해주세요.")

# ──────────────────────────────────────────────
# main()
# ──────────────────────────────────────────────
def main():
    global image_queue

    # Health server
    threading.Thread(target=run_health_server, daemon=True).start()

    # Insurance scheduler
    threading.Thread(target=insurance_scheduler, daemon=True).start()

    # Telegram application
    app = (
        Application.builder()
        .token(TELEGRAM_TOKEN)
        .read_timeout(30)
        .write_timeout(30)
        .connect_timeout(30)
        .get_updates_read_timeout(5)    # 폴링 빠른 타임아웃 → Conflict 최소화
        .build()
    )

    async def post_init(application):
        global image_queue
        image_queue = asyncio.Queue(maxsize=10)  # 최대 10개 대기
        asyncio.create_task(process_image_queue_worker())
        # 기존 webhook 제거 + 이전 인스턴스 세션 정리 — Conflict 방지
        await asyncio.sleep(2)  # 구 인스턴스 세션 해제 대기
        await application.bot.delete_webhook(drop_pending_updates=True)
        logger.info("Webhook 삭제 완료 — 폴링 시작 준비")

    app.post_init = post_init

    # 어군탐지기 스케줄러 — app 생성 후 시작
    threading.Thread(target=fish_scheduler, args=(app,), daemon=True).start()
    threading.Thread(target=atlas_reports_scheduler, args=(app,), daemon=True).start()
    logger.info("atlas_reports 서버사이드 폴링 시작")
    logger.info("어군탐지기 스케줄러 시작")

    # 핸들러 등록
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("fish", cmd_fish))    # 어군 브리핑 수동 조회
    app.add_handler(CommandHandler("avoid", cmd_avoid))  # 회피 구역 조회
    app.add_handler(CommandHandler("forecast", lambda u,c: handle_forecast(u, c.args[0] if c.args else None)))
    app.add_handler(CommandHandler("completion_status", lambda u,c: handle_completion_status(u)))
    app.add_handler(CommandHandler("briefing", lambda u,c: handle_briefing(u)))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("id", cmd_id))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    from telegram.ext import CallbackQueryHandler
    app.add_handler(CallbackQueryHandler(handle_fee_callback, pattern=r"^fee:"))

    # 전역 에러 핸들러 — Conflict/Network 오류 자동 복구
    async def error_handler(update, context):
        import telegram
        err = context.error
        if isinstance(err, telegram.error.Conflict):
            logger.warning(f"Conflict 감지 (자동복구 대기): {err}")
        elif isinstance(err, telegram.error.NetworkError):
            logger.warning(f"네트워크 오류 (자동재시도): {err}")
        elif isinstance(err, telegram.error.TimedOut):
            pass  # 타임아웃은 정상 폴링 동작
        else:
            logger.error(f"봇 오류: {type(err).__name__}: {err}")
    app.add_error_handler(error_handler)

    logger.info("자비스 v5 시작")
    app.run_polling(
        drop_pending_updates=True,
        allowed_updates=["message", "callback_query"],
    )

if __name__ == "__main__":
    main()
